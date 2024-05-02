#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #
#   LOncoG : a software for Longitudinal OncoGenomics analysis
#   Authors: S. Besseau, G. Siekaniec, W. Gouraud
# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import json
import pandas as pd
import requests

def Panther_GOEA(module, dict_para, list_genes: list, panther_name: str, logger=None, pvalue=0.05, maxres=12000):
	###############
	# Pantherdb API
	try:
		if len(list_genes) == 0:
			if logger:
				logger.warning('No gene in list !')
			print('No gene in list !')
		else:
			if dict_para['verbose_prints'].upper() == 'TRUE':
				print(f'Request Pantherdb...\n')
			if logger:
				logger.info('Panthere enrichment from API')
			if dict_para['verbose_prints'].upper() == 'TRUE':
				print('Panthere enrichment from API')
			headers = {'accept': 'application/js'}
			genes = ','.join([str(gene) for gene in list_genes])
			genes = str(genes.split(',')).replace('[', '').replace(']', '').replace(' ', '').replace("'", '')
			if dict_para['verbose_prints'].upper() == 'TRUE':
				print(
					r""" curl -X POST -H 'accept: application/js' "http://pantherdb.org/services/oai/pantherdb/enrich/overrep?geneInputList=""" + genes + r"""&organism=9606&annotDataSet=GO%3A0008150&enrichmentTestType=FISHER&correction=FDR" \n""")

			response = requests.post(
				"http://pantherdb.org/services/oai/pantherdb/enrich/overrep?geneInputList=" + genes + "&organism=9606&annotDataSet=GO%3A0008150&enrichmentTestType=FISHER&correction=FDR",
				headers=headers, timeout=30)

			json_data = json.loads(response.text)  # Get json data
			if dict_para['verbose_prints'].upper() == 'TRUE':
				print('Done.\n')

			if 'search' in json_data.keys() and 'error' in json_data[
				'search'].keys():  # Search for errors in the answer
				raise ValueError('Something went wrong with the Panther request (POST) !')

			# Data processing

			d = {}
			compt = 0
			nb_unclassified_GO = 0
			col = ["Id", "Biological Process name", "P-value", "FDR", "Genes involved"]
			for l in json_data['results']['result']:
				if float(l['pValue']) < pvalue and int(l['number_in_list']) != 0:
					compt += 1
					if l["term"]["label"] != 'UNCLASSIFIED':
						new = [l["term"]["id"], l["term"]["label"], l["pValue"], l["fdr"], l["number_in_list"]]
						d[compt] = new
					elif logger:
						nb_unclassified_GO += 1

			if logger:
				if nb_unclassified_GO > 0:
					logger.warning(f'{nb_unclassified_GO} unclassified gene ontology !')
			df = pd.DataFrame.from_dict(d, orient='index', columns=col)
			df['Genes involved'] = df['Genes involved'].astype(int)
			df = df.sort_values(by='P-value', ascending=True)
			df.index = [i for i in range(len(df.index))]
			df.set_index('Id')

			if module == 'summarise':
				parameter = dict_para['S_Panther_format'].upper()
			elif module == 'compare':
				parameter = dict_para['C_Panther_table_format'].upper()
			elif module == 'merge':
				parameter = dict_para['M_Panther_table_format'].upper()

			if module == 'summarise':
				path = dict_para['output_path_sample']
			elif module == 'compare':
				path = dict_para['output_path_comparison']
			elif module == 'merge':
				path = dict_para['output_path'] + 'merge/'

			df = df.reset_index(drop=True)
			df = df.sort_values(by='Genes involved', ascending=False)
			df = df.drop(columns=['FDR'])

			if 'XLSX' in parameter:
				df.to_excel(path + panther_name + '.xlsx', index=False)
				panther_table = path + panther_name + '.xlsx'
				try:
					workbook = load_workbook(panther_table)
					worksheet = workbook.active

					for column in worksheet.columns:
						max_length = 0
						column = [cell for cell in column]
						for cell in column:
							try:
								if len(str(cell.value)) > max_length:
									max_length = len(cell.value)
							except:
								pass
						adjusted_width = (max_length + 2) * 1.2
						worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

					border_style = Border(left=Side(border_style='thin'),
										  right=Side(border_style='thin'),
										  top=Side(border_style='thin'),
										  bottom=Side(border_style='thin'))

					for row in worksheet.iter_rows():
						for cell in row:
							if cell.value:
								cell.border = border_style
								cell.alignment = Alignment(horizontal='center', vertical='center')

					first_row = worksheet[1]
					for cell in first_row:
						if cell.value:
							cell.font = Font(bold=True)
							cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

					workbook.save(panther_table)
				except:
					print('making Panther GOEA df more lisible failed')

			elif 'TSV' in parameter:
				tsv = path + panther_name + '.tsv'
				df.to_csv(tsv, sep='\t')
			elif 'CSV' in parameter:
				csv = path + panther_name + '.csv'
				df.to_csv(csv, sep=',')
	except:
		print("Panther couldn't find any related significant pathway.")
