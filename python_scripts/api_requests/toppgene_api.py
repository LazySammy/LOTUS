#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #
#   LOncoG : a software for Longitudinal OncoGenomics analysis
#   Authors: S. Besseau, G. Siekaniec, W. Gouraud
# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #

import json
import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

def ToppGene_GOEA(module, dict_parameters, list_genes: list, toppgene_name: str, logger=None, pvalue=0.05, maxres=100):
	if len(list_genes) == 0:
		if logger:
			logger.warning('No gene in list !')
	else:
		try:

			if logger:
				logger.info('Extraction of the biological process from the list of genes')
			if dict_parameters['verbose_prints'].upper() == 'TRUE':
				print('Extraction of the biological process from the list of genes...')

			lookup = "/tmp/lookup.txt"
			ToppGeneEnrichment = '/tmp/ToppGeneEnrichment.txt'

			# ToppGene API
			if logger:
				logger.info('Get gene id using ToppGene API')
			if dict_parameters['verbose_prints'].upper() == 'TRUE':
				print('Get gene id using ToppGene API')

			# Lookup for genes and get their Entrez id
			lookup_genes = "\"" + str("\",\"".join(list_genes)) + "\""
			lookup_genes = lookup_genes.split(',')
			lookup_genes = [symbol.strip('"') for symbol in lookup_genes]
			url = 'https://toppgene.cchmc.org/API/lookup'
			headers = {'Content-Type': 'application/json'}
			data = {'Symbols': lookup_genes}
			response = requests.post(url, headers=headers, data=json.dumps(data))
			json_data = response.json()

			genes = set()
			if json_data['Genes']:
				for l in json_data['Genes']:
					for k, v in l.items():
						if k == 'Entrez':
							genes.add(v)
				genes = list(genes)
				if dict_parameters['verbose_prints'].upper() == 'TRUE':
					print(f'Genes ids for enrichment {genes}')
				if logger:
					logger.info(f'ToppGene enrichment from API')

				# Get info from ToppGene API
				url = 'https://toppgene.cchmc.org/API/enrich'
				# genes = [93953]
				payload = {
					'Genes': genes,
					'Type': 'GeneOntologyBiologicalProcess',
					'PValue': pvalue,
					'MinGenes': 1,
					'MaxGenes': 2000,
					'MaxResults': maxres,
					'Correction': 'FDR'
				}
				data = json.dumps(payload)
				headers = {'Content-Type': 'application/json'}
				response = requests.post(url, data=data, headers=headers)
				json_data = response.json()

				d = {}
				compt = 0
				if json_data:
					for l in json_data['Annotations']:
						l['URL'] = 'https://www.ebi.ac.uk/QuickGO/term/' + l['ID']
						for k, v in l.items():
							if k == 'Category' and v == 'GeneOntologyBiologicalProcess':
								compt += 1
								d[compt] = l

					if d != {}:
						df = pd.DataFrame.from_dict(d).T
						df = df.rename(columns={"ID": "Id", "Name": "Biological Process name", "PValue": "P-value"})
						df = df.rename(columns={'GenesInTerm': 'n genes'})
						columns_to_keep = ['Id', 'Biological Process name', 'P-value', 'n genes', 'URL']
						df = df.drop(columns=[col for col in df.columns if col not in columns_to_keep])
						df = df.set_index('Id')
						df = df.sort_values('P-value', ascending=True)

						if module == 'summarise':
							parameter = dict_parameters['S_ToppGene_table_format'].upper()
						elif module == 'compare':
							parameter = dict_parameters['C_ToppGene_table_format'].upper()
						elif module == 'merge':
							parameter = dict_parameters['M_ToppGene_table_format'].upper()

						if module == 'summarise':
							path = dict_parameters['output_path_sample']
							path = path.split('.')[0] + '/'
						elif module == 'compare':
							path = dict_parameters['output_path_comparison']
						elif module == 'merge':
							path = dict_parameters['output_path'] + 'merge/'

						# Output files verification
						if 'XLSX' in parameter:
							toppgene_path = path + toppgene_name + '.xlsx'
							df.to_excel(toppgene_path)

							workbook = load_workbook(toppgene_path)
							worksheet = workbook.active

							for column in worksheet.columns:
								max_length = 0
								column = [cell for cell in column]
								for cell in column:
									try:
										if len(str(cell.value)) > max_length:
											max_length = len(cell.value)
									except:
										pass
								adjusted_width = (max_length + 2) * 1.2
								worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

							border_style = Border(left=Side(border_style='thin'),
												  right=Side(border_style='thin'),
												  top=Side(border_style='thin'),
												  bottom=Side(border_style='thin'))

							for row in worksheet.iter_rows():
								for cell in row:
									if cell.value:
										cell.border = border_style
										cell.alignment = Alignment(horizontal='center', vertical='center')

							first_row = worksheet[1]
							for cell in first_row:
								if cell.value:
									cell.font = Font(bold=True)
									cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

							workbook.save(toppgene_path)

						elif 'TSV' in parameter:
							tsv = path + toppgene_name + '.tsv'
							df.to_csv(tsv, sep='\t')
						elif 'CSV' in parameter:
							csv = path + toppgene_name + '.csv'
							df.to_csv(csv, sep=',')

					else:
						if logger:
							logger.info('ToppGene could not find any significant biological process corresponding to this gene list!')
						print('ToppGene could not find any significant biological process corresponding to this gene list!')


			else:
				print('No valid Gene !')
		except:
			print("Too many genes are concerned for ToppGene GO to be run.")
