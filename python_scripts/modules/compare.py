#!/usr/bin/env python 3
# -*- coding: utf-8 -*-
# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #
#   LOncoG : a software for Longitudinal OncoGenomics analysis
#   Authors: S. Besseau, G. Siekaniec, W. Gouraud
# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #

import datetime
import logging
import math
import sys
import os
import pandas as pd
import re
import numpy as np
import warnings
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side
from pathlib import Path
from python_scripts.reusable_functions.check_files import verif_input, verif_supplementary_information_file
from python_scripts.api_requests.toppgene_api import ToppGene_GOEA
from python_scripts.api_requests.panther_api import Panther_GOEA
from python_scripts.reusable_functions.read_vcf import read_vcf, get_vcf_header
from python_scripts.reusable_functions.read_gff3 import read_gff3

def get_informations_for_genes(dict_para, info_file, logger):
	df = pd.read_excel(info_file, index_col=1)
	df = df.drop(['Ordre'], axis=1)
	df.set_axis([source.split(' Info')[0] for source in df.columns], axis="columns")
	if dict_para['verbose_prints'].upper() == 'TRUE' or dict_para['verbose_prints'].upper() == 'YES':
		print('Extracting information from {} sources: {}'.format(len(list(df.columns)), ", ".join(list(df.columns))))
		print(f'Extracting information from {len(list(df.columns))} sources: {", ".join(list(df.columns))}')
	logger.info(f'Extract information from {len(list(df.columns))} sources: {", ".join(list(df.columns))}')
	# print('DF',df)
	return (df)


def get_variants(file, variants_save):
	'''
	Take a vcf file and return all its variants in a set (or a dictionnary with the index fields)
	Input : vcf file and empty set to save variants
	Output : set containing the variants
	'''
	try:
		for line in read_vcf(file):
			if type(line) == type({}):
				intfield = line
			else:
				chr = line[intfield['idx_chr']]
				pos = line[intfield['idx_pos']]
				ref = line[intfield['idx_ref']]
				alts = line[intfield['idx_alts']]
				for alt in alts.split(','):
					variants_save.add((chr, pos, ref, alt))
	except:
		# print(file)
		pass
	return variants_save


def add_to_genes_anno(keep_vcf: bool, genes: dict, gene: str, type: str, chr: str, gene_position: (), mg: str, mc: str, mp: str, anno_info: str):
	'''
	Add 1 to a gene in the genes dictionary according to its type
		Input : genes dictionary, gene name, type of the gene ('weak' or 'strong') + add chromosome, gene position, mutation in genomic sequence, mutation in coding sequence, mutation in proteic sequence
	'''

	if keep_vcf:
		if not gene in genes.keys():
			genes[gene] = {}
			genes[gene]['weak'] = 0
			genes[gene]['strong'] = 0
			genes[gene]['common'] = 0
			genes[gene]['mg'] = []
			genes[gene]['mc'] = []
			genes[gene]['mp'] = []
			genes[gene]['anno_info'] = []
		genes[gene][type] += 1
		genes[gene]['chr'] = chr
		genes[gene]['gene_pos'] = gene_position
		genes[gene]['mg'].append(mg)
		genes[gene]['mc'].append(mc)
		genes[gene]['mp'].append(mp)
		genes[gene]['anno_info'].append(anno_info)
	else:
		if not gene in genes.keys():
			genes[gene] = {}
			genes[gene]['all'] = 0
			genes[gene]['mg'] = []
			genes[gene]['mc'] = []
			genes[gene]['mp'] = []
			genes[gene]['anno_info'] = []
		genes[gene]['all'] += 1
		genes[gene]['chr'] = chr
		genes[gene]['gene_pos'] = gene_position
		genes[gene]['mg'].append(mg)
		genes[gene]['mc'].append(mc)
		genes[gene]['mp'].append(mp)
		genes[gene]['anno_info'].append(anno_info)


def add_to_genes_funco(genes: dict, gene: str, type: str, chr: str, gene_position: (), mg: str, mc: str, mp: str):
	'''
	Add 1 to a gene in the genes dictionary according to its type
		Input : genes dictionary, gene name, type of the gene ('weak' or 'strong') + add chromosome, gene position, mutation in genomic sequence, mutation in coding sequence, mutation in proteic sequence
	'''
	if not gene in genes.keys():
		genes[gene] = {}
		genes[gene]['weak'] = 0
		genes[gene]['strong'] = 0
		genes[gene]['mg'] = []
		genes[gene]['mc'] = []
		genes[gene]['mp'] = []
	genes[gene][type] += 1
	genes[gene]['chr'] = chr
	genes[gene]['gene_pos'] = gene_position
	genes[gene]['mg'].append(mg)
	genes[gene]['mc'].append(mc)
	genes[gene]['mp'].append(mp)


def modify_variants_pass_and_get_genes(dict_para, file1, weak, strong, gene_name_dico: {}, transcript_dico: {}, parent_name: str):
	'''
	Get the pass variants lines from a vcf and add the variant type ('weak'/'strong'/'common') + get the count of types weak and strong for every gene
	Input : vcf file 1 to modify, vcf file 2 (just used for the name), set of specific variants from file 1, set of weak variants from file 1, set of strong variants from file 1 and the logger
	Output : dictionary containing genes and their number of weak and strong variants (specific to vcf file 1)
	'''
	if weak == 'no':
		keep_vcf = False

	outfile = dict_para['output_path_comparison'] + 'common_variants.test.vcf'
	genes = {}
	with open(outfile, 'w') as o:
		for line in get_vcf_header(file1):
			o.write(line + '\n')
		for line in read_vcf(file1):
			if type(line) == type({}):
				intfield = line
			else:
				chromosome = line[intfield['idx_chr']]
				ID_ensemble = []
				anno_info = []
				mg = []  # mutation in genomic sequence
				mc = []  # mutation in coding sequence
				mp = []  # mutation in proteic sequence
				infos = [tuple(infos.split('=')) if len(infos.split('=')) > 1 else (infos, '') for infos in line[intfield['idx_info']].split(';')]
				nb_alt = len(line[intfield['idx_alts']].split(','))

				try:  # annovar
					dictionary = {key: value for key, value in infos}
					gene = dictionary['Gene.refGene']
					if '\\' in gene:
						gene = gene.split('\\')[0]
					# print(gene)
					try:
						gene_position = gene_name_dico[gene][0]
					except:
						pass
					g_mutation = "g." + chromosome + ":" + line[intfield['idx_pos']] + line[intfield['idx_ref']] + ">" + line[intfield['idx_alts']]
					for i in range(nb_alt):
						if nb_alt > 1:
							g_mutation = "g." + chromosome + ":" + line[intfield['idx_pos']] + line[intfield['idx_ref']] + ">" + line[intfield['idx_alts']]
							mutation1 = g_mutation.split(',')[0]
							if '>' in mutation1:
								mutation2 = g_mutation.split('>')[0] + '>' + g_mutation.split(',')[1]
							elif 'ins' in mutation1:
								mutation2 = mutation1.split['ins'][0] + 'ins' + mutation1.split('ins')[1]
							elif 'del' in mutation1:
								mutation2 = mutation1.split['del'][0] + 'del' + mutation1.split('del')[1]
							mg.append(mutation1)
							mg.append(mutation2)
						else:
							mg.append(g_mutation)

						if dictionary['ExonicFunc.refGene'] == ".":
							anno_info.append('NO INFO')
							mc.append('NO INFO')
							mp.append('NO INFO')
						elif dictionary['ExonicFunc.refGene'] == "unknown":
							anno_info.append('unknown')
							mc.append('unknown')
							mp.append('unknown')
						else:
							anno_info.append(dictionary['ExonicFunc.refGene'])
						i = 0
						for info in anno_info:
							if info == '.':
								anno_info[i] = 'NO INFO'
							i += 1
						i = 0
						pattern1 = r'NM(.*?:.*?):'
						matches1 = re.findall(pattern1, dictionary['AAChange.refGene'])
						if len(matches1) == 1:
							dictionary['AAChange.refGene'] + ','
						modified_matches = ['NM' + match for match in matches1]
						pattern2 = r'c\.(.*?):'
						matches2 = re.findall(pattern2, dictionary['AAChange.refGene'])
						k = 0
						pattern3 = r'p\.(.*?)(?=,|$)'
						matches3 = re.findall(pattern3, dictionary['AAChange.refGene'])
						dic_mc = {}
						dic_mp = {}
						list1 = []
						list2 = []

						for transcript in modified_matches:
							if 'NMT:' in transcript:
								transcript = transcript.replace('NMT:', '')
							list1.append(matches2[k])
							dic_mc[transcript] = list1
							list2.append(matches3[k])
							dic_mp[transcript] = list2
							if k == 0:  # NOUS MANQUE SUREMENT L AUTRE MOITIE SI NB_ALTS > 1, A VERIFIER
								if k != len(modified_matches) - 1:
									mc.append(transcript + ' -> ' + matches2[k] + ' | ')
									mp.append(transcript + ' -> ' + matches3[k] + ' | ')
								elif k == len(modified_matches) - 1:
									mc.append(transcript + ' -> ' + matches2[k])
									mp.append(transcript + ' -> ' + matches3[k])
							elif k >= 1:
								if k != len(modified_matches):
									mc[i] = mc[i] + transcript + ' -> ' + matches2[k] + ' | '
									mp[i] = mp[i] + transcript + ' -> ' + matches3[k] + ' | '
								elif k == len(modified_matches) - 1:
									mc[i] = mc[i] + transcript + ' -> ' + matches2[k]
									mp[i] = mp[i] + transcript + ' -> ' + matches3[k]
							k += 1

				except:  # funcotator
					idx_funcotation = list(zip(*infos))[0].index('FUNCOTATION')
					info_funcotation = list(zip(*infos))[1][idx_funcotation].split(',')
					gene = info_funcotation[0].split('|')[0].lstrip('[')
					id_ensembl = info_funcotation[0].split('|')[12]
					try:
						if len(gene_name_dico[gene]) > 1:
							gene_position = transcript_dico[id_ensembl.split('.')[0]][0]
						else:
							gene_position = gene_name_dico[gene][0]
					except KeyError:
						try:
							gene_position = transcript_dico[id_ensembl.split('.')[0]][0]
						except KeyError:
							gene_position = ('', chromosome, '')
					for i in range(nb_alt):
						mg.append(info_funcotation[i].split('|')[11])
						mc.append(info_funcotation[i].split('|')[16])
						mp.append(info_funcotation[i].split('|')[18])
						ID_ensemble.append(info_funcotation[i].split('|')[12])

				try:
					if mc[-1].endswith(' | '):
						mc[-1] = mc[-1].rstrip(' | ')
					if mp[-1].endswith(' | '):
						mp[-1] = mp[-1].rstrip(' | ')
				except:
					pass

				if dict_para['keep_filtered_vcf_after_run'].upper() == 'TRUE' or dict_para['keep_filtered_vcf_after_run'].upper() == 'YES':
					keep_vcf = True
				else:
					keep_vcf = False
				variant_type = []
				for i, alt in enumerate(line[intfield['idx_alts']].split(',')):  # For each variants
					variant = (line[intfield['idx_chr']], line[intfield['idx_pos']], line[intfield['idx_ref']], alt)
					# Add variant type to the info field

					try:
						if keep_vcf:
							if variant in weak:
								variant_type.append('weak')
								try:
									add_to_genes_anno(keep_vcf, genes, gene, 'weak', chromosome, gene_position, mg[i], mc[i], mp[i], anno_info[i])
								except:
									pass  # gene is weird, we don't want the variant
							elif variant in strong:
								variant_type.append('strong')
								try:
									add_to_genes_anno(keep_vcf, genes, gene, 'strong', chromosome, gene_position, mg[i], mc[i], mp[i], anno_info[i])
								except:
									pass  # gene is weird, we don't want the variant
							else:
								variant_type.append('common')
								try:
									add_to_genes_anno(keep_vcf, genes, gene, 'common', chromosome, gene_position, mg[i], mc[i], mp[i], anno_info[i])
								except:
									pass  # gene is weird, we don't want the variant
						else:
							try:
								add_to_genes_anno(keep_vcf, genes, gene, 'no', chromosome, gene_position, mg[i], mc[i], mp[i], anno_info[i])
							except:
								print('a')
					except:
						# print("not annotated with ANNOVAR")
						pass
					try:
						if genes[gene]['anno_info'] is list:
							if len(genes[gene]['anno_info']) >= 2:
								print(genes[gene])
						else:
							try:
								variant = (line[intfield['idx_chr']], line[intfield['idx_pos']], line[intfield['idx_ref']], alt)
								# Add variant type to the info field
								if variant in weak:
									variant_type.append('weak')
									add_to_genes_funco(genes, gene, 'weak', chromosome, gene_position, mg[i], mc[i], mp[i])
								elif variant in strong:
									variant_type.append('strong')
									add_to_genes_funco(genes, gene, 'strong', chromosome, gene_position, mg[i], mc[i], mp[i])
								else:
									variant_type.append('common')
							except:
								# print('not annotated by funcotator')
								pass
					except:
						pass  # gene is weird, we don't want the variant

				# Recreate and write the vcf line
				line[intfield['idx_info']] = line[intfield['idx_info']] + ';VARIANT_TYPE=' + ','.join(variant_type)  # Add the variant type to the vcf Info field
				o.write("\t".join(line) + '\n')
	os.remove(outfile)
	return genes


def create_snp_plot(args, df_snp, df_snp2):
	'''
	Snp count plot creation
	Input : snp profile of file 1, snp profile of file 2, output name for the plot and the logger
	'''

	df_snp['Frequency (%)'] = df_snp['Frequency (%)'] / 100
	# df_snp.set_index(['Mutation', 'Context'], inplace=True)
	df_snp = df_snp.drop('Total', axis=1)

	df_snp2['Frequency (%)'] = df_snp2['Frequency (%)'] / 100
	# df_snp2.set_index(['Mutation', 'Context'], inplace=True)
	df_snp2 = df_snp2.drop('Total', axis=1)

	try:
		name1 = args['output_path_comparison'].split('sons/')[1].split("/")[0].split('___')[0]
		name2 = args['output_path_comparison'].split('sons/')[1].split("/")[0].split('___')[1]
	except:
		name1 = args['output_path_sample1']
		name2 = args['output_path_sample2']

	# Colors
	white_96 = ['white'] * 96
	color_6 = ['darkblue', 'blue', 'lightblue', 'darkgreen', 'green', 'lightgreen']
	color_96 = []
	for i in color_6:
		color_96 += [i] * 16
	########

	with warnings.catch_warnings():
		warnings.filterwarnings("ignore")
		bars = [row[1] for index, row in df_snp.iterrows()]
	height = [float(i) for i in df_snp.iloc[:, 2] - df_snp2.iloc[:, 2]]  # height of the bars from the difference between the two snp profiles
	height2 = [float(i) for i in df_snp.iloc[:, 2]]
	height3 = [float(-i) for i in df_snp2.iloc[:, 2]]
	height4 = [i if i > 0 else 0 for i in height]
	height5 = [i if i < 0 else 0 for i in height]
	del height
	group = ['C>A', 'C>G', 'C>T', 'T>A', 'T>C', 'T>G']
	x_pos = np.arange(len(bars))

	# Create bars in two subplot
	fig = plt.figure(figsize=(15, 10))
	plt.gcf().subplots_adjust(left=0.1, bottom=0.1, right=0.9, top=0.9, wspace=0, hspace=0.12)
	ax1 = plt.subplot(211)
	ax1.bar(x_pos, height2, color=white_96, edgecolor=color_96, linestyle="--")
	ax1.bar(x_pos, height4, color=color_96)
	ax1.set_xticks(x_pos)
	ax1.set_xticklabels(bars, rotation=90, fontsize=9)
	ax2 = plt.subplot(212)
	ax2.bar(x_pos, height3, color=white_96, edgecolor=color_96, linestyle="--")
	ax2.bar(x_pos, height5, color=color_96)

	# x-axis and y-axis creation
	ax2.set_xticks(x_pos)
	ax2.set_xticklabels([])
	ax3 = ax2.twiny()
	newpos = [8, 24, 40, 56, 72, 88]
	ax3.set_xticks(newpos)
	ax3.set_xticklabels(group)
	ax2.xaxis.set_ticks_position('top')
	ax3.xaxis.set_ticks_position('bottom')
	ax3.xaxis.set_label_position('bottom')
	ax3.spines['bottom'].set_position(('outward', 5))

	ax3.set_xlabel('Mutation types', labelpad=15, fontsize=12)
	ax3.set_xlim(ax2.get_xlim())

	plt.tick_params(
		axis='x',  # changes apply to the x-axis
		which='both',  # both major and minor ticks are affected
		bottom=False,  # ticks along the bottom edge are off
		labelbottom=True)  # labels along the bottom edge are on

	for xtick, color in zip(ax3.get_xticklabels(), color_6):
		xtick.set_color(color)
		xtick.set_size(12)
	ax3.spines['bottom'].set_visible(False)

	plt.draw()  # populate the yticklabels

	ylabels = [round(float(ytick.get_text()), 2) if (ytick.get_text()[0] == '0' or ytick.get_text()[0] == '1') else round(float(ytick.get_text()[1:]), 2) for ytick in
			   ax1.get_yticklabels()]
	ylabels2 = [round(float(ytick.get_text()), 2) if (ytick.get_text()[0] == '0' or ytick.get_text()[0] == '1') else round(float(ytick.get_text()[1:]), 2) for ytick in
				ax2.get_yticklabels()]
	max_ylabel = max(ylabels + ylabels2)

	ylab = [x for x in np.arange(0, max_ylabel + 0.01, 0.01)]
	# no weird labels (too numerous for example)
	if len(ylab) > 10:
		ylab = ylab[::2]
	while len(ylab) > 20:
		ylab = [round(label, 2) for label in ylab]
		ylab = ylab[::2]

	ax1.set_yticks(ylab)
	ax1.set_yticklabels([str(int(i * 100)) for i in ylab])
	ax2.set_yticks([-i for i in ylab[::-1]])
	ax2.set_yticklabels([str(int(i * 100)) for i in ylab[::-1]])

	name1 = name1.rsplit('/', 1)[-1].split('.vcf')[0]
	name2 = name2.rsplit('/', 1)[-1].split('.vcf')[0]
	ax1.text(102, ax1.get_ylim()[1] - 0.5 * (ax1.get_ylim()[1] - ax1.get_ylim()[0]), name1, fontsize=9,
			 color='black', rotation=270, va='center', ha='right')
	ax2.text(102, ax2.get_ylim()[0] + 0.5 * (ax2.get_ylim()[1] - ax2.get_ylim()[0]), name2, fontsize=9,
			 color='black', rotation=270, va='center', ha='right')

	fig.text(0.05, 0.5, 'Proportion of mutation type (%)', va='center', rotation='vertical', fontsize=11)
	pair_id = args['output_path_comparison'].split('sons/')[1].replace('/', '')
	plt.title(f"Comparison of SNP types between time 1 and time 2,\nfor patient {pair_id}", fontweight='bold', fontsize=13, y=2, pad=50)
	formats = args['C_SNP_profile_plot_format(s)'].upper()
	outname = args['output_path_comparison'] + 'SNP_profile'
	if 'PNG' in formats:
		plt.savefig(outname + '.png', dpi=600)
	if 'PDF' in formats:
		plt.savefig(outname + '.pdf', dpi=600)
	if 'SVG' in formats:
		plt.savefig(outname + '.svg', dpi=600)
	if 'JPG' in formats:
		plt.savefig(outname + '.jpg', dpi=600)

	plt.close()


def create_SNP_profile_table(dict_para, df1, df2):
	df = pd.concat([df1, df2], ignore_index=True)
	dftot = df.rename({'Unnamed: 0': 'Mutation type', 'Unnamed: 1': 'DNA context', 'passed': 'Proportion (%)',
					   'Associated count passed': 'Count'}, axis=1)
	dftot["Frequency (%)"] = dftot["Frequency (%)"].apply(lambda x: x * 100)
	outname = dict_para['output_path_comparison'] + 'SNP_profile'

	if 'TSV' in dict_para['C_SNP_profile_table_format(s)'].upper():
		outname = outname + '.tsv'
		dftot.to_csv(outname, sep='\t', index=False)
	if 'CSV' in dict_para['C_SNP_profile_table_format(s)'].upper():
		outname = outname + '.csv'
		dftot.to_csv(outname, index=False)
	if 'XLSX' in dict_para['C_SNP_profile_table_format(s)'].upper():
		outname = outname + '.xlsx'
		dftot.to_excel(outname, index=False)


def graph_snp(dict_para, out_C_SNP_profile, logger):
	'''
	Create the snp profile plot for each snp profile file (if exist)
	Input : snp profile files, snp profile files, output name for the plot and the logger
	'''
	file_extensions = ['.xlsx', '.csv', '.tsv']
	snp_profile_files = []
	filename = None

	snp1 = dict_para['output_path'] + 'samples/' + dict_para['file1'].replace(".vcf", "") + '/'
	for file in os.listdir(snp1):
		if file.startswith('SNP_profile') and file.endswith(tuple(file_extensions)):
			filename = file
			break
	snp_profile_files.append(snp1 + filename)
	snp2 = dict_para['output_path'] + 'samples/' + dict_para['file2'].replace(".vcf", "") + '/'
	for file in os.listdir(snp2):
		if file.startswith('SNP_profile') and file.endswith(tuple(file_extensions)):
			filename = file
			break
	snp_profile_files.append(snp2 + filename)

	if dict_para['verbose_prints'].upper() == 'TRUE':
		print('Create profile comparison graph...')
	for num in range(1, len(snp_profile_files)):
		if 'xlsx' in filename:
			df1 = pd.read_excel(snp_profile_files[num - 1], engine='openpyxl')
			df2 = pd.read_excel(snp_profile_files[num], engine='openpyxl')
		elif 'csv' in filename:
			df1 = pd.read_csv(snp_profile_files[num - 1])
			df2 = pd.read_csv(snp_profile_files[num])
		elif 'tsv' in filename:
			df1 = pd.read_csv(snp_profile_files[num - 1], sep='\t')
			df2 = pd.read_csv(snp_profile_files[num], sep='\t')

		create_snp_plot(dict_para, df1, df2)
		create_SNP_profile_table(dict_para, df1, df2)


def true_stem(path):
	'''
	Takes a file Path and return the true stem of this path (file name without extensions)
	'''
	stem = Path(path).stem
	return stem if stem == path else true_stem(stem)


def create_graph_indel(deletion1, deletion2, insertion1, insertion2, outname, logger, dict_para):
	'''
	Create the indel count comparison plot
		Input : Insertion count file 1, Insertion count file 2, Deletion count file 1, Deletion count file 2, output name for the plot and the logger
	'''

	insert = True
	delet = True
	if (deletion1 == None or deletion2 == None) and (insertion1 == None or insertion2 == None):  # no insertion nor deltion counts file
		print('Warning ! No indel files ! No graphs will be created !')
		logger.warning(f'No indel files ! No graphs will be created !')
		return None
	elif insertion1 == None or insertion2 == None:  # insertion counts file does not exist
		print('Warning ! No insertion file !')
		logger.warning(f'No insertion file !')
		insert = False
	elif deletion1 == None or deletion2 == None:  # deltion counts file does not exist
		print('Warning ! No deletion file !')
		logger.warning(f'No deletion file !')
		delet = False

	# Get dataframe from tsv file

	if insert:
		ins1 = pd.read_csv(insertion1, sep='\t', header=0, index_col=0).iloc[:, 0]
		sum_ins1 = sum(ins1)
		ins2 = pd.read_csv(insertion2, sep='\t', header=0, index_col=0).iloc[:, 0]
		sum_ins2 = sum(ins2)
		height_ins1 = list([float(i) / sum_ins1 for i in ins1])
		height_ins2 = list([-(float(i) / sum_ins2) for i in ins2])
		bars_ins1 = list(ins1.index)
		bars_ins2 = list(ins2.index)
		name_ins1 = ins1.name
		name_ins2 = ins2.name
	if delet:
		del1 = pd.read_csv(deletion1, sep='\t', header=0, index_col=0).iloc[:, 0]
		del2 = pd.read_csv(deletion2, sep='\t', header=0, index_col=0).iloc[:, 0]
		sum_del1 = sum(del1)
		sum_del2 = sum(del2)
		height_del1 = list([float(i) / sum_del1 for i in del1])
		height_del2 = list([-(float(i) / sum_del2) for i in del2])
		bars_del1 = list(del1.index)
		bars_del2 = list(del2.index)
		name_del1 = del1.name
		name_del2 = del2.name

	if delet and insert:
		maximum = max([max(bars_ins1) + 1, max(bars_ins2) + 1, max(bars_del1) + 1, max(bars_del2) + 1])
		max_y = max(height_del1 + height_del2 + height_ins1 + height_ins2) + 0.05
	elif delet:
		maximum = max([max(bars_del1) + 1, max(bars_del2) + 1])
		max_y = max(height_del1 + height_del2) + 0.05
	elif not delet:
		maximum = max([max(bars_ins1) + 1, max(bars_ins2) + 1])
		max_y = max(height_ins1 + height_ins2) + 0.05

	x1 = [0] + [i + 1 for i in range(maximum)]
	x2 = [0] + [i + 1 for i in range(maximum)]

	width = 0.25

	fig = plt.figure(figsize=(15, 10))
	ax1 = plt.subplot(1, 1, 1)

	# create plot according to deletion and insertion counts

	if delet and insert:
		plt.bar([float(i) + (width * 0.65) for i in bars_ins1], height_ins1, color='r', width=width, edgecolor='r', label='Insertion_' + true_stem(name_ins1))
		plt.bar([float(i) - (width * 0.65) for i in bars_del1], height_del1, color='darkred', width=width, edgecolor='darkred', label='Deletion_' + true_stem(name_del1))
		plt.bar([float(i) + (width * 0.65) for i in bars_ins2], height_ins2, color='b', width=width, edgecolor='b', label='Insertion_' + true_stem(name_ins2))
		plt.bar([float(i) - (width * 0.65) for i in bars_del2], height_del2, color='darkblue', width=width, edgecolor='darkblue', label='Deletion_' + true_stem(name_del2))
		x1 = sorted(list(set(bars_del1).union(set(bars_ins1))))
		x2 = sorted(list(set(bars_ins2).union(set(bars_del2))))
	elif not insert:
		plt.bar([float(i) for i in bars_del1], height_del1, color='darkred', width=width, edgecolor='darkred', label='Deletion_' + true_stem(name_del1))
		plt.bar([float(i) for i in bars_del2], height_del2, color='darkblue', width=width, edgecolor='darkblue', label='Deletion_' + true_stem(name_del2))
		x1 = bars_del1
		x2 = bars_del2
	elif not delet:
		plt.bar([float(i) for i in bars_ins1], height_ins1, color='r', width=width, edgecolor='r', label='Insertion_' + true_stem(name_ins1))
		plt.bar([float(i) for i in bars_ins2], height_ins2, color='b', width=width, edgecolor='b', label='Insertion_' + true_stem(name_ins2))
		x1 = bars_ins1
		x2 = bars_ins2

	plt.legend()
	ax1.grid(axis='x', color='lightgrey', linestyle='-', linewidth=0.5)
	max_x = max(x1 + x2)

	ax1.set_xlim(0, max_x + 1)
	ax1.set_xlabel("Indel size (bp)")
	ax1.set_ylabel("Indel percentage")
	ax1.set_xticks(x2)
	ax1.set_xticklabels(x2, fontsize=10)

	ax2 = ax1.twiny()
	ax2.set_xlim(0, max_x + 1)
	ax2.set_xticks(x1)
	ax2.set_xticklabels(x1, fontsize=10)
	ax2.grid(axis='x', color='lightgrey', linestyle='-', linewidth=0.5)

	fig.set_size_inches(max_x / 3, 10)

	y_pos = list(np.arange(-1, 1.1, 0.1))
	y_value = [round(i, 1) for i in (list(np.arange(1, 0, -0.1)) + list(np.arange(0, 1.1, 0.1)))]
	ax1.set_yticks(y_pos)
	ax1.set_yticklabels(y_value)

	ylabels = [str(round(float(ytick.get_text()), 2)) if (ytick.get_text()[0] == '0' or ytick.get_text()[0] == '1') else str(round(float(ytick.get_text()[1:]), 2)) for ytick in
			   ax1.get_yticklabels()]
	ax1.set_yticks(ax1.get_yticks())
	ax1.set_yticklabels([str(round(float(i) * 100, 1)) for i in ylabels])
	ax1.set_ylim(-max_y, max_y)

	plt.hlines(0, 0, maximum + 1, color='black', linewidth=1)

	logger.info(f'Draw indel size barplot in {outname}')

	formats = dict_para['C_indel_profile_plot_format(s)'].upper()
	if 'PNG' in formats:
		outname = Path(outname).with_suffix('.png')
	if 'JPG' in formats:
		outname = Path(outname).with_suffix('.jpg')
	if 'PDF' in formats:
		outname = Path(outname).with_suffix('.pdf')
	if 'SVG' in formats:
		outname = Path(outname).with_suffix('.svg')
	plt.savefig(outname)

	plt.close()


def graph_indel(dict_para, insertions_count_files, deletions_count_files, out_indel, logger):
	'''
	Create the indel count comparison plot for each indel count file (if exist)
	Input : Insertion count files, Deletion count files, output name for the plot and the logger
	'''
	if len(insertions_count_files) != len(deletions_count_files):
		logger.error(f'Not the same number of insertions and deletions count files. No graph will be produced !')
		raise ValueError(f'Not the same number of insertions and deletions count files. No graph will be produced !')
		exit(1)
	if dict_para['verbose_prints'].upper() == 'TRUE':
		print('Draw indel size barplot...')
	for num in range(1, len(insertions_count_files)):
		create_graph_indel(insertions_count_files[num - 1], insertions_count_files[num], deletions_count_files[num - 1], deletions_count_files[num], out_indel, logger, dict_para)


def add_mutation_type(dic_mutation_types, dic_mutation_subtypes, genes, gene):
	# Mutation type counting
	for mutation in genes[gene]['mg']:
		ref = mutation.split('>')[0].split(':')[1]
		ref = re.sub(r'[^A-Za-z]', '', ref)
		alt = mutation.split('>')[1]
		if len(ref) == len(alt) == 1:
			dic_mutation_types['SNP'] += 1
		elif len(ref) == len(alt) == 2:
			dic_mutation_types['DNP'] += 1
		elif len(ref) == len(alt) == 3:
			dic_mutation_types['TNP'] += 1
		elif len(ref) == len(alt) > 3:
			dic_mutation_types['ONP'] += 1
		elif len(ref) > len(alt):
			dic_mutation_types['DELETION'] += 1
			dic_mutation_types['INDEL'] += 1
		elif len(ref) < len(alt):
			dic_mutation_types['INSERTION'] += 1
			dic_mutation_types['INDEL'] += 1

	# Mutation subtype counting
	for mutation in genes[gene]['anno_info']:
		if mutation == 'synonymous_SNV':
			dic_mutation_subtypes['synonymous SNV'] += 1
		elif mutation == 'nonsynonymous_SNV':
			dic_mutation_subtypes['nonsynonymous SNV'] += 1
		elif mutation == 'frameshift_substitution':
			dic_mutation_subtypes['frameshift substitution'] += 1
		elif mutation == 'nonframeshift_substitution':
			dic_mutation_subtypes['non frameshift substitution'] += 1
		elif mutation == 'stopgain':
			dic_mutation_subtypes['stopgain'] += 1
		elif mutation == 'stoploss':
			dic_mutation_subtypes['stoploss'] += 1
		elif mutation == 'startloss':
			dic_mutation_subtypes['startloss'] += 1
		elif mutation == 'frameshift_insertion':
			dic_mutation_subtypes['frameshift insertion'] += 1
		elif mutation == 'nonframeshift_insertion':
			dic_mutation_subtypes['non frameshift insertion'] += 1
		elif mutation == 'frameshift_deletion':
			dic_mutation_subtypes['frameshift deletion'] += 1
		elif mutation == 'nonframeshift_deletion':
			dic_mutation_subtypes['non frameshift deletion'] += 1
		elif mutation == 'unknown':
			dic_mutation_subtypes['unknown'] += 1

	return dic_mutation_types, dic_mutation_subtypes


def create_mutation_types_barplot(dict_para, dic_mutation_types_t1, dic_mutation_types_t2):
	plt.clf()

	labels = [label for label in dic_mutation_types_t1 if dic_mutation_types_t1[label] != 0 or dic_mutation_types_t2[label] != 0]
	values_t1 = [dic_mutation_types_t1[label] for label in labels]
	values_t2 = [dic_mutation_types_t2[label] for label in labels]
	x = np.arange(len(labels))

	# Calculate the number of bars
	num_bars = len(labels)
	if num_bars <= 2:
		bar_width = 0.2  # Reduce the bar width for one or two bars
		bar_gap = 0.1
	else:
		bar_width = 0.2
		bar_gap = 0

	fig, ax = plt.subplots(figsize=(12, 8))

	rects1 = ax.bar(x - bar_width - bar_gap / 2, values_t1, bar_width, label='time 1', color=(0, 0.5, 0), edgecolor='black', linewidth=1)
	rects2 = ax.bar(x + bar_gap / 2, values_t2, bar_width, label='time 2', color=(0, 0, 0.5), edgecolor='black', linewidth=1)

	for rect in rects1:
		height = rect.get_height()
		ax.annotate('{}'.format(height),
					xy=(rect.get_x() + rect.get_width() / 2, height),
					xytext=(0, 3),
					textcoords="offset points",
					ha='center', va='bottom',
					fontsize=13)

	for rect in rects2:
		height = rect.get_height()
		ax.annotate('{}'.format(height),
					xy=(rect.get_x() + rect.get_width() / 2, height),
					xytext=(0, 3),
					textcoords="offset points",
					ha='center', va='bottom',
					fontsize=13)

	ax.set_ylabel('Count', fontsize=14.5, labelpad=10)
	patient = dict_para['output_path_comparison'].split('isons/')[1].replace("/", "").replace("___", "_")
	ax.set_title('Mutation types comparison between time 1 and time 2\n(patient ' + patient + ')', fontsize=16, pad=10)
	ax.set_xticks(x)
	ax.set_xticklabels(labels, ha='center', fontsize=13)
	ax.set_ylim(0, max(max(values_t1), max(values_t2)) * 1.1)

	offsets = [0.1, 0.05, -0.08, -0.05]
	num_offsets = len(labels)
	proportion = sum(offsets) / len(offsets)
	offsets = [proportion * i for i in range(num_offsets)]
	ax.set_xticks(x - offsets, minor=False)
	if num_bars <= 2:
		ax.set_xticklabels(labels, minor=False)
		ax.set_xlim([-0.5, len(labels) - 0.5])
	else:
		ax.set_xticklabels(labels, minor=False)
	plt.tick_params(bottom=False)

	with warnings.catch_warnings():
		warnings.simplefilter("ignore")
		ax.set_yticklabels(ax.get_yticks().astype(int), fontsize=11)
	plt.tick_params(bottom=False)

	ax.legend(fontsize=15, edgecolor='white')
	plt.tight_layout()

	file_formats = dict_para['C_types_plot_format(s)'].upper()
	path = dict_para['output_path_comparison'] + 'mutation_types'
	if 'PNG' in file_formats:
		plt.savefig(path + '.png', dpi=600)
	if 'PDF' in file_formats:
		plt.savefig(path + '.pdf', dpi=600)
	if 'SVG' in file_formats:
		plt.savefig(path + '.svg', dpi=600)
	if 'JPG' in file_formats:
		plt.savefig(path + '.jpg', dpi=600)

	plt.close()


def create_mutation_subtypes_barplot(dict_para, dic_mutation_subtypes_t1, dic_mutation_subtypes_t2):
	plt.clf()

	labels = [label for label in dic_mutation_subtypes_t1 if dic_mutation_subtypes_t1[label] != 0 or dic_mutation_subtypes_t2[label] != 0]
	values_t1 = [dic_mutation_subtypes_t1[label] for label in labels]
	values_t2 = [dic_mutation_subtypes_t2[label] for label in labels]

	x = np.arange(len(labels))
	if len(values_t1) <= 2:
		bar_width = 0.1  # Reduce the bar width for one or two bars
		bar_gap = 0.05
	else:
		bar_width = 0.2
		bar_gap = 0.1

	fig, ax = plt.subplots(figsize=(13, 8))

	if len(values_t1) != 3:
		rects1 = ax.bar(x - bar_width - bar_gap / 2, values_t1, bar_width, label='time 1', color=(0, 0.5, 0), edgecolor='black', linewidth=1)
		rects2 = ax.bar(x + bar_gap / 2, values_t2, bar_width, label='time 2', color=(0, 0, 0.5), edgecolor='black', linewidth=1)
	else:
		rects2 = ax.bar(x - bar_width - bar_gap / 3.2, values_t2, bar_width, label='time 2', color=(0, 0.5, 0), edgecolor='black', linewidth=1)
		rects1 = ax.bar(x + bar_gap / 3.2, values_t1, bar_width, label='time 1', color=(0, 0, 0.5), edgecolor='black', linewidth=1)

	for rect in rects1:
		height = rect.get_height()
		ax.annotate('{}'.format(height),
					xy=(rect.get_x() + rect.get_width() / 2, height),
					xytext=(0, 2),
					textcoords="offset points",
					ha='center', va='bottom',
					fontsize=13)

	for rect in rects2:
		height = rect.get_height()
		ax.annotate('{}'.format(height),
					xy=(rect.get_x() + rect.get_width() / 2, height),
					xytext=(0, 2),
					textcoords="offset points",
					ha='center', va='bottom',
					fontsize=13)

	ax.set_ylabel('Count', fontsize=14, labelpad=5)
	patient = dict_para['output_path_comparison'].split('isons/')[1].replace("/", "").replace("___", "_")
	ax.set_title('Mutation subtypes comparison between time 1 and time 2 \n(patient ' + patient + ')', fontsize=16, pad=10)
	ax.set_xticks(x)
	if len(values_t1) > 5:
		ax.set_xticklabels(labels, rotation=90, ha='center', fontsize=13)
	try:
		ax.set_ylim(0, max(max(values_t1), max(values_t2)) * 1.1)
	except:
		print('a')

	offset = bar_width / 2

	if len(labels) > 2:
		ax.set_xticks(x - offset, minor=False)
	else:
		ax.set_xticks(x, minor=False)

	ax.set_xticklabels(labels, ha='center', fontsize=14)

	if len(labels) <= 2:
		ax.set_xlim([-0.5, len(labels) - 0.5])
	plt.tick_params(bottom=False)

	handles, legend_labels1 = ax.get_legend_handles_labels()
	handles = handles[::-1]
	legend_labels = legend_labels1[::-1]
	ax.legend(handles, legend_labels, fontsize=15, edgecolor='white')
	plt.tight_layout()

	with warnings.catch_warnings():
		warnings.simplefilter("ignore")
		ax.set_yticklabels(ax.get_yticks().astype(int), fontsize=14)

	file_formats = dict_para['C_subtypes_plot_format(s)'].upper()
	path = dict_para['output_path_comparison'] + 'mutation_subtypes'
	if 'PNG' in file_formats:
		plt.savefig(path + '.png', dpi=400)
	if 'PDF' in file_formats:
		plt.savefig(path + '.pdf', dpi=400)
	if 'SVG' in file_formats:
		plt.savefig(path + '.svg', dpi=400)
	if 'JPG' in file_formats:
		plt.savefig(path + '.jpg', dpi=400)

	plt.close()


def count_genes(df, dict_para):
	common_genes = []
	change_genes = []
	union_genes = []
	test = 0
	test1 = 0
	t1 = True
	t2 = True

	for index, row in df.iterrows():
		if not row['Gene'] in union_genes:
			union_genes.append(row['Gene'])

		if pd.isna(row['g.(t1)']) and pd.isna(row['g.(t2)']):
			print('No mutation found in both t1 and t2!')

		elif (not pd.isna(row['g.(t1)']) and pd.isna(row['g.(t2)'])) or (not pd.isna(row['g.(t2)']) and pd.isna(row['g.(t1)'])):
			if not row['Gene'] in change_genes:
				change_genes.append(row['Gene'])
			elif not row['Gene'] in common_genes:
				common_genes.append(row['Gene'])

		elif not pd.isna(row['g.(t1)']) and not pd.isna(row['g.(t2)']):
			if not '|' in (row['g.(t1)']) and not '|' in (row['g.(t2)']):
				if row['g.(t1)'] != row['g.(t2)']:
					if not row['Gene'] in change_genes:
						change_genes.append(row['Gene'])
				elif row['g.(t1)'] == row['g.(t2)']:
					if not row['Gene'] in common_genes:
						common_genes.append(row['Gene'])

			elif '|' in (row['g.(t1)']) and not '|' in (row['g.(t2)']):
				if not row['Gene'] in change_genes:
					change_genes.append(row['Gene'])

			elif '|' in row['g.(t2)'] and not '|' in (row['g.(t1)']):
				if not row['Gene'] in change_genes:
					change_genes.append(row['Gene'])

			elif '|' in row['g.(t1)'] and '|' in row['g.(t2)']:
				common = True
				for variant in row['g.(t1)'].split('|'):
					if variant not in row['g.(t2)']:
						common = False
						if not row['Gene'] in change_genes:
							change_genes.append(row['Gene'])

				for variant in row['g.(t2)'].split('|'):
					if variant not in row['g.(t1)']:
						common = False
						if not row['Gene'] in change_genes:
							change_genes.append(row['Gene'])
				if common:
					if not row['Gene'] in common_genes:
						common_genes.append(row['Gene'])
	return len(common_genes), len(change_genes)

def create_protein_impacts_plots(dict_para, dict_impacts, pair_id):
	selection_approach = dict_para['variants_selection_approach'].lower()
	sample1 = dict_para['output_path_sample1'].replace(".vcf", "") .split('samples/')[1].split("/")[0]
	sample2 = dict_para['output_path_sample2'].replace(".vcf", "") .split('samples/')[1].split("/")[0]
	path1 = dict_para['output_path'] + 'samples/' + sample1 + '/' + 'passed_variants.xlsx'
	path2 = dict_para['output_path'] + 'samples/' + sample2 + '/' + 'passed_variants.xlsx'

	if 'xlsx' in path1:
		df_variants1 = pd.read_excel(path1, index_col=0)
	if 'xlsx' in path2:
		df_variants2 = pd.read_excel(path2, index_col=0)

	protein_impacts1 = {}
	for index, row in df_variants1.iterrows():
		variant = index
		try:
			sift_score = row['SIFT score']
		except:
			sift_score = 'not found'
		try:
			sift_pred = row['SIFT pred']
		except:
			sift_pred = 'not found'
		try:
			polyphen2_score = row['Polyphen2 score']
		except:
			polyphen2_score = 'not found'
		try:
			polyphen2_pred = row['Polyphen2 pred']
		except:
			polyphen2_pred = 'not found'

		protein_impacts1[variant] = {
			'SIFT score': sift_score,
			'SIFT pred': sift_pred,
			'Polyphen2 score': polyphen2_score,
			'Polyphen2 pred': polyphen2_pred}

	protein_impacts2 = {}
	for index, row in df_variants2.iterrows():
		variant = index
		try:
			sift_score = row['SIFT score']
		except:
			sift_score = 'not found'
		try:
			sift_pred = row['SIFT pred']
		except:
			sift_pred = 'not found'
		try:
			polyphen2_score = row['Polyphen2 score']
		except:
			polyphen2_score = 'not found'
		try:
			polyphen2_pred = row['Polyphen2 pred']
		except:
			polyphen2_pred = 'not found'
		protein_impacts2[variant] = {
			'SIFT score': sift_score,
			'SIFT pred': sift_pred,
			'Polyphen2 score': polyphen2_score,
			'Polyphen2 pred': polyphen2_pred}

	config_path = 'config.txt'
	with open(config_path, "r") as file:
		lines = file.readlines()
		for line in lines:
			if "max_SIFT_score" in line:
				sift_score_threshold = float(line.split("max_SIFT_score =")[1].strip())
			elif "min_PolyPhen2_score" in line:
				polyphen_threshold = float(line.split("min_PolyPhen2_score =")[1].strip())

	sift_thresholds = [-1, 0.05, 0.2, 1.5]
	polyphen_thresholds = [-1, 0.45, 0.95, 1.5]

	protein_impacts1_counts = {
		'SIFT scores counts': {f'{sift_thresholds[i]}-{sift_thresholds[i + 1]}': 0 for i in range(len(sift_thresholds) - 1)},
		'Polyphen2 scores counts': {f'{polyphen_thresholds[i]}-{polyphen_thresholds[i + 1]}': 0 for i in range(len(polyphen_thresholds) - 1)},
		'SIFT predictions': {},
		'Polyphen2 predictions': {}
	}

	if not all(not bool(d) for d in protein_impacts1.values()):
		for impact in protein_impacts1.values():
			try:
				sift_score = impact['SIFT score']
			except:
				sift_score = 'not found'
			try:
				polyphen_score = impact['Polyphen2 score']
			except:
				polyphen_score = 'not found'
			try:
				sift_pred = impact['SIFT pred']
			except:
				sift_pred = 'not found'
			try:
				polyphen_pred = impact['Polyphen2 pred']
			except:
				polyphen_pred = 'not found'

			if sift_score == 'not found':
				if sift_score not in protein_impacts1_counts['SIFT scores counts']:
					protein_impacts1_counts['SIFT scores counts'][sift_score] = 1
				else:
					protein_impacts1_counts['SIFT scores counts'][sift_score] += 1
			else:
				interval = None
				for i in range(len(sift_thresholds) - 1):
					if sift_thresholds[i] <= sift_score < sift_thresholds[i + 1]:
						interval = f'{sift_thresholds[i]}-{sift_thresholds[i + 1]}'
						break
				protein_impacts1_counts['SIFT scores counts'][interval] += 1

			if polyphen_score == 'not found':
				if polyphen_score not in protein_impacts1_counts['Polyphen2 scores counts']:
					protein_impacts1_counts['Polyphen2 scores counts'][polyphen_score] = 1
				else:
					protein_impacts1_counts['Polyphen2 scores counts'][polyphen_score] += 1
			else:
				interval = None
				for i in range(len(polyphen_thresholds) - 1):
					if polyphen_thresholds[i] <= polyphen_score < polyphen_thresholds[i + 1]:
						interval = f'{polyphen_thresholds[i]}-{polyphen_thresholds[i + 1]}'
						break
				protein_impacts1_counts['Polyphen2 scores counts'][interval] += 1

			if sift_pred != 'not found':
				if sift_pred not in protein_impacts1_counts['SIFT predictions']:
					protein_impacts1_counts['SIFT predictions'][sift_pred] = 1
				else:
					protein_impacts1_counts['SIFT predictions'][sift_pred] += 1
			if polyphen_pred != 'not found':
				if polyphen_pred not in protein_impacts1_counts['Polyphen2 predictions']:
					protein_impacts1_counts['Polyphen2 predictions'][polyphen_pred] = 1
				else:
					protein_impacts1_counts['Polyphen2 predictions'][polyphen_pred] += 1

	protein_impacts2_counts = {
		'SIFT scores counts': {f'{sift_thresholds[i]}-{sift_thresholds[i + 1]}': 0 for i in range(len(sift_thresholds) - 1)},
		'Polyphen2 scores counts': {f'{polyphen_thresholds[i]}-{polyphen_thresholds[i + 1]}': 0 for i in range(len(polyphen_thresholds) - 1)},
		'SIFT predictions': {},
		'Polyphen2 predictions': {}
	}

	if not all(not bool(d) for d in protein_impacts2.values()):
		for impact in protein_impacts2.values():
			try:
				sift_score = impact['SIFT score']
			except:
				sift_score = 'not found'
			try:
				polyphen_score = impact['Polyphen2 score']
			except:
				polyphen_score = 'not found'
			try:
				sift_pred = impact['SIFT pred']
			except:
				sift_pred = 'not found'
			try:
				polyphen_pred = impact['Polyphen2 pred']
			except:
				polyphen_pred = 'not found'

			if sift_score == 'not found':
				if sift_score not in protein_impacts2_counts['SIFT scores counts']:
					protein_impacts2_counts['SIFT scores counts'][sift_score] = 1
				else:
					protein_impacts2_counts['SIFT scores counts'][sift_score] += 1
			else:
				interval = None
				for i in range(len(sift_thresholds) - 1):
					if sift_thresholds[i] <= sift_score < sift_thresholds[i + 1]:
						interval = f'{sift_thresholds[i]}-{sift_thresholds[i + 1]}'
						break
				protein_impacts2_counts['SIFT scores counts'][interval] += 1

			if polyphen_score == 'not found':
				if polyphen_score not in protein_impacts2_counts['Polyphen2 scores counts']:
					protein_impacts2_counts['Polyphen2 scores counts'][polyphen_score] = 1
				else:
					protein_impacts2_counts['Polyphen2 scores counts'][polyphen_score] += 1
			else:
				interval = None
				for i in range(len(polyphen_thresholds) - 1):
					if polyphen_thresholds[i] <= polyphen_score < polyphen_thresholds[i + 1]:
						interval = f'{polyphen_thresholds[i]}-{polyphen_thresholds[i + 1]}'
						break
				protein_impacts2_counts['Polyphen2 scores counts'][interval] += 1

			if sift_pred != 'not found':
				if sift_pred not in protein_impacts2_counts['SIFT predictions']:
					protein_impacts2_counts['SIFT predictions'][sift_pred] = 0
				else:
					protein_impacts2_counts['SIFT predictions'][sift_pred] += 1

			if polyphen_pred != 'not found':
				if polyphen_pred not in protein_impacts2_counts['Polyphen2 predictions']:
					protein_impacts2_counts['Polyphen2 predictions'][polyphen_pred] = 0
				else:
					protein_impacts2_counts['Polyphen2 predictions'][polyphen_pred] += 1

	sift_scores1 = [mutation['SIFT score'] for mutation in protein_impacts1.values() if 'SIFT score' in mutation]
	sift_scores2 = [mutation['SIFT score'] for mutation in protein_impacts2.values() if 'SIFT score' in mutation]

	# sift_scores = []
	# for key in dict_impacts:
	# 	sift_scores.append(dict_impacts[key]['SIFT_score'])
	# n_sifts = len(sift_scores)
	# sift_scores = [score for score in sift_scores if score != 'not found']
	# fig, ax = plt.subplots()

	if sift_scores1 == ['not found']:
		sift_scores1 = [0]
	if sift_scores2 == ['not found']:
		sift_scores2 = [0]

	n_sifts1 = len(sift_scores1)
	n_sifts2 = len(sift_scores2)

	sift_scores1 = [score for score in sift_scores1 if score != 'not found']
	sift_scores2 = [score for score in sift_scores2 if score != 'not found']

	fig, ax = plt.subplots(figsize=(12, 6))
	bplot1 = ax.boxplot(sift_scores1, positions=[1], notch=True, vert=True, patch_artist=True, labels=['Protein Impacts: t1'])
	bplot2 = ax.boxplot(sift_scores2, positions=[2], notch=True, vert=True, patch_artist=True, labels=['Protein Impacts: t2'])
	colors = ['forestgreen', 'darkblue']
	for bplot in (bplot1, bplot2):
		for patch, color in zip(bplot['boxes'], colors):
			patch.set_facecolor(color)
	ax.set_title(f'Boxplot comparing SIFT predicted protein impacts between times,\n for patient {pair_id}', fontsize=14, pad=10)
	ax.set_ylabel('SIFT Score', fontsize=12, labelpad=15)
	ax.set_xlim(0, 3)
	ax.set_xticks([1, 2])
	ax.set_xticklabels(['Protein Impacts: t1', 'Protein Impacts: t2'])
	ax.grid(color='lightgray', linestyle='--')

	mean1 = np.mean(sift_scores1)
	mean2 = np.mean(sift_scores2)
	if sift_scores1 != [0]:
		ax.text(0.9, mean1, f"x̄: {mean1:.2f}", ha='right', verticalalignment='center_baseline', fontweight='bold', fontsize=9)
	if sift_scores2 != [0]:
		ax.text(1.9, mean2, f"x̄: {mean2:.2f}", ha='right', verticalalignment='center_baseline', fontweight='bold', fontsize=9)

	plt.tight_layout()

	legend_labels1 = []
	legend_counts1 = []
	for category, sub_dict in protein_impacts1_counts.items():
		if 'SIFT' in category:
			for sub_category, count in sub_dict.items():
				legend_labels1.append(f"{sub_category}: ")
				legend_counts1.append(count)
	# ax.legend(legend_labels, loc='center left', bbox_to_anchor=(0.65, 0.8), edgecolor='black')
	fig.set_size_inches(12, 6)

	height = 0.92
	plt.annotate(f"t1 variants = {len(protein_impacts1.keys())}", xy=(0.75, height), xycoords='axes fraction', bbox=dict(facecolor='white', edgecolor='none'), fontsize=10, fontweight='bold')
	height = height - 0.06
	if legend_counts1 == [] or all(label == 0 for label in legend_counts1):
		plt.annotate(f"Caution: no SIFT info found at t1!", xy=(0.75, height), xycoords='axes fraction', bbox=dict(facecolor='white', edgecolor='none'), fontsize=10, fontstyle='italic', color='red')
		height = height - 0.07
	else:
		plt.annotate(f"t1 scores:\n", xy=(0.75, height - 0.02), xycoords='axes fraction', fontsize=10, bbox=dict(facecolor='white', edgecolor='none'), fontweight='bold')
		height = height - 0.07

	first_not_found = False
	pred_print = 0
	for label, count in zip(legend_labels1, legend_counts1):
		if label == '-1-0.05: ':
			label = '[0-0.05]: '
		if label == '0.05-0.2: ':
			label = '[0.05-0.2]: '
		if label == '0.2-1.5: ':
			label = '[0.2-1]: '
		if 'Deleterious' not in label and 'Tolerated' not in label and 'Damaging' not in label and 'Benign' not in label:
			text_color = 'black'
			if label == '[0-0.05]: ':
				text_color = 'darkgreen'
			elif label == '[0.05-0.2]: ':
				text_color = 'darkorange'
			elif label == '[0.2-1]: ':
				text_color = 'darkred'
			elif 'not found' in label.lower():
				text_color = 'black'
				first_not_found = True
			ratio = round(count / n_sifts1 * 100)
			if ratio > 0:
				plt.annotate(f"{label}{ratio}%\n", xy=(0.75, height), xycoords='axes fraction', fontsize=10, bbox=dict(facecolor='white', edgecolor='none'), color=text_color)
				height = height - 0.05
		else:
			if pred_print == 0:
				plt.annotate(f"t1 predictions:\n", xy=(0.75, height - 0.01), xycoords='axes fraction', fontsize=10, bbox=dict(facecolor='white', edgecolor='none'), fontweight='bold')
				height = height - 0.05
			if 'deleterious' in label.lower():
				text_color = 'darkgreen'
			elif 'tolerated' in label.lower():
				text_color = 'darkorange'
			elif 'not found' in label.lower():
				text_color = 'black'
			else:
				text_color = 'black'
			ratio = round(count / n_sifts1 * 100)
			if ratio > 0:
				plt.annotate(f"{label}{ratio}%\n", xy=(0.75, height), xycoords='axes fraction', fontsize=10, bbox=dict(facecolor='white', edgecolor='none'), color=text_color)
				height = height - 0.05
			pred_print += 1


	height = height + 0.04
	plt.annotate(f"--------------------------------", xy=(0.75, height), xycoords='axes fraction', bbox=dict(facecolor='white', edgecolor='none'), fontsize=13)
	height = height - 0.04
	legend_labels2 = []
	legend_counts2 = []
	for category, sub_dict in protein_impacts2_counts.items():
		if 'SIFT' in category:
			for sub_category, count in sub_dict.items():
				legend_labels2.append(f"{sub_category}: ")
				legend_counts2.append(count)

	plt.annotate(f"t2 variants = {len(protein_impacts2.keys())}", xy=(0.75, height), xycoords='axes fraction', bbox=dict(facecolor='white', edgecolor='none'), fontsize=10, fontweight='bold')
	height = height - 0.05
	if legend_counts2 == [] or all(label == 0 for label in legend_counts2):
		plt.annotate(f"Caution: no SIFT info found at t2!", xy=(0.75, height), xycoords='axes fraction', fontsize=10, bbox=dict(facecolor='white', edgecolor='none'), fontstyle='italic', color='red')
		height = height - 0.07
	else:
		plt.annotate(f"t2 scores:\n", xy=(0.75, height - 0.03), xycoords='axes fraction', fontsize=10, bbox=dict(facecolor='white', edgecolor='none'), fontweight='bold')
		height = height - 0.07

	first_not_found = False
	pred_print = 0
	for label, count in zip(legend_labels2, legend_counts2):
		if label == '-1-0.05: ':
			label = '[0-0.05]: '
		if label == '0.05-0.2: ':
			label = '[0.05-0.2]: '
		if label == '0.2-1.5: ':
			label = '[0.2-1]: '
		if 'deleterious' not in label.lower() and 'tolerated' not in label.lower() and not first_not_found:
			text_color = 'black'
			if label == '[0-0.05]: ':
				text_color = 'darkgreen'
			elif label == '[0.05-0.2]: ':
				text_color = 'darkorange'
			elif label == '[0.2-1]: ':
				text_color = 'darkred'
			elif 'not found' in label:
				text_color = 'black'
				first_not_found = True
			ratio = round(count / n_sifts2 * 100)
			if ratio > 0:
				plt.annotate(f"{label}{ratio}%\n", xy=(0.75, height - 0.01), xycoords='axes fraction', fontsize=10, bbox=dict(facecolor='white', edgecolor='none'), color=text_color)
				height = height - 0.05
		else:
			if pred_print == 0:
				plt.annotate(f"t2 predictions:\n", xy=(0.75, height - 0.015), xycoords='axes fraction', fontsize=10, bbox=dict(facecolor='white', edgecolor='none'), fontweight='bold')
				height = height - 0.06
			if 'deleterious' in label.lower():
				text_color = 'darkgreen'
			elif 'tolerated' in label.lower():
				text_color = 'darkorange'
			elif 'not found' in label.lower():
				text_color = 'black'
			else:
				text_color = 'black'
			ratio = round(count / n_sifts2 * 100)
			if ratio > 0:
				plt.annotate(f"{label}{ratio}%\n", xy=(0.75, height), xycoords='axes fraction', fontsize=10, bbox=dict(facecolor='white', edgecolor='none'), color=text_color)
				height = height - 0.05
			pred_print += 1


	plt.axhline(y=sift_score_threshold, color='darkred', linestyle='--')
	ax = plt.gca()
	custom_padding = max([max(sift_scores1), max(sift_scores2)]) - min([min(sift_scores1), min(sift_scores2)])
	ax.annotate(str(sift_score_threshold), xy=(0.008, sift_score_threshold + custom_padding), xycoords=('axes fraction', 'data'), fontsize=9, color='darkred')

	output_path = dict_para['output_path_comparison'] + 'SIFT_protein_impacts'

	if 'PNG' in dict_para['C_protein_impacts_format(s)'].upper():
		plt.savefig(output_path + '.png', dpi=400)
	if 'PDF' in dict_para['C_protein_impacts_format(s)'].upper():
		plt.savefig(output_path + '.pdf', dpi=400)
	if 'SVG' in dict_para['C_protein_impacts_format(s)'].upper():
		plt.savefig(output_path + '.svg', dpi=400)
	if 'JPG' in dict_para['C_protein_impacts_format(s)'].upper():
		plt.savefig(output_path + '.jpg', dpi=400)
	plt.close()

	polyphen_scores1 = [mutation['Polyphen2 score'] for mutation in protein_impacts1.values() if 'Polyphen2 score' in mutation]
	polyphen_scores2 = [mutation['Polyphen2 score'] for mutation in protein_impacts2.values() if 'Polyphen2 score' in mutation]
	n_polyphens1 = len(polyphen_scores1)
	n_polyphens2 = len(polyphen_scores2)
	polyphen_scores1 = [score for score in polyphen_scores1 if score != 'not found']
	polyphen_scores2 = [score for score in polyphen_scores2 if score != 'not found']

	if polyphen_scores1 == ['not found']:
		polyphen_scores1 = [0]

	fig, ax = plt.subplots(figsize=(12, 6))
	bplot1 = ax.boxplot(polyphen_scores1, positions=[1], notch=True, vert=True, patch_artist=True, labels=['Protein Impacts: t1'])
	bplot2 = ax.boxplot(polyphen_scores2, positions=[2], notch=True, vert=True, patch_artist=True, labels=['Protein Impacts: t2'])
	colors = ['forestgreen', 'darkblue']
	for bplot in (bplot1, bplot2):
		for patch, color in zip(bplot['boxes'], colors):
			patch.set_facecolor(color)
	ax.set_title(f'Boxplot comparing Polyphen2 predicted protein impacts between times,\n for patient {pair_id}', fontsize=14, pad=10)
	ax.set_ylabel('Polyphen2 Score', fontsize=12, labelpad=15)
	ax.set_xlim(0, 4)
	ax.set_xticks([1, 2])
	ax.set_xticklabels(['Protein Impacts: t1', 'Protein Impacts: t2'])
	ax.grid(color='lightgray', linestyle='--')

	mean1 = np.mean(polyphen_scores1)
	mean2 = np.mean(polyphen_scores2)
	if polyphen_scores1 != [0]:
		ax.text(0.9, mean1, f"x̄: {mean1:.2f}", ha='right', verticalalignment='center_baseline', fontweight='bold', fontsize=9)
	if polyphen_scores2 != [0]:
		ax.text(1.9, mean2, f"x̄: {mean2:.2f}", ha='right', verticalalignment='center_baseline', fontweight='bold', fontsize=9)

	plt.tight_layout()

	legend_labels1 = []
	legend_counts1 = []
	for category, sub_dict in protein_impacts1_counts.items():
		if 'Polyphen' in category:
			for sub_category, count in sub_dict.items():
				legend_labels1.append(f"{sub_category}: ")
				legend_counts1.append(count)
	# ax.legend(legend_labels, loc='center left', bbox_to_anchor=(0.65, 0.8), edgecolor='black')
	fig.set_size_inches(12, 6)

	height = 0.91
	x_pos = 0.65
	plt.annotate(f"t1 variants = {len(protein_impacts1.keys())}", xy=(x_pos, height), bbox=dict(facecolor='white', edgecolor='none'), xycoords='axes fraction',
				 fontsize=10, fontweight='bold')
	height = height - 0.06
	if legend_counts1 == [] or all(label == 0 for label in legend_counts1):
		ax.annotate(f"Caution: no Polyphen2 info found at t1!", xy=(x_pos, height), xycoords='axes fraction', fontsize=10, fontstyle='italic', color='red')
		height = height - 0.08
	else:
		height = height - 0.02
		ax.annotate("t1 scores\n", xy=(x_pos, height), xycoords='axes fraction', fontsize=10, fontweight='bold')
		height = height - 0.04

	pred_print = 0
	first_not_found = False
	for label, count in zip(legend_labels1, legend_counts1):
		if label == '-1-0.45: ':
			label = '[0-0.45]: '
		if label == '0.45-0.95: ':
			label = '[0.45-0.95]: '
		if label == '0.95-1.5: ':
			label = '[0.95-1]: '
		if 'damaging' not in label and 'benign' not in label and not first_not_found:
			text_color = 'black'
			if label == '[0-0.45]: ':
				text_color = 'darkred'
			elif label == '[0.45-0.95]: ':
				text_color = 'darkorange'
			elif label == '[0.95-1]: ':
				text_color = 'darkgreen'
			elif 'not found' in label:
				text_color = 'black'
				first_not_found = True
			ratio = round(count / n_polyphens1 * 100)
			if ratio > 0:
				plt.annotate(f"{label}{ratio}%\n", xy=(0.65, height), xycoords='axes fraction', fontsize=10, bbox=dict(facecolor='white', edgecolor='none'), color=text_color)
				height = height - 0.04
		else:
			if pred_print == 0:
				plt.annotate(f"t1 predictions:\n", xy=(0.65, height - 0.005), xycoords='axes fraction', fontsize=10, bbox=dict(facecolor='white', edgecolor='none'), fontweight='bold')
				height = height - 0.05
			if 'probably' in label.lower():
				text_color = 'darkgreen'
			elif 'possibly' in label.lower():
				text_color = 'darkorange'
			elif 'benign' in label.lower():
				text_color = 'darkred'
			elif 'not found' in label:
				text_color = 'black'
			else:
				text_color = 'black'
			ratio = round(count / n_polyphens1 * 100)
			if ratio > 0:
				plt.annotate(f"{label}{ratio}%\n", xy=(0.65, height), xycoords='axes fraction', fontsize=10, bbox=dict(facecolor='white', edgecolor='none'), color=text_color)
				height = height - 0.04
			pred_print += 1

	pred_print = 0
	height = height + 0.03
	ax.annotate(f"--------------------------------", xy=(x_pos, height), xycoords='axes fraction', fontsize=13)
	height = height - 0.04
	legend_labels2 = []
	legend_counts2 = []
	for category, sub_dict in protein_impacts2_counts.items():
		if 'Polyphen' in category:
			for sub_category, count in sub_dict.items():
				legend_labels2.append(f"{sub_category}: ")
				legend_counts2.append(count)

	plt.annotate(f"t2 variants = {len(protein_impacts2.keys())}", xy=(x_pos, height), bbox=dict(facecolor='white', edgecolor='none'), xycoords='axes fraction',
				 fontsize=10, fontweight='bold')
	height = height - 0.07
	if legend_counts2 == [] or all(label == 0 for label in legend_counts2):
		ax.annotate(f"Caution: no Polyphen2 info found at t2!", xy=(x_pos, height), xycoords='axes fraction', fontsize=10, fontstyle='italic', color='red')
		height = height - 0.08
	else:
		height = height - 0.005
		plt.annotate("t2 scores\n", xy=(x_pos, height), xycoords='axes fraction', bbox=dict(facecolor='white', edgecolor='none'), fontsize=10, fontweight='bold')
		height = height - 0.04

	pred_print = 0
	first_not_found = False
	for label, count in zip(legend_labels2, legend_counts2):
		if label == '-1-0.45: ':
			label = '[0-0.45]: '
		if label == '0.45-0.95: ':
			label = '[0.45-0.95]: '
		if label == '0.95-1.5: ':
			label = '[0.95-1]: '
		if 'damaging' not in label and 'benign' not in label and not first_not_found:
			text_color = 'black'
			if label == '[0-0.45]: ':
				text_color = 'darkred'
			elif label == '[0.45-0.95]: ':
				text_color = 'darkorange'
			elif label == '[0.95-1]: ':
				text_color = 'darkgreen'
			elif 'not found' in label:
				text_color = 'black'
				first_not_found = True
			ratio = round(count / n_polyphens2 * 100)
			plt.annotate(f"{label}{ratio}%\n", xy=(0.65, height), xycoords='axes fraction', fontsize=10, bbox=dict(facecolor='white', edgecolor='none'), color=text_color)
		else:
			if pred_print == 0:
				plt.annotate(f"t2 predictions:\n", xy=(0.65, height - 0.01), xycoords='axes fraction', fontsize=10, bbox=dict(facecolor='white', edgecolor='none'), fontweight='bold')
				height = height - 0.05
			if 'probably' in label.lower():
				text_color = 'darkgreen'
			elif 'possibly' in label.lower():
				text_color = 'darkorange'
			elif 'benign' in label.lower():
				text_color = 'darkred'
			elif 'not found' in label:
				text_color = 'black'
			else:
				text_color = 'black'
			ratio = round(count / n_polyphens2 * 100)
			if ratio > 0:
				plt.annotate(f"{label}{ratio}%\n", xy=(0.65, height), xycoords='axes fraction', fontsize=10, color=text_color)
			pred_print += 1
		height = height - 0.04

	plt.axhline(y=polyphen_threshold, color='darkred', linestyle='--')
	ax = plt.gca()
	custom_padding = max(max(polyphen_scores1), max(polyphen_scores2)) - min(min(polyphen_scores1), min(polyphen_scores2))
	ax.annotate(str(polyphen_threshold), xy=(0.008, polyphen_threshold + custom_padding), xycoords=('axes fraction', 'data'), fontsize=9, color='darkred')

	output_path = dict_para['output_path_comparison'] + 'Polyphen2_protein_impacts'

	if 'PNG' in dict_para['C_protein_impacts_format(s)'].upper():
		plt.savefig(output_path + '.png', dpi=400)
	if 'PDF' in dict_para['C_protein_impacts_format(s)'].upper():
		plt.savefig(output_path + '.pdf', dpi=400)
	if 'SVG' in dict_para['C_protein_impacts_format(s)'].upper():
		plt.savefig(output_path + '.svg', dpi=400)
	if 'JPG' in dict_para['C_protein_impacts_format(s)'].upper():
		plt.savefig(output_path + '.jpg', dpi=400)

	# only change variants !
	# categories = []
	# values = []
	#
	# for key in dict_impacts.keys():
	# 	sub_dic = dict_impacts[key]
	# 	for k in sub_dic.keys():
	# 		categories.append((key + k))
	# 		values.append(sub_dic[k])
	#
	# true_categories = []
	# for cat in categories:
	# 	if 'SIFT scores counts' in cat:
	# 		category = 'n SIFT (' + cat.split('SIFT scores counts')[1] + ')'
	# 		true_categories.append(category)
	# 	elif 'Polyphen2 scores counts' in cat:
	# 		category = 'n Polyphen2 (' + cat.split('Polyphen2 scores counts')[1] + ')'
	# 		true_categories.append(category)
	# 	elif 'SIFT predictions' in cat:
	# 		category = 'SIFT (' + cat.split('SIFT predictions')[1] + ')'
	# 		true_categories.append(category)
	# 	elif 'Polyphen2 predictions' in cat:
	# 		category = 'Polyphen2 (' + cat.split('Polyphen2 predictions')[1] + ')'
	# 		true_categories.append(category)
	#
	# sift_color = '#8B0000'
	# polyphen2_color = '#00008B'
	#
	# fig, ax = plt.subplots()
	# bars = ax.bar(true_categories, values)
	#
	# for i, category in enumerate(true_categories):
	# 	if 'SIFT' in category:
	# 		bars[i].set_color(sift_color)
	# 	elif 'Polyphen2' in category:
	# 		bars[i].set_color(polyphen2_color)
	#
	# for bar in bars:
	# 	height = bar.get_height()
	# 	ax.text(bar.get_x() + bar.get_width() / 2, height, str(int(height)), ha='center', va='bottom')
	#
	# ax.set_ylim(0, max(values) * 1.1)  # Adjust the multiplier (1.2) to control the range
	# ax.set_ylabel('Count', labelpad=15)
	# ax.set_title('Mutation Analysis')
	#
	# plt.xticks(rotation=45, ha='right', fontsize=10)
	# plt.subplots_adjust(bottom=0.4)
	# fig.set_size_inches(10, 6)
	#
	# compare_path = dict_para['output_path_comparison'].split('.vcf')[0]
	# if 'CHANGE' in dict_para['variants_selection_approach'].upper():
	# 	plt.title("Protein impact for each mutation that appeared or disappeared \nbetween times, found in patient " + pair_id, pad=12, fontsize=12)
	#
	# if 'PNG' in dict_para['C_protein_impacts_format(s)'].upper():
	# 	plt.savefig(compare_path.split('stats')[0] + 'protein_impacts.png', dpi=400)
	# if 'SVG' in dict_para['C_protein_impacts_format(s)'].upper():
	# 	plt.savefig(compare_path.split('stats')[0] + 'protein_impacts.svg', dpi=400)
	# if 'PDF' in dict_para['C_protein_impacts_format(s)'].upper():
	# 	plt.savefig(compare_path.split('stats')[0] + 'protein_impacts.pdf', dpi=400)
	# if 'JPG' in dict_para['C_protein_impacts_format(s)'].upper():
	# 	plt.savefig(compare_path.split('stats')[0] + 'protein_impacts.jpg', dpi=400)
	#
	# plt.close()


def write_stats(dict_para, df_variants, dict_variants):
	out_stats = dict_para['output_path_comparison'] + dict_para['variants_selection_approach'] + 'd_variants_stats.txt'
	df_variants.set_index('Variant', inplace=True)
	variants_dict = df_variants.to_dict(orient='index')
	t1_dict = {}
	t2_dict = {}
	mutation_counts = {'synonymous_SNV': 0, 'nonsynonymous_SNV': 0, 'stopgain': 0, 'startloss': 0, 'stoploss': 0, 'nonframeshift_insertion': 0,
					   'frameshift_insertion': 0, 'nonframeshift_deletion': 0, 'frameshift_deletion': 0,
					   'nonframeshift_substitution': 0, 'frameshift_substitution': 0}
	chromosome_counts = {'chr1': 0, 'chr2': 0, 'chr3': 0, 'chr4': 0, 'chr5': 0, 'chr6': 0, 'chr7': 0, 'chr8': 0, 'chr9': 0, 'chr10': 0,
						 'chr11': 0, 'chr12': 0, 'chr13': 0, 'chr14': 0, 'chr15': 0, 'chr16': 0, 'chr17': 0, 'chr18': 0, 'chr19': 0, 'chr20': 0,
						 'chr21': 0, 'chr22': 0, 'chrX': 0, 'chrY': 0, 'chrM': 0}
	gene_counts = {}

	config_path = 'config.txt'
	with open(config_path, "r") as file:
		lines = file.readlines()
		for line in lines:
			if "max_SIFT_score" in line:
				sift_score_threshold = float(line.split("max_SIFT_score =")[1].strip())
			elif "min_PolyPhen2_score" in line:
				polyphen_threshold = float(line.split("min_PolyPhen2_score =")[1].strip())

	sift_thresholds = [-1, 0.05, 0.2, 1.5]
	polyphen_thresholds = [-1, 0.45, 0.95, 1.5]

	dict_impacts_counts = {
		'SIFT scores counts': {f'{sift_thresholds[i]}-{sift_thresholds[i + 1]}': 0 for i in range(len(sift_thresholds) - 1)},
		'Polyphen2 scores counts': {f'{polyphen_thresholds[i]}-{polyphen_thresholds[i + 1]}': 0 for i in range(len(polyphen_thresholds) - 1)},
		'SIFT predictions': {},
		'Polyphen2 predictions': {}
	}

	for impact in dict_variants.values():
		try:
			sift_score = impact['SIFT score']
		except:
			sift_score = 'not found'
		try:
			polyphen_score = impact['PolyPhen2 score']
		except:
			polyphen_score = 'not found'
		try:
			sift_pred = impact['SIFT pred']
		except:
			sift_pred = 'not found'
		try:
			polyphen_pred = impact['PolyPhen2 pred']
		except:
			polyphen_pred = 'not found'

		if sift_score == 'not found':
			if sift_score not in dict_impacts_counts['SIFT scores counts']:
				dict_impacts_counts['SIFT scores counts'][sift_score] = 1
			else:
				dict_impacts_counts['SIFT scores counts'][sift_score] += 1
		else:
			interval = None
			for i in range(len(sift_thresholds) - 1):
				if sift_thresholds[i] <= sift_score < sift_thresholds[i + 1]:
					interval = f'{sift_thresholds[i]}-{sift_thresholds[i + 1]}'
					break
			dict_impacts_counts['SIFT scores counts'][interval] += 1

		if polyphen_score == 'not found':
			if polyphen_score not in dict_impacts_counts['Polyphen2 scores counts']:
				dict_impacts_counts['Polyphen2 scores counts'][polyphen_score] = 1
			else:
				dict_impacts_counts['Polyphen2 scores counts'][polyphen_score] += 1
		else:
			interval = None
			for i in range(len(polyphen_thresholds) - 1):
				if polyphen_thresholds[i] <= polyphen_score < polyphen_thresholds[i + 1]:
					interval = f'{polyphen_thresholds[i]}-{polyphen_thresholds[i + 1]}'
					break
			dict_impacts_counts['Polyphen2 scores counts'][interval] += 1

		dict_impacts_counts['SIFT predictions'].setdefault(sift_pred, 0)
		dict_impacts_counts['SIFT predictions'][sift_pred] += 1
		dict_impacts_counts['Polyphen2 predictions'].setdefault(polyphen_pred, 0)
		dict_impacts_counts['Polyphen2 predictions'][polyphen_pred] += 1

	dict_impacts_counts['SIFT scores counts'] = {key: value for key, value in dict_impacts_counts['SIFT scores counts'].items() if value != 0}
	dict_impacts_counts['Polyphen2 scores counts'] = {key: value for key, value in dict_impacts_counts['Polyphen2 scores counts'].items() if value != 0}
	dict_impacts_counts['SIFT predictions'] = {key: value for key, value in dict_impacts_counts['SIFT predictions'].items() if value != 0}
	sift_preds_mapping = {'D': 'damaging', 'T': 'tolerated', 'U': 'unknown'}
	dict_impacts_counts['SIFT predictions'] = {sift_preds_mapping.get(key, key): value for key, value in dict_impacts_counts['SIFT predictions'].items() if value != 0}
	dict_impacts_counts['Polyphen2 predictions'] = {key: value for key, value in dict_impacts_counts['Polyphen2 predictions'].items() if value != 0}
	dict_impacts_counts['Polyphen2 predictions'] = {('deleterious' if key == 'D' else key): value for key, value in dict_impacts_counts['Polyphen2 predictions'].items() if
													value != 0}
	polyphen_preds_mapping = {'B': 'benign', 'D': 'probably damaging', 'P': 'possibly damaging'}
	dict_impacts_counts['Polyphen2 predictions'] = {polyphen_preds_mapping.get(key, key): value for key, value in dict_impacts_counts['Polyphen2 predictions'].items() if
													value != 0}

	for variant_id, variant_info in variants_dict.items():
		if variant_info['time'] == 't1':
			t1_dict[variant_id] = variant_info
		elif variant_info['time'] == 't2':
			t2_dict[variant_id] = variant_info
		for key in mutation_counts.keys():
			try:
				if key == variant_info['Mutation']:
					mutation_counts[key] += 1
			except:
				# print('not annotated by ANNOVAR')
				pass
		for key in chromosome_counts.keys():
			if key == variant_info['chr']:
				chromosome_counts[key] += 1
		if variant_info['gene(s)'] not in gene_counts.keys():
			gene_counts[variant_info['gene(s)']] = 1
		else:
			gene_counts[variant_info['gene(s)']] += 1
	sorted_genes = sorted(gene_counts.items(), key=lambda x: x[1], reverse=True)
	sorted_chromosomes = sorted(chromosome_counts.items(), key=lambda x: x[1], reverse=True)

	n_t1 = len(t1_dict.keys())
	n_t2 = len(t2_dict.keys())
	avg_DP = round(sum([float(variant_info['DP'])for variant_info in variants_dict.values()]) / len(variants_dict.keys()))
	valid_AD_values = [float(variant_info['AD']) for variant_info in variants_dict.values() if variant_info['AD'] != '']
	try:
		avg_AD = round(sum(valid_AD_values) / len(valid_AD_values))
	except:
		print('no ad in df_variants')

	try:  # ANNOVAR
		avg_QUAL = round(sum([float(variant_info['QUALITY'])for variant_info in variants_dict.values()]) / len(variants_dict.keys()))
		avg_SB = round(sum([float(variant_info['SB'])for variant_info in variants_dict.values()]) / len(variants_dict.keys()), 2)
	except:
		# print('not annotated by ANNOVAR')
		pass
	try:
		avg_VAF_sample = round(sum([float(variant_info['VAF sample'])for variant_info in variants_dict.values()]) / len(variants_dict.keys()), 2)
	except:
		# print('no VAF sample found')
		pass

	total_vaf = 0
	total_variants = 0

	for variant_info in variants_dict.values():
		try:
			total_vaf += float(variant_info['VAF pop'])
			total_variants += 1
		except:
			pass

	avg_VAF_pop = round(total_vaf / total_variants, 6)

	with open(out_stats, 'w') as o:
		if int(n_t1) == 0 or int(n_t1) == 1:
			plural_t1 = ''
		else:
			plural_t1 = 's'
		if int(n_t2) == 0 or int(n_t2) == 1:
			plural_t2 = ''
		else:
			plural_t2 = 's'
		o.write(f'{n_t1} mutation{plural_t1} disappeared and {n_t2} mutation{plural_t2} appeared between 2 times.\n\n')
		o.write('--- Mutation subtypes ---\n')
		if mutation_counts['synonymous_SNV'] > 0:
			o.write(f'synonymous SNV: {mutation_counts["synonymous_SNV"]}\t')
		if mutation_counts['nonsynonymous_SNV'] > 0:
			o.write(f'nonsynonymous SNV: {mutation_counts["nonsynonymous_SNV"]}\n')
		if mutation_counts['stopgain'] > 0:
			o.write(f'stopgain: {mutation_counts["stopgain"]}\t')
		if mutation_counts['startloss'] > 0:
			o.write(f'startloss: {mutation_counts["startloss"]}\t')
		if mutation_counts['stoploss'] > 0:
			o.write(f'stoploss: {mutation_counts["stoploss"]}\n')
		if mutation_counts['nonframeshift_insertion'] > 0:
			o.write(f'nonframeshift insertion: {mutation_counts["nonframeshift_insertion"]}\t')
		if mutation_counts['frameshift_insertion'] > 0:
			o.write(f'frameshift insertion: {mutation_counts["frameshift_insertion"]}')
		if mutation_counts['nonframeshift_deletion'] > 0:
			o.write(f'\nnonframeshift deletion: {mutation_counts["nonframeshift_deletion"]}\t')
		if mutation_counts['frameshift_deletion'] > 0:
			o.write(f'frameshift deletion: {mutation_counts["frameshift_deletion"]}')
		if mutation_counts['nonframeshift_substitution'] > 0:
			o.write(f'\nnonframeshift substitution: {mutation_counts["nonframeshift_substitution"]}\t')
		if mutation_counts['frameshift_substitution'] > 0:
			o.write(f'frameshift substitution: {mutation_counts["frameshift_substitution"]}')

		o.write('\n\n--- Characteristics ---\n')
		o.write(f'Average DP: {avg_DP}\t')
		try:
			o.write(f'\tAverage AD: {avg_AD}\n')
		except:
			pass
		try:
			o.write(f'Average QUAL: {avg_QUAL}\t')
			o.write(f'Average SB: {avg_SB}\n')
		except:
			# print('not annotated by ANNOVAR')
			pass
		try:
			o.write(f'Average VAF sample: {avg_VAF_sample}\t')
		except:
			print('no VAF sample found')
		try:
			o.write(f'   Average VAF pop: {avg_VAF_pop}')
		except:
			# print('no VAF pop found')
			pass

		o.write(f'\n\n----- Impact on protein -----\n')
		sift_scores = str(dict_impacts_counts["SIFT scores counts"]).replace("'", "").replace("{", "").replace("}", "").replace(":", " =")
		sift_preds = str(dict_impacts_counts["SIFT predictions"]).replace("'", "").replace("{", "").replace("}", "").replace(":", " =")
		polyphen_scores = str(dict_impacts_counts["Polyphen2 scores counts"]).replace("'", "").replace("{", "").replace("}", "").replace(":", " =")
		polyphen_preds = str(dict_impacts_counts["Polyphen2 predictions"]).replace("'", "").replace("{", "").replace("}", "").replace(":", " =")
		o.write(f'SIFT scores counts: \n{sift_scores}\n\n')
		o.write(f'SIFT predictions: \n{sift_preds}\n\n')
		o.write(f'Polyphen2 scores counts: \n{polyphen_scores}\n\n')
		o.write(f'Polyphen predictions: \n{polyphen_preds}')

		o.write('\n\n--- Locations ---\n')
		for chr, count in sorted_chromosomes:
			if count > 0:
				o.write(f'{chr}: {count}\n')

		o.write('\n--- Mutated genes ---\n')
		for gene, count in sorted_genes:
			o.write(f'{gene}: {count}\n')

		return dict_impacts_counts


def compare_vcf(out_gene, dict_para, gene_name_dico: {}, gene_id_dico: {}, transcript_dico: {}, infos: str, enrichment: bool, df_dataset, logger):
	'''
	Compare each vcf files with the n-1 vcf file (starting from the second file)
	Input : vcf files (pass and filtered from the filter module), boolean to know if GOEA need to be done and the logger
	'''

	time1_column = []
	time2_column = []
	pair_names_column = []

	for value in df_dataset[dict_para['time1_column_name']]:
		if pd.notnull(value) and value not in ('NaN', 'naN') and value.strip() != '':
			time1_column.append(value)

	for value in df_dataset[dict_para['time2_column_name']]:
		if pd.notnull(value) and value not in ('NaN', 'naN') and value.strip() != '':
			time2_column.append(value)

	for value in df_dataset[dict_para['pair_names_column_name']]:
		if pd.notnull(value) and value not in ('NaN', 'naN') and str(value).strip() != '':
			try:
				pair_names_column.append(str(int(value)))
			except:
				pair_names_column.append(str(value))

	time1_column = [element.replace('.vcf', '') for element in time1_column]
	time2_column = [element.replace('.vcf', '') for element in time2_column]
	dict_pairs = {}
	for i in range(len(time1_column)):
		dict_pairs[time1_column[i] + '___' + time2_column[i]] = pair_names_column[i]

	vcf_pass_files = []
	vcf_pass_files.append(dict_para['output_path'] + 'samples/' + dict_para['passed1'].split('passed')[0].split('samples/')[1].replace('/', "").replace(".vcf","") + '/' +
						  dict_para['passed1'].split('passed')[0].split('samples/')[1].replace('/', "").replace('/', "").replace(".vcf","") + '_passed.vcf')
	vcf_pass_files.append(dict_para['output_path'] + 'samples/' + dict_para['passed2'].split('passed')[0].split('samples/')[1].replace('/', "").replace(".vcf","") + '/' +
						  dict_para['passed2'].split('passed')[0].split('samples/')[1].replace('/', "").replace(".vcf","") + '_passed.vcf')
	if dict_para['keep_filtered_vcf_after_run'].upper() == 'YES' or dict_para['keep_filtered_vcf_after_run'].upper() == 'TRUE':
		vcf_filtered_files = []
		vcf_filtered_files.append(
			dict_para['filtered1'].split('filtered.')[0] + dict_para['filtered1'].split('filtered.')[0].split('samples/')[1].replace('/', "") + '_filtered.vcf')
		vcf_filtered_files.append(
			dict_para['filtered2'].split('filtered.')[0] + dict_para['filtered2'].split('filtered.')[0].split('samples/')[1].replace('/', "") + '_filtered.vcf')

	# Get genes infos if not None
	if infos:
		infos_df = get_informations_for_genes(dict_para, infos, logger)

	print('Creating comparison genes and variants mutations tables, and VCF file ...')

	if dict_para['verbose_prints'].upper() == 'TRUE':
		print('Save genes lists from comparisons...')

	color = '\033[38;2;255;215;0m'
	change_counts = []
	common_counts = []

	for num in range(1, len(vcf_pass_files)):  # Comparison file by file starting from the second with is n-1
		logger.info(f'Start comparing {true_stem(vcf_pass_files[num - 1])} and {true_stem(vcf_pass_files[num])} !')
		variants_pass_save = set()
		variants_pass_save2 = set()
		variants_filtered_save = set()
		variants_filtered_save2 = set()
		variants_pass_save = get_variants(vcf_pass_files[num - 1], variants_pass_save)
		variants_pass_save2 = get_variants(vcf_pass_files[num], variants_pass_save2)
		keep_vcf = False
		if dict_para['keep_filtered_vcf_after_run'].upper() == 'YES' or dict_para['keep_filtered_vcf_after_run'].upper() == 'TRUE':
			variants_filtered_save = get_variants(vcf_filtered_files[num - 1], variants_filtered_save)
			variants_filtered_save2 = get_variants(vcf_filtered_files[num], variants_filtered_save2)
			keep_vcf = True

		# Get variants sets specific to each file and separate this sets in weak and strong variants using the filtered file from the other vcf
		croissant1 = variants_pass_save - variants_pass_save2  # variants from pass 1 without pass 2
		croissant2 = variants_pass_save2 - variants_pass_save  # variants from pass 2 without pass 1
		if dict_para['keep_filtered_vcf_after_run'].upper() == 'YES' or dict_para['keep_filtered_vcf_after_run'].upper() == 'TRUE':
			strong1 = croissant1 - variants_filtered_save2  # variants from pass 1 without pass 2 and filter 2
			strong2 = croissant2 - variants_filtered_save  # variants from pass 2 without pass 1 and filter 1
			weak1 = croissant1.intersection(variants_filtered_save2)  # variants from pass 1 without pass 2 but present in filter 2
			weak2 = croissant2.intersection(variants_filtered_save)  # variants from pass 2 without pass 1 but present in filter 1

		# Creation of the dataframe containing impacted gene and creation of the new vcf files containing the genes type ('strong'/'weak'/'common')
		dfGenes = pd.DataFrame(columns=['Gene', 'Chromosome', 'Gene start position', 'Gene end position', 'Unique variants', 'Total variants', 'Variation (%)',
										'|Variants delta|', 'Mutations (t1)', 'g.(t1)', 'c.(t1)', 'p.(t1)', 'Mutations (t2)', 'g.(t2)',
										'c.(t2)', 'p.(t2)'])
		chromosomes_list = []
		gene_position_list = []
		weakness_list = []
		charge_list = []
		dic_unique_variants = {}
		dic_variants_delta = {}
		tumor1_list = []
		tumor2_list = []
		genomic_variant_annotation_list = []
		coding_variant_annotation_list = []
		proteomic_variant_annotation_list = []
		genomic_variant2_annotation_list = []
		coding_variant2_annotation_list = []
		proteomic_variant2_annotation_list = []
		dic_anno_info = {}

		if keep_vcf:
			genes1 = modify_variants_pass_and_get_genes(dict_para, vcf_pass_files[num - 1], weak1, strong1, gene_name_dico, transcript_dico, str(Path(out_gene).resolve().parent))
			genes2 = modify_variants_pass_and_get_genes(dict_para, vcf_pass_files[num], weak2, strong2, gene_name_dico, transcript_dico, str(Path(out_gene).resolve().parent))
		else:
			genes1 = modify_variants_pass_and_get_genes(dict_para, vcf_pass_files[num - 1], 'no', 'no', gene_name_dico, transcript_dico, str(Path(out_gene).resolve().parent))
			genes2 = modify_variants_pass_and_get_genes(dict_para, vcf_pass_files[num], 'no', 'no', gene_name_dico, transcript_dico, str(Path(out_gene).resolve().parent))

		selection = dict_para['variants_selection_approach'].upper()
		union = False
		change = False
		common = False
		if selection == 'UNION':
			union = True
			genes_list = list(set(genes1.keys()).union(set(genes2.keys())))
		elif selection == 'CHANGE':
			change = True
			genes_list = []
			total_count = 0
			unique_genes = []
			for gene in genes1:
				if gene not in unique_genes:
					unique_genes.append(gene)
					total_count += 1
				# print('gene : ' + str(gene))
				# print('1 : ' + str(genes1[gene]['mg']))
				# try:
				# 	print('2 : ' + str(genes2[gene]['mg']))
				# except:
				# 	print('not in genes2')
				# print(genes_list)
				if ',' in str(genes1[gene]['mg']):
					l_mutations = genes1[gene]['mg']
					for mutation in l_mutations:
						try:
							if mutation not in str(genes2[gene]['mg']):
								if not gene in genes_list:
									genes_list.append(gene)
						except:
							# print('not in genes2')
							if not gene in genes_list:
								genes_list.append(gene)
				try:
					if str(genes1[gene]['mg']) not in str(genes2[gene]['mg']):
						if not gene in genes_list:
							genes_list.append(gene)
				except:
					# print('not in genes2')
					if not gene in genes_list:
						genes_list.append(gene)

			for gene in genes2:
				if gene not in unique_genes:
					unique_genes.append(gene)
					total_count += 1
				# print('gene : ' + str(gene))
				# try:
				# 	print('2 : ' + str(genes1[gene]['mg']))
				# except:
				# 	print('not in genes2')
				# print('1 : ' + str(genes2[gene]['mg']))
				# print(genes_list)
				if ',' in str(genes2[gene]['mg']):
					l_mutations = genes2[gene]['mg']
					for mutation in l_mutations:
						try:
							if mutation not in str(genes1[gene]['mg']):
								if not gene in genes_list:
									genes_list.append(gene)
						except:
							# print('not in genes1')
							if not gene in genes_list:
								genes_list.append(gene)
				try:
					if str(genes2[gene]['mg']) not in str(genes1[gene]['mg']):
						if not gene in genes_list:
							genes_list.append(gene)
				except:
					# print('not in genes1')
					if not gene in genes_list:
						genes_list.append(gene)

			change_counts.append(total_count)

		elif selection == 'COMMON':
			common = True
			genes_list = []
			total_count = 0
			unique_genes = []
			for gene in genes1:
				if gene not in unique_genes:
					unique_genes.append(gene)
					total_count += 1
				try:
					if str(genes1[gene]['mg']) == str(genes2[gene]['mg']):
						if not gene in genes_list:
							genes_list.append(gene)
				except:
					pass
			# print('not in genes2')
			for gene in genes2:
				if gene not in unique_genes:
					unique_genes.append(gene)
					total_count += 1
				try:
					if str(genes1[gene]['mg']) == str(genes2[gene]['mg']):
						if not gene in genes_list:
							genes_list.append(gene)
				except:
					pass
			common_counts.append(total_count)
		# print('not in genes1')

		dic_mutation_types_t1 = {'SNP': 0, 'DNP': 0, 'TNP': 0, 'ONP': 0, 'INDEL': 0, 'INSERTION': 0, 'DELETION': 0}
		dic_mutation_types_t2 = {'SNP': 0, 'DNP': 0, 'TNP': 0, 'ONP': 0, 'INDEL': 0, 'INSERTION': 0, 'DELETION': 0}
		dic_mutation_subtypes_t1 = {'synonymous SNV': 0, 'nonsynonymous SNV': 0, 'frameshift substitution': 0,
									'non frameshift substitution': 0, 'stopgain': 0, 'stoploss': 0,
									'startloss': 0, 'frameshift insertion': 0,
									'non frameshift insertion': 0, 'frameshift deletion': 0,
									'non frameshift deletion': 0, 'unknown': 0}
		dic_mutation_subtypes_t2 = {'synonymous SNV': 0, 'nonsynonymous SNV': 0, 'frameshift substitution': 0,
									'non frameshift substitution': 0, 'stopgain': 0, 'stoploss': 0,
									'startloss': 0, 'frameshift insertion': 0,
									'non frameshift insertion': 0, 'frameshift deletion': 0,
									'non frameshift deletion': 0, 'unknown': 0}

		# unique variants counting for each gene found in at least one of the two samples of current pair
		for gene in genes1:
			# Mutation types and subtypes counting
			try:  # annovar
				dic_mutation_types_t1, dic_mutation_subtypes_t1 = add_mutation_type(dic_mutation_types_t1, dic_mutation_subtypes_t1, genes1, gene)
			except:
				# print('not annotated with annovar')
				pass

			try:
				if gene not in dic_unique_variants.keys():
					dic_unique_variants[gene] = {}
					dic_unique_variants[gene]['t1'] = []
				for variant in genes1[gene]['mg']:
					if variant not in dic_unique_variants[gene]['t1']:
						dic_unique_variants[gene]['t1'].append(variant)
			except:
				pass
		for gene in genes2:
			try:  # annovar
				dic_mutation_types_t2, dic_mutation_subtypes_t2 = add_mutation_type(dic_mutation_types_t2, dic_mutation_subtypes_t2, genes2, gene)
			except:
				# print('not annotated with annovar')
				pass
			try:
				if gene not in dic_unique_variants.keys():
					dic_unique_variants[gene] = {}
					dic_unique_variants[gene]['t2'] = []
				if 't2' not in dic_unique_variants[gene].keys():
					dic_unique_variants[gene]['t2'] = []
				for variant in genes2[gene]['mg']:
					if variant not in dic_unique_variants[gene]['t2']:
						dic_unique_variants[gene]['t2'].append(variant)
			except:
				pass

		if change or common:
			dic_unique_variants_copy = dic_unique_variants.copy()
			for key in dic_unique_variants_copy.keys():
				if key not in genes_list:
					del dic_unique_variants[key]

		for gene in genes_list:
			# |Variants delta| counting for each gene found in at least one of the two samples of current pair
			if gene in genes1.keys() and gene in genes2.keys():
				unique_set = set(genes1[gene]['mg']).union(set(genes2[gene]['mg']))
				unique_set1 = set(genes1[gene]['mg'])
				unique_set2 = set(genes2[gene]['mg'])
				dic_variants_delta[gene] = len(unique_set2) - len(unique_set1)
				dic_unique_variants[gene]['set'] = unique_set

			elif gene in genes1.keys() and not gene in genes2.keys() and not common:
				try:
					dic_variants_delta[gene] = -len(genes1[gene]['mg'])
					dic_unique_variants[gene]['set'] = set(genes1[gene]['mg'])
				except:
					pass
			elif gene in genes2.keys() and not gene in genes1.keys() and not common:
				try:
					dic_variants_delta[gene] = len(genes2[gene]['mg'])
					dic_unique_variants[gene]['set'] = set(genes2[gene]['mg'])
				except:
					pass

			try:  # ANNOVAR
				if gene in genes1.keys() and gene in genes2.keys():
					chromosomes_list.append(genes1[gene]['chr'])
					gene_position_list.append(genes1[gene]['gene_pos'])
					if keep_vcf:
						charge_list.append(genes1[gene]['weak'] + genes2[gene]['weak'] + genes1[gene]['strong'] + genes2[gene][
							'strong'] + genes1[gene]['common'] + genes2[gene]['common'])
						tumor1_list.append(genes1[gene]['weak'] + genes1[gene]['strong'] + genes1[gene]['common'])
						tumor2_list.append(genes2[gene]['weak'] + genes2[gene]['strong'] + genes2[gene]['common'])
					else:
						charge_list.append(genes1[gene]['all'])
						tumor1_list.append(genes1[gene]['all'])
						tumor2_list.append(0)
					dic_anno_info[gene] = []
					try:
						# if len(genes1[gene]['anno_info']) > 15:
						# 	print(genes1[gene]['anno_info'])
						for info in genes1[gene]['anno_info']:
							dic_anno_info[gene].append(info)
						for info in genes2[gene]['anno_info']:
							dic_anno_info[gene].append(info)
					except:
						print('error anno info')

					genomic_variant_annotation_list.append('|'.join(genes1[gene]['mg']))
					coding_variant_annotation_list.append('|'.join(genes1[gene]['mc']))
					proteomic_variant_annotation_list.append('|'.join(genes1[gene]['mp']))
					genomic_variant2_annotation_list.append('|'.join(genes2[gene]['mg']))
					coding_variant2_annotation_list.append('|'.join(genes2[gene]['mc']))
					proteomic_variant2_annotation_list.append('|'.join(genes2[gene]['mp']))

				elif gene in genes1.keys() and not common:
					chromosomes_list.append(genes1[gene]['chr'])
					gene_position_list.append(genes1[gene]['gene_pos'])
					if keep_vcf:
						charge_list.append(genes1[gene]['weak'] + genes1[gene]['strong'] + genes1[gene]['common'])
						tumor1_list.append(genes1[gene]['weak'] + genes1[gene]['strong'] + genes1[gene]['common'])
					else:
						charge_list.append(genes1[gene]['all'])
						tumor1_list.append(genes1[gene]['all'])
					dic_anno_info[gene] = []
					for info in genes1[gene]['anno_info']:
						dic_anno_info[gene].append(info)
					tumor2_list.append(0)
					genomic_variant_annotation_list.append('|'.join(genes1[gene]['mg']))
					coding_variant_annotation_list.append('|'.join(genes1[gene]['mc']))
					proteomic_variant_annotation_list.append('|'.join(genes1[gene]['mp']))
					genomic_variant2_annotation_list.append('')
					coding_variant2_annotation_list.append('')
					proteomic_variant2_annotation_list.append('')

				elif gene in genes2.keys() and not common:
					chromosomes_list.append(genes2[gene]['chr'])
					gene_position_list.append(genes2[gene]['gene_pos'])
					tumor1_list.append(0)
					dic_anno_info[gene] = []
					for info in genes2[gene]['anno_info']:
						dic_anno_info[gene].append(info)
					if keep_vcf:
						charge_list.append(genes2[gene]['weak'] + genes2[gene]['strong'] + genes2[gene]['common'])
						tumor2_list.append(genes2[gene]['weak'] + genes2[gene]['strong'] + genes2[gene]['common'])
					else:
						charge_list.append(genes2[gene]['all'])
						tumor2_list.append(genes2[gene]['all'])
					genomic_variant_annotation_list.append('')
					coding_variant_annotation_list.append('')
					proteomic_variant_annotation_list.append('')
					genomic_variant2_annotation_list.append('|'.join(genes2[gene]['mg']))
					coding_variant2_annotation_list.append('|'.join(genes2[gene]['mc']))
					proteomic_variant2_annotation_list.append('|'.join(genes2[gene]['mp']))
			except:  # funcotator
				if gene in genes1.keys() and gene in genes2.keys():
					chromosomes_list.append(genes1[gene]['chr'])
					gene_position_list.append(genes1[gene]['gene_pos'])
					if keep_vcf:
						charge_list.append(genes1[gene]['weak'] + genes2[gene]['weak'] + genes1[gene]['strong'] + genes2[gene]['strong'])
						tumor1_list.append(genes1[gene]['weak'] + genes1[gene]['strong'])
						tumor2_list.append(genes2[gene]['weak'] + genes2[gene]['strong'])
					else:
						charge_list.append(genes1[gene]['all'])
						tumor1_list.append(genes1[gene]['all'])
						tumor2_list.append(genes2['all'])
					genomic_variant_annotation_list.append('|'.join(genes1[gene]['mg']))
					coding_variant_annotation_list.append('|'.join(genes1[gene]['mc']))
					proteomic_variant_annotation_list.append('|'.join(genes1[gene]['mp']))
					genomic_variant2_annotation_list.append('|'.join(genes2[gene]['mg']))
					coding_variant2_annotation_list.append('|'.join(genes2[gene]['mc']))
					proteomic_variant2_annotation_list.append('|'.join(genes2[gene]['mp']))
				elif gene in genes1.keys() and not common:
					chromosomes_list.append(genes1[gene]['chr'])
					gene_position_list.append(genes1[gene]['gene_pos'])
					if keep_vcf:
						charge_list.append(genes1[gene]['weak'] + genes1[gene]['strong'])
						tumor1_list.append(genes1[gene]['weak'] + genes1[gene]['strong'])
					else:
						charge_list.append(genes1[gene]['all'])
						tumor1_list.append(genes1[gene]['all'])
					tumor2_list.append(0)
					genomic_variant_annotation_list.append('|'.join(genes1[gene]['mg']))
					coding_variant_annotation_list.append('|'.join(genes1[gene]['mc']))
					proteomic_variant_annotation_list.append('|'.join(genes1[gene]['mp']))
					genomic_variant2_annotation_list.append('')
					coding_variant2_annotation_list.append('')
					proteomic_variant2_annotation_list.append('')
				elif gene in genes2.keys() and not common:
					chromosomes_list.append(genes2[gene]['chr'])
					gene_position_list.append(genes2[gene]['gene_pos'])
					if keep_vcf:
						charge_list.append(genes2[gene]['weak'] + genes2[gene]['strong'])
						tumor2_list.append(genes2[gene]['weak'] + genes2[gene]['strong'])
					else:
						charge_list.append(genes2[gene]['all'])
						tumor2_list.append(genes2[gene]['all'])
					tumor1_list.append(0)
					genomic_variant_annotation_list.append('')
					coding_variant_annotation_list.append('')
					proteomic_variant_annotation_list.append('')
					genomic_variant2_annotation_list.append('|'.join(genes2[gene]['mg']))
					coding_variant2_annotation_list.append('|'.join(genes2[gene]['mc']))
					proteomic_variant2_annotation_list.append('|'.join(genes2[gene]['mp']))

		dfGenes['Gene'] = genes_list
		dfGenes['Chromosome'] = chromosomes_list
		# some gene names are empty, error out of range -> try/except to replace empty gene name by empty string
		for i in range(len(gene_position_list)):
			if not all(gene_position_list[i]):
				gene_position_list[i] = tuple('empty' if not element else element for element in gene_position_list[i])
		gene_position_start_list = [element[2][0] for element in gene_position_list]
		dfGenes['Gene start position'] = gene_position_start_list
		gene_position_end_list = [element[2][1] for element in gene_position_list]
		dfGenes['Gene end position'] = gene_position_end_list
		for gene in dic_unique_variants.keys():
			try:
				len(dic_unique_variants[gene]['t1'])
			except:
				dic_unique_variants[gene]['t1'] = []
			try:
				len(dic_unique_variants[gene]['t2'])
			except:
				dic_unique_variants[gene]['t2'] = []
			try:
				dfGenes.loc[dfGenes['Gene'] == gene, 'Unique variants'] = len(dic_unique_variants[gene]['set'])
			except:
				pass
			if not common:
				if dic_variants_delta[gene] == 0:
					dfGenes.loc[dfGenes['Gene'] == gene, '|Variants delta|'] = 0
				else:
					dfGenes.loc[dfGenes['Gene'] == gene, '|Variants delta|'] = abs(dic_variants_delta[gene])  # |abs| helps sorting dataframe
			if gene in genes1.keys() and gene in genes2.keys():
				dfGenes.loc[dfGenes['Gene'] == gene, 'Total variants'] = len(genes1[gene]['mg']) + len(genes2[gene]['mg'])
				gene_times_max_variants = max(len(genes2[gene]['mg']), len(genes1[gene]['mg']))
				if not change and not common:
					dfGenes.loc[dfGenes['Gene'] == gene, 'Variation (%)'] = round((len(genes2[gene]['mg']) - len(genes1[gene]['mg'])) / gene_times_max_variants * 100)
			elif gene in genes1.keys() and not common:
				dfGenes.loc[dfGenes['Gene'] == gene, 'Total variants'] = len(genes1[gene]['mg'])
				if not change and not common:
					dfGenes.loc[dfGenes['Gene'] == gene, 'Variation (%)'] = -100
			elif gene in genes2.keys() and not common:
				dfGenes.loc[dfGenes['Gene'] == gene, 'Total variants'] = len(genes2[gene]['mg'])
				if not change:
					dfGenes.loc[dfGenes['Gene'] == gene, 'Variation (%)'] = 100
			try:
				dfGenes.loc[dfGenes['Gene'] == gene, 'Mutations (t1)'] = len(genes1[gene]['mg'])
			except:
				dfGenes.loc[dfGenes['Gene'] == gene, 'Mutations (t1)'] = 0
			try:
				dfGenes.loc[dfGenes['Gene'] == gene, 'Mutations (t2)'] = len(genes2[gene]['mg'])
			except:
				dfGenes.loc[dfGenes['Gene'] == gene, 'Mutations (t2)'] = 0

		dfGenes['g.(t1)'] = genomic_variant_annotation_list
		dfGenes['c.(t1)'] = coding_variant_annotation_list
		dfGenes['p.(t1)'] = proteomic_variant_annotation_list
		dfGenes['g.(t2)'] = genomic_variant2_annotation_list
		dfGenes['c.(t2)'] = coding_variant2_annotation_list
		dfGenes['p.(t2)'] = proteomic_variant2_annotation_list

		dfGenes = dfGenes.sort_values(by=['|Variants delta|', 'Unique variants', 'Variation (%)', 'Gene'], ascending=[False, False, False, True])
		dfGenes = dfGenes[(dfGenes['Gene'] != 'NONE') & (dfGenes['Gene'] != 'Unknown')]  # skipping variants with no corresponding gene name by the annotator
		# dfGenes = dfGenes.set_index('Gene')

		pair_id = dict_para['output_path_comparison'].split('sons/')[1].replace('/', '')
		try:
			out_gene = dict_para['output_path_comparison'] + dict_pairs[pair_id] + '_' + dict_para['variants_selection_approach'] + '_genes'
		except:
			out_gene = dict_para['output_path_comparison'] + pair_id + '_' + dict_para['variants_selection_approach'] + '_genes'

		# Save impacted genes information
		if infos:
			dfGenes = dfGenes.join(infos_df)
			# Drop empty informational columns
			dfGenes = dfGenes.dropna(axis=1, how='all')
		# dfGenes.to_excel('comparison_test.xlsx', index=False)

		logger.info(f'Save genes list in {out_gene}')

		if dict_para['verbose_prints'].upper() == 'TRUE':
			print(f'Save genes list in {out_gene}')
		if 'TSV' in dict_para['C_mutated_genes_table_format(s)'].upper():
			dfGenes.to_csv(out_gene + '.tsv', sep='\t')
		elif 'CSV' in dict_para['C_mutated_genes_table_format(s)'].upper():
			dfGenes.to_csv(out_gene + '.csv', sep=',')
		elif 'XLSX' in dict_para['C_mutated_genes_table_format(s)'].upper():
			dfGenes.to_excel(out_gene + '.xlsx', index=False)
		logger.info(f'Save genes list in {Path(out_gene).with_suffix(".xlsx")}')
		if dict_para['verbose_prints'].upper() == 'TRUE':
			print(f'Save genes list in {Path(out_gene).with_suffix(".xlsx")}')

		# passed vcfs parsing for DP, etc
		try:
			output_path_sample1 = (dict_para['output_path_comparison'].split('___')[0].replace('comparisons', 'samples') +
							   dict_para['output_path_comparison'].split('___')[0].replace('comparisons', 'samples').split('samples')[1] + '_passed.vcf')
			output_path_sample2 = (dict_para['output_path_comparison'].split('___')[0].split('/')[0] + '/samples/' + dict_para['output_path_comparison'].split('___')[1] +
							   dict_para['output_path_comparison'].split('___')[1].replace("/", "_") + 'passed.vcf')
		except:
			reversed_pairs_dict = {value: key for key, value in dict_pairs.items()}
			output_path_sample1 = (dict_para['output_path'] + 'samples/' + reversed_pairs_dict[pair_id].split('___')[0] + '/' +
								   reversed_pairs_dict[pair_id].split('___')[0] + '_passed.vcf')
			output_path_sample2 = (dict_para['output_path'] + 'samples/' + reversed_pairs_dict[pair_id].split('___')[1] + '/' +
								   reversed_pairs_dict[pair_id].split('___')[1] + '_passed.vcf')
			dict_para['output_path_sample1'] = output_path_sample1
			dict_para['output_path_sample2'] = output_path_sample2

		dict_passed1 = {}
		dict_passed2 = {}

		k = 0
		for output_path in [output_path_sample1, output_path_sample2]:
			dict_passed = {}
			with open(output_path, "r") as vcf_file:
				for line in vcf_file:
					if not line.startswith("#"):
						format_names = line.split('\t')[-2].split(':')
						format_values = line.split('\t')[-1].split(':')
						if not 'GT' in format_names:
							format_names = line.split('\t')[-3].split(':')
						line_list = line.split('\t')
						chr_value = line_list[0]
						pos_value = line_list[1]
						ref_value = line_list[3]
						alt_value = line_list[4]
						alt_values = alt_value.split(',')
						for alt_value in alt_values:
							key = 'g.' + chr_value + ':' + pos_value + ref_value + '>' + alt_value

							if key not in dict_passed.keys():
								dict_passed[key] = {}

							try:
								format_names = line.split('\t')[-2]
								if not 'GT' in format_names:
									format_names = line.split('\t')[-3]
								index_VAF_sample = format_names.index('AF')
								format_values = line.split('\t')[-1]
								VAF_sample = format_values[index_VAF_sample]
								dict_passed[key]['VAF sample'] = VAF_sample
							except:
								print('problem with VAF sample finding')

							if 'VAF pop' not in dict_passed[key].keys():
								match_VAF_pop = re.search(r';gnomad40_exome_AF=([^;]+)', line)
								try:
									VAF_pop = match_VAF_pop.group(1)
									dict_passed[key]['VAF pop'] = VAF_pop
								except:
									print('VAF pop not found')

							if 'Mutation' not in dict_passed[key].keys():
								match_subtype = re.search(r';ExonicFunc.refGene=([^;]+)', line)
								try:
									subtype = match_subtype.group(1)
									dict_passed[key]['Mutation'] = subtype
								except:
									print('mutation subtype not found, no ANNOVAR refgene')

							if 'SIFT score' not in dict_passed[key].keys():
								match_SIFT = re.search(r'SIFT_score=([^;]+)', line)
								try:
									SIFT_score = match_SIFT.group(1)
									dict_passed[key]['SIFT score'] = SIFT_score
									sift_score_bool = True
								except:
									print('no SIFT score found')
									sift_score_bool = False
							if 'SIFT pred' not in dict_passed[key].keys():
								match_SIFT_pred = re.search(r'SIFT_pred=([^;]+)', line)
								try:
									SIFT_pred = match_SIFT_pred.group(1)
									dict_passed[key]['SIFT pred'] = SIFT_pred
									sift_pred_bool = True
								except:
									print('no SIFT pred found')
									sift_pred_bool = False
							if 'PolyPhen2 score' not in dict_passed[key].keys():
								match_PolyPhen2 = re.search(r'Polyphen2_HDIV_score=([^;]+)', line)
								try:
									PolyPhen2_score = match_PolyPhen2.group(1)
									dict_passed[key]['PolyPhen2 score'] = PolyPhen2_score
									poly_score_bool = True
								except:
									print('no PolyPhen2 score found')
									poly_score_bool = False
							if 'PolyPhen2 pred' not in dict_passed[key].keys():
								match_PolyPhen2_pred = re.search(r'Polyphen2_HDIV_pred=([^;]+)', line)
								try:
									Polyphen2_pred = match_PolyPhen2_pred.group(1)
									dict_passed[key]['PolyPhen2 pred'] = Polyphen2_pred
									poly_pred_bool = True
								except:
									print('no PolyPhen2 pred found')
									poly_pred_bool = False

							if 'FORMAT' not in dict_passed[key].keys():
								dict_passed[key]['FORMAT'] = []
							dict_passed[key]['FORMAT'].append(line_list[9])
							if 'DP' not in dict_passed[key].keys():
								dict_passed[key]['DP'] = []
							dict_passed[key]['DP'].append(round(int(line_list[7].split('DP=')[1].split(';')[0]), 0))
							if 'AD' not in dict_passed[key].keys():
								dict_passed[key]['AD'] = []
							AD_index = format_names.split(':').index('AD')
							AD_values = format_values.split(':')[AD_index].split(',')
							AD_values = [int(value) for value in AD_values]
							max_AD = max(AD_values)
							dict_passed[key]['AD'].append(max_AD)
							if 'info' not in dict_passed[key].keys():
								dict_passed[key]['info'] = []
							try:  # ANNOVAR
								if 'QUALITY' not in dict_passed[key].keys():
									dict_passed[key]['QUALITY'] = []
								if line_list[5] != '.':
									dict_passed[key]['QUALITY'].append(round(float(line_list[5]), 0))
								if 'SB' not in dict_passed[key].keys():
									dict_passed[key]['MQSBZ'] = []
								try:
									dict_passed[key]['MQSBZ'].append(round(float(line_list[7].split('MQSBZ=')[1].split(';')[0]), 2))
								except:
									try:
										dict_passed[key]['MQBZ'].append(round(float(line_list[7].split('MQBZ=')[1].split(';')[0]), 2))
									except:
										if 'STRANDQ' not in dict_passed[key].keys():
											match_SB = re.search(r'STRANDQ=([^;]+)', line)
											try:
												SB_score = match_SB.group(1)
												dict_passed[key]['STRANDQ'] = SB_score
											except:
												try:
													SB_index = format_names.split(':').index('SB')
													SB_values = format_values.split(':')[SB_index].replace('\n', "").split(',')
													SB_values = [int(sb_value) for sb_value in SB_values]
													refFw = SB_values[0] + 1
													refRv = SB_values[1] + 1
													altFw = SB_values[2] + 1
													altRv = SB_values[3] + 1
													symmetricalRatio = (refFw * altRv) / (altFw * refRv)
													refRatio = refRv / refFw
													altRatio = altFw / altRv
													SOR = round(math.log(symmetricalRatio) + refRatio - altRatio, 3)
													dict_passed[key]['SOR'] = SOR
												except:
													print("MQSBZ score not provided by ANNOVAR")  # ajouter FUNCO FILTRES en dessous
							except:
								pass

			if k == 0:
				dict_passed1 = dict_passed.copy()
			# print('dict_passed1')
			else:
				dict_passed2 = dict_passed.copy()
			# print('dict_passed2')
			k += 1

		# table creation
		dict_variants = {}
		try:
			a = dfGenes['Gene']
		except:
			print('a')
		for gene in dfGenes['Gene']:
			times = ["g.(t1)", "g.(t2)"]
			for time in times:
				pattern = r"\((.*?)\)"
				content = re.findall(pattern, time)
				time_of_variant = content[0]
				variants = dfGenes.loc[dfGenes['Gene'] == gene, time].iloc[0]
				# print(variants)
				if '|' in variants:
					variants = variants.split('|')
					for variant in variants:
						# print(variant)
						pattern = r"(\w{3,})\1{3,}"
						match = re.search(pattern, variant)
						if not match:
							if dict_para['variants_selection_approach'].upper() == 'CHANGE' and not (variant in dict_passed1.keys() and variant in dict_passed2.keys()):
								if variant not in dict_variants.keys():
									dict_variants[variant] = {}
									dict_variants[variant]['info'] = ''
									dict_variants[variant]['chr'] = re.search(r'chr[^:]+', variant).group()
									dict_variants[variant]['pos'] = re.search(r':(\d+)', variant).group(1)
									dict_variants[variant]['gene(s)'] = gene
									dict_variants[variant]['ref'] = re.search(r':\d+([A-Za-z]+)>', variant).group(1)
									dict_variants[variant]['alt'] = re.search(r'(\w+)$', variant).group()
									dict_variants[variant]['time'] = time_of_variant

									try:
										dict_variants[variant]['SIFT score'] = str(dict_passed1[variant]['SIFT score']).replace('[', '').replace(']', '')
									except:
										dict_variants[variant]['SIFT score'] = str(dict_passed2[variant]['SIFT score']).replace('[', '').replace(']', '')

									try:
										dict_variants[variant]['PolyPhen2 score'] = str(dict_passed1[variant]['PolyPhen2 score']).replace('[', '').replace(']', '')
									except:
										dict_variants[variant]['PolyPhen2 score'] = str(dict_passed2[variant]['PolyPhen2 score']).replace('[', '').replace(']', '')

									try:
										dict_variants[variant]['SIFT pred'] = str(dict_passed1[variant]['SIFT pred']).replace('[', '').replace(']', '')
									except:
										dict_variants[variant]['SIFT pred'] = str(dict_passed2[variant]['SIFT pred']).replace('[', '').replace(']', '')

									try:
										dict_variants[variant]['PolyPhen2 pred'] = str(dict_passed1[variant]['PolyPhen2 pred']).replace('[', '').replace(']', '')
									except:
										dict_variants[variant]['PolyPhen2 pred'] = str(dict_passed2[variant]['PolyPhen2 pred']).replace('[', '').replace(']', '')

									try:
										dict_variants[variant]['DP'] = str(dict_passed1[variant]['DP']).replace('[', '').replace(']', '')
									except:
										try:
											dict_variants[variant]['DP'] = str(dict_passed2[variant]['DP']).replace('[', '').replace(']', '')
										except:
											print('dict_variants')
									try:  # ANNOVAR
										if not 'info' in dict_variants[variant].keys():
											if dict_para['variants_selection_approach'].upper() == 'CHANGE':
												try:
													dict_variants[variant]['info'] = str(dict_passed1[variant]['info']).replace('[', '').replace(']', '').replace("'", '')
												except:
													try:
														dict_variants[variant]['info'] = str(dict_passed2[variant]['info']).replace('[', '').replace(']', '').replace("'", '')
													except:
														print('dict_variants')
										try:
											AD = str(dict_passed1[variant]['AD']).replace('[', '').replace(']', '').replace("'", '')
											if 'CONTQ=93;DP=96;ECNT=2;GERMQ=93;MBQ=41' in AD:
												print('a')
											dict_variants[variant]['AD'] = AD.replace(',', '|')
										except:
											try:
												AD = str(dict_passed2[variant]['AD']).replace('[', '').replace(']', '').replace("'", '')
												if 'CONTQ=93;DP=96;ECNT=2;GERMQ=93;MBQ=41' in AD:
													print('a')
												dict_variants[variant]['AD'] = AD.replace(',', '|')
											except:
												print('dict_variants')
										if 'QUALITY' not in dict_variants[variant].keys():
											dict_variants[variant]['QUALITY'] = ''
										try:
											dict_variants[variant]['QUALITY'] = str(dict_passed1[variant]['QUALITY']).replace('[', '').replace(']', '')
										except:
											try:
												dict_variants[variant]['QUALITY'] = str(dict_passed2[variant]['QUALITY']).replace('[', '').replace(']', '')
											except:
												print('dict_variants')
										if 'MQSBZ' not in dict_variants[variant].keys():
											dict_variants[variant]['MQSBZ'] = ''
										try:
											dict_variants[variant]['MQSBZ'] = str(dict_passed1[variant]['MQSBZ']).replace('[', '').replace(']', '')
										except:
											try:
												dict_variants[variant]['MQSBZ'] = str(dict_passed2[variant]['MQSBZ']).replace('[', '').replace(']', '')
											except:
												print('dict_variants: no strand bias')
									except:
										# print('not annotated by ANNOVAR')
										pass
							elif dict_para['variants_selection_approach'].upper() != 'CHANGE':
								dict_variants[variant]['gene(s)'] += f' | {gene}'
							else:
								pass

				else:
					variant = variants
					if variant not in dict_variants.keys() and variant != '':
						# print('test')
						pattern = r"(\w{3,})\1{3,}"
						match = re.search(pattern, variant)
						if not match:
							if dict_para['variants_selection_approach'].upper() == 'CHANGE' and not (variant in dict_passed1.keys() and variant in dict_passed2.keys()):
								dict_variants[variant] = {}
								dict_variants[variant]['info'] = ''
								dict_variants[variant]['chr'] = re.search(r'chr[^:]+', variant).group()
								dict_variants[variant]['pos'] = re.search(r':(\d+)', variant).group(1)
								dict_variants[variant]['gene(s)'] = gene
								dict_variants[variant]['ref'] = re.search(r':\d+([A-Za-z]+)>', variant).group(1)
								dict_variants[variant]['alt'] = re.search(r'(\w+)$', variant).group()
								dict_variants[variant]['time'] = time_of_variant

								try:
									dict_variants[variant]['DP'] = str(dict_passed1[variant]['DP']).replace('[', '').replace(']', '')
								except:
									dict_variants[variant]['DP'] = str(dict_passed2[variant]['DP']).replace('[', '').replace(']', '')
								try:
									try:
										dict_variants[variant]['AD'] = str(dict_passed1[variant]['AD']).replace('[', '').replace(']', '')
									except:
										dict_variants[variant]['AD'] = str(dict_passed2[variant]['AD']).replace('[', '').replace(']', '')
									try:
										dict_variants[variant]['QUALITY'] = str(dict_passed1[variant]['QUALITY']).replace('[', '').replace(']', '')
									except:
										dict_variants[variant]['QUALITY'] = str(dict_passed2[variant]['QUALITY']).replace('[', '').replace(']', '')
									try:
										dict_variants[variant]['MQSBZ'] = str(dict_passed1[variant]['MQSBZ']).replace('[', '').replace(']', '')
									except:
										dict_variants[variant]['MQSBZ'] = str(dict_passed2[variant]['MQSBZ']).replace('[', '').replace(']', '')
								except:
									# print('not annotated by ANNOVAR')
									pass
					else:
						if dict_para['variants_selection_approach'].upper() != 'CHANGE':
							dict_variants[variant]['gene(s)'] += f' | {gene}'
						else:
							pass

		dict_variants_final = {}
		table1_path = re.sub(r'[^/]+$', '', dict_para['output_path_sample1'])
		files_list1 = os.listdir(table1_path)
		for file in files_list1:
			if 'variants' in file:
				table1_path = table1_path + file
		table2_path = re.sub(r'[^/]+$', '', dict_para['output_path_sample2'])
		files_list2 = os.listdir(table2_path)
		for file in files_list2:
			if 'variants' in file:
				table2_path = table2_path + file

		if 'xlsx' in table1_path:
			df_variants1 = pd.read_excel(table1_path)
		elif 'csv' in table1_path:
			df_variants1 = pd.read_csv(table1_path)
		elif 'tsv' in table1_path:
			df_variants1 = pd.read_csv(table1_path, sep='\t')

		if 'xlsx' in table2_path:
			df_variants2 = pd.read_excel(table2_path)
		elif 'csv' in table2_path:
			df_variants2 = pd.read_csv(table2_path)
		elif 'tsv' in table2_path:
			df_variants2 = pd.read_csv(table2_path, sep='\t')

		dict_variants_1 = {}
		dict_variants_2 = {}

		for index, row in df_variants1.iterrows():
			key = row.iloc[0]
			value = row.iloc[1:].to_dict()
			dict_variants_1[key] = value
		for index, row in df_variants2.iterrows():
			key = row.iloc[0]
			value = row.iloc[1:].to_dict()
			dict_variants_2[key] = value

		for variant in dict_variants.keys():
			if not len(dict_variants[variant]['gene(s)'].split('|')) > 20:
				dict_variants_final[variant] = dict_variants[variant]
				try:
					dict_variants_final[variant]['Mutation'] = dict_variants_1[variant]['Mutation']
					dict_variants_final[variant]['VAF sample'] = dict_variants_1[variant]['VAF sample']
				except:
					try:
						dict_variants_final[variant]['Mutation'] = dict_variants_2[variant]['Mutation']
						dict_variants_final[variant]['VAF sample'] = dict_variants_2[variant]['VAF sample']
					except:
						print('a')
				try:
					dict_variants_final[variant]['VAF sample'] = dict_variants_1[variant]['VAF sample']
				except:
					try:
						dict_variants_final[variant]['VAF sample'] = dict_variants_2[variant]['VAF sample']
					except:
						print('b')
				try:
					dict_variants_final[variant]['VAF pop'] = dict_variants_1[variant]['VAF pop']
				except:
					try:
						dict_variants_final[variant]['VAF pop'] = dict_variants_2[variant]['VAF pop']
					except:
						print('c')
				try:
					dict_variants_final[variant]['SIFT score'] = dict_variants_1[variant]['SIFT score']
				except:
					try:
						dict_variants_final[variant]['SIFT score'] = dict_variants_2[variant]['SIFT score']
					except:
						print('no SIFT score for this variant')
				try:
					dict_variants_final[variant]['PolyPhen2 score'] = dict_variants_1[variant]['Polyphen2 score']
				except:
					try:
						dict_variants_final[variant]['PolyPhen2 score'] = dict_variants_2[variant]['Polyphen2 score']
					except:
						print('no PolyPhen2 score for this variant')
				try:
					dict_variants_final[variant]['SIFT pred'] = dict_variants_1[variant]['SIFT pred']
				except:
					try:
						dict_variants_final[variant]['SIFT pred'] = dict_variants_2[variant]['SIFT pred']
					except:
						print('no SIFT pred for this variant')
				try:
					dict_variants_final[variant]['PolyPhen2 pred'] = dict_variants_1[variant]['Polyphen2 pred']
				except:
					try:
						dict_variants_final[variant]['PolyPhen2 pred'] = dict_variants_2[variant]['Polyphen2 pred']
					except:
						print('no PolyPhen2 pred for this variant')

		empty_info = True
		for dic in dict_variants_final.values():
			if dic.get('info') != '':
				# print('info')
				empty_info = False
				break
		if empty_info:
			# print('no info')
			for dic in dict_variants_final.values():
				del dic['info']

		# sorting dict so we can order variants among pos withing each chr
		dict_sorting = {}
		for key in dict_variants_final.keys():
			match = re.search(r'chr(\d+|X)', key)
			if match:
				chr = match.group(1)
			if not key in dict_sorting.keys():
				dict_sorting[key] = {}
			dict_sorting[key]['chr'] = chr
			dict_sorting[key]['pos'] = re.search(r':(\d+)', key).group(1)

		grouped_dict = {}
		for key, value in dict_sorting.items():
			chr_value = value['chr']
			if chr_value in grouped_dict:
				grouped_dict[chr_value].append(key)
			else:
				grouped_dict[chr_value] = [key]

		for key in grouped_dict:
			grouped_dict[key] = sorted(grouped_dict[key], key=lambda x: int(dict_sorting[x]['pos']))
		order = ['1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12', '13', '14', '15', '16', '17', '18', '19', '20', '21', '22', 'X']
		merged_list = []
		for key in order:
			if key in grouped_dict:
				merged_list.extend(grouped_dict[key])

		dict_variants_final_ordered = {}
		for key in merged_list:
			if key in dict_variants_final:
				dict_variants_final_ordered[key] = dict_variants_final[key]

		df_variants = pd.DataFrame.from_dict(dict_variants_final_ordered, orient='index')
		df_variants.reset_index(inplace=True)
		df_variants.rename(columns={'index': 'Variant'}, inplace=True)
		df_variants = df_variants.sort_values(by='VAF sample', axis=0)
		df_variants.rename(columns={'MQSBZ': 'SB'}, inplace=True)
		if all(value == '' for value in df_variants['QUALITY'].tolist()):
			df_variants.drop('QUALITY', axis=1, inplace=True)

		if 'QUALITY' in df_variants.columns and 'VAF sample' in df_variants.columns and 'VAF pop' in df_variants.columns and 'SIFT score' in df_variants.columns \
				and 'PolyPhen2 score' in df_variants.columns and 'SIFT pred' in df_variants.columns and 'PolyPhen2 pred' in df_variants.columns:
			df_variants = df_variants[['Variant', 'chr', 'pos', 'gene(s)', 'ref', 'alt', 'time', 'Mutation', 'DP', 'QUALITY', 'AD', 'VAF sample', 'VAF pop', \
									   'SIFT score', 'PolyPhen2 score', 'SIFT pred', 'PolyPhen2 pred']]
		elif 'QUALITY' not in df_variants.columns and 'Mutation' in df_variants.columns and 'VAF pop' in df_variants.columns and 'VAF sample' in df_variants.columns\
			and 'SIFT score' in df_variants.columns and 'PolyPhen2 score' in df_variants.columns and 'SIFT pred' in df_variants.columns and 'PolyPhen2 pred' in df_variants.columns:
			df_variants = df_variants[['Variant', 'chr', 'pos', 'gene(s)', 'ref', 'alt', 'time', 'Mutation', 'DP', 'AD', 'VAF sample', 'VAF pop', 'SIFT score', \
									   'PolyPhen2 score', 'SIFT pred', 'PolyPhen2 pred']]
		elif 'QUALITY' in df_variants.columns and 'VAF sample' in df_variants.columns:
			df_variants = df_variants[['Variant', 'chr', 'pos', 'gene(s)', 'ref', 'alt', 'time', 'Mutation', 'DP', 'QUALITY', 'AD', 'VAF sample']]
		elif 'QUALITY' in df_variants.columns and 'VAF pop' in df_variants.columns:
			df_variants = df_variants[['Variant', 'chr', 'pos', 'gene(s)', 'ref', 'alt', 'time', 'Mutation', 'DP', 'QUALITY', 'AD', 'VAF pop']]
		elif 'VAF sample' in df_variants.columns and 'VAF pop' in df_variants.columns:
			df_variants = df_variants[['Variant', 'chr', 'pos', 'gene(s)', 'ref', 'alt', 'time', 'DP', 'AD', 'VAF sample', 'VAF pop']]
		else:
			df_variants = df_variants[['Variant', 'chr', 'pos', 'gene(s)', 'ref', 'alt', 'time', 'DP', 'AD']]

		try:
			variants_output = dict_para['output_path_comparison'] + dict_pairs[pair_id] + '_' + dict_para['variants_selection_approach'] + '_variants'
		except:
			variants_output = dict_para['output_path_comparison'] + pair_id + '_' + dict_para['variants_selection_approach'] + '_variants'

		df_variants['VAF pop'] = pd.to_numeric(df_variants['VAF pop'], errors='coerce')
		df_variants['VAF sample'] = pd.to_numeric(df_variants['VAF sample'], errors='coerce')
		df_variants['SIFT score'] = pd.to_numeric(df_variants['SIFT score'], errors='coerce')
		df_variants['PolyPhen2 score'] = pd.to_numeric(df_variants['PolyPhen2 score'], errors='coerce')
		if 'SIFT pred' in df_variants.columns:
			df_variants.sort_values(['SIFT score', 'PolyPhen2 score', 'VAF pop', 'VAF sample', 'SIFT pred', 'PolyPhen2 pred', 'DP', 'AD'], ascending=[True, False, True, True, True, False, False, False], inplace=True)
		else:
			df_variants.sort_values(['VAF pop', 'VAF sample', 'DP', 'AD'], ascending=[True, True, False, False], inplace=True)
		df_variants['VAF pop'].fillna('not found', inplace=True)
		df_variants['VAF sample'].fillna('not found', inplace=True)
		df_variants['SIFT score'].fillna('not found', inplace=True)
		df_variants['PolyPhen2 score'].fillna('not found', inplace=True)
		df_variants['VAF pop'] = df_variants['VAF pop'].apply(lambda x: '{:.2e}'.format(x) if isinstance(x, float) else x)

		if 'Mutation' in df_variants.columns:
			try:
				df_variants['Mutation'] = df_variants['Mutation'].apply(lambda s: s.replace('.', 'unknown'))
			except:
				print('d')

		formats = dict_para['C_variants_table_format(s)'].upper()
		if 'TSV' in formats:
			df_variants.to_csv(variants_output + '.tsv', sep='\t', index=False)
		if 'CSV' in formats:
			df_variants.to_csv(variants_output + '.csv', sep=',', index=False)
		if 'XLSX' in formats:
			variants_output = variants_output + '.xlsx'
			df_variants.to_excel(variants_output, index=False)

			workbook = load_workbook(variants_output)
			worksheet = workbook.active

			mutation_keywords = ['nonsynonymous', 'stopgain', 'stoploss', 'frameshift']

			highlight_colors = {
				'green': "FFA6FF00",  # Light green
				'yellow': "FFFFFF00",  # Light yellow
				'orange': "FFFFC000",  # Light orange
				'red': "FFFF4D4D",  # Light red
			}

			# DP ranges
			values = df_variants['DP'].tolist()
			numeric_values = [float(value) for value in values]
			quartile1 = np.percentile(numeric_values, 25)
			quartile2 = np.percentile(numeric_values, 50)
			quartile3 = np.percentile(numeric_values, 75)
			DP_range1 = [0, quartile1]
			DP_range2 = [quartile1, quartile2]
			DP_range3 = [quartile2, quartile3]
			DP_range4 = [quartile3, max(numeric_values)]

			# AD ranges
			values = df_variants['AD'].tolist()
			numeric_values = [float(value) for value in values]
			quartile1 = np.percentile(numeric_values, 25)
			quartile2 = np.percentile(numeric_values, 50)
			quartile3 = np.percentile(numeric_values, 75)
			AD_range1 = [0, quartile1]
			AD_range2 = [quartile1, quartile2]
			AD_range3 = [quartile2, quartile3]
			AD_range4 = [quartile3, max(numeric_values)]

			# VAF sample ranges
			values = df_variants['VAF sample'].tolist()
			values = [value for value in values if value != 'not found']
			numeric_values = [float(value) for value in values]
			quartile1 = np.percentile(numeric_values, 25)
			quartile2 = np.percentile(numeric_values, 50)
			quartile3 = np.percentile(numeric_values, 75)
			VAF_sample_range1 = [0, quartile1]
			VAF_sample_range2 = [quartile1, quartile2]
			VAF_sample_range3 = [quartile2, quartile3]
			VAF_sample_range4 = [quartile3, max(numeric_values)]

			# VAF pop ranges
			values = df_variants['VAF pop'].tolist()
			values = [value for value in values if value != 'not found']
			numeric_values = [float(value) for value in values]
			quartile1 = np.percentile(numeric_values, 25)
			quartile2 = np.percentile(numeric_values, 50)
			quartile3 = np.percentile(numeric_values, 75)
			VAF_pop_range1 = [0, quartile1]
			VAF_pop_range2 = [quartile1, quartile2]
			VAF_pop_range3 = [quartile2, quartile3]
			VAF_pop_range4 = [quartile3, max(numeric_values)]

			# SIFT ranges
			SIFT_range1 = [-1, 0.05]
			SIFT_range2 = [0.05, 0.1]
			SIFT_range3 = [0.1, 0.5]
			SIFT_range4 = [0.5, 1.5]

			# PolyPhen2 score ranges
			PolyPhen2_range1 = [-1, 0.453]
			PolyPhen2_range2 = [0.453, 0.7]
			PolyPhen2_range3 = [0.7, 0.956]
			PolyPhen2_range4 = [0.956, 1.5]

			header_row = worksheet[1]
			column_names = [cell.value for cell in header_row]
			for row in worksheet.iter_rows(min_row=2):
				for cell, column_name in zip(row, column_names):
					if column_name == 'Mutation':
						if any(keyword in cell.value for keyword in mutation_keywords):
							cell.fill = PatternFill(start_color=highlight_colors['green'], end_color=highlight_colors['green'], fill_type="solid")
						else:
							cell.fill = PatternFill(start_color=highlight_colors['red'], end_color=highlight_colors['red'], fill_type="solid")
					elif column_name == 'DP':
						if DP_range1[0] <= int(cell.value) <= DP_range1[1]:
							cell.fill = PatternFill(start_color=highlight_colors['red'], end_color=highlight_colors['red'], fill_type="solid")
						elif DP_range2[0] <= int(cell.value) <= DP_range2[1]:
							cell.fill = PatternFill(start_color=highlight_colors['orange'], end_color=highlight_colors['orange'], fill_type="solid")
						elif DP_range3[0] <= int(cell.value) <= DP_range3[1]:
							cell.fill = PatternFill(start_color=highlight_colors['yellow'], end_color=highlight_colors['yellow'], fill_type="solid")
						elif DP_range4[0] <= int(cell.value) <= DP_range4[1]:
							cell.fill = PatternFill(start_color=highlight_colors['green'], end_color=highlight_colors['green'], fill_type="solid")
					elif column_name == 'AD':
						if AD_range1[0] <= int(cell.value) <= AD_range1[1]:
							cell.fill = PatternFill(start_color=highlight_colors['red'], end_color=highlight_colors['red'], fill_type="solid")
						elif AD_range2[0] <= int(cell.value) <= AD_range2[1]:
							cell.fill = PatternFill(start_color=highlight_colors['orange'], end_color=highlight_colors['orange'], fill_type="solid")
						elif AD_range3[0] <= int(cell.value) <= AD_range3[1]:
							cell.fill = PatternFill(start_color=highlight_colors['yellow'], end_color=highlight_colors['yellow'], fill_type="solid")
						elif AD_range4[0] <= int(cell.value) <= AD_range4[1]:
							cell.fill = PatternFill(start_color=highlight_colors['green'], end_color=highlight_colors['green'], fill_type="solid")
					elif column_name == 'VAF sample':
						if 'not found' in str(cell.value):
							pass
						elif VAF_sample_range1[0] <= float(cell.value) <= VAF_sample_range1[1]:
							cell.fill = PatternFill(start_color=highlight_colors['green'], end_color=highlight_colors['green'], fill_type="solid")
						elif VAF_sample_range2[0] <= float(cell.value) <= VAF_sample_range2[1]:
							cell.fill = PatternFill(start_color=highlight_colors['yellow'], end_color=highlight_colors['yellow'], fill_type="solid")
						elif VAF_sample_range3[0] <= float(cell.value) <= VAF_sample_range3[1]:
							cell.fill = PatternFill(start_color=highlight_colors['orange'], end_color=highlight_colors['orange'], fill_type="solid")
						elif VAF_sample_range4[0] <= float(cell.value) <= VAF_sample_range4[1]:
							cell.fill = PatternFill(start_color=highlight_colors['red'], end_color=highlight_colors['red'], fill_type="solid")
					elif column_name == 'VAF pop':
						if cell.value == 'not found':
							pass
						elif VAF_pop_range1[0] <= float(cell.value) <= VAF_pop_range1[1]:
							cell.fill = PatternFill(start_color=highlight_colors['green'], end_color=highlight_colors['green'], fill_type="solid")
						elif VAF_pop_range2[0] <= float(cell.value) <= VAF_pop_range2[1]:
							cell.fill = PatternFill(start_color=highlight_colors['yellow'], end_color=highlight_colors['yellow'], fill_type="solid")
						elif VAF_pop_range3[0] <= float(cell.value) <= VAF_pop_range3[1]:
							cell.fill = PatternFill(start_color=highlight_colors['orange'], end_color=highlight_colors['orange'], fill_type="solid")
						elif VAF_pop_range4[0] <= float(cell.value) <= VAF_pop_range4[1]:
							cell.fill = PatternFill(start_color=highlight_colors['red'], end_color=highlight_colors['red'], fill_type="solid")
					elif column_name == 'SIFT score':
						if cell.value == 'not found':
							pass
						elif SIFT_range1[0] <= float(cell.value) <= SIFT_range1[1]:
							cell.fill = PatternFill(start_color=highlight_colors['green'], end_color=highlight_colors['green'], fill_type="solid")
						elif SIFT_range2[0] <= float(cell.value) <= SIFT_range2[1]:
							cell.fill = PatternFill(start_color=highlight_colors['yellow'], end_color=highlight_colors['yellow'], fill_type="solid")
						elif SIFT_range3[0] <= float(cell.value) <= SIFT_range3[1]:
							cell.fill = PatternFill(start_color=highlight_colors['orange'], end_color=highlight_colors['orange'], fill_type="solid")
						elif SIFT_range4[0] <= float(cell.value) <= SIFT_range4[1]:
							cell.fill = PatternFill(start_color=highlight_colors['red'], end_color=highlight_colors['red'], fill_type="solid")
					elif column_name == 'PolyPhen2 score':
						if cell.value == 'not found':
							pass
						elif PolyPhen2_range1[0] <= float(cell.value) <= PolyPhen2_range1[1]:
							cell.fill = PatternFill(start_color=highlight_colors['red'], end_color=highlight_colors['green'], fill_type="solid")
						elif PolyPhen2_range2[0] <= float(cell.value) <= PolyPhen2_range2[1]:
							cell.fill = PatternFill(start_color=highlight_colors['orange'], end_color=highlight_colors['yellow'], fill_type="solid")
						elif PolyPhen2_range3[0] <= float(cell.value) <= PolyPhen2_range3[1]:
							cell.fill = PatternFill(start_color=highlight_colors['yellow'], end_color=highlight_colors['orange'], fill_type="solid")
						elif PolyPhen2_range4[0] <= float(cell.value) <= PolyPhen2_range4[1]:
							cell.fill = PatternFill(start_color=highlight_colors['green'], end_color=highlight_colors['red'], fill_type="solid")
					elif column_name == 'SIFT pred':
						if 'deleterious' in cell.value.lower():
							cell.fill = PatternFill(start_color=highlight_colors['green'], end_color=highlight_colors['green'], fill_type="solid")
						elif 'tolerated' in cell.value.lower():
							cell.fill = PatternFill(start_color=highlight_colors['red'], end_color=highlight_colors['red'], fill_type="solid")
					elif column_name == 'PolyPhen2 pred':
						if 'probably' in cell.value.lower():
							cell.fill = PatternFill(start_color=highlight_colors['green'], end_color=highlight_colors['green'], fill_type="solid")
						elif 'possibly' in cell.value.lower():
							cell.fill = PatternFill(start_color=highlight_colors['yellow'], end_color=highlight_colors['yellow'], fill_type="solid")
						elif 'benign' in cell.value.lower():
							cell.fill = PatternFill(start_color=highlight_colors['orange'], end_color=highlight_colors['orange'], fill_type="solid")

			alignment = Alignment(horizontal="center")
			border = Border(left=Side(style='thin'),
							right=Side(style='thin'),
							top=Side(style='thin'),
							bottom=Side(style='thin'))

			for cell in worksheet[1]:
				if cell.value:
					cell.alignment = alignment
					cell.border = border

			for column in worksheet.columns:
				max_length = 0
				column_letter = column[0].column_letter
				for cell in column:
					try:
						if len(str(cell.value)) > max_length:
							max_length = len(cell.value)
					except:
						pass
				adjusted_width = (max_length + 2) * 1.2
				worksheet.column_dimensions[column_letter].width = adjusted_width

			for row in worksheet.iter_rows(min_row=1):
				for cell in row:
					cell.border = border
					cell.alignment = alignment
			workbook.save(variants_output)
		try:
			output_file = (dict_para['output_path_comparison'] + dict_pairs[pair_id] + '_' + dict_para['variants_selection_approach'] + '_variants.vcf')
		except:
			output_file = (dict_para['output_path_comparison'] + pair_id + '_' + dict_para['variants_selection_approach'] + '_variants.vcf')
		vcf_id = dict_para['output_path_comparison'].split("/")[2]

		dict_impacts_counts = write_stats(dict_para, df_variants, dict_variants)
		create_protein_impacts_plots(dict_para, dict_impacts_counts, pair_id)

		with open(output_file, "w") as f:
			f.write("##fileformat=VCFv4.3\n")
			current_datetime = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

			f.write("##fileDate=" + current_datetime + "\n")

			# chr definitions
			contig_lines = [
				"##contig=<ID=chr1,length=249250621>\n",
				"##contig=<ID=chr2,length=242193529>\n",
				"##contig=<ID=chr3,length=198295559>\n",
				"##contig=<ID=chr4,length=190214555>\n",
				"##contig=<ID=chr5,length=181538259>\n",
				"##contig=<ID=chr6,length=170805979>\n",
				"##contig=<ID=chr7,length=159345973>\n",
				"##contig=<ID=chr8,length=145138636>\n",
				"##contig=<ID=chr9,length=138394717>\n",
				"##contig=<ID=chr10,length=133797422>\n",
				"##contig=<ID=chr11,length=135086622>\n",
				"##contig=<ID=chr12,length=133275309>\n",
				"##contig=<ID=chr13,length=114364328>\n",
				"##contig=<ID=chr14,length=107043718>\n",
				"##contig=<ID=chr15,length=101991189>\n",
				"##contig=<ID=chr16,length=90338345>\n",
				"##contig=<ID=chr17,length=83257441>\n",
				"##contig=<ID=chr18,length=80373285>\n",
				"##contig=<ID=chr19,length=58617616>\n",
				"##contig=<ID=chr20,length=64444167>\n",
				"##contig=<ID=chr21,length=46709983>\n",
				"##contig=<ID=chr22,length=50818468>\n",
				"##contig=<ID=chrX,length=156040895>\n",
				"##contig=<ID=chrY,length=57227415>\n",
				"##contig=<ID=chrM,length=16569>\n",
			]
			for line in contig_lines:
				f.write(line)

			# FORMAT definitons
			f.write("##FORMAT=<ID=GT,Number=1,Type=String,Description=\"Genotype\">\n")
			f.write("##FORMAT=<ID=PL,Number=G,Type=Integer,Description=\"Phred-scaled Genotype Likelihoods\">\n")
			f.write("##FORMAT=<ID=DP,Number=1,Type=Integer,Description=\"Total Depth\">\n")
			f.write("##FORMAT=<ID=AD,Number=.,Type=Integer,Description=\"Allelic Depths\">\n")
			f.write("##FORMAT=<ID=GP,Number=1,Type=String,Description=\"Genotype Probability\">\n")
			f.write("##FORMAT=<ID=GQ,Number=1,Type=Integer,Description=\"Genotype Quality\">\n")

			if dict_para.get("variants_selection_approach", "").upper() == "CHANGE":
				f.write("##INFO=<ID=DP,Number=1,Type=Integer,Description=\"Total Depth\">\n")
				f.write("##INFO=<ID=AD,Number=.,Type=Integer,Description=\"Allelic Depths\">\n")
				f.write("##INFO=<ID=MQSBZ,Number=1,Type=Float,Description=\"MQSBZ Score\">\n")
			else:
				print('code to right : header for funco and snpeff')

			f.write(f"#CHROM\tPOS\tID\tREF\tALT\tQUAL\tFILTER\tINFO\tFORMAT\t{vcf_id}\n")

			for _, row in df_variants.iterrows():
				chrom = row['chr']
				pos = row['pos']
				ref = row['ref']
				alt = row['alt']
				chrom = chrom.replace(' ', '')
				ref = ref.replace(' ', '')
				alt = alt.replace(' ', '')

				DP = row['DP']
				if 'QUALITY' in df_variants.columns:
					QUALITY = row['QUALITY']
				else:
					QUALITY = '.'
				try:
					AD = row['AD'].replace('|', ',')
				except:
					pass
				# SB = row['SB']
				SB = ''

				try:
					SIFT_score = row['SIFT score']
					SIFT_pred = row['SIFT pred']
				except:
					print('no SIFT annotation')

				try:
					Polyphen2_score = row['PolyPhen2 score']
					Polyphen2_pred = row['PolyPhen2 pred']
				except:
					print('no PolyPhen2 annotation')

				try:
					VAF_pop = row['VAF pop']
				except:
					print('no VAF pop')

				key = 'g.' + str(chrom) + ':' + str(pos) + ref + '>' + alt
				if dict_para['variants_selection_approach'].upper() == 'CHANGE':
					try:
						format_values = dict_passed1[key]['FORMAT']
					except KeyError:
						format_values = dict_passed2[key]['FORMAT']
					format_values = str(format_values).replace('[', '').replace(']', '').replace("'", '').replace("\\n", '')

				if 'VAF pop' not in df_variants.columns:
					if 'Mutation' and 'SIFT score' and 'SIFT pred' and 'PolyPhen2 score' and 'PolyPhen2 pred' in df_variants.columns:
						f.write(f"{chrom}\t{pos}\t.\t{ref}\t{alt}\t{QUALITY}\tPASS\tDP={DP};AD={AD};SB={SB};mutation={row['Mutation']};SIFT_score={SIFT_score}"
								f"; SIFT_pred={SIFT_pred}; Polyphen2_score={Polyphen2_score}; Polyphen2_pred={Polyphen2_pred}\t"
							f"GT:PL:DP:AD:GP:GQ:VAF\t{format_values}\n")
					elif 'Mutation' and not ('SIFT score' and 'SIFT pred' and 'PolyPhen2 score' and 'PolyPhen2 pred' in df_variants.columns):
						f.write(f"{chrom}\t{pos}\t.\t{ref}\t{alt}\t{QUALITY}\tPASS\tDP={DP};AD={AD};SB={SB};mutation={row['Mutation']}\t"
								f"GT:PL:DP:AD:GP:GQ:VAF\t{format_values}\n")
				else:
					if 'Mutation' and 'SIFT score' and 'SIFT pred' and 'PolyPhen2 score' and 'PolyPhen2 pred' in df_variants.columns:
						f.write(f"{chrom}\t{pos}\t.\t{ref}\t{alt}\t{QUALITY}\tPASS\tDP={DP};AD={AD};SB={SB};AF={VAF_pop};mutation={row['Mutation']};SIFT_score={SIFT_score}"
								f"; SIFT_pred={SIFT_pred}; Polyphen2_score={Polyphen2_score}; Polyphen2_pred={Polyphen2_pred}\t"
								f"GT:PL:DP:AD:GP:GQ:VAF\t{format_values}\n")
					elif 'Mutation' and not ('SIFT score' and 'SIFT pred' and 'PolyPhen2 score' and 'PolyPhen2 pred' in df_variants.columns):
						f.write(f"{chrom}\t{pos}\t.\t{ref}\t{alt}\t{QUALITY}\tPASS\tDP={DP};AD={AD};SB={SB};AF={VAF_pop};mutation={row['Mutation']}\t"
								f"GT:PL:DP:AD:GP:GQ:VAF\t{format_values}\n")


		print("VCF file created successfully.")

		common_genes_count = 0
		change_genes_count = 0
		df = dfGenes
		variants_count = df_variants.shape[0]

		if dict_para['colored_execution'].upper() == 'TRUE' or dict_para['colored_execution'].upper() == 'YES':
			colored = True
		else:
			colored = False

		if dict_para['variants_selection_approach'].upper() == 'CHANGE':
			change = len(dfGenes)
			union = int(change_counts[0])
			common = union - change

			if colored:
				print(f'\033[1m{change} genes\033[0m\033[38;2;255;193;7m contain at least one variant that appeared or disappeared between t1 and t2.\n'
					  f'{common} genes contain the same variant(s) in t1 and t2, {union} genes contain at least one variant in t1 or t2.\n'
					  f'\033[1m{variants_count} variants appeared or disappeared between t1 and t2.\033[0m\033[38;2;255;193;7m')
			else:
				print(f'{change} genes contain at least one variant that appeared or disappeared between t1 and t2.\n'
					  f'{common} genes contain the same variant(s) in t1 and t2, {union} genes contain at least one variant in t1 or t2.'
					  f'{variants_count} variants appeared or disappeared between t1 and t2.')

		elif dict_para['variants_selection_approach'].upper() == 'UNION':
			common_count, change_count = count_genes(df, dict_para)
			if colored:
				print(f'\033[1m{len(dfGenes)} genes\033[0m\033[38;2;255;193;7m contain at least one variant in t1 or t2.\n'
					  f'{common_count} genes contain the same variant(s) in t1 and t2, {change_count}'
					  f'genes contain at least one variant that appeared or disappeared between t1 and t2.')
			else:
				print(f'{len(dfGenes)} genes contain at least one variant in t1 or t2.\n'
					  f'{common_count} genes contain the same variant(s) in t1 and t2, {change_count}'
					  f'genes contain at least one variant that appeared or disappeared between t1 and t2.')

		elif dict_para['variants_selection_approach'].upper() == 'COMMON':
			common = len(dfGenes)
			union = int(common_counts[0])
			change = union - common
			if colored:
				print(f'\033[1m{common} genes\033[0m\033[38;2;255;193;7m the same variant(s) in t1 and t2.\n'
					  f'{change} genes contain at least one variant that appeared or disappeared between t1 and t2, {union} genes contain at least one variant in t1 or t2.')
			else:
				print(f'{common} genes contain the same variant(s) in t1 and t2.\n'
					  f'{change} genes contain at least one variant that appeared or disappeared between t1 and t2, {union} genes contain at least one variant in t1 or t2.')

		print('Creating mutation types and subtypes plots...')
		if dict_para['C_types_plot'].strip().lower() == 'barplot':
			create_mutation_types_barplot(dict_para, dic_mutation_types_t1, dic_mutation_types_t2)
		#TODO: create_mutation_types_piechart(dict_para, dic_mutation_types_t1, dic_mutation_types_t2)
		if dict_para['C_subtypes_plot'].strip().lower() == 'barplot':
			create_mutation_subtypes_barplot(dict_para, dic_mutation_subtypes_t1, dic_mutation_subtypes_t2)
		#TODO: create_mutation_subtypes_piechart(dict_para, dic_mutation_subtypes_t1, dic_mutation_subtypes_t2)

		################################################
		# Biological process enrichment using the genes list with the ToppGene and Panther API
		if enrichment:
			if len(genes_list) < 1500:
				print('Computing ToppGene GOEA analysis...')
				toppgene_name = 'GO_ToppGene'
				ToppGene_GOEA('summarise', dict_para, genes_list, toppgene_name, logger)
			else:
				print("The VCF is heavy, too many genes are concerned for ToppGene GO to be run.")
			if len(genes_list) < 2000:
				print('Computing Panther GOEA analysis...')
				panther_name = 'GO_Panther'
				Panther_GOEA('compare', dict_para, genes_list, panther_name, logger)
			else:
				print("The VCF is heavy, too many genes are concerned for Panther GO to be run.")


def main(args, df):
	output_path = args['output_path_comparison']
	enrichment = args['C_enrichment']
	if 'YES' in enrichment or 'TRUE' in enrichment or 'PANTHER' in enrichment or 'TOPPGENE' in enrichment:
		enrichment = True
	if args['variants_selection_approach'].upper() == 'CHANGE':
		mode = 'change'
	elif args['variants_selection_approach'].upper() == 'UNION':
		mode = 'union'
	elif args['variants_selection_approach'].upper() == 'COMMON':
		mode = 'common'
	out_gene = output_path + "Mutated_Genes_" + mode + "." + args['C_mutated_genes_table_format(s)'].lower()
	gff3 = args['human_chromosomes_reference_file']
	current_directory = os.getcwd()

	# Logger configuration
	logger = logging.getLogger('LOTUS compare')
	logger.setLevel(logging.DEBUG)
	fh = logging.FileHandler(args['log'])
	fh.setLevel(logging.DEBUG)
	formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
	fh.setFormatter(formatter)
	logger.addHandler(fh)
	try:
		logger.info('Verification of input in config file')
		# vcf_filtered_files, vcf_pass_files, snp_profile_files, insertions_count_files, deletions_count_files = verif_input_config(args.config)
		logger.info('- Inputs file ok -')
	except ValueError:
		print('Problem with config file : ', sys.exc_info()[1])
		logger.error('- Problem with config file -')
		exit(1)

	try:
		verif_input(gff3)
		genes_positions = read_gff3(gff3)
		transcript_dico = genes_positions[2]
		gene_id_dico = genes_positions[1]
		gene_name_dico = genes_positions[0]
	except UnicodeDecodeError:
		print(f'{gff3} can not be read as a classic gff3 file : ', sys.exc_info()[1])
		logger.error(f'- {gff3} can not be read as a classic gff3 file ! -')
		exit(1)
	except ValueError:
		print(f'{gff3} is not a gff3 file (or the first # line missing) : ', sys.exc_info()[1])
		logger.error(f'- {gff3} is not a gff3 file ! -')
		exit(1)

	# Verification of given arguments
	vcf_filtered1 = args['filtered1']
	vcf_filtered2 = args['filtered2']
	vcf_strong_weak1 = args['passed1']
	vcf_strong_weak2 = args['passed2']

	out_indel = 'indel_profile'
	if "." in out_indel:
		out_indel = re.sub(r"\..*", "", out_indel)

	infos = None
	if args['verbose_prints'].upper() == 'TRUE':
		print('Search for file Lotus_ExternalBases_202301.xlsx ...')
	infos = 'input/resources/Lotus_ExternalBases_202301.xlsx'
	logger.info('Verification of {infos}')
	try:
		infos = verif_supplementary_information_file(infos, current_directory)
	except ValueError:
		print(f'Problem with {infos}: ', sys.exc_info()[0])
		logger.error('- Problem with {infos} -')
		exit(1)

	# Start
	logger.info('********************************************************************************************************************')
	logger.info('*** LOTUS compare module ***')
	no_argument = ''
	if enrichment:
		no_argument += ' --enrichment'
	configuration = args['dataset_path']
	logger.info(f'** cmd line : python lotus.py compare -c {str(configuration)} ' + str(no_argument) + ' **')
	logger.info('* Start comparing *')
	logger.info(f'Current directory : {Path().absolute()}')

	##############################
	# Comparison of both vcf files

	compare_vcf(out_gene, args, gene_name_dico, gene_id_dico, transcript_dico, infos, enrichment, df, logger)

	####################################################
	# Create snp mutation types plot and indel size plot

	insertions_count_files = []
	if os.path.exists(args['insert1']):
		insertions_count_files.append(args['insert1'])
		insertions_count_files.append(args['insert2'])
	else:
		print("No common insertion mutation found between both files.")
	deletions_count_files = []
	deletions_count_files.append(args['indel1'])
	deletions_count_files.append(args['indel2'])

	# kiwi
	if args['verbose_prints'].upper() == 'FALSE':
		print('Creating plots...')
	out_C_SNP_profile = output_path + 'SNP_profile'
	graph_snp(args, out_C_SNP_profile, logger)
	if len(insertions_count_files) > 0:
		graph_indel(args, insertions_count_files, deletions_count_files, out_indel, logger)

	logger.info('* End comparing *')
	logger.info('********************************************************************************************************************')

# End
