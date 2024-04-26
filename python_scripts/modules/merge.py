#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #
#   LOncoG : a software for Longitudinal OncoGenomics analysis
#   Authors: S. Besseau, G. Siekaniec, W. Gouraud
# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #

from math import isnan
from matplotlib.ticker import ScalarFormatter
from pathlib import Path
import sys
import datetime
import scipy.stats as stats
import os
import numpy as np
import logging
import re
import warnings
from tqdm import tqdm
import pandas as pd
from copy import deepcopy
from more_itertools import powerset, unique_everseen
from openpyxl.styles import Border, Side, PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from upsetplot import from_memberships
from upsetplot import UpSet
from python_scripts.reusable_functions.check_files import verif_output, verif_input_config_merge, verif_input_xlsx, verif_input_tsv, \
	verif_input, verif_supplementary_information_file
from python_scripts.api_requests.toppgene_api import ToppGene_GOEA
from python_scripts.api_requests.panther_api import Panther_GOEA
from python_scripts.reusable_functions.chromosomes_plot import create_chromosomes_plot
import matplotlib.pyplot as plt


def get_informations_for_genes(args, info_file, logger):
	'''
	Extract informations from the external cancer databases file.
	'''
	df = pd.read_excel(info_file, index_col=1)
	df = df.drop(['Ordre'], axis=1)
	df.set_axis([source.split(' Info')[0] for source in df.columns], axis="columns")
	if args['verbose_prints'].upper() == 'TRUE':
		print(f'Extracting informations from {len(list(df.columns))} sources: {", ".join(list(df.columns))}')
	logger.info(f'Extract informations from {len(list(df.columns))} sources: {", ".join(list(df.columns))}')
	return (df)


def read_config(args, config_file):
	'''
	Read the merge config file (composed of one gene.tsv compare file per line).
	'''
	with open(config_file, 'r') as f:
		for line in f:
			line = line.strip()
			if line != '':
				line = line.split(',')
				if verif_input_xlsx(line[0]):
					data = pd.read_excel(line[0], index_col=0)
				elif verif_input_tsv(line[0]):
					data = pd.read_csv(line[0], sep='\t', index_col=0)
				if len(line) > 1:
					yield line[1], data
				else:
					yield line[0], data


def add_variants_to_dictionnary(df_all_genes_info: {}, gene_name: str, pos: int, field: str, row):
	'''
	Add a variant in count dictionnary.
	'''
	if field == 'gb1' or field == 'gb2' or field == 'cb1' or field == 'cb2' or field == 'pb1' or field == 'pb2':
		if df_all_genes_info[gene_name][field] == {}:
			df_all_genes_info[gene_name][field] = []
		if '|' in str(row.iloc[pos]):
			l_variants = row.iloc[pos].split('|')
			for variant in l_variants:
				df_all_genes_info[gene_name][field].append(variant)
		elif str(row.iloc[pos]).startswith('g.'):
			df_all_genes_info[gene_name][field].append(row.iloc[pos])
		elif 'nan' in str(row.iloc[pos]):
			df_all_genes_info[gene_name][field] = []
	elif field != 'mutation types':
		if not isnan(row.iloc[pos]):
			for i in row.iloc[pos]:
				try:
					if not i in df_all_genes_info[gene_name][field].keys():
						df_all_genes_info[gene_name][field][i] = 1
					else:
						df_all_genes_info[gene_name][field][i] += 1
				except:
					if not i in df_all_genes_info[gene_name][field]:
						df_all_genes_info[gene_name][field].append(row.iloc[pos])


def count_mutation_types(args, mutation_types_counting_method):
	# Which file belongs do what time?
	df_path = args['dataset_path']
	if df_path.endswith('xlsx'):
		df = pd.read_excel(df_path)
	elif df_path.endswith('csv'):
		try:
			df = pd.read_csv(df_path)
		except:
			df = pd.read_excel(df_path)
	time1_column_values = df[args['time1_column_name']].tolist()
	time2_column_values = df[args['time2_column_name']].tolist()
	time1_column_values = [value for value in time1_column_values if not pd.isna(value)]
	time2_column_values = [value for value in time2_column_values if not pd.isna(value)]
	new_values_t1 = []
	for value in time1_column_values:
		if '.vcf' in value:
			new_values_t1.append(value.split('.vcf')[0])
		else:
			new_values_t1.append(value)
	time1_column_values = new_values_t1
	new_values_t2 = []
	for value in time2_column_values:
		if '.vcf' in value:
			new_values_t2.append(value.split('.vcf')[0])
		else:
			new_values_t2.append(value)
	time2_column_values = new_values_t2

	# Get passed files paths
	samples_path = args['output_path'] + 'samples/'
	vcf_passed_files = []
	for root, dirs, files in os.walk(samples_path):
		for file in files:
			if file.endswith('passed.vcf'):
				folder = os.path.basename(root)
				vcf_passed_files.append(samples_path + folder + '/' + file)

	t1_files = []
	t2_files = []
	for file in vcf_passed_files:
		if file.split('samples/')[1].split('/')[0] in time1_column_values:
			t1_files.append(file)
		for name in time2_column_values:
			if file.split('samples/')[1].split('/')[0] in name:
				t2_files.append(file)

	# Mutation types and subtypes counting
	l_unique_variants_t1 = []
	l_unique_variants_t2 = []
	dic_unique_mutation_types_t1 = {}
	dic_unique_mutation_types_t2 = {}
	dic_unique_mutation_subtypes_t1 = {}
	dic_unique_mutation_subtypes_t2 = {}
	dic_total_mutation_types_t1 = {}
	dic_total_mutation_types_t2 = {}
	dic_total_mutation_subtypes_t1 = {}
	dic_total_mutation_subtypes_t2 = {}

	for file in t1_files:
		with open(file, "r") as vcf_file:
			for line in vcf_file:
				if not line.startswith("#"):
					line = line.split("\t")
					chr = line[0]
					pos = line[1]
					ref = line[3]
					alt = line[4]
					variant = chr + ':' + pos + ':' + ref + ':' + alt

					if mutation_types_counting_method == 'UNIQUE':
						if not variant in l_unique_variants_t1:
							l_unique_variants_t1.append(variant)
							if len(ref) == len(alt):
								if len(ref) == 1:
									mutation_type = 'SNP'
								elif len(ref) == 2:
									mutation_type = 'DNP'
								elif len(ref) == 3:
									mutation_type = 'TNP'
								elif len(ref) > 3:
									mutation_type = 'ONP'
							elif len(ref) > len(alt):
								mutation_type = 'Deletion'
							elif len(ref) < len(alt):
								mutation_type = 'Insertion'
							if not mutation_type in dic_unique_mutation_types_t1.keys():
								dic_unique_mutation_types_t1[mutation_type] = 0
							dic_unique_mutation_types_t1[mutation_type] += 1

							try:  # ANNOVAR
								match = re.search(r'ExonicFunc\.refGene=(.*?);', str(line))
								if match:
									mutation_subtype = match.group(1)
									if not mutation_subtype in dic_unique_mutation_subtypes_t1.keys():
										dic_unique_mutation_subtypes_t1[mutation_subtype] = 0
									dic_unique_mutation_subtypes_t1[mutation_subtype] += 1
							except:
								print('not annotated with ANNOVAR')

					elif mutation_types_counting_method == 'TOTAL':
						if len(ref) == len(alt):
							if len(ref) == 1:
								mutation_type = 'SNP'
							elif len(ref) == 2:
								mutation_type = 'DNP'
							elif len(ref) == 3:
								mutation_type = 'TNP'
							elif len(ref) > 3:
								mutation_type = 'ONP'
						elif len(ref) > len(alt):
							mutation_type = 'Deletion'
						elif len(ref) < len(alt):
							mutation_type = 'Insertion'
						if not mutation_type in dic_total_mutation_types_t1.keys():
							dic_total_mutation_types_t1[mutation_type] = 0
						dic_total_mutation_types_t1[mutation_type] += 1

						try:  # ANNOVAR
							match = re.search(r'ExonicFunc\.refGene=(.*?);', str(line))
							if match:
								mutation_subtype = match.group(1)
								if not mutation_subtype in dic_total_mutation_subtypes_t1.keys():
									dic_total_mutation_subtypes_t1[mutation_subtype] = 0
								dic_total_mutation_subtypes_t1[mutation_subtype] += 1
						except:
							print('not annotated with ANNOVAR')

	for file in t2_files:
		with open(file, "r") as vcf_file:
			for line in vcf_file:
				if not line.startswith("#"):
					line = line.split("\t")
					chr = line[0]
					pos = line[1]
					ref = line[3]
					alt = line[4]
					variant = chr + ':' + pos + ':' + ref + ':' + alt

					if mutation_types_counting_method == 'UNIQUE':
						if not variant in l_unique_variants_t2:
							l_unique_variants_t2.append(variant)
							if len(ref) == len(alt):
								if len(ref) == 1:
									mutation_type = 'SNP'
								elif len(ref) == 2:
									mutation_type = 'DNP'
								elif len(ref) == 3:
									mutation_type = 'TNP'
								elif len(ref) > 3:
									mutation_type = 'ONP'
							elif len(ref) > len(alt):
								mutation_type = 'Deletion'
							elif len(ref) < len(alt):
								mutation_type = 'Insertion'
							if not mutation_type in dic_unique_mutation_types_t2.keys():
								dic_unique_mutation_types_t2[mutation_type] = 0
							dic_unique_mutation_types_t2[mutation_type] += 1

							try:  # ANNOVAR
								match = re.search(r'ExonicFunc\.refGene=(.*?);', str(line))
								if match:
									mutation_subtype = match.group(1)
									if not mutation_subtype in dic_unique_mutation_subtypes_t2.keys():
										dic_unique_mutation_subtypes_t2[mutation_subtype] = 0
									dic_unique_mutation_subtypes_t2[mutation_subtype] += 1
							except:
								print('not annotated with ANNOVAR')

					elif mutation_types_counting_method == 'TOTAL':
						if len(ref) == len(alt):
							if len(ref) == 1:
								mutation_type = 'SNP'
							elif len(ref) == 2:
								mutation_type = 'DNP'
							elif len(ref) == 3:
								mutation_type = 'TNP'
							elif len(ref) > 3:
								mutation_type = 'ONP'
						elif len(ref) > len(alt):
							mutation_type = 'Deletion'
						elif len(ref) < len(alt):
							mutation_type = 'Insertion'
						if not mutation_type in dic_total_mutation_types_t2.keys():
							dic_total_mutation_types_t2[mutation_type] = 0
						dic_total_mutation_types_t2[mutation_type] += 1

						try:  # ANNOVAR
							match = re.search(r'ExonicFunc\.refGene=(.*?);', str(line))
							if match:
								mutation_subtype = match.group(1)
								if not mutation_subtype in dic_total_mutation_subtypes_t2.keys():
									dic_total_mutation_subtypes_t2[mutation_subtype] = 0
								dic_total_mutation_subtypes_t2[mutation_subtype] += 1
						except:
							print('not annotated with ANNOVAR')

	if mutation_types_counting_method == 'UNIQUE':
		try:  # ANNOVAR
			return dic_unique_mutation_types_t1, dic_unique_mutation_types_t2, dic_unique_mutation_subtypes_t1, dic_unique_mutation_subtypes_t2
		except:
			return dic_unique_mutation_types_t1, dic_unique_mutation_types_t2
	elif mutation_types_counting_method == 'TOTAL':
		try:  # ANNOVAR
			return dic_total_mutation_types_t1, dic_total_mutation_types_t2, dic_total_mutation_subtypes_t1, dic_total_mutation_subtypes_t2
		except:
			return dic_total_mutation_types_t1, dic_total_mutation_types_t2


def create_mutation_types_barplot(args, dic_mutation_types_t1, dic_mutation_types_t2, counting_method):
	plt.clf()
	values_t1 = []
	values_t2 = []
	labels = []
	for key in dic_mutation_types_t1.keys():
		if key in dic_mutation_types_t2.keys():
			values_t1.append(dic_mutation_types_t1[key])
			values_t2.append(dic_mutation_types_t2[key])
			if key not in labels:
				labels.append(key)
		elif key not in dic_mutation_types_t2.keys():
			values_t1.append(dic_mutation_types_t1[key])
			values_t2.append(0)
			if key not in labels:
				labels.append(key)

	x = np.arange(len(labels))
	num_bars = len(labels)
	if num_bars <= 2:
		bar_width = 0.1  # Reduce the bar width for one or two bars
		bar_gap = 0.1
	else:
		bar_width = 0.2
		bar_gap = 0.1

	fig, ax = plt.subplots(figsize=(13, 8))

	rects1 = ax.bar(x - bar_width - bar_gap / 2, values_t1, bar_width, label='time 1', color=(0, 0.5, 0), edgecolor='black', linewidth=1, align="center")
	rects2 = ax.bar(x + bar_gap / 2, values_t2, bar_width, label='time 2', color=(0, 0, 0.5), edgecolor='black', linewidth=1, align="center")

	for rect in rects1:
		height = rect.get_height()
		ax.annotate('{}'.format(height),
					xy=(rect.get_x() + rect.get_width() / 2, height),
					xytext=(0, 3),
					textcoords="offset points",
					ha='center', va='bottom',
					fontsize=11)

	for rect in rects2:
		height = rect.get_height()
		ax.annotate('{}'.format(height),
					xy=(rect.get_x() + rect.get_width() / 2, height),
					xytext=(0, 3),
					textcoords="offset points",
					ha='center', va='bottom',
					fontsize=11)

	ax.set_ylabel('Count', fontsize=13, labelpad=10)
	mode = counting_method.lower()
	mode = mode.capitalize()
	ax.set_title(mode + ' mutation types comparison between time 1 and time 2\n(crossover of all patients from the cohort)', fontsize=16, pad=10)
	ax.set_xticks(x)
	ax.set_xticklabels(labels, fontsize=13)
	ax.set_ylim(0, max(max(values_t1), max(values_t2)) * 1.1)

	offset = bar_width / 2
	ax.set_xticks(x - offset, minor=False)
	ax.set_xticklabels(labels, minor=False)
	plt.tick_params(bottom=False)

	ax.legend(fontsize=14, edgecolor='white')
	plt.tight_layout()

	file_formats = args['M_types_plot_format(s)'].upper()
	path = args['output_path'] + 'merge/mutation_types'
	if 'PNG' in file_formats:
		plt.savefig(path + '.png', dpi=400)
	if 'PDF' in file_formats:
		plt.savefig(path + '.pdf', dpi=400)
	if 'SVG' in file_formats:
		plt.savefig(path + '.svg', dpi=400)
	if 'JPG' in file_formats:
		plt.savefig(path + '.jpg', dpi=400)

	plt.close()


def create_mutation_subtypes_barplot(args, dic_mutation_subtypes_t1, dic_mutation_subtypes_t2, counting_method):
	if '.' in dic_mutation_subtypes_t1.keys():
		dic_mutation_subtypes_t1['unknown'] += dic_mutation_subtypes_t1['.']
		del dic_mutation_subtypes_t1['.']
	if '.' in dic_mutation_subtypes_t2.keys():
		if 'unknown' not in dic_mutation_subtypes_t2.keys():
			dic_mutation_subtypes_t2['unknown'] = 0
		dic_mutation_subtypes_t2['unknown'] += dic_mutation_subtypes_t2['.']
		del dic_mutation_subtypes_t2['.']

	plt.clf()
	for key in dic_mutation_subtypes_t1.keys():
		if key not in dic_mutation_subtypes_t2.keys():
			dic_mutation_subtypes_t2[key] = 0
	labels = [label for label in dic_mutation_subtypes_t1 if dic_mutation_subtypes_t1[label] != 0 and dic_mutation_subtypes_t2[label] != 0]
	values_t1 = [dic_mutation_subtypes_t1[label] for label in labels]
	values_t2 = [dic_mutation_subtypes_t2[label] for label in labels]

	num_bars = len(labels)
	if num_bars <= 2:
		bar_width = 0.1  # Reduce the bar width for one or two bars
		bar_gap = 0.1
	else:
		bar_width = 0.2
		bar_gap = 0.1

	fig, ax = plt.subplots(figsize=(13, 8))

	x_indexes = np.arange(len(labels))
	x1 = x_indexes - bar_width / 2 - bar_gap / 2
	x2 = x_indexes + bar_width / 2 + bar_gap / 2

	rects1 = ax.bar(x1, values_t1, bar_width, label='time 1', color=(0, 0.5, 0), edgecolor='black', linewidth=1, align="center")
	rects2 = ax.bar(x2, values_t2, bar_width, label='time 2', color=(0, 0, 0.5), edgecolor='black', linewidth=1, align="center")

	for rect in rects1:
		height = rect.get_height()
		ax.annotate('{}'.format(height),
					xy=(rect.get_x() + rect.get_width() / 2, height),
					xytext=(0, 3),
					textcoords="offset points",
					ha='center', va='bottom',
					fontsize=11)

	for rect in rects2:
		height = rect.get_height()
		ax.annotate('{}'.format(height),
					xy=(rect.get_x() + rect.get_width() / 2, height),
					xytext=(0, 3),
					textcoords="offset points",
					ha='center', va='bottom',
					fontsize=11)

	ax.set_ylabel('Count', fontsize=14.5, labelpad=15)
	ax.set_title('Mutation subtypes comparison between time 1 and time 2\n(crossover of all patients from the cohort)', fontsize=16, pad=10)
	ax.set_xticks(x_indexes)

	if num_bars <= 7:
		ax.set_xticklabels(labels, rotation=0, fontsize=11, ha='center')
	else:
		ax.set_xticklabels(labels, rotation=90, fontsize=11, ha='center')  # Adjust ha='right'
	ax.set_ylim(0, max(max(values_t1), max(values_t2)) * 1.1)

	ax.tick_params(axis='y', labelsize=13)
	ax.legend(fontsize=14, edgecolor='black')
	plt.tight_layout()

	file_formats = args['M_subtypes_plot_format(s)'].upper()
	path = args['output_path'] + 'merge/mutation_subtypes'
	if 'PNG' in file_formats:
		plt.savefig(path + '.png', dpi=400)
	if 'PDF' in file_formats:
		plt.savefig(path + '.pdf', dpi=400)
	if 'SVG' in file_formats:
		plt.savefig(path + '.svg', dpi=400)
	if 'JPG' in file_formats:
		plt.savefig(path + '.jpg', dpi=400)

	plt.close()

def improve_df_style(df_variants_merge, path):
	wb = Workbook()
	ws = wb.active

	for r_idx, row in enumerate(dataframe_to_rows(df_variants_merge, index=False, header=True), 1):
		for c_idx, value in enumerate(row, 1):
			cell = ws.cell(row=r_idx, column=c_idx, value=value)

			cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
			cell.alignment = Alignment(horizontal='center', vertical='center')

			if r_idx == 1 and value:
				cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
				cell.font = Font(bold=True)

			length = len(str(value))
			if ws.column_dimensions[get_column_letter(c_idx)].width < length + 2:
				ws.column_dimensions[get_column_letter(c_idx)].width = length + 2

	if 'variant' in path.lower():
		mutation_keywords = ['nonsynonymous', 'stopgain', 'stoploss', 'frameshift']

		highlight_colors = {
			'green': "FFA6FF00",  # Light green
			'yellow': "FFFFFF00",  # Light yellow
			'orange': "FFFFC000",  # Light orange
			'red': "FFFF4D4D",  # Light red
		}

		# DP ranges
		values = df_variants_merge['DP'].tolist()
		numeric_values = [float(value) for value in values]
		quartile1 = np.percentile(numeric_values, 25)
		quartile2 = np.percentile(numeric_values, 50)
		quartile3 = np.percentile(numeric_values, 75)
		DP_range1 = [0, quartile1]
		DP_range2 = [quartile1, quartile2]
		DP_range3 = [quartile2, quartile3]
		DP_range4 = [quartile3, max(numeric_values)]

		# AD ranges
		values = df_variants_merge['AD'].tolist()
		numeric_values = [float(value) for value in values]
		quartile1 = np.percentile(numeric_values, 25)
		quartile2 = np.percentile(numeric_values, 50)
		quartile3 = np.percentile(numeric_values, 75)
		AD_range1 = [0, quartile1]
		AD_range2 = [quartile1, quartile2]
		AD_range3 = [quartile2, quartile3]
		AD_range4 = [quartile3, max(numeric_values)]

		# VAF sample ranges
		values = df_variants_merge['VAF sample'].tolist()
		values = [value for value in values if value != 'not found']
		numeric_values = [float(value) for value in values]
		quartile1 = np.percentile(numeric_values, 25)
		quartile2 = np.percentile(numeric_values, 50)
		quartile3 = np.percentile(numeric_values, 75)
		VAF_sample_range1 = [0, quartile1]
		VAF_sample_range2 = [quartile1, quartile2]
		VAF_sample_range3 = [quartile2, quartile3]
		VAF_sample_range4 = [quartile3, max(numeric_values)]

		# VAF pop ranges
		values = df_variants_merge['VAF pop'].tolist()
		values = [value for value in values if value != 'not found']
		numeric_values = [float(value) for value in values]
		quartile1 = np.percentile(numeric_values, 25)
		quartile2 = np.percentile(numeric_values, 50)
		quartile3 = np.percentile(numeric_values, 75)
		VAF_pop_range1 = [0, quartile1]
		VAF_pop_range2 = [quartile1, quartile2]
		VAF_pop_range3 = [quartile2, quartile3]
		VAF_pop_range4 = [quartile3, max(numeric_values)]

		# SIFT ranges
		SIFT_range1 = [-1, 0.05]
		SIFT_range2 = [0.05, 0.1]
		SIFT_range3 = [0.1, 0.5]
		SIFT_range4 = [0.5, 1.5]

		# PolyPhen2 score ranges
		PolyPhen2_range1 = [-1, 0.453]
		PolyPhen2_range2 = [0.453, 0.7]
		PolyPhen2_range3 = [0.7, 0.956]
		PolyPhen2_range4 = [0.956, 1.5]

		header_row = ws[1]
		column_names = [cell.value for cell in header_row]
		for row in ws.iter_rows(min_row=2):
			for cell, column_name in zip(row, column_names):
				if column_name == 'Mutation':
					if any(keyword in cell.value for keyword in mutation_keywords):
						cell.fill = PatternFill(start_color=highlight_colors['green'], end_color=highlight_colors['green'], fill_type="solid")
					else:
						cell.fill = PatternFill(start_color=highlight_colors['red'], end_color=highlight_colors['red'], fill_type="solid")
				elif column_name == 'DP':
					if DP_range1[0] <= int(cell.value) <= DP_range1[1]:
						cell.fill = PatternFill(start_color=highlight_colors['red'], end_color=highlight_colors['red'], fill_type="solid")
					elif DP_range2[0] <= int(cell.value) <= DP_range2[1]:
						cell.fill = PatternFill(start_color=highlight_colors['orange'], end_color=highlight_colors['orange'], fill_type="solid")
					elif DP_range3[0] <= int(cell.value) <= DP_range3[1]:
						cell.fill = PatternFill(start_color=highlight_colors['yellow'], end_color=highlight_colors['yellow'], fill_type="solid")
					elif DP_range4[0] <= int(cell.value) <= DP_range4[1]:
						cell.fill = PatternFill(start_color=highlight_colors['green'], end_color=highlight_colors['green'], fill_type="solid")
				elif column_name == 'AD':
					if AD_range1[0] <= int(cell.value) <= AD_range1[1]:
						cell.fill = PatternFill(start_color=highlight_colors['red'], end_color=highlight_colors['red'], fill_type="solid")
					elif AD_range2[0] <= int(cell.value) <= AD_range2[1]:
						cell.fill = PatternFill(start_color=highlight_colors['orange'], end_color=highlight_colors['orange'], fill_type="solid")
					elif AD_range3[0] <= int(cell.value) <= AD_range3[1]:
						cell.fill = PatternFill(start_color=highlight_colors['yellow'], end_color=highlight_colors['yellow'], fill_type="solid")
					elif AD_range4[0] <= int(cell.value) <= AD_range4[1]:
						cell.fill = PatternFill(start_color=highlight_colors['green'], end_color=highlight_colors['green'], fill_type="solid")
				elif column_name == 'VAF sample':
					if 'not found' in str(cell.value):
						pass
					elif VAF_sample_range1[0] <= float(cell.value) <= VAF_sample_range1[1]:
						cell.fill = PatternFill(start_color=highlight_colors['green'], end_color=highlight_colors['green'], fill_type="solid")
					elif VAF_sample_range2[0] <= float(cell.value) <= VAF_sample_range2[1]:
						cell.fill = PatternFill(start_color=highlight_colors['yellow'], end_color=highlight_colors['yellow'], fill_type="solid")
					elif VAF_sample_range3[0] <= float(cell.value) <= VAF_sample_range3[1]:
						cell.fill = PatternFill(start_color=highlight_colors['orange'], end_color=highlight_colors['orange'], fill_type="solid")
					elif VAF_sample_range4[0] <= float(cell.value) <= VAF_sample_range4[1]:
						cell.fill = PatternFill(start_color=highlight_colors['red'], end_color=highlight_colors['red'], fill_type="solid")
				elif column_name == 'VAF pop':
					if cell.value == 'not found':
						pass
					elif VAF_pop_range1[0] <= float(cell.value) <= VAF_pop_range1[1]:
						cell.fill = PatternFill(start_color=highlight_colors['green'], end_color=highlight_colors['green'], fill_type="solid")
					elif VAF_pop_range2[0] <= float(cell.value) <= VAF_pop_range2[1]:
						cell.fill = PatternFill(start_color=highlight_colors['yellow'], end_color=highlight_colors['yellow'], fill_type="solid")
					elif VAF_pop_range3[0] <= float(cell.value) <= VAF_pop_range3[1]:
						cell.fill = PatternFill(start_color=highlight_colors['orange'], end_color=highlight_colors['orange'], fill_type="solid")
					elif VAF_pop_range4[0] <= float(cell.value) <= VAF_pop_range4[1]:
						cell.fill = PatternFill(start_color=highlight_colors['red'], end_color=highlight_colors['red'], fill_type="solid")
				elif column_name == 'SIFT score':
					if cell.value == 'not found':
						pass
					elif SIFT_range1[0] <= float(cell.value) <= SIFT_range1[1]:
						cell.fill = PatternFill(start_color=highlight_colors['green'], end_color=highlight_colors['green'], fill_type="solid")
					elif SIFT_range2[0] <= float(cell.value) <= SIFT_range2[1]:
						cell.fill = PatternFill(start_color=highlight_colors['yellow'], end_color=highlight_colors['yellow'], fill_type="solid")
					elif SIFT_range3[0] <= float(cell.value) <= SIFT_range3[1]:
						cell.fill = PatternFill(start_color=highlight_colors['orange'], end_color=highlight_colors['orange'], fill_type="solid")
					elif SIFT_range4[0] <= float(cell.value) <= SIFT_range4[1]:
						cell.fill = PatternFill(start_color=highlight_colors['red'], end_color=highlight_colors['red'], fill_type="solid")
				elif column_name == 'Polyphen2 score':
					if cell.value == 'not found':
						pass
					elif PolyPhen2_range1[0] <= float(cell.value) <= PolyPhen2_range1[1]:
						cell.fill = PatternFill(start_color=highlight_colors['red'], end_color=highlight_colors['green'], fill_type="solid")
					elif PolyPhen2_range2[0] <= float(cell.value) <= PolyPhen2_range2[1]:
						cell.fill = PatternFill(start_color=highlight_colors['orange'], end_color=highlight_colors['yellow'], fill_type="solid")
					elif PolyPhen2_range3[0] <= float(cell.value) <= PolyPhen2_range3[1]:
						cell.fill = PatternFill(start_color=highlight_colors['yellow'], end_color=highlight_colors['orange'], fill_type="solid")
					elif PolyPhen2_range4[0] <= float(cell.value) <= PolyPhen2_range4[1]:
						cell.fill = PatternFill(start_color=highlight_colors['green'], end_color=highlight_colors['red'], fill_type="solid")
				elif column_name == 'SIFT pred':
					if 'deleterious' in cell.value.lower():
						cell.fill = PatternFill(start_color=highlight_colors['green'], end_color=highlight_colors['green'], fill_type="solid")
					elif 'tolerated' in cell.value.lower():
						cell.fill = PatternFill(start_color=highlight_colors['red'], end_color=highlight_colors['red'], fill_type="solid")
				elif column_name == 'Polyphen2 pred':
					if 'probably' in cell.value.lower():
						cell.fill = PatternFill(start_color=highlight_colors['green'], end_color=highlight_colors['green'], fill_type="solid")
					elif 'possibly' in cell.value.lower():
						cell.fill = PatternFill(start_color=highlight_colors['yellow'], end_color=highlight_colors['yellow'], fill_type="solid")
					elif 'benign' in cell.value.lower():
						cell.fill = PatternFill(start_color=highlight_colors['orange'], end_color=highlight_colors['orange'], fill_type="solid")
	wb.save(path)

def write_stats_merge(args, df_variants_M, pairs):
	out_stats = args['output_path'] + 'merge/merged_stats.txt'
	df_variants_M.set_index('Variant', inplace=True)
	variants_dict = df_variants_M.to_dict(orient='index')
	mutation_counts = {'synonymous_SNV': 0, 'nonsynonymous_SNV': 0, 'stopgain': 0, 'startloss': 0, 'stoploss': 0, 'nonframeshift_insertion': 0,
					   'frameshift_insertion': 0, 'nonframeshift_deletion': 0, 'frameshift_deletion': 0,
					   'nonframeshift_substitution': 0, 'frameshift_substitution': 0}
	chromosome_counts = {'chr1': 0, 'chr2': 0, 'chr3': 0, 'chr4': 0, 'chr5': 0, 'chr6': 0, 'chr7': 0, 'chr8': 0, 'chr9': 0, 'chr10': 0,
						 'chr11': 0, 'chr12': 0, 'chr13': 0, 'chr14': 0, 'chr15': 0, 'chr16': 0, 'chr17': 0, 'chr18': 0, 'chr19': 0, 'chr20': 0,
						 'chr21': 0, 'chr22': 0, 'chrX': 0, 'chrY': 0, 'chrM': 0}
	gene_counts = {}
	pair_counts = {elem: 0 for elem in pairs}
	for variant_id, variant_info in variants_dict.items():
		for key in mutation_counts.keys():
			try:
				if key == variant_info['Mutation']:
					mutation_counts[key] += 1
			except:
				print('not annotated by ANNOVAR')
		for key in chromosome_counts.keys():
			if key == variant_info['Chr']:
				chromosome_counts[key] += 1
		for key in pair_counts.keys():
			if key in variant_info['Patient(s)']:
				pair_counts[key] += 1
		if variant_info['Gene'] not in gene_counts.keys():
			gene_counts[variant_info['Gene']] = 1
		else:
			gene_counts[variant_info['Gene']] += 1

	sorted_gene_counts = dict(sorted(gene_counts.items(), key=lambda x: x[1], reverse=True))
	top_10_genes_with_values = {gene: sorted_gene_counts[gene] for gene in list(sorted_gene_counts.keys())[:10]}
	sorted_chromosome_counts = dict(sorted(chromosome_counts.items(), key=lambda x: x[1], reverse=True))
	top_10_chromosomes_with_values = {chromosome: sorted_chromosome_counts[chromosome] for chromosome in list(sorted_chromosome_counts.keys())[:10]}
	sorted_pair_counts = dict(sorted(pair_counts.items(), key=lambda x: x[1], reverse=True))

	min_var_patients = args['min_patients_threshold_for_variants_upset_plot']
	found_in_all_patients = []
	found_in_at_least_x_patients = []
	for index, row in df_variants_M.iterrows():
		if row['Count'] == len(pairs):
			found_in_all_patients.append(index)
		if row['Count'] >= len(min_var_patients):
			found_in_at_least_x_patients.append(index)

	dict_variants_counts = {}
	for index, row in df_variants_M.iterrows():
		variant = index
		count = int(row['Count'])
		if count > 1:
			dict_variants_counts[variant] = count

	avg_DP = round(df_variants_M['DP'].mean())
	try:
		avg_QUALITY = round(df_variants_M['QUALITY'].mean())
	except:
		pass
		# avg_QUALITY = 'not found'
		# print('not annotated by ANNOVAR')
	avg_AD = round(df_variants_M['AD'].mean())
	if not all(value == 'not found' for value in df_variants_M['SB']):
		avg_SB = round(df_variants_M['SB'].mean(), 3)
	try:
		avg_VAF_sample = round(df_variants_M['VAF sample'].mean(), 3)
	except:
		pass
		# avg_VAF_sample = 'not found'
		print('no VAF sample found')
	try:
		avg_VAF_pop = round(df_variants_M['VAF pop'].mean(), 5)
	except:
		pass
		# avg_VAF_pop = 'not found'
		# print('no VAF pop found')

	with open(out_stats, 'w') as o:
		o.write(f'{len(found_in_all_patients)} mutations appeared or disappeared in all patients of the cohort.\n')
		o.write(f'{len(found_in_at_least_x_patients)} mutations appeared or disappeared in at least {min_var_patients} patients of the cohort.\n\n')
		o.write('--- Mutation subtypes ---\n')
		if mutation_counts['synonymous_SNV'] > 0:
			o.write(f'synonymous SNV: {mutation_counts["synonymous_SNV"]}\t')
		if mutation_counts['nonsynonymous_SNV'] > 0:
			o.write(f'nonsynonymous SNV: {mutation_counts["nonsynonymous_SNV"]}\n')
		if mutation_counts['stopgain'] > 0:
			o.write(f'stopgain: {mutation_counts["stopgain"]}\t')
		if mutation_counts['startloss'] > 0:
			o.write(f'startloss: {mutation_counts["startloss"]}\t')
		if mutation_counts['stoploss'] > 0:
			o.write(f'stoploss: {mutation_counts["stoploss"]}\n')
		if mutation_counts['nonframeshift_insertion'] > 0:
			o.write(f'nonframeshift insertion: {mutation_counts["nonframeshift_insertion"]}\t')
		if mutation_counts['frameshift_insertion'] > 0:
			o.write(f'frameshift insertion: {mutation_counts["frameshift_insertion"]}')
		if mutation_counts['nonframeshift_deletion'] > 0:
			o.write(f'\nnonframeshift deletion: {mutation_counts["nonframeshift_deletion"]}\t')
		if mutation_counts['frameshift_deletion'] > 0:
			o.write(f'frameshift deletion: {mutation_counts["frameshift_deletion"]}')
		if mutation_counts['nonframeshift_substitution'] > 0:
			o.write(f'\nnonframeshift substitution: {mutation_counts["nonframeshift_substitution"]}\t')
		if mutation_counts['frameshift_substitution'] > 0:
			o.write(f'frameshift substitution: {mutation_counts["frameshift_substitution"]}')

		o.write('\n\n--- Number of variants per patient ---\n')
		for pair, count in sorted_pair_counts.items():
			o.write(f'{pair}: {count}\n')

		o.write('\n--- Variants characteristics ---\n')
		o.write(f'Average DP: {avg_DP}\t')
		o.write(f'\tAverage AD: {avg_AD}\n')
		try:
			o.write(f'Average QUAL: {avg_QUALITY}\t')
			o.write(f'Average SB: {avg_SB}\n')
		except:
			print('not annotated by ANNOVAR')
		try:
			o.write(f'Average VAF sample: {avg_VAF_sample}\t')
		except:
			print('no VAF sample found')
		try:
			o.write(f'Average VAF pop: {avg_VAF_pop}')
		except:
			# print('no VAF pop found')
			pass

		o.write('\n\n--- 10 most mutated chromosomes ---\n')
		for chr, count in top_10_chromosomes_with_values.items():
			if count > 0:
				o.write(f'{chr}: {count}\n')

		o.write('\n--- 10 most mutated genes ---\n')
		for gene, count in top_10_genes_with_values.items():
			o.write(f'{gene}: {count}\n')

def create_variants_table(args):
	comparison_path = args['output_path'] + 'comparisons/'
	file_paths_with_variants = []
	allowed_extensions = ['.xlsx', '.tsv', '.csv']
	for folder_name in os.listdir(comparison_path):
		folder_path = os.path.join(comparison_path, folder_name)
		if os.path.isdir(folder_path):
			for file_name in os.listdir(folder_path):
				if "variants" in file_name and any(ext in file_name for ext in allowed_extensions):
					file_path = os.path.join(folder_path, file_name)
					file_paths_with_variants.append(file_path)

	dict_dataframes = {}
	for file_path in file_paths_with_variants:
		file_name = os.path.splitext(os.path.basename(file_path))[0]
		_, extension = os.path.splitext(file_path)
		if extension == '.xlsx':
			df = pd.read_excel(file_path, index_col='Variant')
		elif extension == '.csv':
			df = pd.read_csv(file_path, index_col='Variant')
		elif extension == '.tsv':
			df = pd.read_csv(file_path, sep='\t', index_col='Variant')
		dict_from_df = df.to_dict(orient='index')
		dict_dataframes[file_name] = dict_from_df
	common_suffix = os.path.commonprefix([key[::-1] for key in dict_dataframes.keys()])[::-1]
	selection = '_' + args['variants_selection_approach']
	dict_dataframes = {key.split(selection)[0]: value for key, value in dict_dataframes.items()}

	dict_variants_count = {}
	dict_variants_pairs = {}

	# Create the table with all variants
	for key in dict_dataframes.keys():
		for variant in dict_dataframes[key]:
			old_columns = list(dict_dataframes[key][variant].keys())
			break

	variants, chr, pos, ref, alt, genes, time, patients, mutation, DP, QUALITY, AD, SB, VAF_sample, VAF_pop, sift_scores, sift_preds, polyphen_scores, polyphen_preds = \
		([] for _ in range(19))

	i = 0
	for patient in dict_dataframes:
		for variant in dict_dataframes[patient]:
			if variant not in dict_variants_count:
				dict_variants_count[variant] = 1
				dict_variants_pairs[variant] = [patient]
			else:
				dict_variants_count[variant] += 1
				dict_variants_pairs[variant].append(patient)
			variants.append(variant)
			chr.append(dict_dataframes[patient][variant]['chr'])
			pos.append(dict_dataframes[patient][variant]['pos'])
			ref.append(dict_dataframes[patient][variant]['ref'])
			alt.append(dict_dataframes[patient][variant]['alt'])
			time.append(dict_dataframes[patient][variant]['time'])
			genes.append(dict_dataframes[patient][variant]['gene(s)'])
			patients.append(patient)
			if 'Mutation' in old_columns:
				mutation.append(dict_dataframes[patient][variant]['Mutation'])
			DP.append(dict_dataframes[patient][variant]['DP'])
			try:
				QUALITY.append(dict_dataframes[patient][variant]['QUALITY'])
			except:
				# print('no quality in vcf')
				pass
			AD.append(dict_dataframes[patient][variant]['AD'])
			try:
				SB.append(dict_dataframes[patient][variant]['SB'])
			except:
				SB.append('not found')
				# print('no SB found')
			try:
				sift_scores.append(dict_dataframes[patient][variant]['SIFT score'])
			except:
				sift_scores.append('not found')
				print('no sift score')
			try:
				sift_preds.append(dict_dataframes[patient][variant]['SIFT pred'])
			except:
				sift_preds.append('not found')
				print('no sift pred')
			try:
				polyphen_scores.append(dict_dataframes[patient][variant]['PolyPhen2 score'])
			except:
				polyphen_scores.append('not found')
				print('no polyphen score')
			try:
				polyphen_preds.append(dict_dataframes[patient][variant]['PolyPhen2 pred'])
			except:
				polyphen_preds.append('not found')
				print('no polyphen pred')

			if 'VAF sample' in old_columns:
				VAF_sample.append(dict_dataframes[patient][variant]['VAF sample'])
			if 'VAF pop' in old_columns:
				VAF_pop.append(dict_dataframes[patient][variant]['VAF pop'])
			i += 1

	dict_for_table = {col: lst for col, lst in zip(['Variant', 'Chr', 'Pos', 'Ref', 'Alt', 'Gene', 'Time', 'Patient(s)', 'Mutation', 'DP', 'QUALITY', 'AD', 'SB', 'VAF sample', 'VAF pop',
													'SIFT score', 'SIFT pred', 'Polyphen2 score', 'Polyphen2 pred'],
												   [variants, chr, pos, ref, alt, genes, time, patients, mutation, DP, QUALITY, AD, SB, VAF_sample, VAF_pop, sift_scores,
													sift_preds, polyphen_scores, polyphen_preds]) if lst}

	l_variants = dict_for_table['Variant']
	duplicates = []

	element_count = {}
	for item in l_variants:
		if item in element_count:
			element_count[item] += 1
		else:
			element_count[item] = 1
	for item, count in element_count.items():
		if count > 1:
			duplicates.append(item)

	dict_for_table_with_counts = {}
	j = 0
	for i in range(len(dict_for_table['Variant'])):
		variant = dict_for_table['Variant'][i]
		if variant not in dict_for_table_with_counts.keys():
			dict_for_table_with_counts[variant] = {key: value[i] for key, value in dict_for_table.items() if key != 'Variant' or key != 'Patient(s)'}
			dict_for_table_with_counts[variant]['Count'] = 1
			dict_for_table_with_counts[variant]['Patient(s)'] = []
			dict_for_table_with_counts[variant]['Patient(s)'].append(dict_for_table['Patient(s)'][i])
			try:
				dict_for_table_with_counts[variant]['Mutation'] = []
				dict_for_table_with_counts[variant]['Mutation'].append(dict_for_table['Mutation'][i])
			except:
				print('not annotated by ANNOVAR')
		else:
			dict_for_table_with_counts[variant]['Count'] += 1
			dict_for_table_with_counts[variant]['Patient(s)'].append(dict_for_table['Patient(s)'][i])
			try:
				dict_for_table_with_counts[variant]['Mutation'].append(dict_for_table['Mutation'][i])
			except:
				print('not annotated by ANNOVAR')
		j += 1

	df_variants_merge = pd.DataFrame(dict_for_table_with_counts).transpose()
	i = 0
	for i in range(len(df_variants_merge['Patient(s)'])):
		df_variants_merge['Patient(s)'][i] = str(df_variants_merge['Patient(s)'][i]).replace('[', '').replace(']', '').replace("'", '')
		try:
			df_variants_merge['Mutation'][i] = str(df_variants_merge['Mutation'][i]).replace('[', '').replace(']', '').replace("'", '')
		except:
			print('not annotated by ANNOVAR')

	column_order_priority = [
		['Variant', 'Chr', 'Pos', 'Ref', 'Alt', 'Gene', 'Mutation', 'Count', 'Time', 'DP', 'QUALITY', 'AD', 'SB', 'VAF sample', 'VAF pop', 'SIFT score', 'SIFT pred',
		 'Polyphen2 score', 'Polyphen2 pred', 'Patient(s)'],
		['Variant', 'Chr', 'Pos', 'Ref', 'Alt', 'Gene', 'Mutation', 'Count', 'Time', 'DP', 'AD', 'SB', 'VAF sample', 'VAF pop', 'SIFT score', 'SIFT pred',
		 'Polyphen2 score', 'Polyphen2 pred', 'Patient(s)'],
		['Variant', 'Chr', 'Pos', 'Ref', 'Alt', 'Gene', 'Mutation', 'Count', 'Time', 'DP', 'QUALITY', 'AD', 'SB', 'VAF sample', 'VAF pop', 'Patient(s)'],
		['Variant', 'Chr', 'Pos', 'Ref', 'Alt', 'Gene', 'Mutation', 'Count', 'Time', 'DP', 'AD', 'SB', 'VAF sample', 'VAF pop', 'Patient(s)'],
		['Variant', 'Chr', 'Pos', 'Ref', 'Alt', 'Gene', 'Mutation', 'Count', 'Time', 'DP', 'QUALITY', 'AD', 'SB', 'VAF sample', 'Patient(s)'],
		['Variant', 'Chr', 'Pos', 'Ref', 'Alt', 'Gene', 'Mutation', 'Count', 'Time', 'DP', 'QUALITY', 'AD', 'SB', 'VAF pop', 'Patient(s)'],
		['Variant', 'Chr', 'Pos', 'Ref', 'Alt', 'Gene', 'Count', 'Time', 'DP', 'AD', 'SB', 'VAF sample', 'VAF pop', 'Patient(s)'],
		['Variant', 'Chr', 'Pos', 'Ref', 'Alt', 'Gene', 'Count', 'Time', 'DP', 'AD', 'SB', 'VAF sample', 'Patient(s)'],
		['Variant', 'Chr', 'Pos', 'Ref', 'Alt', 'Gene', 'Count', 'Time', 'DP', 'AD', 'SB', 'VAF pop', 'Patient(s)'],
		['Variant', 'Chr', 'Pos', 'Ref', 'Alt', 'Gene', 'Count', 'Time', 'DP', 'AD', 'SB', 'Patient(s)'],
	]

	for new_order in column_order_priority:
		if all(col in df_variants_merge.columns.tolist() for col in new_order):
			df_variants_merge = df_variants_merge[new_order]
			break

	try:
		if pd.isna(df_variants_merge['QUALITY']).all():
			df_variants_merge = df_variants_merge.drop(columns=['QUALITY'])
	except:
		pass
	try:
		df_variants_merge['VAF pop'] = df_variants_merge['VAF pop'].replace('not found', np.nan)
		df_variants_merge['VAF pop'] = df_variants_merge['VAF pop'].astype(float)
		df_variants_merge = df_variants_merge.sort_values('VAF pop', ascending=True)
		df_variants_merge['VAF pop'] = df_variants_merge['VAF pop'].fillna('not found')
		df_variants_merge['VAF pop'] = df_variants_merge['VAF pop'].apply(lambda x: "{:.2e}".format(x) if isinstance(x, float) else x)
	except:
		print('no VAF pop found')

	# df_variants_merge = df_variants_merge.sort_values(by='Count', ascending=False)
	formats = args['M_variants_table_format(s)'].upper()
	if 'XLSX' in formats:
		improve_df_style(df_variants_merge, args['output_path'] + 'merge/merged_variants' + '.xlsx')
		# df_variants_merge.to_excel(args['output_path'] + 'merge/' + 'variants' + '.xlsx', index=False)
	if 'CSV' in formats:
		df_variants_merge.to_csv(args['output_path'] + 'merge/merged_variants' + '.csv', index=False, sep='\t')
	if 'TSV' in formats:
		df_variants_merge.to_csv(args['output_path'] + 'merge/merged_variants' + '.tsv', index=False, sep='\t')

	l_variants = dict_for_table['Variant']

	return df_variants_merge


def create_merged_vcf(args, df_variants_M):
	df_variants_M = df_variants_M.sort_values(by=['Chr', 'Pos'])
	output_file = args['output_path'] + 'merge/merged_variants.vcf'
	with open(output_file, "w") as f:
		f.write("##fileformat=VCFv4.3\n")

		current_datetime = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
		f.write("##fileDate=" + current_datetime + "\n")

		# chr definitions
		contig_lines = [
			"##contig=<ID=chr1,length=249250621>\n",
			"##contig=<ID=chr2,length=242193529>\n",
			"##contig=<ID=chr3,length=198295559>\n",
			"##contig=<ID=chr4,length=190214555>\n",
			"##contig=<ID=chr5,length=181538259>\n",
			"##contig=<ID=chr6,length=170805979>\n",
			"##contig=<ID=chr7,length=159345973>\n",
			"##contig=<ID=chr8,length=145138636>\n",
			"##contig=<ID=chr9,length=138394717>\n",
			"##contig=<ID=chr10,length=133797422>\n",
			"##contig=<ID=chr11,length=135086622>\n",
			"##contig=<ID=chr12,length=133275309>\n",
			"##contig=<ID=chr13,length=114364328>\n",
			"##contig=<ID=chr14,length=107043718>\n",
			"##contig=<ID=chr15,length=101991189>\n",
			"##contig=<ID=chr16,length=90338345>\n",
			"##contig=<ID=chr17,length=83257441>\n",
			"##contig=<ID=chr18,length=80373285>\n",
			"##contig=<ID=chr19,length=58617616>\n",
			"##contig=<ID=chr20,length=64444167>\n",
			"##contig=<ID=chr21,length=46709983>\n",
			"##contig=<ID=chr22,length=50818468>\n",
			"##contig=<ID=chrX,length=156040895>\n",
			"##contig=<ID=chrY,length=57227415>\n",
			"##contig=<ID=chrM,length=16569>\n",
		]
		for line in contig_lines:
			f.write(line)

		# FORMAT definitons
		f.write("##FORMAT=<ID=GT,Number=1,Type=String,Description=\"Genotype\">\n")
		f.write("##FORMAT=<ID=PL,Number=G,Type=Integer,Description=\"Phred-scaled Genotype Likelihoods\">\n")
		f.write("##FORMAT=<ID=DP,Number=1,Type=Integer,Description=\"Total Depth\">\n")
		f.write("##FORMAT=<ID=AD,Number=.,Type=Integer,Description=\"Allelic Depths\">\n")
		f.write("##FORMAT=<ID=GP,Number=1,Type=String,Description=\"Genotype Probability\">\n")
		f.write("##FORMAT=<ID=GQ,Number=1,Type=Integer,Description=\"Genotype Quality\">\n")

		if args.get("variants_selection_approach", "").upper() == "CHANGE":
			f.write("##INFO=<ID=DP,Number=1,Type=Integer,Description=\"Total Depth\">\n")
			f.write("##INFO=<ID=AD,Number=.,Type=Integer,Description=\"Allelic Depths\">\n")
			f.write("##INFO=<ID=MQSBZ,Number=1,Type=Float,Description=\"MQSBZ Score\">\n")
		else:
			print('code to right : header for funco and snpeff')

		f.write(f"#CHROM\tPOS\tID\tREF\tALT\tQUAL\tFILTER\tINFO\tFORMAT\t{'merged_variants.vcf'}\n")

		for _, row in df_variants_M.iterrows():
			chrom = row['Chr']
			pos = row['Pos']
			ref = row['Ref']
			alt = row['Alt']
			chrom = chrom.replace(' ', '')
			ref = ref.replace(' ', '')
			alt = alt.replace(' ', '')
			DP = row['DP']

			anno, mutation, VAF_sample, VAF_pop = True, True, True, True

			try:
				QUALITY = row['QUALITY']
				mutation = row['Mutation']
			except:
				anno = False
				pass
			try:
				AD = str(int(row['AD'])).replace('|', ',')
			except:
				pass
			SB = row['SB']
			try:
				VAF_sample = row['VAF sample']
			except:
				pass
			try:
				VAF_pop = row['VAF pop']
			except:
				VAF_pop = False
				pass

			key = 'g.' + str(chrom) + ':' + str(pos) + ref + '>' + alt

			if anno and VAF_sample and VAF_pop:
				format_values = f"{DP}:{AD}:{VAF_sample}"
				f.write(f"{chrom}\t{pos}\t.\t{ref}\t{alt}\t{QUALITY}\tPASS\tg={key};DP={DP};SB={SB};AF={VAF_pop};mutation={mutation}\tDP:AD:VAF\t{format_values}\n")
			elif anno and VAF_sample and not VAF_pop:
				format_values = f"{DP}:{AD}:{VAF_sample}"
				f.write(f"{chrom}\t{pos}\t.\t{ref}\t{alt}\t{QUALITY}\tPASS\tg={key};DP={DP};SB={SB};mutation={mutation}\tDP:AD:VAF\t{format_values}\n")
			elif anno and not VAF_sample and VAF_pop:
				format_values = f"{DP}:{AD}"
				f.write(f"{chrom}\t{pos}\t.\t{ref}\t{alt}\t{QUALITY}\tPASS\tg={key};DP={DP};SB={SB};AF={VAF_pop};mutation={mutation}\tDP:AD\t{format_values}\n")
			elif anno and not VAF_sample and not VAF_pop:
				format_values = f"{DP}:{AD}"
				f.write(f"{chrom}\t{pos}\t.\t{ref}\t{alt}\t{QUALITY}\tPASS\tg={key};DP={DP};SB={SB};mutation={mutation}\tDP:AD\t{format_values}\n")
			elif not anno and VAF_sample and VAF_pop:
				format_values = f"{DP}:{AD}:{VAF_sample}"
				f.write(f"{chrom}\t{pos}\t.\t{ref}\t{alt}\t.\tPASS\tg={key};DP={DP};SB={SB};AF={VAF_pop}\tDP:AD:VAF\t{format_values}\n")
			elif not anno and VAF_sample and not VAF_pop:
				format_values = f"{DP}:{AD}:{VAF_sample}"
				f.write(f"{chrom}\t{pos}\t.\t{ref}\t{alt}\t.\tPASS\tg={key};DP={DP};SB={SB}\tDP:AD:VAF\t{format_values}\n")
	# print("VCF file created successfully.")


def create_VAF_pop_scatter_plot(args, df_variants_M):
	t1 = df_variants_M[df_variants_M['Time'] == 't1']
	t1_VAF_pop = t1['VAF pop'].tolist()
	t1_VAF_pop = [float(x) for x in t1_VAF_pop if x != 'not found']
	t2 = df_variants_M[df_variants_M['Time'] == 't2']
	t2_VAF_pop = t2['VAF pop'].tolist()
	t2_VAF_pop = [float(x) for x in t2_VAF_pop if x != 'not found']

	plt.clf()
	data = [t1_VAF_pop, t2_VAF_pop]
	fig, ax = plt.subplots()

	boxplot_colors = ['forestgreen', 'darkblue']
	boxplot = ax.boxplot(data, patch_artist=True, labels=['t1', 't2'])

	for patch, color in zip(boxplot['boxes'], boxplot_colors):
		patch.set_facecolor(color)

	ax.set_ylabel('VAF pop', labelpad=10)
	ax.set_title('Comparison of VAF pop between two times')

	no_stats = False

	# Perform t-test if n>10 in each time
	if len(t1_VAF_pop) > 10 and len(t2_VAF_pop) > 10:
		statistic, p_value = stats.ttest_ind(t1_VAF_pop, t2_VAF_pop)
	else:
		no_stats = True

	if not no_stats:
		if p_value < 0.05:
			if np.mean(t1_VAF_pop) > np.mean(t2_VAF_pop):
				conclusion = 't1 > t2'
			else:
				conclusion = 't1 < t2'
			p_value = 'p < 0.05 (*)'
			significance = 'significant difference'
		else:
			conclusion = 't1 = t2'
			p_value = 'p > 0.05'
			significance = 'no significant difference'

		significance_text = f"{conclusion}\n{significance}\np-value: {p_value}\n(t-test)"
		ax.text(0.5, 0.95, significance_text, transform=ax.transAxes,
				verticalalignment='top', horizontalalignment='center',
				bbox={'facecolor': 'white', 'edgecolor': 'black', 'pad': 5},
				fontsize=8)

	file_formats = args['M_VAF_plot_format(s)'].upper()
	path = args['output_path'] + 'merge/VAF_pop_scatter'

	formatter = ScalarFormatter(useMathText=True)
	formatter.set_powerlimits((0, 0))
	plt.gca().yaxis.set_major_formatter(formatter)

	if 'PNG' in file_formats:
		path = path + '.png'
		plt.savefig(path, dpi=400)
	if 'PDF' in file_formats:
		path = path + '.pdf'
		plt.savefig(path, dpi=400)
	if 'SVG' in file_formats:
		path = path + '.svg'
		plt.savefig(path, dpi=400)
	if 'JPG' in file_formats:
		path = path + '.jpg'
		plt.savefig(path, dpi=400)

	plt.close()


def create_VAF_sample_scatter_plot(args, df_variants_M):
	t1 = df_variants_M[df_variants_M['Time'] == 't1']
	t1_VAF_sample = t1['VAF sample'].tolist()
	t2 = df_variants_M[df_variants_M['Time'] == 't2']
	t2_VAF_sample = t2['VAF sample'].tolist()

	plt.clf()
	data = [t1_VAF_sample, t2_VAF_sample]
	fig, ax = plt.subplots()

	boxplot_colors = ['forestgreen', 'darkblue']
	boxplot = ax.boxplot(data, patch_artist=True, labels=['t1', 't2'])

	for patch, color in zip(boxplot['boxes'], boxplot_colors):
		patch.set_facecolor(color)

	ax.set_ylabel('VAF sample', labelpad=10)
	ax.set_title('Comparison of VAF sample between two times')

	no_stats = False

	# Perform t-test if n>10 in each time
	if len(t1_VAF_sample) > 10 and len(t2_VAF_sample) > 10:
		statistic, p_value = stats.ttest_ind(t1_VAF_sample, t2_VAF_sample)
	else:
		no_stats = True

	if not no_stats:
		if p_value < 0.05:
			if np.mean(t1_VAF_sample) > np.mean(t2_VAF_sample):
				conclusion = 't1 > t2'
			else:
				conclusion = 't1 < t2'
			p_value = 'p < 0.05 (*)'
			significance = 'significant difference'
		else:
			conclusion = 't1 = t2'
			p_value = 'p > 0.05'
			significance = 'no significant difference'

		significance_text = f"{conclusion}\n{significance}\np-value: {p_value}\n(t-test)"
		ax.text(0.5, 0.95, significance_text, transform=ax.transAxes,
				verticalalignment='top', horizontalalignment='center',
				bbox={'facecolor': 'white', 'edgecolor': 'black', 'pad': 5},
				fontsize=8)

	file_formats = args['M_VAF_plot_format(s)'].upper()
	path = args['output_path'] + 'merge/VAF_sample_scatter'

	formatter = ScalarFormatter(useMathText=True)
	formatter.set_powerlimits((0, 0))
	plt.gca().yaxis.set_major_formatter(formatter)

	if 'PNG' in file_formats:
		path = path + '.png'
		plt.savefig(path, dpi=400)
	if 'PDF' in file_formats:
		path = path + '.pdf'
		plt.savefig(path, dpi=400)
	if 'SVG' in file_formats:
		path = path + '.svg'
		plt.savefig(path, dpi=400)
	if 'JPG' in file_formats:
		path = path + '.jpg'
		plt.savefig(path, dpi=400)

	plt.close()


def merge_results(args, file_paths, category, output, infos, cytoband_file, chromosomes_output, step, nb_files, enrichment, logger):
	# Count mutation types and subtypes (and plot them)

	# mutation_types_counting_method = args['mutations_types_counting'].upper()
	mutation_types_counting_method = 'UNIQUE'
	if mutation_types_counting_method == 'UNIQUE':
		print('Counting and plotting unique mutations types and subtypes...')
		try:  # ANNOVAR
			dic_unique_mutation_types_t1, dic_unique_mutation_types_t2, dic_unique_mutation_subtypes_t1, dic_unique_mutation_subtypes_t2 = count_mutation_types(args, 'UNIQUE')
			create_mutation_types_barplot(args, dic_unique_mutation_types_t1, dic_unique_mutation_types_t2, 'UNIQUE')
			create_mutation_subtypes_barplot(args, dic_unique_mutation_subtypes_t1, dic_unique_mutation_subtypes_t2, 'UNIQUE')
		except:
			dic_unique_mutation_types_t1, dic_unique_mutation_types_t2 = count_mutation_types(args, 'UNIQUE')
			create_mutation_types_barplot(args, dic_unique_mutation_types_t1, dic_unique_mutation_types_t2, 'UNIQUE')
	elif mutation_types_counting_method == 'TOTAL':
		try:  # ANNOVAR
			dic_total_mutation_types_t1, dic_total_mutation_types_t2, dic_total_mutation_subtypes_t1, dic_total_mutation_subtypes_t2 = count_mutation_types(args, 'TOTAL')
			create_mutation_types_barplot(args, dic_total_mutation_types_t1, dic_total_mutation_types_t2, 'TOTAL')
			create_mutation_subtypes_barplot(args, dic_total_mutation_subtypes_t1, dic_total_mutation_subtypes_t2, 'TOTAL')
		except:
			dic_total_mutation_types_t1, dic_total_mutation_types_t2 = count_mutation_types(args, 'TOTAL')
			create_mutation_types_barplot(args, dic_total_mutation_types_t1, dic_total_mutation_types_t2, 'TOTAL', )

	# Get genes infos if not None from the 6 databases
	if infos:
		df_databases_info = get_informations_for_genes(args, infos, logger)

	names = set()
	df_all_genes_info = {}  # gene dictionary to create tsv union file (list format)

	dic_gene_fields = {'chr': '', 'start': '', 'end': '', 'weakness': [], 'mutation types': [], 'gb1': {}, 'cb1': {}, 'pb1': {}, 'gb2': {}, 'cb2': {}, 'pb2': {}, 'samples': []}
	if args['verbose_prints'].upper() == 'TRUE':
		print('Reading files...')

	if args['colored_execution'].upper() == 'TRUE' or args['colored_execution'].upper() == 'YES':
		color = '\033[38;2;255;152;0m'
		no_tqdm_bar = False
	else:
		color = ''
		no_tqdm_bar = True

	pbar_files = tqdm(total=int(nb_files / 2), disable=no_tqdm_bar, bar_format='{l_bar}{bar:30}{r_bar}', ncols=150, smoothing=1)
	pbar_files.set_description(color + f' -> Extracting information from \033[3mCompare\033[0m' + color + ' results (' + str(round(nb_files / 2)) + ' patients)')
	dic_unique_variants = {}
	dic_genes_in_pairs = {}

	for name in file_paths:
		if 'genes' in name:
			element = name.split("/")[-1]
			folder = name.split(element)[0]
			files = [f for f in os.listdir(folder) if os.path.isfile(os.path.join(folder, f))]
			for file in files:
				if 'genes' in file and 'xlsx' in file:
					name = folder + file
			id = name.split('sons/')[1].split('/')[0]
			dic_genes_in_pairs[id] = []
			if 'tsv' in name:
				df_from_compare = pd.read_csv(name, sep='\t', index_col=0)
			elif 'xlsx' in name:
				df_from_compare = pd.read_excel(name, index_col=0)
			elif 'csv' in name:
				df_from_compare = pd.read_csv(name, index_col=0)
			# if args['verbose_prints'].upper() == 'TRUE':
			#     print(name)
			if not name in names:
				names.add(name)

		col_passed1 = df_from_compare.columns.values.tolist()[5]
		col_passed2 = df_from_compare.columns.values.tolist()[9]
		logger.info(f'Processing of samples B1: {col_passed1} and B2: {col_passed2}')  # b1 and b2 are the names of the samples
		# row is an object: doesn't have gene column (used as index)
		selection = args['variants_selection_approach'].upper()
		union = False
		change = False
		common = False
		if selection == 'UNION':
			union = True
		elif selection == 'CHANGE':
			change = True
		elif selection == 'COMMON':
			common = True

		# if args['vcf_annotation_method'].upper() == 'FUNCOTATOR':
		# 	passed_index = 5  # because we removed index column
		# 	gb_index = 6
		# 	cb_index = 7
		# 	pb_index = 8
		# 	passed1_index = 9
		# 	gb1_index = 10
		# 	cb1_index = 11
		# 	pb1_index = 12

		# elif args['vcf_annotation_method'].upper() == 'ANNOVAR':
		if union:
			passed_index = 6
			mutation_type = 7
			gb_index = 8
			cb_index = 9
			pb_index = 10
			passed1_index = 11
			gb1_index = 12
			cb1_index = 13
			pb1_index = 14
		elif change:
			passed_index = 5
			mutation_type = 6
			gb_index = 7
			cb_index = 8
			pb_index = 9
			passed1_index = 10
			gb1_index = 11
			cb1_index = 12
			pb1_index = 13
		elif common:
			passed_index = 4
			mutation_type = 5
			gb_index = 6
			cb_index = 7
			pb_index = 8
			passed1_index = 9
			gb1_index = 10
			cb1_index = 11
			pb1_index = 12

		variants_count = []

		for index, row in df_from_compare.iterrows():
			if not index in dic_genes_in_pairs[id]:
				dic_genes_in_pairs[id].append(index)

			if index not in df_all_genes_info.keys():
				df_all_genes_info[index] = deepcopy(dic_gene_fields)  # index is the gene name

			if df_all_genes_info[index]['chr'] == '':
				df_all_genes_info[index]['chr'] = row['Chromosome']

			if df_all_genes_info[index]['start'] == '':
				try:
					df_all_genes_info[index]['start'] = int(row['Gene start position'])
				except:
					df_all_genes_info[index]['start'] = 'error'
			if df_all_genes_info[index]['end'] == '':
				try:
					df_all_genes_info[index]['end'] = int(row['Gene end position'])
				except:
					df_all_genes_info[index]['start'] = 'error'

			# df_all_genes_info[index]['weakness'].append(float(row['Gene weakness (in %)']))

			try:
				add_variants_to_dictionnary(df_all_genes_info, index, mutation_type, 'mutation types', row)
			except:
				print('not annotated with ANNOVAR')

			add_variants_to_dictionnary(df_all_genes_info, index, gb_index, 'gb1', row)
			add_variants_to_dictionnary(df_all_genes_info, index, cb_index, 'cb1', row)
			add_variants_to_dictionnary(df_all_genes_info, index, pb_index, 'pb1', row)
			add_variants_to_dictionnary(df_all_genes_info, index, gb1_index, 'gb2', row)
			add_variants_to_dictionnary(df_all_genes_info, index, cb1_index, 'cb2', row)
			add_variants_to_dictionnary(df_all_genes_info, index, pb1_index, 'pb2', row)
			df_all_genes_info[index]['samples'].append(name)

		pbar_files.update(1)
	pbar_files.close()

	dic_intermediate_genes = {}  # gene dictionary to create tsv union file (mean computation for each list)
	dic_patients_genes_lists = {}  # dictionary containing genes list for each sample - use to create the upset plot
	genes_pos_for_chromosomes = {}  # dictionary containing genes position for each chromosome - use to create the chromosomes plot
	genes_pos_for_chromosomes_t1 = {}  # s1 = sample 1
	genes_pos_for_chromosomes_t2 = {}  # s2 = sample 2

	if args['colored_execution'].upper() == 'TRUE' or args['colored_execution'].upper() == 'YES':
		color = '\033[38;2;255;152;0m'
		no_tqdm_bar = False
	else:
		color = ''
		no_tqdm_bar = True

	pbar_files = tqdm(total=len(df_all_genes_info.items()), disable=no_tqdm_bar, bar_format='{l_bar}{bar:30}{r_bar}', ncols=150, smoothing=1)
	pbar_files.set_description(color + f' -> Preparing information for \033[3mgenes_union \033[0m' + color + 'dataframe')

	dic_mutation_types = {}
	dic_mutation_subtypes = {}

	# counting_method = args['mutations_types_counting'].upper()
	counting_method = 'UNIQUE'
	if counting_method == 'UNIQUE':
		unique = True
	elif counting_method == 'TOTAL':
		total = True

	df_dataset = pd.read_excel(args['dataset_path'])

	time1_column = []
	time2_column = []
	pair_names_column = []

	for value in df_dataset[args['time1_column_name']]:
		if pd.notnull(value) and value not in ('NaN', 'naN') and value.strip() != '':
			time1_column.append(value)

	for value in df_dataset[args['time2_column_name']]:
		if pd.notnull(value) and value not in ('NaN', 'naN') and value.strip() != '':
			time2_column.append(value)

	for value in df_dataset[args['pair_names_column_name']]:
		if pd.notnull(value) and value not in ('NaN', 'naN') and str(value).strip() != '':
			try:
				pair_names_column.append(str(int(value)))
			except:
				pair_names_column.append(str(value))

	time1_column = [element.replace('.vcf', '') for element in time1_column]
	time2_column = [element.replace('.vcf', '') for element in time2_column]
	dict_pairs = {}
	for i in range(len(time1_column)):
		dict_pairs[time1_column[i] + '___' + time2_column[i]] = pair_names_column[i]
	reversed_pairs_dict = {value: key for key, value in dict_pairs.items()}

	patients_threshold = args['min_patients_threshold_for_dataframes']
	df_names = pd.read_excel(args['dataset_path'])
	ids_column = args['pair_names_column_name']
	is_id_column_filled = False
	if ids_column.upper() != 'NONE' and ids_column != 'NO' and ids_column != '' and ids_column != 'FALSE':
		is_id_column_filled = True
		ids = df_names[ids_column].dropna().tolist()
		names1 = df_names[args['time1_column_name']].dropna().tolist()
		names2 = df_names[args['time2_column_name']].dropna().tolist()
		names1 = [name.replace(".vcf", "") for name in names1]
		names2 = [name.replace(".vcf", "") for name in names2]
		dic_names = {}
		j = 0
		for i in range(len(ids)):
			key = f"{names1[j]}___{names2[j]}"
			dic_names[key] = str(ids[j])
			j += 1

	for k, v in df_all_genes_info.items():
		if not v['chr'] in genes_pos_for_chromosomes_t1.keys():
			genes_pos_for_chromosomes_t1[v['chr']] = []
		if not v['chr'] in genes_pos_for_chromosomes_t2.keys():
			genes_pos_for_chromosomes_t2[v['chr']] = []
		genes_pos_for_chromosomes[v['chr']] = [(v['start'], v['end'])] * int(len(v['samples']))

		genes_pos_for_chromosomes_t1[v['chr']] += [(v['start'], v['end'])] * len(v['gb1'])
		genes_pos_for_chromosomes_t2[v['chr']] += [(v['start'], v['end'])] * len(v['gb2'])
		# A CORRIGER : remettre gb1 en dic et si ca change rien en tournant, vérifier positions
		# * int(sum(list(v['gb1'].values())))

		if not k in dic_intermediate_genes.keys():
			dic_intermediate_genes[k] = {}
		reversed_pairs_dict = {value: key for key, value in dic_names.items()}
		in_pairs = []
		for id in dic_genes_in_pairs.keys():
			if k in dic_genes_in_pairs[id]:
				in_pairs.append(id)
		if is_id_column_filled:
			final_in_pairs = []
			for pair in in_pairs:
				try:
					final_in_pairs.append(dic_names[pair])
				except:
					final_in_pairs.append(reversed_pairs_dict[pair])
		else:
			final_in_pairs = in_pairs

		short_pair_names = []
		for pair in final_in_pairs:
			short_pair_names.append(dict_pairs[pair])
		dic_intermediate_genes[k]['Sample pair(s)'] = str(short_pair_names).replace('[', '').replace(']', '').replace('\'', '')

		for sample in v['samples']:
			if not sample in dic_patients_genes_lists.keys():  # i
				dic_patients_genes_lists[sample] = []
			dic_patients_genes_lists[sample].append(k)  # add only adds an element to the set if the element is not already present

		dic_intermediate_genes[k]['Chromosome'] = v['chr']
		dic_intermediate_genes[k]['Gene start position'] = v['start']
		dic_intermediate_genes[k]['Gene end position'] = v['end']
		dic_intermediate_genes[k]['Nb samples'] = int(len(v['samples']))

		# unique variants counting for each gene found in all samples from the whole dataset provided
		try:
			if k not in dic_unique_variants.keys():
				dic_unique_variants[k] = {}
				dic_unique_variants[k]['gb1'] = []
			for variant in df_all_genes_info[k]['gb1']:
				if variant not in dic_unique_variants[k]:
					dic_unique_variants[k]['gb1'].append(variant)
		except:
			if k not in dic_unique_variants.keys():
				dic_unique_variants[k] = {}
				dic_unique_variants[k]['gb1'] = []
		try:
			if 'gb2' not in dic_unique_variants[k].keys():
				dic_unique_variants[k]['gb2'] = []
			elif k not in dic_unique_variants.keys():
				dic_unique_variants[k] = {}
				dic_unique_variants[k]['gb2'] = []
			for variant in df_all_genes_info[k]['gb2']:
				if variant not in dic_unique_variants[k]:
					dic_unique_variants[k]['gb2'].append(variant)
		except:
			print(k)
			if k not in dic_unique_variants.keys():
				dic_unique_variants[k] = {}
				dic_unique_variants[k]['gb2'] = []

		# dic_intermediate_genes[k]['Total variants'] = len(set(v['gb1'].keys()).union(set(v['gb2'].keys())))  # gb1 and gb2 are dict with the number of mutation for each gene
		# dic_intermediate_genes[k]['Mean gene weakness'] = np.mean(v['weakness'])
		dic_intermediate_genes[k]['Unique variants'] = len(set(dic_unique_variants[k]['gb1']).union(set(dic_unique_variants[k]['gb2'])))
		dic_intermediate_genes[k]['Total variants'] = len(dic_unique_variants[k]['gb1']) + len(dic_unique_variants[k]['gb2'])
		max_times_variants_number = max(len(set(dic_unique_variants[k]['gb1'])), len(set(dic_unique_variants[k]['gb2'])))
		if not common and not change:
			dic_intermediate_genes[k]['Unique variation (%)'] = round(
				(len(set(dic_unique_variants[k]['gb2'])) - len(set(dic_unique_variants[k]['gb1']))) / max_times_variants_number * 100)
			dic_intermediate_genes[k]['Total variation (%)'] = round((len(dic_unique_variants[k]['gb2']) - len(dic_unique_variants[k]['gb1'])) / max_times_variants_number * 100)
		if not common:
			dic_intermediate_genes[k]['Δ Unique variants'] = len(set(dic_unique_variants[k]['gb2'])) - len(set(dic_unique_variants[k]['gb1']))
			dic_intermediate_genes[k]['Δ total variants'] = len(dic_unique_variants[k]['gb2']) - len(dic_unique_variants[k]['gb1'])
		dic_intermediate_genes[k]['Unique mutations (t1)'] = len(set(dic_unique_variants[k]['gb1']))
		dic_intermediate_genes[k]['Total mutations (t1)'] = len(dic_unique_variants[k]['gb1'])
		dic_intermediate_genes[k]['g.(t1)'] = '|'.join([str(variant) for variant in v['gb1']])
		dic_intermediate_genes[k]['c.(t1)'] = '|'.join([str(variant) for variant in v['cb1']])
		dic_intermediate_genes[k]['p.(t1)'] = '|'.join([str(variant) for variant in v['pb1']])
		dic_intermediate_genes[k]['Unique mutations (t2)'] = len(set(dic_unique_variants[k]['gb2']))
		dic_intermediate_genes[k]['Total mutations (t2)'] = len(dic_unique_variants[k]['gb2'])
		dic_intermediate_genes[k]['g.(t2)'] = '|'.join([str(variant) for variant in v['gb2']])
		dic_intermediate_genes[k]['c.(t2)'] = '|'.join([str(variant) for variant in v['cb2']])
		dic_intermediate_genes[k]['p.(t2)'] = '|'.join([str(variant) for variant in v['pb2']])
		# dic_intermediate_genes[k]['g.TPn union'] = '|'.join([str(k) + '(' + str(v) + ')' for k, v in v['gb1'].items()])
		# dic_intermediate_genes[k]['c.TPn union'] = '|'.join([str(k) + '(' + str(v) + ')' for k, v in v['cb1'].items()])
		# dic_intermediate_genes[k]['p.TPn union'] = '|'.join([str(k) + '(' + str(v) + ')' for k, v in v['pb1'].items()])
		# dic_intermediate_genes[k]['Nb Mutation TPn+1'] = len(set(v['gb2'].keys()))
		# dic_intermediate_genes[k]['g.TPn+1 union'] = '|'.join([str(k) + '(' + str(v) + ')' for k, v in v['gb2'].items()])
		# dic_intermediate_genes[k]['c.TPn+1 union'] = '|'.join([str(k) + '(' + str(v) + ')' for k, v in v['cb2'].items()])
		# dic_intermediate_genes[k]['p.TPn+1 union'] = '|'.join([str(k) + '(' + str(v) + ')' for k, v in v['pb2'].items()])

		pbar_files.update(1)
	pbar_files.close()

	for gene in dic_unique_variants.keys():
		dic_unique_variants[gene]['set'] = set(dic_unique_variants[gene]['gb1']).union(dic_unique_variants[gene]['gb2'])

	print('Creating genes and variants dataframes ...')
	# del df_all_genes_info
	df_final_genes = pd.DataFrame(dic_intermediate_genes).T  # get pandas Dataframe from dic_intermediate_genes
	# del dic_intermediate_genes
	df_final_genes.index.name = 'Gene'  # add a name to the index column

	df_final_genes = df_final_genes.sort_values(by=['Nb samples', 'Unique variants', 'Gene'],
												ascending=[False, False, True])  # sort columns
	df_final_genes['Nb samples'] = df_final_genes['Nb samples'].astype('int')
	# if args['vcf_annotation_method'].upper() == 'FUNCOTATOR':
	#     df_final_genes[['Unique variants', 'Total variants', 'g.(t1)', 'c.(t1)', 'p.(t1)',
	#         'Mutations (t1)', 'g.(t2)', 'c.(t2)', 'p.(t2)', 'Mutations (t2)']] = (
	#         np.round(df_final_genes[['Total variants', 'g.(t1)', 'c.(t1)', 'p.(t1)',
	#             'Mutations (t1)', 'g.(t2)', 'c.(t2)', 'p.(t2)', 'Mutations (t2)']], 2))
	# elif args['vcf_annotation_method'].upper() == 'ANNOVAR':
	#     df_final_genes[['Unique variants', 'Total variants', 'g.(t1)', 'c.(t1)', 'p.(t1)',
	#         'Mutations (t1)', 'g.TPn+1 union', 'c.(t2)', 'p.(t2)', 'Mutations (t2)']] = np.round(
	#         df_final_genes[['Unique variants', 'Total variants', 'g.(t1)', 'c.(t1)', 'p.(t1)',
	#             'Mutations (t1)', 'g.(t2)', 'c.(t2)', 'p.(t2)', 'Mutations (t2)']], 2)
	patients_threshold = args['min_patients_threshold_for_dataframes']
	if infos:
		# Add additional cancer centric genes information
		df_final_genes = df_final_genes.join(df_databases_info)

		rows_to_remove = []
		for index, row in df_final_genes.iterrows():
			sample_pairs = row['Nb samples']
			if int(sample_pairs) < int(patients_threshold):
				rows_to_remove.append(index)

		# Drop empty informational columns
		df_final_genes = df_final_genes.drop(rows_to_remove)
		empty_cols = [col for col in df_final_genes.columns if df_final_genes[col].isnull().all()]
		df_final_genes.drop(empty_cols, axis=1, inplace=True)

	genes_list = list(df_final_genes.index)

	##########################################
	# Save genes list files (.xlsx and .tsv) #
	logger.info(f'Save genes list files in {Path(output).with_suffix(".MutatedGenes.xlsx")} and {Path(output).with_suffix(".MutatedGenes.tsv")}')

	if 'XLSX' in args['M_mutated_genes_table_format(s)'].upper():
		df_final_genes = pd.DataFrame(df_final_genes)
		improve_df_style(df_final_genes, args['output_path'] + 'merge/merged_genes.xlsx')
	elif 'TSV' in args['M_mutated_genes_table_format(s)'].upper():
		df_final_genes.to_csv(output, sep='\t')
	elif 'CSV' in args['M_mutated_genes_table_format(s)'].upper():
		df_final_genes.to_csv(output, sep=',')

	df_variants_M = create_variants_table(args)
	write_stats_merge(args, df_variants_M, pair_names_column)

	# Biological process enrichment using the genes list with the ToppGene and Panther API
	if enrichment:
		# discard = args['discard_weak_variants'].upper()
		discard = 'YES'
		if enrichment and discard == 'TRUE' or discard == 'YES':
			print(f"\033[1m{len(genes_list)}\033[0m{color} genes are concerned.")
			if len(genes_list) < 1500:
				print('Computing ToppGene and Panther GOEA analysis...')
				toppgene_name = 'GO_Toppgene'
				ToppGene_GOEA('summarise', args, genes_list, toppgene_name, logger)
			else:
				print("The VCF is heavy, too many genes are concerned for ToppGene GO to be run.")
			if len(genes_list) < 2000:
				panther_name = 'GO_Panther'
				Panther_GOEA('summarise', args, genes_list, panther_name, logger)
			else:
				print("The VCF is too heavy, too many genes are concerned for Panther GO to be run.")

	print('Creating mutations cartography plot...')
	if cytoband_file.upper() != 'NONE' or cytoband_file == '':
		##### Create the Chromosome plot
		create_chromosomes_plot(args, genes_pos_for_chromosomes_t1, genes_pos_for_chromosomes_t2,
								genes_pos_for_chromosomes, cytoband_file, chromosomes_output, step, logger)

	create_merged_vcf(args, df_variants_M)

	if 'VAF sample' in df_variants_M.columns:
		create_VAF_sample_scatter_plot(args, df_variants_M)
	if 'VAF pop' in df_variants_M.columns:
		create_VAF_pop_scatter_plot(args, df_variants_M)

	# Create the UpSetPlots
	print('Creating upset plots...')
	if dic_patients_genes_lists == {}:
		print('a')
	create_genes_upsetplot(args, dic_patients_genes_lists, category, names, logger)
	category = [list(i) for i in list(powerset(unique_everseen(file_paths))) if i != ()]  # create all possible combinations of samples
	create_variants_upsetplot(args, df_variants_M, category, names, logger)


def create_protein_impacts_plots(args, df_variants_M):
	print('a')  # UTILISER LE DF ou y a tout et les TIMES POUR LES FAIRE!!!!

def create_genes_upsetplot(args, data, category, names, logger):
	'''
	Create an Upsetplot showing the number of common genes to the different sample set
	'''
	# data = list of genes for each pair
	plt.clf()
	old_names = names
	new_names = set()
	for file_name in names:
		start_index = file_name.find('sons/') + len('sons/')
		end_index = file_name.find('/compared')
		if start_index != -1 and end_index != -1 and start_index < end_index:
			extracted_name = file_name[start_index:end_index].replace('___', '-')
			new_names.add(extracted_name)

	i = 0
	for cat in category:
		j = 0
		for file_name in cat:
			start_index = file_name.find('sons/') + len('sons/')
			end_index = file_name.find('/compared')
			if start_index != -1 and end_index != -1 and start_index < end_index:
				category[i][j] = file_name[start_index:end_index].replace('___', '-')
			j += 1
		i += 1

	upset_min_genes = args['min_number_of_genes_for_upset_plot']
	if upset_min_genes == 0:
		upset_min_genes = max([len(v) for v in data.values()])

	no_gene = old_names - set([k for k in data.keys()])
	if no_gene != set():
		for sample in no_gene:
			data[sample] = set()
			print(f'Warning : sample {sample} doesn\'t have gene !')
			logger.warning(f'Sample {sample} doesn\'t have gene !')

	category.sort(key=len)

	all_genes = set()
	for v in data.values():
		if not all_genes:
			all_genes = set(v)
		else:
			all_genes = all_genes.union(set(v))

	data1 = {}
	for key in list(data.keys()):
		start_index = key.find('sons/') + len('sons/')
		element = key.split("/")[-1]
		end_index = key.find(element)
		if start_index != -1 and end_index != -1 and start_index < end_index:
			new_name = key[start_index:end_index].replace('___', '-').replace('/','')  # new edit
			data1[new_name] = data[key]
	del data
	data = data1

	data2 = []

	for sublist in category:
		sublist[0] = sublist[0].rsplit('/', 1)[0]
	updated_data = {}
	for key in data.keys():
		updated_key = key.rstrip('/')
		updated_data[updated_key] = data[key]
	data = updated_data

	cat_count = 0
	for cat in category:
		# print(cat)
		value_count = 0
		for value in cat:
			new_value = value.split('sons/')[1]
			if '___' in new_value:
				new_value = new_value.replace('___', '-')
			if 'genes' in new_value:
				new_value = new_value.split('/genes')[0]
			# print(value + ' -> ' + new_value)
			if '/' in new_value:
				new_value = new_value.split('/')[0]
			category[cat_count][value_count] = new_value
			value_count += 1
		cat_count += 1

	for c in category[::-1]:
		save_set = None
		for v in c:
			if save_set is None:
				save_set = set(data[v])
			else:
				save_set = save_set.intersection(set(data[v]))  # important
		data2.append(len(save_set))

	# remove categories that we don't want according to parameters
	category.reverse()
	patients_upset_threshold = int(args['min_patients_threshold_for_genes_upset_plot'])
	l_len_cat = []
	for cat in category:
		j = 0
		for x in cat:
			for o in x.split(','):
				j += 1
		l_len_cat.append(j)
	k = 0
	for i in l_len_cat:
		if i < patients_upset_threshold:
			del category[k]
			del data2[k]
			k = k - 1
		k += 1

	upset_min_genes = int(args['min_number_of_genes_for_upset_plot'])
	j = 0
	new_data = []
	new_list = []
	for i in data2:
		if int(i) >= upset_min_genes:
			new_data.append(i)
			new_list.append(category[j])
		j += 1

	input_table = args['dataset_path']
	if 'CSV' in input_table.upper():
		try:
			input_table = pd.read_csv(input_table, sep=',', index_col=0)
		except:
			input_table = pd.read_excel(input_table, index_col=0)
	elif 'TSV' in input_table.upper():
		input_table = pd.read_csv(input_table, sep='\t', index_col=0)
	elif 'XLSX' in input_table.upper():
		input_table = pd.read_excel(input_table, index_col=0)
	col1_name = args['time1_column_name']
	col2_name = args['time2_column_name']
	col_id_name = args['pair_names_column_name']

	dict_names_ids = {}
	for i in range(len(input_table)):
		try:
			row = input_table.iloc[i]
			key = row[col1_name] + '-' + row[col2_name]
			value = row[col_id_name]
			dict_names_ids[key] = value
		except:
			pass

	for sublist in new_list:
		for i in range(len(sublist)):
			if sublist[i] in dict_names_ids:
				sublist[i] = dict_names_ids[sublist[i]]

	df = from_memberships(new_list, data=new_data)

	# UPSET PLOT
	with warnings.catch_warnings():
		warnings.filterwarnings("ignore")
		try:
			UpSet(df, show_counts=True, element_size=40).plot()
		except:
			print('Error in genes UpSetPlot creation : upset fonction')

	plt.ylabel('Number of mutated genes')
	info = f"min patients in subset: {patients_upset_threshold} | min genes in subset: {upset_min_genes}"
	if len(data2) < 4:
		plt.title('Number of mutated genes ', y=1.05)
		plt.figtext(0.55, 0.05, info, ha="center", fontsize=7,
					bbox={"facecolor": "lightgray", "alpha": 0.5, "pad": 3})  # Adjust fontsize parameter

	if 'SVG' in args['upset_plots_format(s)'].upper():
		upset_output = args['output_path'] + 'merge/upset_genes' + '.svg'
		if args['verbose_prints'].upper() == 'TRUE':
			print(f'Create UpSetPlot in {upset_output} !')
		logger.info(f'Create UpSetPlot in {upset_output} !')
		try:
			plt.savefig(upset_output, format='svg', dpi=500)
		except:
			plt.savefig(upset_output, format='svg', dpi=200)

	if 'PNG' in args['upset_plots_format(s)'].upper():
		upset_output = args['output_path'] + 'merge/upset_genes' + '.png'
		if args['verbose_prints'].upper() == 'TRUE':
			print(f'Create UpSetPlot in {upset_output} !')
		logger.info(f'Create UpSetPlot in {upset_output} !')
		try:
			plt.savefig(upset_output, format='png', dpi=500)
		except:
			plt.savefig(upset_output, format='png', dpi=200)
	if 'PDF' in args['upset_plots_format(s)'].upper():
		upset_output = args['output_path'] + 'merge/upset_genes' + '.pdf'
		if args['verbose_prints'].upper() == 'TRUE':
			print(f'Create UpSetPlot in {upset_output} !')
		logger.info(f'Create UpSetPlot in {upset_output} !')
		try:
			plt.savefig(upset_output, format='pdf', dpi=500)
		except:
			plt.savefig(upset_output, format='pdf', dpi=200)

def create_variants_upsetplot(args, df_variants_M_upset, category, names, logger):
	data = {}
	for index, row in df_variants_M_upset.iterrows():
		variant = index
		patients = str(row['Patient(s)']).replace(" ","").split(',')
		for patient in patients:
			if patient not in data.keys():
				data[patient] = []
			if variant not in data[patient]:
				data[patient].append(variant)
	plt.clf()

	i = 0
	for cat in category:
		j = 0
		for file_name in cat:
			start_index = file_name.find('sons/') + len('sons/')
			end_index = file_name.find('/compared')
			if start_index != -1 and end_index != -1 and start_index < end_index:
				category[i][j] = file_name[start_index:end_index].replace('___', '-')
			j += 1
		i += 1

	upset_min_variants = args['min_number_of_variants_for_upset_plot']
	if upset_min_variants == 0:
		upset_min_variants = max([len(v) for v in data.values()])

	category.sort(key=len)

	all_variants = set()
	for v in data.values():
		if not all_variants:
			all_variants = set(v)
		else:
			all_variants = all_variants.union(set(v))

	data2 = []
	cat = []
	to_suppr = set()
	to_suppr2 = set()
	# remove genes that are not found in all samples comparisons

	for sublist in category:
		sublist[0] = sublist[0].rsplit('/', 1)[0]

	cat_count = 0
	for cat in category:
		# print(cat)
		value_count = 0
		for value in cat:
			try:
				new_value = value.split('sons/')[1]
				if '___' in new_value:
					new_value = new_value.replace('___', '-')
				if 'genes' in new_value:
					new_value = new_value.split('/genes')[0]
				# print(value + ' -> ' + new_value)
				new_value = new_value.split('/')[0]
				category[cat_count][value_count] = new_value
			except:
				pass
			value_count += 1
		cat_count += 1

	ko = 0
	for c in category[::-1]:
		if len(category) == 1:
			print('a')
		save_set = None
		for v in c:
			if save_set is None:
				save_set = set(data[v])
			else:
				save_set = save_set.intersection(set(data[v]))  # important
				if len(save_set) == 2:
					ko += 1
		data2.append(len(save_set))

	# remove categories that we don't want according to parameters
	category.reverse()
	patients_upset_threshold = int(args['min_patients_threshold_for_variants_upset_plot'])
	l_len_cat = []
	for cat in category:
		j = 0
		for x in cat:
			for o in x.split(','):
				j += 1
		l_len_cat.append(j)
	k = 0
	for i in l_len_cat:
		if i < patients_upset_threshold:
			del category[k]
			del data2[k]
			k = k - 1
		k += 1

	upset_min_variants = int(args['min_number_of_variants_for_upset_plot'])
	j = 0
	new_data = []
	new_list = []
	for i in data2:
		if int(i) >= upset_min_variants:
			new_data.append(i)
			new_list.append(category[j])
		j += 1

	input_table = args['dataset_path']
	if 'CSV' in input_table.upper():
		try:
			input_table = pd.read_csv(input_table, sep=',', index_col=0)
		except:
			input_table = pd.read_excel(input_table, index_col=0)
	elif 'TSV' in input_table.upper():
		input_table = pd.read_csv(input_table, sep='\t', index_col=0)
	elif 'XLSX' in input_table.upper():
		input_table = pd.read_excel(input_table, index_col=0)
	col1_name = args['time1_column_name']
	col2_name = args['time2_column_name']
	col_id_name = args['pair_names_column_name']

	dict_names_ids = {}
	for i in range(len(input_table)):
		try:
			row = input_table.iloc[i]
			key = row[col1_name] + '-' + row[col2_name]
			value = row[col_id_name]
			dict_names_ids[key] = value
		except:
			pass

	for sublist in new_list:
		for i in range(len(sublist)):
			if sublist[i] in dict_names_ids:
				sublist[i] = dict_names_ids[sublist[i]]

	df_error = False
	try:
		df = from_memberships(new_list, data=new_data)
	except:
		df_error = True
		print('Error in variants UpSetPlot creation : nothing to plot, be less restrictive in parameters')

	# UPSET PLOT
	with warnings.catch_warnings():
		warnings.filterwarnings("ignore")

		if not df_error:
			UpSet(df, show_counts=True, element_size=40).plot()

	plt.ylabel('Number of variants')
	info = f"min patients in subset: {patients_upset_threshold} | min variants in subset: {upset_min_variants}"
	if len(data2) < 4:
		plt.title('Number of mutated variants ', y=1.05)
		plt.figtext(0.55, 0.05, info, ha="center", fontsize=7,
					bbox={"facecolor": "lightgray", "alpha": 0.5, "pad": 3})
	try:
		if 'SVG' in args['upset_plots_format(s)'].upper():
			upset_output = args['output_path'] + 'merge/upset_variants' + '.svg'
			if args['verbose_prints'].upper() == 'TRUE':
				print(f'Create UpSetPlot in {upset_output} !')
			logger.info(f'Create UpSetPlot in {upset_output} !')
			try:
				plt.savefig(upset_output, format='svg', dpi=500)
			except:
				plt.savefig(upset_output, format='svg', dpi=200)

		if 'PNG' in args['upset_plots_format(s)'].upper():
			upset_output = args['output_path'] + 'merge/upset_variants' + '.png'
			if args['verbose_prints'].upper() == 'TRUE':
				print(f'Create UpSetPlot in {upset_output} !')
			logger.info(f'Create UpSetPlot in {upset_output} !')
			try:
				plt.savefig(upset_output, format='png', dpi=500)
			except:
				plt.savefig(upset_output, format='png', dpi=200)

		if 'PDF' in args['upset_plots_format(s)'].upper():
			upset_output = args['output_path'] + 'merge/upset_variants' + '.pdf'
			if args['verbose_prints'].upper() == 'TRUE':
				print(f'Create UpSetPlot in {upset_output} !')
			logger.info(f'Create UpSetPlot in {upset_output} !')
			try:
				plt.savefig(upset_output, format='pdf', dpi=500)
			except:
				plt.savefig(upset_output, format='pdf', dpi=200)
	except:
		print('Error in variants UpSetPlot creation : plot too large, be more restrictive in parameters')
def main(args, output_path):
	output = args['M_MutatedGenes_name']
	enrichment = args['M_enrichment'].upper()
	if 'YES' in enrichment or 'TRUE' in enrichment or 'PANTHER' in enrichment or 'TOPPGENE' in enrichment:
		enrichment = True

	# chromosomes plot variables
	cytoband_file = args['cytoband_file']
	chromosomes_output = args['chromosomes_plot_name']
	step = args['chromosome_step']
	current_directory = os.getcwd()

	# Logger configuration
	logger = logging.getLogger('LOTUS merge')
	logger.setLevel(logging.DEBUG)
	fh = logging.FileHandler(args['log'])
	fh.setLevel(logging.DEBUG)
	formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
	fh.setFormatter(formatter)
	logger.addHandler(fh)

	if cytoband_file.upper() != 'NONE':
		try:
			logger.info('Verification of {cytoband_file}')
			verif_input(cytoband_file)
			try:
				logger.info('Verification of {chromosomes_output}')
				verif_output(chromosomes_output)
				logger.info('- Output file ok -')
			except ValueError:
				print(f'Problem with {chromosomes_output}: ', sys.exc_info()[0])
				logger.error('- Problem with {chromosomes_output} -')
				exit(1)
			logger.info('- Input file ok -')
		except ValueError:
			print(f'Problem with {cytoband_file}: ', sys.exc_info()[0])
			logger.error('- Problem with {cytoband_file} -')
			exit(1)

	if int(step) < 500000:
		warnings.warn("A step value below 500000 may cause a long calculation time !", RuntimeWarning)
		logger.warning('A step value below 500000 may cause a long calculation time !')

	if args['verbose_prints'].upper() == 'TRUE':
		print('Search for file Lotus_ExternalBases_202301.xlsx ...')
	infos = 'Lotus_ExternalBases_202301.xlsx'
	logger.info('Verification of {infos}')
	try:
		infos = verif_supplementary_information_file(infos, current_directory)
	except ValueError:
		print(f'Problem with {infos}: ', sys.exc_info()[0])
		logger.error('- Problem with {infos} -')
		exit(1)

	logger.info(
		'**************************************************************************************************************')
	logger.info('*** LOTUS merging module ***')
	no_argument = ''
	if enrichment != '' or enrichment.upper() != 'FALSE':
		no_argument += ' --enrichment'
	logger.info('* Start merging *')
	logger.info(f'Current directory : {Path().absolute()}')

	df_dataset = pd.read_excel(args['dataset_path'])
	l_col1_names = df_dataset[args['time1_column_name']].dropna().str.strip().replace('', np.nan).dropna().tolist()
	l_col1_new = [name.split('.funco')[0] for name in l_col1_names]
	l_col2_names = df_dataset[args['time2_column_name']].dropna().str.strip().replace('', np.nan).dropna().tolist()
	l_col2_new = [name.split('.funco')[0] for name in l_col2_names]
	nb_files = len(l_col1_new) + len(l_col2_new)
	names = [col1 + "___" + col2 for col1, col2 in zip(l_col1_new, l_col2_new)]
	path = "config.txt"
	with open(path, 'r') as file:
		for line in file:
			if line.startswith('#') or line == '\n' or line.startswith('------'):
				pass
			else:
				try:
					key, value = line.strip().replace(" ", "").split('=')
				except:
					pass
				if key == "C_MutatedGenes_name":
					C_MutatedGenes_name = value.lower()
				elif key == "C_MutatedGenes":
					C_MutatedGenes_format = value.lower()
				elif key == 'variants_selection_approach':
					variants_selection_approach = value.lower()

	df_dataset = pd.read_excel(args['dataset_path'])
	time1_column = []
	time2_column = []
	pair_names_column = []

	for value in df_dataset[args['time1_column_name']]:
		if pd.notnull(value) and value not in ('NaN', 'naN') and value.strip() != '':
			time1_column.append(value)

	for value in df_dataset[args['time2_column_name']]:
		if pd.notnull(value) and value not in ('NaN', 'naN') and value.strip() != '':
			time2_column.append(value)

	for value in df_dataset[args['pair_names_column_name']]:
		if pd.notnull(value) and value not in ('NaN', 'naN') and str(value).strip() != '':
			try:
				pair_names_column.append(str(int(value)))
			except:
				pair_names_column.append(str(value))

	time1_column = [element.replace('.vcf', '') for element in time1_column]
	time2_column = [element.replace('.vcf', '') for element in time2_column]
	dict_pairs = {}
	for i in range(len(time1_column)):
		dict_pairs[time1_column[i] + '___' + time2_column[i]] = pair_names_column[i]

	if pair_names_column != [] or pair_names_column != [''] or pair_names_column != ['NaN']:
		names = pair_names_column
	M_MutatedGenes_name = args['variants_selection_approach'] + '_genes'
	M_MutatedGenes_format = args['M_mutated_genes_table_format(s)'].lower()
	file_paths = [output_path + 'comparisons/' + name + '/' + name + '_' + M_MutatedGenes_name + "." + M_MutatedGenes_format for name in names]
	file_paths = [path.replace(".vcf", "") for path in file_paths]

	category = [list(i) for i in list(powerset(unique_everseen(file_paths))) if i != ()]  # create all possible combinations of samples

	logger.info(f'Merging {nb_files} files !')
	if args['verbose_prints'].upper() == 'TRUE':
		print('Start merging...')

	merge_results(args, file_paths, category, output, infos, cytoband_file, chromosomes_output, int(step), nb_files, enrichment, logger)

	logger.info('* End merging *')
	logger.info(
		'**************************************************************************************************************')

# End
