#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #
#   LOncoG : a software for Longitudinal OncoGenomics analysis
#   Authors: S. Besseau, G. Siekaniec, W. Gouraud
# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #

import math
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill, Border, Side
import os
import pyfastx
from openpyxl.styles import Alignment
import pickle as pk
from pathlib import Path
import re
import sys
import logging
from tqdm import tqdm
from itertools import dropwhile
from collections import OrderedDict, Counter
from copy import deepcopy
import warnings
import pandas as pd
from python_scripts.reusable_functions.check_files import verif_output
from python_scripts.api_requests.toppgene_api import ToppGene_GOEA
from python_scripts.api_requests.panther_api import Panther_GOEA
from python_scripts.reusable_functions.read_vcf import read_vcf
import matplotlib.pyplot as plt

TRANSLATE = {'A>C': 'T>G', 'A>G': 'T>C', 'A>T': 'T>A', 'C>A': 'C>A', 'C>G': 'C>G', 'C>T': 'C>T',
			 'G>A': 'C>T', 'G>C': 'C>G', 'G>T': 'C>A', 'T>A': 'T>A', 'T>C': 'T>C', 'T>G': 'T>G'}
TAB = str.maketrans("ACGT", "TGCA")


def create_snp_dict():
	'''
	Create the dictionnary to count the SNP variant and create the vcf SNP profile. Only the C>* and T>* value are used because the variant sens isn't known.
	Input : None:
	Output : dictionnary with empty counts
	'''
	c_words = {}
	t_words = {}
	for i in 'ATCG':
		word = i
		for j in 'CT':
			word2 = word + j
			for k in 'ATCG':
				word3 = word2 + k
				if word3[1] == 'C':
					c_words[word3] = 0
				else:
					t_words[word3] = 0
	save = {}
	for snp in set(TRANSLATE.values()):
		if snp[0] == 'C':
			save[snp] = deepcopy(c_words)
		else:
			save[snp] = deepcopy(t_words)
	return save


def create_dataframe_from_gene(dict_para, d):
	'''
	Take a dictionary containing the genes list with information for each genes and transform it in pandas dataframe
	Input : dictionary containing genes (d)
	Output : dataframe corresponding to the dictionary
	'''

	col = ['n variants', 'snp, dnp, tnp, onp, ins, del', 'Ref(s)', 'Alt(s)', 'Chr', 'Position(s)',
		   'avg DP', 'avg QUAL', 'avg VAF sample', 'avg VAF pop', 'Mutation subtypes', 'Subtype_pos']  # Columns names

	for key in d.keys():
		d[key][10] = str(d[key][10]).replace('{', '').replace('}', '').replace('.', 'unknown')

	d = OrderedDict(sorted(d.items()))
	df = pd.DataFrame.from_dict(d, orient='index', columns=col)

	id = 'Gene'
	df.index.name = id

	# Modification of the dataframe using a copy of it
	df2 = pd.DataFrame()
	df2['snp, dnp, tnp, onp, ins, del'] = [','.join(map(str, l)) for l in df['snp, dnp, tnp, onp, ins, del']]
	df['snp, dnp, tnp, onp, ins, del'] = df2['snp, dnp, tnp, onp, ins, del'].values
	del df2
	df2 = pd.DataFrame()
	df2['Ref(s)'] = [','.join(map(str, l)) for l in df['Ref(s)']]
	df['Ref(s)'] = df2['Ref(s)'].values
	del df2
	df2 = pd.DataFrame()
	df2['Alt(s)'] = [','.join([str(l2) if len(l2) > 1 else str(l2[0]) for l2 in l]) for l in
							 df['Alt(s)']]
	df['Alt(s)'] = df2['Alt(s)'].values
	del df2
	df2 = pd.DataFrame()
	df2['avg DP'] = [','.join(map(str, l)) for l in df['avg DP']]
	df['avg DP'] = df2['avg DP'].values
	del df2

	df2 = pd.DataFrame()
	df2['Position(s)'] = [','.join(map(str, l)) for l in df['Position(s)']]
	df['Position(s)'] = df2['Position(s)'].values
	del df2

	for key in d.keys():
		# print(d[key])
		df.loc[df['avg VAF sample'] == key, 'avg VAF sample'] = [d[key][8]] * df.loc[df['avg VAF sample'] == key, 'avg VAF sample'].shape[0]

	sorted_df = df
	complete_df = df.copy()

	any_VAF_sample = []
	any_VAF_pop = []
	for index, row in sorted_df.iterrows():
		try:
			l_qual = row['avg QUAL']
			if len(l_qual) == 1:
				qual_value = round(float(l_qual[0]),0)
			else:
				qual_value = round(np.mean([float(val) for val in l_qual]),0)
			sorted_df.at[index, 'avg QUAL'] = qual_value
		except:
			pass

		try:
			l_VAF_sample = row['avg VAF sample']
			if len(l_VAF_sample) == 1:
				VAF_sample = round(l_VAF_sample[0], 3)
			else:
				VAF_sample = round(np.mean([float(val) for val in l_VAF_sample]), 3)
			sorted_df.at[index, 'avg VAF sample'] = VAF_sample
			any_VAF_sample.append(VAF_sample)
		except:
			pass

		try:
			l_VAF_pop = row['avg VAF pop']
			if len(l_VAF_pop) == 1:
				try:
					VAF_pop = "{:.2e}".format(l_VAF_pop[0])
					mean_vaf = VAF_pop
				except:
					mean_vaf = 'not found'
				any_VAF_pop.append(l_VAF_pop[0])
			else:
				if all(element == 'not found' for element in l_VAF_pop):
					mean_vaf = 'not found'
				else:
					for vaf in l_VAF_pop:
						any_VAF_pop.append(vaf)
					filtered_values = [float(v) for v in l_VAF_pop if v != 'not found']
					mean_vaf = "{:.2e}".format(sum(filtered_values) / len(filtered_values))
			sorted_df.at[index, 'avg VAF pop'] = mean_vaf
		except:
			pass

		try:
			l_DP = row['avg DP'].split(',')
			if len(l_DP) == 1:
				DP = l_DP[0]
			else:
				DP = int(round(np.mean([float(val) for val in l_DP]), 0))
			sorted_df.at[index, 'avg DP'] = DP
		except:
			pass
	if all(value == '.' for value in l_qual):
		new_col = ['n variants', 'snp, dnp, tnp, onp, ins, del', 'Mutation subtypes', 'Chr', 'Position(s)', 'Ref(s)', 'Alt(s)',
				   'avg DP', 'avg VAF sample', 'avg VAF pop']
	else:
		new_col = ['n variants', 'snp, dnp, tnp, onp, ins, del', 'Mutation subtypes', 'Chr', 'Position(s)', 'Ref(s)', 'Alt(s)',
		   'avg DP', 'avg QUAL', 'avg VAF sample', 'avg VAF pop']  # Columns name
	sorted_df = sorted_df.reindex(columns=new_col)

	if any_VAF_sample == []:
		sorted_df.drop(columns=['avg VAF sample'], inplace=True)
	if any_VAF_pop == [] or all(element == 'not found' for element in any_VAF_pop):
		sorted_df.drop(columns=['avg VAF pop'], inplace=True)

	sorted_df['avg DP'] = sorted_df['avg DP'].astype(float)
	try:
		sorted_df['avg VAF pop'] = sorted_df['avg VAF pop'].replace('0.00e+00', 0)
		sorted_df['avg VAF pop'] = sorted_df['avg VAF pop'].replace(['not found', 'nan'], np.nan)
		sorted_df['avg VAF pop'] = sorted_df['avg VAF pop'].apply(lambda x: np.nan if isinstance(x, list) and len(x) == 0 else x)
		sorted_df['avg VAF pop'] = sorted_df['avg VAF pop'].replace(['not found', 'nan'], np.nan)
		sorted_df = sorted_df.sort_values(by='avg VAF pop', ascending=True)
		sorted_df['avg VAF pop'] = sorted_df['avg VAF pop'].fillna('not found')
		sorted_df = sorted_df.sort_values(by=['avg DP', 'n variants'], ascending=False)
		sorted_df['avg VAF pop'] = sorted_df['avg VAF pop'].apply(lambda x: "{:.2e}".format(float(x)) if x != 'not found' else np.nan)
	except:
		print('no vaf pop in sample')

	for index, row in sorted_df.iterrows():
		l_refs = row['Ref(s)'].split(',')
		l_alts = row["Alt(s)"].split(',')

		if len(l_refs) != len(l_alts):
			merged_alts = []
			for i in range(1, len(l_alts)):
				if ']' in l_alts[i]:
					merged_value = (l_alts[i-1].replace(" ", "") + '/' + l_alts[i].replace(" ", "")).replace("'","").replace("[","").replace("]","")
					merged_alts.append(merged_value)
				else:
					merged_alts.append(l_alts[i])
			new_merged_alts = []
			for i in merged_alts:
				i = i.replace('[', '').replace(']', '').replace("'", "").replace(" ", "")
				new_merged_alts.append(i)
			sorted_df.at[index, "Alt(s)"] = ','.join(new_merged_alts)

	for index, row in complete_df.iterrows():
		l_refs = row['Ref(s)'].split(',')
		l_alts = row["Alt(s)"].split(',')

		if len(l_refs) != len(l_alts):
			merged_alts = []
			for i in range(1, len(l_alts)):
				if ']' in l_alts[i]:
					merged_value = (l_alts[i-1].replace(" ", "") + '/' + l_alts[i].replace(" ", "")).replace("'","").replace("[","").replace("]","")
					merged_alts.append(merged_value)
				else:
					merged_alts.append(l_alts[i])
			new_merged_alts = []
			for i in merged_alts:
				i = i.replace('[', '').replace(']', '').replace("'", "").replace(" ", "")
				new_merged_alts.append(i)
			complete_df.at[index, "Alt(s)"] = ','.join(new_merged_alts)

	return complete_df, sorted_df


def create_variants_table(dict_para, complete_df, dict_impacts, dict_multiple_alts):
	df_variants = pd.DataFrame()
	any_vaf_sample = []
	any_vaf_pop = []
	for index, row in complete_df.iterrows():
		chr = row['Chr']
		gene = index
		if ',' in row['Alt(s)']:
			positions = row['Position(s)'].split(',')
			alts = row['Alt(s)'].split(',')
			refs = row['Ref(s)'].split(',')
			DPs = str(row['avg DP']).split(',')
			quals = str(row['avg QUAL']).replace("[",'').replace("]",'').replace("'","").replace('"',"").replace(" ","").split(',')

			old_VAF_samples = str(row['avg VAF sample']).split(',')
			new_VAF_samples = []
			for vaf_sample in old_VAF_samples:
				vaf_sample = vaf_sample.replace('[',"").replace("'","").replace(" ","")
				if vaf_sample == 'nan' or vaf_sample == '':
					new_VAF_samples.append('no')
				else:
					vaf_sample = vaf_sample.replace(']','').replace(',', '.')
					new_VAF_samples.append(vaf_sample)
					any_vaf_sample.append(vaf_sample)

			old_VAF_pops = str(row['avg VAF pop']).split(',')
			new_VAF_pops = []
			if old_VAF_pops == ['[]']:
				for k in DPs:
					new_VAF_pops.append('no')
			else:
				for vaf_pop in old_VAF_pops:
					vaf_pop = vaf_pop.replace('[',"").replace("'","").replace(" ","")
					if vaf_pop == 'nan' or vaf_pop == '':
						new_VAF_pops.append('no')
					else:
						vaf_pop = vaf_pop.replace(']','')
						new_VAF_pops.append(vaf_pop)
						any_vaf_pop.append(vaf_pop)
			mutation_subtypes = row['Subtype_pos']

			dict_subtypes = {}
			for key in mutation_subtypes.keys():
				pos_count = 0
				for pos in mutation_subtypes[key]:
					dict_subtypes[str(mutation_subtypes[key][pos_count])] = key
					pos_count += 1

			i = 0
			new_alts = []
			alts_variants = False
			for k in alts:
				if '/' in k:
					alts_variants = True
					multiple_alts_elements = k

			if alts_variants:
				new_refs = []
				new_alts = []
				new_positions = []
				new_genes = []
				new_new_DPs = []
				new_quals = []
				new_new_VAF_samples = []
				new_new_VAF_pops = []
				new_mutation_subtypes = []

				j = 0
				for nuc in multiple_alts_elements.split('/'):
					new_refs.append(refs[j])
					new_alts.append(nuc)
					new_positions.append(positions[j])
					new_genes.append(gene)
					new_new_DPs.append(DPs[j])
					new_quals.append(quals[j])
					new_mutation_subtypes.append(dict_subtypes[positions[j]])
					j += 1

				ref_counter = 0
				for ref in new_refs:
					variant = 'g.' + chr + ':' + new_positions[0] + ref + '>' + new_alts[ref_counter]
					try:
						df_variants[variant] = {'Gene': new_genes[0], 'Chr': row['Chr'], 'Position': new_positions[0], 'Ref': ref, 'Alt': new_alts[ref_counter],
											'DP': new_new_DPs[0], 'QUAL': new_quals[0], 'VAF sample': dict_multiple_alts[variant]['VAF_sample'],
											'VAF pop': dict_multiple_alts[variant]['VAF_pop'], 'Mutation': new_mutation_subtypes[0]}
					except:
						print('a')
					ref_counter += 1

			else:
				for k in range(len(alts)):
					try:
						df_variants['g.' + chr + ':' + str(positions[k]) + str(refs[k]) + '>' + str(alts[k])] = \
						{'Gene': gene, 'Chr': row['Chr'], 'Position': row['Position(s)'].split(',')[k], 'Ref': refs[k], 'Alt': alts[k],
						'DP': DPs[k], 'QUAL': round(float(quals[k]), 0), 'VAF sample': new_VAF_samples[k], 'VAF pop': row['avg VAF pop'][k],
						'Mutation': dict_subtypes[positions[k]]}
					except:
						try:
							df_variants['g.' + chr + ':' + str(positions[k]) + str(refs[k]) + '>' + str(alts[k])] = \
							{'Gene': gene, 'Chr': row['Chr'], 'Position': str(positions[k]), 'Ref': refs[k], 'Alt': alts[k],
							 'DP': DPs[k], 'VAF sample': new_VAF_samples[k], 'VAF pop': row['avg VAF pop'][k],
							 'Mutation': dict_subtypes[positions[k]]}
						except:
							print('a')
						# if 'AAAGTGAGAGCTAAACATTGG' in alts:
						# 	print(i)
						# 	print(chr, positions[i], refs[i], alts[i], DPs[i], quals[i], new_VAF_samples[i], row['avg VAF pop'][i], dict_subtypes[positions[i]])
		else:
			try:
				df_variants['g.' + chr + ':' + str(row['Position(s)']) + str(row['Ref(s)']) + '>' + str(row['Alt(s)'])] = \
					{'Gene': gene, 'Chr': row['Chr'], 'Position': row['Position(s)'], 'Ref': row['Ref(s)'], 'Alt': row['Alt(s)'],
					 'DP': row['avg DP'], 'QUAL': round(float(row['avg QUAL'][0]),0), 'VAF sample': row['avg VAF sample'][0],
					 'VAF pop': row['avg VAF pop'], 'Mutation': row['Mutation subtypes'].replace("'","").split(':')[0]}
			except:
				df_variants['g.' + chr + ':' + str(row['Position(s)']) + str(row['Ref(s)']) + '>' + str(row['Alt(s)'])] = \
					{'Gene': gene, 'Chr': row['Chr'], 'Position': row['Position(s)'], 'Ref': row['Ref(s)'], 'Alt': row['Alt(s)'],
					 'DP': row['avg DP'], 'VAF sample': row['avg VAF sample'][0],
					 'VAF pop': row['avg VAF pop'], 'Mutation': row['Mutation subtypes'].replace("'", "").split(':')[0]}

	df_variants = df_variants.transpose()
	dict_impact_meanings_SIFT = {"D": "Deleterious", "T": "Tolerated", "U": "Unknown", "not found": "not found"}
	dict_impact_meanings_polyphen = {"B": "Benign", "P": "Possibly damaging", "D": "Probably damaging", "U": "Unknown", "not found": "not found"}

	if not all(not bool(d) for d in dict_impacts.values()):
		df_variants["SIFT score"] = pd.Series(dtype=float)
		df_variants["SIFT pred"] = pd.Series(dtype=str)
		df_variants["Polyphen2 score"] = pd.Series(dtype=float)
		df_variants["Polyphen2 pred"] = pd.Series(dtype=str)

		for index, row in df_variants.iterrows():
			if '/' in index:
				index = index.replace('[', '').replace(']', '').replace("'", "").replace(" ", "")
			try:
				df_variants.at[index, "SIFT score"] = dict_impacts[index]['SIFT score']
				df_variants.at[index, "SIFT pred"] = dict_impact_meanings_SIFT[dict_impacts[index]['SIFT pred']]
				df_variants.at[index, "Polyphen2 score"] = dict_impacts[index]['Polyphen score']
				df_variants.at[index, "Polyphen2 pred"] = dict_impact_meanings_polyphen[dict_impacts[index]['Polyphen pred']]
			except:
				try:
					df_variants.at[index, "SIFT score"] = dict_impacts[index]['SIFT_score']
					df_variants.at[index, "SIFT pred"] = dict_impact_meanings_SIFT[dict_impacts[index]['SIFT_pred']]
					df_variants.at[index, "Polyphen2 score"] = dict_impacts[index]['Polyphen_score']
					df_variants.at[index, "Polyphen2 pred"] = dict_impact_meanings_polyphen[dict_impacts[index]['Polyphen_pred']]
				except:
					try:
						df_variants.at[index, "SIFT score"] = dict_multiple_alts[index]['SIFT_score']
						df_variants.at[index, "SIFT pred"] = dict_impact_meanings_SIFT[dict_multiple_alts[index]['SIFT_pred']]
						df_variants.at[index, "Polyphen2 score"] = dict_multiple_alts[index]['Polyphen_score']
						df_variants.at[index, "Polyphen2 pred"] = dict_impact_meanings_polyphen[dict_multiple_alts[index]['Polyphen_pred']]
					except:
						print('b')

	is_vaf_sample = False
	is_vaf_pop = False

	try:
		df_variants["VAF sample"] = df_variants["VAF sample"].astype(float)
		any_vaf_sample = df_variants["VAF sample"].tolist()
	except:
		pass

	try:
		df_variants['VAF pop'] = df_variants['VAF pop'].astype(str)
		df_variants['VAF pop'] = df_variants['VAF pop'].str.replace('[', '').str.replace(']', '').replace("'", "")
		df_variants['VAF pop'] = df_variants['VAF pop'].str.strip("'")
		l_vaf_pop = df_variants["VAF pop"].tolist()
	except:
		pass

	if any_vaf_sample == [] or all(value == '' for value in any_vaf_sample) or all(value == 'nan' for value in any_vaf_sample):
		df_variants.drop(columns=['VAF sample'], inplace=True)
	else:
		is_vaf_sample = True

	try:
		l_no_vaf_pop = []
		for element in l_vaf_pop:
			element = element.replace("'", "")
			if element in ['not found', 'nan', 'no']:
				l_no_vaf_pop.append('no')
			else:
				l_no_vaf_pop.append('yes')
		if all(element == 'no' for element in l_no_vaf_pop):
			df_variants.drop(columns=['VAF pop'], inplace=True)
		else:
			is_vaf_pop = True
	except:
		print('No VAF pop')

	is_qual = False
	try:
		if row['avg QUAL'] == [] or all(value == '' for value in row['avg QUAL']) or all(value == 'nan' for value in row['avg QUAL']):
			df_variants.drop(columns=['QUAL'], inplace=True)
		else:
			is_qual = True
	except:
		is_qual = False

	is_subtype = False
	try:
		if row['Mutation subtypes'] == [] or all(value == '' for value in row['Mutation subtypes']) or all(value == 'nan' for value in row['Mutation subtypes']):
			df_variants.drop(columns=['Mutation'], inplace=True)
		else:
			is_subtype = True
	except:
		if row['Mutation'] == [] or all(value == '' for value in row['Mutation']) or all(value == 'nan' for value in row['Mutation']):
			df_variants.drop(columns=['Mutation'], inplace=True)
		else:
			is_subtype = True

	if is_vaf_pop and is_vaf_sample:
		df_variants['VAF pop'] = pd.to_numeric(df_variants['VAF pop'], errors='coerce')
		df_variants_sorted = df_variants.sort_values(by=["VAF pop", "VAF sample", "DP"], ascending=[True, True, False])
		for i in range(len(df_variants_sorted['VAF pop'])):
			if pd.isna(df_variants_sorted['VAF pop'].iloc[i]):
				df_variants_sorted['VAF pop'].iloc[i] = 'not found'
			else:
				df_variants_sorted['VAF pop'].iloc[i] = "{:.2e}".format(float(df_variants_sorted['VAF pop'].iloc[i]))

	elif is_vaf_sample:
		df_variants_sorted = df_variants.sort_values(by=["VAF sample", "DP"], ascending=[True, False])
	elif is_vaf_pop:
		df_variants_sorted = df_variants.sort_values(by=["VAF pop", "DP"], ascending=[True, False])
	else:
		df_variants_sorted = df_variants.sort_values(by=["DP"], ascending=False)

	try:
		df_variants_sorted.sort_values(['SIFT score', 'Polyphen2 score', 'VAF pop', 'VAF sample', 'SIFT pred', 'Polyphen2 pred', 'DP'],
							ascending=[True, False, True, True, True, False, False], inplace=True)
	except:
		pass

	formats = dict_para['S_variants_table_format'].upper()
	output_path = dict_para['output_path_sample'].replace(".vcf", "") + 'passed_variants'
	if 'XLSX' in formats:
		df_variants_sorted.to_excel(output_path + '.xlsx', index_label='Variant')
		workbook = openpyxl.load_workbook(output_path + '.xlsx')
		worksheet = workbook.active

		mutation_keywords = ['nonsynonymous', 'stopgain', 'stoploss', 'frameshift']

		highlight_colors = {
			'green': "FFA6FF00",  # Light green
			'yellow': "FFFFFF00",  # Light yellow
			'orange': "FFFFC000",  # Light orange
			'red': "FFFF4D4D",  # Light red
		}

		# DP ranges
		values = df_variants['DP'].tolist()
		numeric_values = [float(value) for value in values]
		quartile1 = np.percentile(numeric_values, 25)
		quartile2 = np.percentile(numeric_values, 50)
		quartile3 = np.percentile(numeric_values, 75)
		DP_range1 = [0, quartile1]
		DP_range2 = [quartile1, quartile2]
		DP_range3 = [quartile2, quartile3]
		DP_range4 = [quartile3, max(numeric_values)]

		try:
			# AD ranges
			no_ad = False
			values = df_variants['AD'].tolist()
			numeric_values = [float(value) for value in values]
			quartile1 = np.percentile(numeric_values, 25)
			quartile2 = np.percentile(numeric_values, 50)
			quartile3 = np.percentile(numeric_values, 75)
			AD_range1 = [0, quartile1]
			AD_range2 = [quartile1, quartile2]
			AD_range3 = [quartile2, quartile3]
			AD_range4 = [quartile3, max(numeric_values)]
		except:
			no_ad = True

		# VAF sample ranges
		values = df_variants['VAF sample'].tolist()
		values = [value for value in values if value != 'not found']
		numeric_values = [float(value) for value in values]
		quartile1 = np.percentile(numeric_values, 25)
		quartile2 = np.percentile(numeric_values, 50)
		quartile3 = np.percentile(numeric_values, 75)
		VAF_sample_range1 = [0, quartile1]
		VAF_sample_range2 = [quartile1, quartile2]
		VAF_sample_range3 = [quartile2, quartile3]
		VAF_sample_range4 = [quartile3, max(numeric_values)]

		# VAF pop ranges
		values = df_variants['VAF pop'].tolist()
		values = [value for value in values if value != 'not found']
		values = [value for value in values if not math.isnan(value)]
		numeric_values = [float(value) for value in values]
		quartile1 = np.percentile(numeric_values, 25)
		quartile2 = np.percentile(numeric_values, 50)
		quartile3 = np.percentile(numeric_values, 75)
		VAF_pop_range1 = [0, quartile1]
		VAF_pop_range2 = [quartile1, quartile2]
		VAF_pop_range3 = [quartile2, quartile3]
		VAF_pop_range4 = [quartile3, max(numeric_values)]

		# SIFT ranges
		SIFT_range1 = [-1, 0.05]
		SIFT_range2 = [0.05, 0.1]
		SIFT_range3 = [0.1, 0.5]
		SIFT_range4 = [0.5, 1.5]

		# PolyPhen2 score ranges
		PolyPhen2_range1 = [-1, 0.453]
		PolyPhen2_range2 = [0.453, 0.7]
		PolyPhen2_range3 = [0.7, 0.956]
		PolyPhen2_range4 = [0.956, 1.5]

		header_row = worksheet[1]
		column_names = [cell.value for cell in header_row]
		for row in worksheet.iter_rows(min_row=2):
			for cell, column_name in zip(row, column_names):
				if column_name == 'Mutation':
					if any(keyword in cell.value for keyword in mutation_keywords):
						cell.fill = PatternFill(start_color=highlight_colors['green'], end_color=highlight_colors['green'], fill_type="solid")
					else:
						cell.fill = PatternFill(start_color=highlight_colors['red'], end_color=highlight_colors['red'], fill_type="solid")
				elif column_name == 'DP':
					if DP_range1[0] <= int(cell.value) <= DP_range1[1]:
						cell.fill = PatternFill(start_color=highlight_colors['red'], end_color=highlight_colors['red'], fill_type="solid")
					elif DP_range2[0] <= int(cell.value) <= DP_range2[1]:
						cell.fill = PatternFill(start_color=highlight_colors['orange'], end_color=highlight_colors['orange'], fill_type="solid")
					elif DP_range3[0] <= int(cell.value) <= DP_range3[1]:
						cell.fill = PatternFill(start_color=highlight_colors['yellow'], end_color=highlight_colors['yellow'], fill_type="solid")
					elif DP_range4[0] <= int(cell.value) <= DP_range4[1]:
						cell.fill = PatternFill(start_color=highlight_colors['green'], end_color=highlight_colors['green'], fill_type="solid")
				elif column_name == 'AD':
					if AD_range1[0] <= int(cell.value) <= AD_range1[1]:
						cell.fill = PatternFill(start_color=highlight_colors['red'], end_color=highlight_colors['red'], fill_type="solid")
					elif AD_range2[0] <= int(cell.value) <= AD_range2[1]:
						cell.fill = PatternFill(start_color=highlight_colors['orange'], end_color=highlight_colors['orange'], fill_type="solid")
					elif AD_range3[0] <= int(cell.value) <= AD_range3[1]:
						cell.fill = PatternFill(start_color=highlight_colors['yellow'], end_color=highlight_colors['yellow'], fill_type="solid")
					elif AD_range4[0] <= int(cell.value) <= AD_range4[1]:
						cell.fill = PatternFill(start_color=highlight_colors['green'], end_color=highlight_colors['green'], fill_type="solid")
				elif column_name == 'VAF sample':
					if 'not found' in str(cell.value):
						pass
					elif VAF_sample_range1[0] <= float(cell.value) <= VAF_sample_range1[1]:
						cell.fill = PatternFill(start_color=highlight_colors['green'], end_color=highlight_colors['green'], fill_type="solid")
					elif VAF_sample_range2[0] <= float(cell.value) <= VAF_sample_range2[1]:
						cell.fill = PatternFill(start_color=highlight_colors['yellow'], end_color=highlight_colors['yellow'], fill_type="solid")
					elif VAF_sample_range3[0] <= float(cell.value) <= VAF_sample_range3[1]:
						cell.fill = PatternFill(start_color=highlight_colors['orange'], end_color=highlight_colors['orange'], fill_type="solid")
					elif VAF_sample_range4[0] <= float(cell.value) <= VAF_sample_range4[1]:
						cell.fill = PatternFill(start_color=highlight_colors['red'], end_color=highlight_colors['red'], fill_type="solid")
				elif column_name == 'VAF pop':
					if cell.value == 'not found':
						pass
					elif VAF_pop_range1[0] <= float(cell.value) <= VAF_pop_range1[1]:
						cell.fill = PatternFill(start_color=highlight_colors['green'], end_color=highlight_colors['green'], fill_type="solid")
					elif VAF_pop_range2[0] <= float(cell.value) <= VAF_pop_range2[1]:
						cell.fill = PatternFill(start_color=highlight_colors['yellow'], end_color=highlight_colors['yellow'], fill_type="solid")
					elif VAF_pop_range3[0] <= float(cell.value) <= VAF_pop_range3[1]:
						cell.fill = PatternFill(start_color=highlight_colors['orange'], end_color=highlight_colors['orange'], fill_type="solid")
					elif VAF_pop_range4[0] <= float(cell.value) <= VAF_pop_range4[1]:
						cell.fill = PatternFill(start_color=highlight_colors['red'], end_color=highlight_colors['red'], fill_type="solid")
				elif column_name == 'SIFT score':
					if cell.value == 'not found':
						pass
					elif SIFT_range1[0] <= float(cell.value) <= SIFT_range1[1]:
						cell.fill = PatternFill(start_color=highlight_colors['green'], end_color=highlight_colors['green'], fill_type="solid")
					elif SIFT_range2[0] <= float(cell.value) <= SIFT_range2[1]:
						cell.fill = PatternFill(start_color=highlight_colors['yellow'], end_color=highlight_colors['yellow'], fill_type="solid")
					elif SIFT_range3[0] <= float(cell.value) <= SIFT_range3[1]:
						cell.fill = PatternFill(start_color=highlight_colors['orange'], end_color=highlight_colors['orange'], fill_type="solid")
					elif SIFT_range4[0] <= float(cell.value) <= SIFT_range4[1]:
						cell.fill = PatternFill(start_color=highlight_colors['red'], end_color=highlight_colors['red'], fill_type="solid")
				elif column_name == 'Polyphen2 score':
					if cell.value == 'not found':
						pass
					elif PolyPhen2_range1[0] <= float(cell.value) <= PolyPhen2_range1[1]:
						cell.fill = PatternFill(start_color=highlight_colors['red'], end_color=highlight_colors['green'], fill_type="solid")
					elif PolyPhen2_range2[0] <= float(cell.value) <= PolyPhen2_range2[1]:
						cell.fill = PatternFill(start_color=highlight_colors['orange'], end_color=highlight_colors['yellow'], fill_type="solid")
					elif PolyPhen2_range3[0] <= float(cell.value) <= PolyPhen2_range3[1]:
						cell.fill = PatternFill(start_color=highlight_colors['yellow'], end_color=highlight_colors['orange'], fill_type="solid")
					elif PolyPhen2_range4[0] <= float(cell.value) <= PolyPhen2_range4[1]:
						cell.fill = PatternFill(start_color=highlight_colors['green'], end_color=highlight_colors['red'], fill_type="solid")
				elif column_name == 'SIFT pred':
					if 'deleterious' in cell.value.lower():
						cell.fill = PatternFill(start_color=highlight_colors['green'], end_color=highlight_colors['green'], fill_type="solid")
					elif 'tolerated' in cell.value.lower():
						cell.fill = PatternFill(start_color=highlight_colors['red'], end_color=highlight_colors['red'], fill_type="solid")
				elif column_name == 'Polyphen2 pred':
					if 'probably' in cell.value.lower():
						cell.fill = PatternFill(start_color=highlight_colors['green'], end_color=highlight_colors['green'], fill_type="solid")
					elif 'possibly' in cell.value.lower():
						cell.fill = PatternFill(start_color=highlight_colors['yellow'], end_color=highlight_colors['yellow'], fill_type="solid")
					elif 'benign' in cell.value.lower():
						cell.fill = PatternFill(start_color=highlight_colors['orange'], end_color=highlight_colors['orange'], fill_type="solid")

		alignment = Alignment(horizontal="center")
		border = Border(left=Side(style='thin'),
						right=Side(style='thin'),
						top=Side(style='thin'),
						bottom=Side(style='thin'))

		for cell in worksheet[1]:
			if cell.value:
				cell.alignment = alignment
				cell.border = border

		for column in worksheet.columns:
			max_length = 0
			column_letter = column[0].column_letter
			for cell in column:
				try:
					if len(str(cell.value)) > max_length:
						max_length = len(cell.value)
				except:
					pass
			adjusted_width = (max_length + 2) * 1.2
			worksheet.column_dimensions[column_letter].width = adjusted_width

		for row in worksheet.iter_rows(min_row=1):
			for cell in row:
				cell.border = border
				cell.alignment = alignment
		workbook.save(output_path + '.xlsx')

	if 'CSV' in formats:
		df_variants_sorted.to_csv(output_path + '.csv', index=True)
	if 'TSV' in formats:
		df_variants_sorted.to_csv(output_path + '.tsv', sep='\t', index=True)


def create_ordered_dataframe(d):
	'''
	Take a dictionary contaning snp counts, ordered it using it and return the pandas dataframe corresponding
	Input : dictionary contaning snp counts
	Output : dataframe corresponding to the dictionary
	'''
	for k, v in d.items():
		d[k] = OrderedDict(sorted(v.items(), key=lambda x: x[0]))
	d = OrderedDict(sorted(d.items(), key=lambda t: t[0]))
	df = pd.DataFrame.from_dict(d, orient="index").stack().to_frame()
	return df


def true_stem(path):
	'''
	Takes a file Path and return the true stem of this path (file name without extensions)
	'''
	stem = Path(path).stem
	return stem if stem == path else true_stem(stem)


def create_snp_plot(dict_para, d, dcount, name, vcf_name, logger, output_path):
	'''
		Creation of the profile graph
		Input : a dictionnary containing the count for all kind of SNP (d), same dictionnary but with counts instead of percentage, the output file name and the logger
		Output : write a .svg file
		'''
	if dict_para['verbose_prints'].upper() == 'TRUE':
		print(f'Draw profile in {name}...')
	logger.info(f'Draw profile in {name}')


	df = create_ordered_dataframe(d)
	df.columns = [str(true_stem(vcf_name))]
	df.rename(columns={'R01-1_passed': 'Frequency (%)', 'Count': 'Count'}, inplace=True)

	# color for graph

	white_96 = ['white'] * 96
	color_6 = ['darkblue', 'blue', 'lightblue', 'darkgreen', 'green', 'lightgreen']
	color_96 = []
	for i in color_6:
		color_96 += [i] * 16
	bars = [index[1] for index, row in df.iterrows()]
	height = [float(ser.iloc[0]) for _, ser in df.iterrows()]
	group = []
	for index, row in df.iterrows():
		if index[0] not in group:
			group.append(index[0])
	x_pos = np.arange(len(bars))

	# Create bars
	plt.figure(figsize=(15, 10))
	ax1 = plt.subplot(1, 1, 1)
	plt.bar(x_pos, height, color=color_96)

	# Create names on the x-axis
	plt.xticks(x_pos, bars, rotation=90, fontsize=8)
	ax1.set_ylabel('Frequency of each mutation-type (%)')

	# Set scond x-axis
	ax2 = ax1.twiny()
	newlabel = group
	newpos = [8, 24, 40, 56, 72, 88]
	ax2.set_xticks(newpos)
	ax2.set_xticklabels(newlabel)
	# set the position of the second x-axis to bottom
	ax2.xaxis.set_ticks_position('bottom')
	ax2.xaxis.set_label_position('bottom')
	ax2.spines['bottom'].set_position(('outward', 36))
	ax2.set_xlabel('Mutation types')
	ax2.set_xlim(ax1.get_xlim())
	plt.tick_params(
		axis='x',  # changes apply to the x-axis
		which='both',  # both major and minor ticks are affected
		bottom=False,  # ticks along the bottom edge are off
		top=False,  # ticks along the top edge are off
		labelbottom=True)  # label along the bottom edge are on
	for xtick, color in zip(ax2.get_xticklabels(), color_6):
		xtick.set_color(color)
	ax2.spines['bottom'].set_visible(False)

	plt.draw()
	ylabels = [
		str(round(float(ytick.get_text()), 2)) if (ytick.get_text()[0] == '0' or ytick.get_text()[0] == '1') else str(
			round(float(ytick.get_text()[1:]), 2)) for ytick in ax1.get_yticklabels()]
	ax1.set_yticks(ax1.get_yticks())
	ax1.set_yticklabels([str(round(float(i) * 100, 1)) for i in ylabels])

	if 'PNG' in dict_para['S_SNP_profile_formats(s)'].upper():
		png = name + '.png'
		plt.savefig(png)
	if 'SVG' in dict_para['S_SNP_profile_formats(s)'].upper():
		svg = name + '.svg'
		plt.savefig(svg)
	if 'JPG' in dict_para['S_SNP_profile_formats(s)'].upper():
		jpg = name + '.jpg'
		plt.savefig(jpg)

	plt.close()

	counts = []
	for i in df.index:
		counts.append(dcount[i[0]][i[1]])
	df["Count"] = counts
	df.rename(columns=lambda x: re.sub(r'.*_.*', 'Frequency (%)', x), inplace=True)
	df["Frequency (%)"] = df["Frequency (%)"].apply(lambda x: round(x * 100, 2) if x % 1 != 0 else int(x))
	total_count = df['Count'].sum()

	df.reset_index(inplace=True)
	df = df.rename(columns={df.columns[0]: "Mutation", df.columns[1]: "Context"})
	df.at[0, 'Total'] = int(total_count)

	if "XLSX" in dict_para['S_SNP_table_format'].upper():
		file_path = Path(name).with_suffix(".xlsx")
		df.to_excel(file_path, index=False)
	if "CSV" in dict_para['S_SNP_table_format'].upper():
		file_path = Path(name).with_suffix(".csv")
		df.to_csv(file_path, index=False)
	if "TSV" in dict_para['S_SNP_table_format'].upper():
		new_path = os.path.join(os.path.dirname(name), 'SNP_profile.tsv')
		df.to_csv(new_path, sep='\t', index=False)


def create_indel_plot(dict_para, deletion, insertion, name, vcf_name, logger):
	'''
	Creation of the indel size graph
	Input : two dictionnaries containing the count for insertion/deletion of different size, the output file name and the logger
	Output : write a .svg file
	'''

	insert = True
	delet = True
	if deletion == Counter() and insertion == Counter():  # If no Indel
		print('Warning ! No indel in vcf !')
		logger.warning(f'No indel in vcf !')
		return None
	elif insertion == Counter():  # If no insertion
		print('Warning ! No insertion in vcf !')
		logger.warning(f'No insertion in vcf !')
		insert = False
	elif deletion == Counter():  # If no deletion
		print('Warning ! No deletion in vcf !')
		logger.warning(f'No deletion in vcf !')
		delet = False
	if dict_para['verbose_prints'].upper() == 'TRUE':
		print(f'Draw indel size barplot in {name}...')
	logger.info(f'Draw indel size barplot in {name}')

	width = 0.25

	sum_del = 0
	sum_ins = 0
	if delet:
		sum_del = sum(deletion.values())
		df_del = pd.DataFrame.from_dict(deletion, orient='index').sort_index()
		height_del = list([float(i) / sum_del for i in df_del[0]])
		r_del = list(df_del.index)
		max_del = max([float(i) for i in r_del])
	if insert:
		sum_ins = sum(insertion.values())
		df_ins = pd.DataFrame.from_dict(insertion, orient='index').sort_index()
		height_ins = list([float(i) / sum_ins for i in df_ins[0]])
		r_ins = list(df_ins.index)
		max_ins = max([float(i) for i in r_ins])

	if delet and insert:
		maximum = max([int(max_del) + 1, int(max_ins) + 1])
	elif delet:
		maximum = int(max_del) + 1
	elif not delet:
		maximum = int(max_ins) + 1
	x = [0] + [i + 1 for i in range(maximum)]

	ax = plt.axes()
	if delet and insert:
		plt.bar([float(i) - (width * 0.65) for i in r_del], height_del, color='k', width=width, edgecolor='k',
				label='Deletion')
		plt.bar([float(i) + (width * 0.65) for i in r_ins], height_ins, color='r', width=width, edgecolor='r',
				label='Insertion')
	elif not insert:
		plt.bar([float(i) for i in r_del], height_del, color='k', width=width, edgecolor='k', label='Deletion')
	elif not delet:
		plt.bar([float(i) for i in r_ins], height_ins, color='r', width=width, edgecolor='r', label='Insertion')

	plt.xticks(x, fontsize=9)
	plt.yticks()

	if maximum > 10:
		##############################################
		# Adaptation of the figure size to the x range

		plt.gca().margins(x=0)
		plt.gcf().canvas.draw()
		tl = plt.gca().get_xticklabels()
		maxsize = max([t.get_window_extent().width for t in tl]) + 1
		m = 0.5  # inch margin
		s = maxsize / plt.gcf().dpi * maximum + 2 * m
		margin = m / plt.gcf().get_size_inches()[0]
		plt.gcf().subplots_adjust(left=margin, right=1. - margin)
		plt.gcf().set_size_inches(s * 1.5, plt.gcf().get_size_inches()[1])

		plot_margin = 1
		x0, x1, y0, y1 = plt.axis()
		plt.axis((x0 - plot_margin, x1 + plot_margin * 2, y0, y1))

	plt.draw()
	ylabels = [
		str(round(float(ytick.get_text()), 2)) if (ytick.get_text()[0] == '0' or ytick.get_text()[0] == '1') else str(
			round(float(ytick.get_text()[1:]), 2)) for ytick in ax.get_yticklabels()]
	ax.set_yticks(ax.get_yticks())
	ax.set_yticklabels([str(round(float(i) * 100, 1)) for i in ylabels])
	plt.xlabel("Indel size (bp)", labelpad=10)
	plt.ylabel("Indel percentage", labelpad=10)
	plt.legend(frameon=True, edgecolor='black')

	if "SVG" in dict_para['S_indel_profile_format(s)'].upper():
		svg = name + '.svg'
		plt.savefig(svg, dpi=400)
	if "PNG" in dict_para['S_indel_profile_format(s)'].upper():
		png = name + '.png'
		plt.savefig(png, dpi=400)
	if "JPG" in dict_para['S_indel_profile_format(s)'].upper():
		jpg = name + '.jpg'
		plt.savefig(jpg, dpi=400)

	plt.close()

	if dict_para['verbose_prints'].upper() == 'TRUE':
		print(f'Saving indel counts in {Path(name).with_suffix(".tsv")}...')
	logger.info(f'Save indel counts in {Path(name).with_suffix(".tsv")}')

	# if delet and insert:
	# 	df_del.columns = [str(Path(vcf_name).stem)]
	# 	df_ins.columns = [str(Path(vcf_name).stem)]
	#
	# 	tsv = name.rsplit('/', 1)[0] + '/deletion_profile'
	# 	if 'TSV' in dict_para['indel_tables_format'].upper():
	# 		df_del.to_csv(Path(tsv).with_suffix(".tsv"), sep='\t')
	# 	if 'CSV' in dict_para['indel_tables_format'].upper():
	# 		df_del.to_csv(Path(tsv).with_suffix(".csv"), sep=',')
	# 	if 'XLSX' in dict_para['indel_tables_format'].upper():
	# 		df_del.to_excel(Path(tsv).with_suffix(".xlsx"), index=False)
	#
	# 	csv = name.rsplit('/', 1)[0] + '/insertion_profile'
	# 	if 'TSV' in dict_para['indel_tables_format'].upper():
	# 		df_ins.to_csv(Path(csv).with_suffix(".tsv"), sep='\t')
	# 	if 'CSV' in dict_para['indel_tables_format'].upper():
	# 		df_ins.to_csv(Path(csv).with_suffix(".csv"), sep=',')
	# 	if 'XLSX' in dict_para['indel_tables_format'].upper():
	# 		df_ins.to_excel(Path(csv).with_suffix(".xlsx"), index=False)


def is_fasta(filename: str) -> bool:
	'''
	Is the file a fasta file?
	Input : path to the file
	Output : True, False (if the file is a pickle file for example) or raise an error if file doesn't exist
	'''
	try:
		fa = pyfastx.Fastx(filename)
		fasta = [content[0] for content in fa]
		return any(fasta)
	except RuntimeError as runerr:
		return False
	except UnicodeDecodeError as uderr:
		return False
	except FileNotFoundError as fnferr:
		print(f'\nFile {filename} doesn\'t exists: {fnferr}\n')
		raise FileNotFoundError(f'\nFile {filename} doesn\'t exists\n')
	except FileExistsError as feerr:
		print(f'\nFile {filename} doesn\'t exists: {feerr}\n')
		raise FileExistsError(f'\nFile {filename} doesn\'t exists\n')
	except:
		print(f'\nUnexpected error: {sys.exc_info()[0]}\n')
		raise


def is_pickle(filename: str) -> bool:
	'''
	Is the file a pickle file?
	Input : path to the file
	Output : True, False (if the file is a fatsa file for example) or raise an error if file doesn't exist
	'''
	try:
		with open(filename, 'rb') as f:
			genome = pk.load(f)
			return any(genome)
	except EOFError as eoferr:
		return False
	except UnicodeDecodeError as uderr:
		return False
	except FileNotFoundError as fnferr:
		print(f'\nFile {filename} doesn\'t exist: {fnferr}\n')
		raise FileNotFoundError(f'\nFile {filename} doesn\'t exists\n')
	except pk.UnpicklingError:
		return False
	except:
		print(f'\nUnexpected error: {sys.exc_info()[0]}\n')
		raise


def get_genome_dict(dict_parameters, genome_file, logger):
	'''
	Check the genome file and (1) read the fasta file and create the pickle dictionary or (2) load the pickle dictionary and return this dictionary
	Input : path to the genome file and logger
	Output : dictionnary containing the genome
	'''
	if dict_parameters['verbose_prints'].upper() == 'TRUE':
		print(f'Check input genome file {genome_file}...')
	logger.info(f'Check input genome file {genome_file}')
	if is_fasta(genome_file):
		genome = {}
		path = Path(genome_file).with_suffix('.pk')
		fa = pyfastx.Fastx(genome_file)
		print("chr", "length")
		for content in fa:
			name = content[0]
			seq = content[1]
			genome[name] = seq
			print(name, len(seq))

		if not path.exists():
			print(f'Create the pickle genome dictionary in {path}')
			logger.info(f'Create the pickle genome dictionary in {path}')
			with open(path, 'wb') as out_pickle:
				pk.dump(genome, out_pickle)
		else:
			print(f'File {path} already exists !')
			logger.warning(f'File {path} already exists !')
	elif is_pickle(genome_file):
		if dict_parameters['verbose_prints'].upper() == 'TRUE':
			print('Charge the pickle genome dictionary...')
		with open(genome_file, 'rb') as f:
			genome = pk.load(f)
	else:
		raise ValueError(f'{genome_file} is not a fasta or a pickle file !')

	return genome


def add_snp(snp_count, ref, alt, triplet):
	'''
	Add 1 to the SNP in counter
	As in VCF the variant is be given for the leading (forward, 5'-3') strand, if we get G>* or A>* the reference, the variant and the triplet are reversed complement
	Input : SNP counter, the reference allele, the variant and the triplet containing the reference
	Output : SNP counter
	'''
	ref = ref.upper()
	alt = alt.upper()
	triplet = triplet.upper()
	if TRANSLATE[str(ref + ">" + alt)] == str(ref + ">" + alt):
		snp_count[str(ref + ">" + alt)][triplet] += 1
	else:
		try:
			snp_count[TRANSLATE[str(ref + ">" + alt)]][triplet.translate(TAB)[::-1]] += 1
		except:
			if 'N' in triplet.upper():
				print('N in triplet')
				if not TRANSLATE[str(ref + ">" + alt)] in snp_count.keys():
					snp_count[TRANSLATE[str(ref + ">" + alt)]] = {}
					snp_count[TRANSLATE[str(ref + ">" + alt)]][triplet.translate(TAB)[::-1]] = 0
				snp_count[TRANSLATE[str(ref + ">" + alt)]][triplet.translate(TAB)[::-1]] += 1

	return snp_count


def size_indel(ref, alt):
	'''
	Get the indel size
	Input : reference allele and the alternative variant
	Output : size of the indel variant
	'''
	size = 0
	if len(ref) > len(alt):
		size = len(ref) - len(alt)
	elif len(alt) > len(ref):
		size = len(alt) - len(ref)
	else:
		print('Not an indel !')
		raise ValueError('Not an indel !')
	return size


def count_variants(dict_parameters, dict_colors, vcf_file_filter, vcf_file_pass, genome, logger):
	'''
	Read the vcf file(s) and create the different counters to summary the files
	Input : path to the vcf file from filter module (optional), path to the vcf file containing only pass variants from filter module, genome in a dictionnary and the output path file and the logger
	Output : counters of (1) the deletion size, (2) the insertion size, (3) the snp (in tripet), the impacted genes list and a dictiionnary contaning general stats from the vcf
	'''
	if dict_parameters['verbose_prints'] == 'TRUE':
		print('Counting variants and get genes...')
	logger.info(f'Compute stats on variants and genes')

	##################
	# Counter/list creation
	stats = {}
	stats['Total'] = 0
	stats['germline'] = 0
	stats['PON'] = 0
	stats['non functional'] = 0
	stats['germline+PON'] = 0
	stats['germline+non functional'] = 0
	stats['PON+non functional'] = 0
	stats['germline+PON+non functional'] = 0
	stats['PASS'] = 0
	stats['SNP'] = [0, set()]  # Single Nucleotide Polymorphism
	stats['DNP'] = [0, set()]  # Double Nucleotide Polymorphism
	stats['TNP'] = [0, set()]  # Triple Nucleotide Polymorphism
	stats['ONP'] = [0, set()]  # Quadruple or more Nucleotide Polymorphism
	stats['INSERTION'] = [0, set()]
	stats['DELETION'] = [0, set()]
	genes_list = {}
	gene_info = [0, [0, 0, 0, 0, 0, 0], [], [], '', [], [], [], [], [], {}, {}]
	dict_impacts = {}
	dict_multiple_alts = {}
	# [gene, [variants number (tumor burden),[snp, dnp, tnp, onp, insertion, deletion]], chromosome, reference, alternative, position]
	idx = {}
	snp_count = create_snp_dict()
	counter_deletion_size = Counter()
	counter_insertion_size = Counter()

	f = open(vcf_file_pass, 'r', encoding='latin1')
	nb_lines_pass = len(
		list(dropwhile(lambda x: x[0] == '#', (line for line in f))))  # passed.vcf content (number of lines / variants)

	####################
	# Read pass vcf file
	if dict_parameters['verbose_prints'] == 'TRUE':
		print(f'Read {vcf_file_pass}...')
	logger.info(f'Read {vcf_file_pass}')

	if dict_parameters['colored_execution'].upper() == 'FALSE' or dict_parameters['colored_execution'].upper() == 'NONE':
		color = ''
		no_tqdm_bar = True
	else:
		color = dict_colors['yellow3']
		no_tqdm_bar = False
	with tqdm(total=nb_lines_pass, disable=no_tqdm_bar, bar_format='{l_bar}{bar:15}{r_bar}', ncols=130, smoothing=1) as pbar:
		pbar.set_description(color + f' -> parsing passed.vcf file:')  # progress bar
		vcf_reader_pass = read_vcf(vcf_file_pass)  # create the read file generator
		for line in vcf_reader_pass:
			if type(line) == type({}):  # get index position
				idx = line
			else:
				# Get values
				chr = line[idx['idx_chr']]
				ref = line[idx['idx_ref']]
				alts = line[idx['idx_alts']].split(',')
				pos = int(line[idx['idx_pos']])
				match_DP = re.search(r'DP=(\d+)', str(line))
				DP_value = match_DP.group(1)

				# predicted impact on protein
				if len(alts) > 1:
					for k in range(len(alts)):
						g_variant = 'g.' + chr + ':' + str(pos) + ref + '>' + alts[k]
						if g_variant not in dict_multiple_alts.keys():
							dict_multiple_alts[g_variant] = {}
						try:
							index_VAF = None
							for i, item in enumerate(line):
								if item.find(':VAF') != -1:
									index_VAF = i
							l_formats = line[index_VAF].split(':')
							index_VAF_value = None
							for i, item in enumerate(l_formats):
								if item.find('VAF') != -1:
									index_VAF_value = i
							l_formats_values = line[index_VAF + 1].split(':')
							VAF_sample = l_formats_values[index_VAF_value]
						except:
							try:
								l_format_names = line[-2].split(':')
								af_index = l_format_names.index('AF')
								VAF_sample = line[-1].split(':')[af_index].split(',')[k]
								dict_multiple_alts[g_variant]['VAF_sample'] = VAF_sample
							except:
								print('problem with VAF sample finding')

						try:
							matches_VAF_pop = re.findall(r';AF=([^;]+)', str(line))
							VAF_pops = []
							for match in matches_VAF_pop:
								VAF_pops.append(match)
							if VAF_pops == []:
								matches_VAF_pop = re.findall(r'gnomad40_exome_AF=([^;]+)', str(line))
								VAF_pops = []
								for match in matches_VAF_pop:
									VAF_pops.append(float(match))
							dict_multiple_alts[g_variant]['VAF_pop'] = VAF_pops[k]
						except:
							print('no VAF pop found in VCF')

						try:
							matches_SIFT_score = re.findall(r';SIFT_score=([^;]+)', str(line))
							SIFT_scores = []
							for match in matches_SIFT_score:
								SIFT_scores.append(float(match))
							dict_multiple_alts[g_variant]['SIFT_score'] = SIFT_scores[k]
						except:
							print('no SIFT_score found in VCF')
						try:
							matches_Polyphen_score = re.findall(r';Polyphen2_HDIV_score=([^;]+)', str(line))
							Polyphen_scores = []
							for match in matches_Polyphen_score:
								Polyphen_scores.append(float(match))
							dict_multiple_alts[g_variant]['Polyphen_score'] = Polyphen_scores[k]
						except:
							print('no Polyphen2_HDIV_score found in VCF')
						try:
							matches_SIFT_pred = re.findall(r';SIFT_pred=([^;]+)', str(line))
							SIFT_preds = []
							for match in matches_SIFT_pred:
								SIFT_preds.append(match)
							dict_multiple_alts[g_variant]['SIFT_pred'] = SIFT_preds[k]
						except:
							print('no SIFT_pred found in VCF')
						try:
							matches_Polyphen_pred = re.findall(r';Polyphen2_HDIV_pred=([^;]+)', str(line))
							Polyphen_preds = []
							for match in matches_Polyphen_pred:
								Polyphen_preds.append(match)
							dict_multiple_alts[g_variant]['Polyphen_pred'] = Polyphen_preds[k]
						except:
							print('no Polyphen2_HDIV_pred found in VCF')


				g_variant = 'g.' + chr + ':' + str(pos) + ref + '>' + alts[0]
				if g_variant not in dict_impacts.keys():
					dict_impacts[g_variant] = {}
				try:
					match_SIFT_score = re.search(r';SIFT_score=([^;]+)', str(line)).group(1)
					if match_SIFT_score == '.':
						SIFT_score = 'not found'
					else:
						SIFT_score = float(match_SIFT_score)
					dict_impacts[g_variant]['SIFT_score'] = SIFT_score
				except:
					print('SIFT_score not found in VCF')
					pass

				try:
					match_Polyphen_score = re.search(r';Polyphen2_HDIV_score=([^;]+)', str(line)).group(1)
					if match_Polyphen_score == '.':
						Polyphen_score = 'not found'
					else:
						Polyphen_score = float(match_Polyphen_score)
					dict_impacts[g_variant]['Polyphen_score'] = Polyphen_score
				except:
					print('Polyphen2_HDIV_score not found in VCF')
					pass

				try:
					match_SIFT_pred = re.search(r';SIFT_pred=([^;]+)', str(line)).group(1)
					if match_SIFT_pred == '.':
						SIFT_pred = 'not found'
					else:
						SIFT_pred = match_SIFT_pred
					dict_impacts[g_variant]['SIFT_pred'] = SIFT_pred
				except:
					print('SIFT_pred not found in VCF')
					pass

				try:
					match_Polyphen_pred = re.search(r';Polyphen2_HDIV_pred=([^;]+)', str(line)).group(1)
					if match_Polyphen_pred == '.':
						Polyphen_pred = 'not found'
					else:
						Polyphen_pred = match_Polyphen_pred
					dict_impacts[g_variant]['Polyphen_pred'] = Polyphen_pred
				except:
					print('Polyphen2_HDIV_pred not found in VCF')
					pass

				try:
					match_VAF_pop = re.search(r';AF=(\d+)', str(line))
					VAF_pop = float(match_VAF_pop.group(1))
				except:
					match_VAF_pop = re.search(r';gnomad40_exome_AF=([^;]+)', str(line))
					try:
						VAF_pop = float(match_VAF_pop.group(1))
						if VAF_pop == 0.0:
							match_VAF_pop = re.search(r';gnomad40_exome_AF_raw=([^;]+)', str(line))
							VAF_pop = float(match_VAF_pop.group(1))
					except:
						VAF_pop = 'not found'
				try:
					index_VAF = None
					for i, item in enumerate(line):
						if item.find(':VAF') != -1:
							index_VAF = i

					l_formats = line[index_VAF].split(':')
					index_VAF_value = None
					for i, item in enumerate(l_formats):
						if item.find('VAF') != -1:
							index_VAF_value = i

					l_formats_values = line[index_VAF + 1].split(':')
					VAF_sample = float(l_formats_values[index_VAF_value])
				except:
					try:
						l_format_names = line[-2].split(':')
						af_index = l_format_names.index('AF')
						VAF_sample = float(line[-1].split(':')[af_index])
					except:
						try:
							l_format_names = line[-3].split(':')
							af_index = l_format_names.index('AF')
							VAF_sample = float(line[-1].split(':')[af_index])
						except:
							print('problem with VAF sample finding')

				triplet = str(genome[chr][pos - 2:pos + 1])  # Get triplet by using the genome and variant position (better definition of SNP with its environment)
				infos = [tuple(infos.split('=')) if len(infos.split('=')) > 1 else (infos, '') for infos in
						 line[idx['idx_info']].split(';')]

				try:  # si le variant est annotée par Funcotator
					idx_funcotation = list(zip(*infos))[0].index('FUNCOTATION')
					gene = list(zip(*infos))[1][idx_funcotation].split('|')[0].lstrip('[')
					l_format = line[-2].split(':')
					l_format_values = line[-1].split(':')
					format_dict = dict(zip(l_format, l_format_values))

				except:  # si le variant est annotée par ANNOVAR
					str_infos = str(infos)
					pattern = r"'Gene.refGene', '(.*?)'"
					match = re.search(pattern, str_infos)
					if match:  # enlever - et après???
						gene = match.group(1)
						if ')' in gene:
							gene = gene.split(')')[0]
						if "'" in gene:
							gene = gene.replace("'", "")
						# if gene.startswith('LOC'):
						#     gene = 'error'
						if '\\' in gene:
							gene = gene.split("\\")[0]
					else:
						print("error : no gene name found")
				# print(gene)
				# if len(gene) > 10:
				# 	print('a')

				if gene != 'NONE' and not gene.startswith('LINC'):  # if ANNOVAR doesn't find any gene for this variant or if lncRNA
					if gene not in genes_list.keys():
						if len(gene) > 15:
							# print('gene name length > 15: ' + gene)
							pass
						genes_list[gene] = deepcopy(gene_info)
					if genes_list[gene][4] == '':
						genes_list[gene][4] = chr
					genes_list[gene][2].append(ref)
					genes_list[gene][3].append(alts)
					genes_list[gene][5].append(pos)
					qual = line[5]
					genes_list[gene][6].append(DP_value)
					genes_list[gene][7].append(qual)
					try:
						genes_list[gene][8].append(VAF_sample)
						genes_list[gene][9].append(VAF_pop)
					except:
						pass

					# exonic function / mutation subtype
					match = re.search(r'ExonicFunc\.refGene=(.*?);', str(line))
					if match:
						mutation_subtype = match.group(1)
						if mutation_subtype == '.':
							mutation_subtype = 'unknown'
						if mutation_subtype not in genes_list[gene][10].keys():
							genes_list[gene][10][mutation_subtype] = 0
						if mutation_subtype not in genes_list[gene][11].keys():
							genes_list[gene][11][mutation_subtype] = []
						genes_list[gene][11][mutation_subtype].append(pos)
						genes_list[gene][10][mutation_subtype] += 1

					for i, alt in enumerate(alts):
						stats['PASS'] += 1
						genes_list[gene][0] += 1

						if is_snp(len(ref), len(alt)):
							stats['SNP'][0] += 1

							try:
								snp_count = add_snp(snp_count, ref, alt, triplet)
							except:
								print('b')
							genes_list[gene][1][0] += 1
							stats['SNP'][1].add(gene)

						elif is_dnp(len(ref), len(alt)):
							stats['DNP'][0] += 1
							genes_list[gene][1][1] += 1
							stats['DNP'][1].add(gene)

						elif is_tnp(len(ref), len(alt)):
							stats['TNP'][0] += 1
							genes_list[gene][1][2] += 1
							stats['TNP'][1].add(gene)

						elif is_onp(len(ref), len(alt)):
							stats['ONP'][0] += 1
							genes_list[gene][1][3] += 1
							stats['ONP'][1].add(gene)

						elif is_insertion(len(ref), len(alt)):
							counter_insertion_size.update([size_indel(ref, alt)])
							stats['INSERTION'][0] += 1
							genes_list[gene][1][4] += 1
							stats['INSERTION'][1].add(gene)

						elif is_deletion(len(ref), len(alt)):
							counter_deletion_size.update([size_indel(ref, alt)])
							stats['DELETION'][0] += 1
							genes_list[gene][1][5] += 1
							stats['DELETION'][1].add(gene)
				pbar.update(1)

	#####################################################
	# If whole vcf (not just pass variants) file is given
	if dict_parameters['keep_filtered_vcf_after_run'].upper() == 'TRUE' or dict_parameters['keep_filtered_vcf_after_run'] == 'YES':
		f = open(vcf_file_filter, 'r', encoding='latin1')
	nb_lines = len(list(dropwhile(lambda x: x[0] == '#', (line for line in f))))  # filtered.vcf content (number of lines / variants)
	f.close()
	if vcf_file_filter:
		if dict_parameters['verbose_prints'] == 'TRUE':
			print(f'Read {vcf_file_filter}...')
		logger.info(f'Read {vcf_file_filter}')
		if dict_parameters['colored_execution'].upper() == 'FALSE' or dict_parameters['colored_execution'].upper() == 'NONE':
			color = ''
			no_tqdm_bar = True
		else:
			color = dict_colors['yellow3']
			no_tqdm_bar = False

	# Get percentage from count for pass SNP
	snp_count_pct = deepcopy(snp_count)
	tot = sum([sum(val.values()) for key, val in snp_count.items()])
	if tot != 0:
		for key, val in snp_count.items():
			for k, v in val.items():
				snp_count_pct[key][k] = round((v / tot), 8)  # get the percentage of each SNP in the total number of SNP
	else:
		print('NO SNP IN SAMPLE ' + vcf_file_pass + ' !')
		for key, val in snp_count.items():
			for k, v in val.items():
				snp_count_pct[key][k] = 0

	return counter_deletion_size, counter_insertion_size, snp_count_pct, snp_count, genes_list, stats, dict_impacts, dict_multiple_alts


def is_snp(ref_length, alt_length):
	'''ex: A > T '''
	if ref_length == 1:
		return ref_length == alt_length
	else:
		return False


def is_dnp(ref_length, alt_length):
	'''ex: AT > GA '''
	if ref_length == 2:
		return ref_length == alt_length
	else:
		return False


def is_tnp(ref_length, alt_length):
	'''ex: AAG > TGC '''
	if ref_length == 3:
		return ref_length == alt_length
	else:
		return False


def is_onp(ref_length, alt_length):
	'''ex: AATGG > GTCAA '''
	if ref_length > 3:
		return ref_length == alt_length
	else:
		return False


def is_deletion(ref_length, alt_length):
	'''ex: TAAG > T '''
	return ref_length > alt_length


def is_insertion(ref_length, alt_length):
	'''ex: T > TAAG '''
	return ref_length < alt_length


def write_stats(dict_subtypes, dict_parameters, vcf_file_filter: str, vcf_file_pass: str, out_stats: str, stats: Counter, dict_impacts, logger):
	'''
	Writes a simple statistics file on the variants of the VCF file
	Input : path to the vcf file from filter module (optional), path to the vcf file containing only pass variants from filter module, counter containing vcf stats and the output path file
	Output : Write the vcf stats file
	'''
	config_path = dict_parameters['config_path']
	with open(config_path, "r") as file:
		lines = file.readlines()
		for line in lines:
			if "max_SIFT_score" in line:
				sift_score_threshold = float(line.split("max_SIFT_score =")[1].strip())
			elif "min_PolyPhen2_score" in line:
				polyphen_threshold = float(line.split("min_PolyPhen2_score =")[1].strip())

	if not all(not bool(d) for d in dict_impacts.values()):
		sift_thresholds = [-1, 0.05, 0.2, 1]
		polyphen_thresholds = [-1, 0.45, 0.95, 1.5]

		dict_impacts_counts = {
			'SIFT scores counts': {f'{sift_thresholds[i]}-{sift_thresholds[i + 1]}': 0 for i in range(len(sift_thresholds) - 1)},
			'Polyphen2 scores counts': {f'{polyphen_thresholds[i]}-{polyphen_thresholds[i + 1]}': 0 for i in range(len(polyphen_thresholds) - 1)},
			'SIFT predictions': {},
			'Polyphen2 predictions': {}
		}

		if not all(not bool(d) for d in dict_impacts.values()):
			for impact in dict_impacts.values():
				sift_score = impact['SIFT_score']
				polyphen_score = impact['Polyphen_score']
				sift_pred = impact['SIFT_pred']
				polyphen_pred = impact['Polyphen_pred']

				if sift_score == 'not found':
					if sift_score not in dict_impacts_counts['SIFT scores counts']:
						dict_impacts_counts['SIFT scores counts'][sift_score] = 1
					else:
						dict_impacts_counts['SIFT scores counts'][sift_score] += 1
				else:
					interval = None
					for i in range(len(sift_thresholds) - 1):
						if sift_thresholds[i] <= sift_score < sift_thresholds[i + 1]:
							interval = f'{sift_thresholds[i]}-{sift_thresholds[i + 1]}'
							break
					dict_impacts_counts['SIFT scores counts'][interval] += 1

				if polyphen_score == 'not found':
					if polyphen_score not in dict_impacts_counts['Polyphen2 scores counts']:
						dict_impacts_counts['Polyphen2 scores counts'][polyphen_score] = 1
					else:
						dict_impacts_counts['Polyphen2 scores counts'][polyphen_score] += 1
				else:
					interval = None
					for i in range(len(polyphen_thresholds) - 1):
						if polyphen_thresholds[i] <= polyphen_score < polyphen_thresholds[i + 1]:
							interval = f'{polyphen_thresholds[i]}-{polyphen_thresholds[i + 1]}'
							break
					dict_impacts_counts['Polyphen2 scores counts'][interval] += 1

				dict_impacts_counts['SIFT predictions'].setdefault(sift_pred, 0)
				dict_impacts_counts['SIFT predictions'][sift_pred] += 1
				dict_impacts_counts['Polyphen2 predictions'].setdefault(polyphen_pred, 0)
				dict_impacts_counts['Polyphen2 predictions'][polyphen_pred] += 1

		dict_impacts_counts['SIFT scores counts'] = {key: value for key, value in dict_impacts_counts['SIFT scores counts'].items() if value != 0}
		dict_impacts_counts['Polyphen2 scores counts'] = {key: value for key, value in dict_impacts_counts['Polyphen2 scores counts'].items() if value != 0}
		dict_impacts_counts['SIFT predictions'] = {key: value for key, value in dict_impacts_counts['SIFT predictions'].items() if value != 0}
		sift_preds_mapping = {'D': 'deleterious', 'T': 'tolerated', 'U': 'unknown'}
		dict_impacts_counts['SIFT predictions'] = {sift_preds_mapping.get(key, key): value for key, value in dict_impacts_counts['SIFT predictions'].items() if value != 0}
		dict_impacts_counts['Polyphen2 predictions'] = {key: value for key, value in dict_impacts_counts['Polyphen2 predictions'].items() if value != 0}
		dict_impacts_counts['Polyphen2 predictions'] = {('probably damaging' if key == 'D' else key): value for key, value in dict_impacts_counts['Polyphen2 predictions'].items() if value != 0}
		polyphen_preds_mapping = {'B': 'benign', 'D': 'probably damaging', 'P': 'possibly damaging'}
		dict_impacts_counts['Polyphen2 predictions'] = {polyphen_preds_mapping.get(key, key): value for key, value in dict_impacts_counts['Polyphen2 predictions'].items() if value != 0}

	out_stats = out_stats + ".txt"
	filter_stats = out_stats.replace('passed_stats.txt', 'filtered_stats.txt')
	if dict_parameters['verbose_prints'].upper() == 'TRUE':
		print('Write stats file...')
	logger.info('Write stats file')

	with open(out_stats, 'w') as o:
		o.write('----- MUTATIONS -----\n')
		mutation_counts = {'synonymous_SNV', 'nonsynonymous_SNV', 'stopgain', 'startloss', 'stoploss', 'nonframeshift_insertion',
						   'frameshift_insertion', 'nonframeshift_deletion', 'frameshift_deletion',
						   'nonframeshift_substitution', 'frameshift_substitution', 'unknown'}
		mutation_counts = {mutation_type: 0 for mutation_type in mutation_counts}
		for values in dict_subtypes.values():
			mutations = values[-2]  # Obtient le dictionnaire des mutations pour chaque clé
			for mutation_type, count in mutations.items():
				if mutation_type == 'unknown' or mutation_type == '.':
					print('mutation subtype not found')
					pass
				if mutation_type == '.':
					mutation_counts['unknown'] = mutation_counts['unknown'] + count
				else:
					mutation_type = mutation_type.replace(" ", "")
					mutation_counts[mutation_type] = mutation_counts[mutation_type] + count
		o.write(vcf_file_pass.split('/')[-1])
		o.write(f'\nTotal variants (after filtering): {stats["PASS"]}\n')
		o.write(f'\n----- MUTATION TYPES -----\n')
		o.write(f'SNP: {stats["SNP"][0]}\t\tDNP: {stats["DNP"][0]}\tTNP: {stats["TNP"][0]}\tONP: {stats["ONP"][0]}\n')
		o.write(
			f'INDEL: {stats["INSERTION"][0] + stats["DELETION"][0]}\tINSERTION: {stats["INSERTION"][0]}, DELETION: {stats["DELETION"][0]}\n')
		o.write(f'\n----- MUTATION SUBTYPES -----\n')

		o.write(
			f'synonymous SNV: {mutation_counts["synonymous_SNV"]}\tnon synonymous SNV: {mutation_counts["nonsynonymous_SNV"]}\n')
		o.write(
			f'frameshift substitution: {mutation_counts["frameshift_substitution"]}\tnon frameshift substitution: {mutation_counts["nonframeshift_substitution"]}\n')
		o.write(
			f'stopgain: {mutation_counts["stopgain"]}\tstoploss: {mutation_counts["stoploss"]}\tstartloss: {mutation_counts["startloss"]}\n')
		o.write(
			f'frameshift insertion: {mutation_counts["frameshift_insertion"]}     non frameshift insertion: {mutation_counts["nonframeshift_insertion"]}\n')
		o.write(
			f'frameshift deletion: {mutation_counts["frameshift_deletion"]}     non frameshift deletion: {mutation_counts["nonframeshift_deletion"]}\n')
		try:
			o.write(f'unknown: {mutation_counts["unknown"]}\n')
		except:
			o.write(f'unknown: 0\n')

		if not all(not bool(d) for d in dict_impacts.values()):
			o.write(f'\n----- IMPACT ON PROTEIN -----\n')
			sift_scores = str(dict_impacts_counts["SIFT scores counts"]).replace("'", "").replace("{", "").replace("}", "").replace(":"," =")
			if '-1-' in sift_scores:
				sift_scores = sift_scores.replace('-1-', '0-')
			if '-1.5' in sift_scores:
				sift_scores = sift_scores.replace('-1.5', '-1')
			sift_preds = str(dict_impacts_counts["SIFT predictions"]).replace("'", "").replace("{", "").replace("}", "").replace(":"," =")
			polyphen_scores = str(dict_impacts_counts["Polyphen2 scores counts"]).replace("'", "").replace("{", "").replace("}", "").replace(":"," =")
			if '-1-' in polyphen_scores:
				polyphen_scores = polyphen_scores.replace('-1-', '0-')
			if '-1.5' in polyphen_scores:
				polyphen_scores = polyphen_scores.replace('-1.5', '-1')
			polyphen_preds = str(dict_impacts_counts["Polyphen2 predictions"]).replace("'", "").replace("{", "").replace("}", "").replace(":"," =")
			o.write(f'SIFT scores counts: {sift_scores}\n')
			o.write(f'SIFT predictions: {sift_preds}\n')
			o.write(f'Polyphen2 scores counts: {polyphen_scores}\n')
			o.write(f'Polyphen predictions: {polyphen_preds}\n')

		# o.write(f'---\nUnique impacted genes GS: { len(stats["SNP"][1]) + len(stats["DNP"][1]) + len(stats["TNP"][1]) + len(stats["ONP"][1]) + len(stats["DELETION"][1]) + len(stats["INSERTION"][1]) } (list in {out_genes})\n')
		concatenated = [stats[key][1] for key in ["SNP", "DNP", "TNP", "ONP", "DELETION", "INSERTION"]]  # get the list of impacted genes for each category
		concatenated_list = [value for sublist in concatenated for value in sublist]  # flatten the list
		unique_genes = set(concatenated_list)  # get the unique genes
		l_duplicate_genes = [gene for gene, count in Counter(concatenated_list).items() if count > 1]
		duplicate_genes = (', '.join(l_duplicate_genes))
		o.write(f'\n----- IMPACTED GENES -----\n')
		o.write(f'-> number of impacted genes for each mutation category: \n')
		o.write(f'SNP: {len(stats["SNP"][1])}\t\tDNP: {len(stats["DNP"][1])}\tTNP: {len(stats["TNP"][1])}\tONP: {len(stats["ONP"][1])}\n')
		o.write(f'INDEL: {len(stats["INSERTION"][1]) + len(stats["DELETION"][1])} \tINSERTION: {len(stats["INSERTION"][1])}\tDELETION: {len(stats["DELETION"][1])}\n')
		o.write(f'-> number of impacted genes: {len(unique_genes)}\n')  # write the number of unique genes (not in two or more categories of mutations)
		if len(l_duplicate_genes) > 1:
			o.write(f'{len(l_duplicate_genes)} of those genes are impacted by 2 types of mutations (or more): \n{duplicate_genes}\n')
		elif len(l_duplicate_genes) == 1:
			o.write(f'{len(l_duplicate_genes)} of those genes is impacted by 2 types of mutations (or more): {duplicate_genes}\n')


def create_mutation_types_tables(dict_para):
	TMB, SNP, DNP, TNP, ONP, INDEL, INS, DEL = [], [], [], [], [], [], [], []
	dict_info = {'Variants': TMB, 'SNP': SNP, 'DNP': DNP, 'TNP': TNP, 'ONP': ONP, 'INDEL': INDEL, 'INSERTION': INS,
				 'DELETION': DEL}
	samples_path = os.path.join(dict_para['output_path'], 'samples/')
	for folder in os.listdir(samples_path):
		with (open(samples_path + folder + '/passed_stats.txt', 'r',  encoding="utf8") as file):
			for line in file:  # avant d'arriver à 'Impacted genes'
				line = line.strip().replace('\t', ' ')
				# print(line)
				if 'impacted' in line:
					break
				elif not line.startswith("#") and not line.startswith("---"):
					if 'SNP: ' in line:
						match = re.search(r'SNP: (\d+)', line)
						value = match.group(1) if match else None
						SNP.append(value)
					if 'DNP: ' in line:
						match = re.search(r'DNP: (\d+)', line)
						value = match.group(1) if match else None
						DNP.append(value)
					if 'TNP: ' in line:
						match = re.search(r'TNP: (\d+)', line)
						value = match.group(1) if match else None
						TNP.append(value)
					if 'ONP: ' in line:
						match = re.search(r'ONP: (\d+)', line)
						value = match.group(1) if match else None
						ONP.append(value)
					if 'INSERTION: ' in line:
						match = re.search(r'INSERTION: (\d+)', line)
						value = match.group(1) if match else None
						INS.append(value)
					if 'DELETION: ' in line:
						match = re.search(r'DELETION: (\d+)', line)
						value = match.group(1) if match else None
						DEL.append(value)
					if 'INDEL: ' in line:
						match = re.search(r'INDEL: (\d+)', line)
						value = match.group(1) if match else None
						INDEL.append(value)
					if 'Total variants (after filtering): ' in line:
						match = re.search(r"Total variants \(after filtering\): (\d+)", line)
						value = match.group(1) if match else None
						TMB.append(value)
	samples = os.listdir(samples_path)
	try:
		df = pd.DataFrame.from_dict(dict_info)
	except:
		print('a')
	df = df.set_index(pd.Index(samples))

	if 'TSV' in dict_para['S_mutations_types_table_format'].upper():
		table_path = dict_para['output_path'] + 'mutation_types.tsv'
		df.to_csv(table_path, sep='\t')
	if 'CSV' in dict_para['S_mutations_types_table_format'].upper():
		table_path = dict_para['output_path'] + 'mutation_types.csv'
		df.to_csv(table_path, sep=',')
	if 'XLSX' in dict_para['S_mutations_types_table_format'].upper():
		table_path = dict_para['output_path'] + 'mutation_types.xlsx'
		df.to_excel(table_path)

	workbook = openpyxl.load_workbook(table_path)
	sheet = workbook.active

	for row in sheet.iter_rows():
		for cell in row:
			cell.alignment = openpyxl.styles.Alignment(horizontal='center')
			cell.border = openpyxl.styles.Border(
				left=openpyxl.styles.Side(style='thin'),
				right=openpyxl.styles.Side(style='thin'),
				top=openpyxl.styles.Side(style='thin'),
				bottom=openpyxl.styles.Side(style='thin')
			)
			if cell.column == 1 or cell.row == 1:
				if cell.column == 1 and cell.row == 1:
					cell.fill = openpyxl.styles.PatternFill(fill_type="none")
				else:
					cell.fill = openpyxl.styles.PatternFill(start_color="FFFFCC", end_color="FFFFCC",
															fill_type="solid")

	for column_cells in sheet.columns:
		max_length = 0
		column = column_cells[0].column_letter
		for cell in column_cells:
			try:
				if len(str(cell.value)) > max_length:
					max_length = len(cell.value)
			except:
				pass
		adjusted_width = (max_length + 2) * 1.2
		sheet.column_dimensions[column].width = adjusted_width

	workbook.save(table_path)

	s_SNV, ns_SNV, f_sub, nf_sub, stopgain, stoploss, startloss, f_ins, nf_ins, f_del, nf_del, unknown = [], [], [], [], [], [], [], [], [], [], [], []
	dic_subtypes = {'synonymous SNV': s_SNV, 'non synonymous SNV': ns_SNV, 'frameshift substitution': f_sub,
					'non frameshift substitution': nf_sub, 'stopgain': stopgain, 'stoploss': stoploss,
					'startloss': startloss, 'frameshift insertion': f_ins,
					'non frameshift insertion': nf_ins, 'frameshift deletion': f_del,
					'non frameshift deletion': nf_del, 'unknown': unknown}
	for folder in os.listdir(samples_path):
		for file in os.listdir(samples_path + folder + '/'):
			if file == 'passed_stats.txt':
				with (open(samples_path + folder + '/' + file, 'r') as file):
					for line in file:
						line = line.strip().replace('\t', ' ')
						if 'impacted' in line:
							break
						elif not line.startswith("#") and not line.startswith("---"):
							if 'synonymous SNV' in line:
								match = re.search(r'synonymous SNV: (\d+)', line)
								value = match.group(1) if match else None
								s_SNV.append(value)
							if 'non synonymous SNV' in line:
								match = re.search(r'non synonymous SNV: (\d+)', line)
								value = match.group(1) if match else None
								ns_SNV.append(value)
							if 'frameshift substitution' in line:
								match = re.search(r'frameshift substitution: (\d+)', line)
								value = match.group(1) if match else None
								f_sub.append(value)
							if 'non frameshift substitution' in line:
								match = re.search(r'non frameshift substitution: (\d+)', line)
								value = match.group(1) if match else None
								nf_sub.append(value)
							if 'stopgain' in line:
								match = re.search(r'stopgain: (\d+)', line)
								value = match.group(1) if match else None
								stopgain.append(value)
							if 'stoploss' in line:
								match = re.search(r'stoploss: (\d+)', line)
								value = match.group(1) if match else None
								stoploss.append(value)
							if 'startloss' in line:
								match = re.search(r'startloss: (\d+)', line)
								value = match.group(1) if match else None
								startloss.append(value)
							if 'frameshift insertion' in line:
								match = re.search(r'frameshift insertion: (\d+)', line)
								value = match.group(1) if match else None
								f_ins.append(value)
							if 'non frameshift insertion' in line:
								match = re.search(r'non frameshift insertion: (\d+)', line)
								value = match.group(1) if match else None
								nf_ins.append(value)
							if 'frameshift deletion' in line:
								match = re.search(r'frameshift deletion: (\d+)', line)
								value = match.group(1) if match else None
								f_del.append(value)
							if 'non frameshift deletion' in line:
								match = re.search(r'non frameshift deletion: (\d+)', line)
								value = match.group(1) if match else None
								nf_del.append(value)
							if 'unknown' in line:
								match = re.search(r'unknown: (\d+)', line)
								value = match.group(1) if match else None
								unknown.append(value)
	samples = os.listdir(samples_path)
	try:
		df = pd.DataFrame.from_dict(dic_subtypes)
	except:
		print('a')
	try:
		df = df.set_index(pd.Index(samples))
	except:
		print('a')

	if 'TSV' in dict_para['S_mutations_subtypes_table_format'].upper():
		table_path = dict_para['output_path'] + 'mutation_subtypes.tsv'
		df.to_csv(table_path, sep='\t')
	if 'CSV' in dict_para['S_mutations_subtypes_table_format'].upper():
		table_path = dict_para['output_path'] + 'mutation_subtypes.csv'
		df.to_csv(table_path, sep=',')
	if 'XLSX' in dict_para['S_mutations_subtypes_table_format'].upper():
		table_path = dict_para['output_path'] + 'mutation_subtypes.xlsx'
		df.to_excel(table_path)


def write_impacted_genes(dict_para, out_genes: str, genes_list: pd.DataFrame, logger):
	'''
	Writes an xlsx file containing the list of genes impacted by the variants from the VCF file
	Input : name use for the output files (.xlsx), the genes list and the logger
	Output : Write the genes list in two different format
	'''

	genes_list = genes_list.sort_values(by='n variants', ascending=False)
	if dict_para['verbose_prints'].upper() == 'TRUE':
		print(f'Write genes list file... {format})')
	logger.info(f'Write genes list file... {format})')
	for path in out_genes:
		if path.endswith('.xlsx'):
			genes_list = genes_list.applymap(str)
			pd.set_option('display.max_columns', None)

			gray_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
			center_alignment = Alignment(horizontal='center', vertical='center')
			thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

			with pd.ExcelWriter(Path(path).with_suffix('.xlsx'), engine='openpyxl') as writer:
				genes_list.to_excel(writer, index=True, sheet_name='Sheet1')
				workbook = writer.book
				worksheet = writer.sheets['Sheet1']

				for cell in worksheet[1]:
					cell.fill = gray_fill

				for row in worksheet.iter_rows():
					for cell in row:
						cell.alignment = center_alignment

						if cell.value and cell.value != 'nan':
							cell.border = thin_border

				for column in worksheet.columns:
					max_length = 0
					for cell in column:
						try:
							if len(str(cell.value)) > max_length:
								max_length = len(cell.value)
						except:
							pass
					adjusted_width = (max_length + 2) * 1.2
					worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

		elif path.endswith('.tsv'):
			genes_list.to_csv(Path(path).with_suffix('.tsv'), sep='\t')
		elif path.endswith('.csv'):
			genes_list.to_csv(Path(path).with_suffix('.csv'), sep=',')


def create_mutations_types_barplot(dict_para):
	dict_info = {'SNP': 0, 'DNP': 0, 'TNP': 0, 'ONP': 0, 'INDEL': 0, 'INSERTION': 0, 'DELETION': 0}
	sample_path = dict_para['output_path_sample'].replace(".vcf","") + '/passed_stats.txt'
	with (open(sample_path, 'r') as file):
		for line in file:
			line = line.strip().replace('\t', ' ')
			if 'impacted' in line:
				break
			elif not line.startswith("#") and not line.startswith("---"):
				if 'SNP: ' in line:
					match = re.search(r'SNP: (\d+)', line)
					value = match.group(1) if match else None
					dict_info['SNP'] = value
				if 'DNP: ' in line:
					match = re.search(r'DNP: (\d+)', line)
					value = match.group(1) if match else None
					dict_info['DNP'] = value
				if 'TNP: ' in line:
					match = re.search(r'TNP: (\d+)', line)
					value = match.group(1) if match else None
					dict_info['TNP'] = value
				if 'ONP: ' in line:
					match = re.search(r'ONP: (\d+)', line)
					value = match.group(1) if match else None
					dict_info['ONP'] = value
				if 'INSERTION: ' in line:
					match = re.search(r'INSERTION: (\d+)', line)
					value = match.group(1) if match else None
					dict_info['INSERTION'] = value
				if 'DELETION: ' in line:
					match = re.search(r'DELETION: (\d+)', line)
					value = match.group(1) if match else None
					dict_info['DELETION'] = value
				if 'INDEL: ' in line:
					match = re.search(r'INDEL: (\d+)', line)
					value = match.group(1) if match else None
					dict_info['INDEL'] = value

	categories = list(dict_info.keys())
	values = [0 if value is None else int(value) for value in dict_info.values()]
	categories_filtered = [cat for cat, val in zip(categories, values) if val != 0]
	colors = {'SNP': '#77c3ec',
			  'DNP': '#89cff0',
			  'TNP': '#9dd9f3',
			  'ONP': '#b8e2f2',
			  'INDEL': '#95b89b',
			  'INSERTION': '#aec9aa',
			  'DELETION': '#bed8c0'}

	values_filtered = [val for val in values if val != 0]

	if len(values_filtered) > 1:
		plt.figure(figsize=(12, 6))  # Add this line to set the figure size
		plt.bar(categories_filtered, values_filtered, color=[colors[cat] for cat in categories_filtered])
	else:
		plt.figure(figsize=(12, 6))  # Add this line to set the figure size
		plt.bar(categories_filtered, values_filtered, color=[colors[cat] for cat in categories_filtered])

	sample_name = sample_path.split('/passed_sta')[0].split('samples/')[1]
	plt.title("Mutation counts for each mutation type found in " + sample_name, pad=10)
	plt.ylabel("Count", labelpad=10)

	for i, val in enumerate(values_filtered):
		plt.text(i, val, str(val), ha='center', va='bottom', fontsize=10.5)

	plt.tight_layout()

	if len(categories_filtered) > 1:
		# Save the plot
		if 'PNG' in dict_para['S_types_barplot_format(s)'].upper():
			plt.savefig(sample_path.split('stats')[0] + 'mutation_types.png', dpi=400)
		if 'SVG' in dict_para['S_types_barplot_format(s)'].upper():
			plt.savefig(sample_path.split('stats')[0] + 'mutation_types.svg', dpi=400)
		if 'PDF' in dict_para['S_types_barplot_format(s)'].upper():
			plt.savefig(sample_path.split('stats')[0] + 'mutation_types.pdf', dpi=400)
		if 'JPG' in dict_para['S_types_barplot_format(s)'].upper():
			plt.savefig(sample_path.split('stats')[0] + 'mutation_types.jpg', dpi=400)

	plt.close()


def create_mutations_subtypes_barplot(dict_para):
	dic_subtypes = {'synonymous SNV': 0, 'non synonymous SNV': 0, 'frameshift substitution': 0,
					'non frameshift substitution': 0, 'stopgain': 0, 'stoploss': 0,
					'startloss': 0, 'frameshift insertion': 0,
					'non frameshift insertion': 0, 'frameshift deletion': 0,
					'non frameshift deletion': 0, 'unknown': 0}
	sample_path = dict_para['output_path_sample'].replace(".vcf","") + '/passed_stats.txt'
	with open(sample_path, 'r') as file:
		for line in file:
			line = line.strip().replace('\t', ' ')
			if 'impacted' in line:
				break
			elif not line.startswith("#") and not line.startswith("---"):
				if 'synonymous SNV' in line:
					match = re.search(r'synonymous SNV: (\d+)', line)
					value = match.group(1) if match else None
					dic_subtypes['synonymous SNV'] = value
				if 'non synonymous SNV' in line:
					match = re.search(r'non synonymous SNV: (\d+)', line)
					value = match.group(1) if match else None
					dic_subtypes['non synonymous SNV'] = value
				if 'frameshift substitution' in line:
					match = re.search(r'frameshift substitution: (\d+)', line)
					value = match.group(1) if match else None
					dic_subtypes['frameshift substitution'] = value
				if 'non frameshift substitution' in line:
					match = re.search(r'non frameshift substitution: (\d+)', line)
					value = match.group(1) if match else None
					dic_subtypes['non frameshift substitution'] = value
				if 'stopgain' in line:
					match = re.search(r'stopgain: (\d+)', line)
					value = match.group(1) if match else None
					dic_subtypes['stopgain'] = value
				if 'stoploss' in line:
					match = re.search(r'stoploss: (\d+)', line)
					value = match.group(1) if match else None
					dic_subtypes['stoploss'] = value
				if 'startloss' in line:
					match = re.search(r'startloss: (\d+)', line)
					value = match.group(1) if match else None
					dic_subtypes['startloss'] = value
				if 'frameshift insertion' in line:
					match = re.search(r'frameshift insertion: (\d+)', line)
					value = match.group(1) if match else None
					dic_subtypes['frameshift insertion'] = value
				if 'non frameshift insertion' in line:
					match = re.search(r'non frameshift insertion: (\d+)', line)
					value = match.group(1) if match else None
					dic_subtypes['non frameshift insertion'] = value
				if 'frameshift deletion' in line:
					match = re.search(r'frameshift deletion: (\d+)', line)
					value = match.group(1) if match else None
					dic_subtypes['frameshift deletion'] = value
				if 'non frameshift deletion' in line:
					match = re.search(r'non frameshift deletion: (\d+)', line)
					value = match.group(1) if match else None
					dic_subtypes['non frameshift deletion'] = value
				if 'unknown' in line:
					match = re.search(r'unknown: (\d+)', line)
					value = match.group(1) if match else None
					dic_subtypes['unknown'] = value

	categories = list(dic_subtypes.keys())
	values = [0 if value is None else int(value) for value in dic_subtypes.values()]
	categories_filtered = [cat for cat, val in zip(categories, values) if val != 0]
	colors = {'synonymous SNV': '#77c3ec',  # Blue
			  'non synonymous SNV': '#89cff0',  # Light Blue
			  'frameshift substitution': '#aec9aa',  # Green
			  'non frameshift substitution': '#95b89b',  # Light Green
			  'stopgain': '#e68a00',  # Dark Orange
			  'stoploss': '#ff9933',  # Light Orange
			  'startloss': '#ffcc66',  # Lighter Orange
			  'frameshift insertion': '#ff6666',  # Red
			  'non frameshift insertion': '#ff9999',  # Light Red
			  'frameshift deletion': '#b366ff',  # Purple
			  'non frameshift deletion': '#cc99ff',  # Light Purple
			  'unknown': '#4d4d4d'}

	values_filtered = [val for val in values if val != 0]
	plt.clf()
	plt.figure(figsize=(12, 6))  # Adjust the figure size as needed
	plt.bar(categories_filtered, values_filtered, color=[colors[cat] for cat in categories_filtered], label='Mutation Subtypes', width=0.3)
	sample_name = sample_path.split('/passed_sta')[0].split('samples/')[1]
	plt.title("Mutation counts for each mutation subtype found in " + sample_name, pad=10)
	plt.ylabel("Count", labelpad=10)

	for i, val in enumerate(values_filtered):
		plt.text(i, val, str(val), ha='center', va='bottom', fontsize=9)

	if len(values_filtered) > 7:
		plt.xticks(rotation=90)

	with warnings.catch_warnings():
		warnings.simplefilter("ignore")
		plt.tight_layout()

	if len(categories_filtered) > 1:
		if 'PNG' in dict_para['S_subtypes_barplot_format(s)'].upper():
			plt.savefig(sample_path.split('stats')[0] + 'mutation_subtypes.png', dpi=400)
		if 'SVG' in dict_para['S_subtypes_barplot_format(s)'].upper():
			plt.savefig(sample_path.split('stats')[0] + 'mutation_subtypes.svg', dpi=400)
		if 'PDF' in dict_para['S_subtypes_barplot_format(s)'].upper():
			plt.savefig(sample_path.split('stats')[0] + 'mutation_subtypes.pdf', dpi=400)
		if 'JPG' in dict_para['S_subtypes_barplot_format(s)'].upper():
			plt.savefig(sample_path.split('stats')[0] + 'mutation_subtypes.jpg', dpi=400)
	plt.close()


def create_protein_impacts_plots(dict_para, dict_impacts):
	config_path = dict_para['config_path']
	with open(config_path, "r") as file:
		lines = file.readlines()
		for line in lines:
			if "max_SIFT_score" in line:
				sift_score_threshold = float(line.split("max_SIFT_score =")[1].strip())
			elif "min_PolyPhen2_score" in line:
				polyphen_threshold = float(line.split("min_PolyPhen2_score =")[1].strip())
	sift_thresholds = [-1, 0.05, 0.2, 1.5]
	polyphen_thresholds = [-1, 0.45, 0.95, 1.5]

	dict_impacts_counts = {
		'SIFT scores counts': {f'{sift_thresholds[i]}-{sift_thresholds[i + 1]}': 0 for i in range(len(sift_thresholds) - 1)},
		'Polyphen2 scores counts': {f'{polyphen_thresholds[i]}-{polyphen_thresholds[i + 1]}': 0 for i in range(len(polyphen_thresholds) - 1)},
		'SIFT predictions': {},
		'Polyphen2 predictions': {}
	}

	if not all(not bool(d) for d in dict_impacts.values()):
		for impact in dict_impacts.values():
			sift_score = impact['SIFT_score']
			polyphen_score = impact['Polyphen_score']
			sift_pred = impact['SIFT_pred']
			polyphen_pred = impact['Polyphen_pred']

			if sift_score == 'not found':
				if sift_score not in dict_impacts_counts['SIFT scores counts']:
					dict_impacts_counts['SIFT scores counts'][sift_score] = 1
				else:
					dict_impacts_counts['SIFT scores counts'][sift_score] += 1
			else:
				interval = None
				for i in range(len(sift_thresholds) - 1):
					if sift_thresholds[i] <= sift_score < sift_thresholds[i + 1]:
						interval = f'{sift_thresholds[i]}-{sift_thresholds[i + 1]}'
						break
				dict_impacts_counts['SIFT scores counts'][interval] += 1

			if polyphen_score == 'not found':
				if polyphen_score not in dict_impacts_counts['Polyphen2 scores counts']:
					dict_impacts_counts['Polyphen2 scores counts'][polyphen_score] = 1
				else:
					dict_impacts_counts['Polyphen2 scores counts'][polyphen_score] += 1
			else:
				interval = None
				for i in range(len(polyphen_thresholds) - 1):
					if polyphen_thresholds[i] <= polyphen_score < polyphen_thresholds[i + 1]:
						interval = f'{polyphen_thresholds[i]}-{polyphen_thresholds[i + 1]}'
						break
				dict_impacts_counts['Polyphen2 scores counts'][interval] += 1

			dict_impacts_counts['SIFT predictions'].setdefault(sift_pred, 0)
			dict_impacts_counts['SIFT predictions'][sift_pred] += 1
			dict_impacts_counts['Polyphen2 predictions'].setdefault(polyphen_pred, 0)
			dict_impacts_counts['Polyphen2 predictions'][polyphen_pred] += 1

	dict_impacts_counts['SIFT scores counts'] = {key: value for key, value in dict_impacts_counts['SIFT scores counts'].items() if value != 0}
	dict_impacts_counts['Polyphen2 scores counts'] = {key: value for key, value in dict_impacts_counts['Polyphen2 scores counts'].items() if value != 0}
	dict_impacts_counts['SIFT predictions'] = {key: value for key, value in dict_impacts_counts['SIFT predictions'].items() if value != 0}
	sift_preds_mapping = {'D': 'deleterious', 'T': 'tolerated', 'U': 'unknown'}
	dict_impacts_counts['SIFT predictions'] = {sift_preds_mapping.get(key, key): value for key, value in dict_impacts_counts['SIFT predictions'].items() if value != 0}
	dict_impacts_counts['Polyphen2 predictions'] = {key: value for key, value in dict_impacts_counts['Polyphen2 predictions'].items() if value != 0}
	dict_impacts_counts['Polyphen2 predictions'] = {('probably damaging' if key == 'D' else key): value for key, value in dict_impacts_counts['Polyphen2 predictions'].items() if
													value != 0}
	polyphen_preds_mapping = {'B': 'benign', 'D': 'probably damaging', 'P': 'possibly damaging'}
	dict_impacts_counts['Polyphen2 predictions'] = {polyphen_preds_mapping.get(key, key): value for key, value in dict_impacts_counts['Polyphen2 predictions'].items() if
													value != 0}

	categories = []
	values = []

	for key in dict_impacts_counts.keys():
		sub_dic = dict_impacts_counts[key]
		for k in sub_dic.keys():
			categories.append((key + k))
			values.append(sub_dic[k])

	true_categories = []
	for cat in categories:
		if 'SIFT scores counts' in cat:
			category = 'n SIFT (' + cat.split('SIFT scores counts')[1] + ')'
			true_categories.append(category)
		elif 'Polyphen2 scores counts' in cat:
			category = 'n Polyphen2 (' + cat.split('Polyphen2 scores counts')[1] + ')'
			true_categories.append(category)
		elif 'SIFT predictions' in cat:
			category = 'SIFT (' + cat.split('SIFT predictions')[1] + ')'
			true_categories.append(category)
		elif 'Polyphen2 predictions' in cat:
			category = 'Polyphen2 (' + cat.split('Polyphen2 predictions')[1] + ')'
			true_categories.append(category)

	sift_scores = []
	for key in dict_impacts:
		sift_scores.append(dict_impacts[key]['SIFT_score'])
	n_sifts = len(sift_scores)
	sift_scores = [score for score in sift_scores if score != 'not found']

	fig, ax = plt.subplots()
	ax.boxplot([sift_scores], labels=['SIFT method'], patch_artist=True,boxprops=dict(facecolor='tan'), medianprops=dict(color='black'),whiskerprops=dict(color='black'), notch=True)
	ax.set_ylabel('Scores', labelpad=15, fontsize=12)

	try:
		mean_sift = np.mean(sift_scores)
		ax.text(0.9, mean_sift, f"x̄: {mean_sift:.2f}", ha='right', verticalalignment='center_baseline', fontweight='bold', fontsize=9)
	except:
		print('no SIFT score found')

	legend_labels = []
	legend_counts = []
	for category, sub_dict in dict_impacts_counts.items():
		if 'SIFT' in category:
			for sub_category, count in sub_dict.items():
				legend_labels.append(f"{sub_category}: ")
				legend_counts.append(count)
	# ax.legend(legend_labels, loc='center left', bbox_to_anchor=(0.65, 0.8), edgecolor='black')
	fig.set_size_inches(12, 6)

	sample_path = dict_para['output_path_sample'].split('.vcf')[0] + '/passed_stats.txt'
	sample_name = sample_path.split('/passed_sta')[0].split('samples/')[1]
	plt.title("SIFT predicted impacts on protein, for mutations found in \n" + sample_name, pad=12, fontsize=12)
	ax.annotate(f"number of variants: {len(dict_impacts.keys())}", xy=(0.7, 0.7), xycoords='axes fraction', fontsize=11, fontweight='bold')
	ax.annotate(f"Scores:\n", xy=(0.7, 0.6), xycoords='axes fraction', fontsize=11, fontweight='bold')
	height = 0.55
	pred_print = 0
	first_not_found = False
	for label, count in zip(legend_labels, legend_counts):
		if label == '-1-0.05: ':
			label = '0-0.05: '
		if label == '0.2-1.5: ':
			label = '0.2-1: '
		if 'deleterious' not in label and 'tolerated' not in label and not first_not_found:
			text_color = 'black'
			if label == '0-0.05: ':
				text_color = 'darkgreen'
			elif label == '0.05-0.2: ':
				text_color = 'darkorange'
			elif label == '0.2-1: ':
				text_color = 'darkred'
			elif 'not found' in label:
				text_color = 'black'
				first_not_found = True
			ratio = round(count / n_sifts * 100)
			plt.annotate(f"{label}{ratio}%\n", xy=(0.7, height), xycoords='axes fraction', fontsize=11, bbox=dict(facecolor='white', edgecolor='none'), color=text_color)
		else:
			if pred_print == 0:
				ax.annotate(f"Predictions:\n", xy=(0.7, height - 0.02), xycoords='axes fraction', fontsize=11, fontweight='bold')
				height = height - 0.07
			if 'deleterious' in label:
				text_color = 'darkgreen'
			elif 'tolerated' in label:
				text_color = 'darkorange'
			elif 'not found' in label:
				text_color = 'black'
			else:
				text_color = 'black'
			ratio = round(count / n_sifts * 100)
			ax.annotate(f"{label}{ratio}%\n", xy=(0.7, height), xycoords='axes fraction', fontsize=11, color=text_color)
			pred_print += 1
		height = height - 0.05

	plt.axhline(y=sift_score_threshold, color='darkred', linestyle='--')
	ax = plt.gca()
	custom_padding = max(sift_scores) - min(sift_scores)
	ax.annotate(str(sift_score_threshold), xy=(0.008, sift_score_threshold + custom_padding), xycoords=('axes fraction', 'data'), fontsize=9, color='darkred')

	if 'PNG' in dict_para['S_sift_protein_impacts_boxplot_format(s)'].upper():
		plt.savefig(sample_path.split('stats')[0] + 'SIFT_protein_impact_scores.png', dpi=400)
	if 'SVG' in dict_para['S_sift_protein_impacts_boxplot_format(s)'].upper():
		plt.savefig(sample_path.split('stats')[0] + 'SIFT_protein_impact_scores.svg', dpi=400)
	if 'PDF' in dict_para['S_sift_protein_impacts_boxplot_format(s)'].upper():
		plt.savefig(sample_path.split('stats')[0] + 'SIFT_protein_impact_scores.pdf', dpi=400)
	if 'JPG' in dict_para['S_sift_protein_impacts_boxplot_format(s)'].upper():
		plt.savefig(sample_path.split('stats')[0] + 'SIFT_protein_impact_scores.jpg', dpi=400)
	plt.close()

	polyphen_scores = [data['Polyphen_score'] for data in dict_impacts.values()]
	n_polyphens = len(polyphen_scores)
	polyphen_scores = [score for score in polyphen_scores if score != 'not found']

	fig, ax = plt.subplots()
	ax.boxplot([polyphen_scores], labels=['Polyphen2 method'], patch_artist=True,boxprops=dict(facecolor='tan'), medianprops=dict(color='black'), whiskerprops=dict(color='black'), notch=True)
	ax.set_ylabel('Scores', labelpad=10)

	try:
		mean_polyphen = np.mean(polyphen_scores)
		ax.text(0.9, mean_polyphen, f"x̄: {mean_polyphen:.2f}", ha='right', verticalalignment='center_baseline', fontweight='bold', fontsize=9)
	except:
		print('no SIFT score found')

	legend_labels = []
	legend_counts = []
	for category, sub_dict in dict_impacts_counts.items():
		if 'Poly' in category:
			for sub_category, count in sub_dict.items():
				legend_labels.append(f"{sub_category}: ")
				legend_counts.append(count)
	# ax.legend(legend_labels, loc='center left', bbox_to_anchor=(0.65, 0.8), edgecolor='black')
	fig.set_size_inches(12, 6)

	sample_path = dict_para['output_path_sample'].split('.vcf')[0] + '/passed_stats.txt'
	sample_name = sample_path.split('/passed_sta')[0].split('samples/')[1]
	plt.title("Polyphen2 predicted impacts on protein, for mutations found in \n" + sample_name, pad=12, fontsize=12)
	ax.annotate(f"number of variants: {len(dict_impacts.keys())}", xy=(0.65, 0.7), xycoords='axes fraction', fontsize=11, fontweight='bold')
	ax.annotate(f"Scores:\n", xy=(0.65, 0.6), xycoords='axes fraction', fontsize=11, fontweight='bold')
	height = 0.55

	pred_print = 0
	first_not_found = False
	for label, count in zip(legend_labels, legend_counts):
		if label == '-1-0.45: ':
			label = ('0-0.45: ')
		if label == '0.95-1.5: ':
			label = '0.95-1: '
		if 'damaging' not in label and 'benign' not in label and not first_not_found:
			text_color = 'black'
			if label == '0-0.45: ':
				text_color = 'darkred'
			elif label == '0.45-0.95: ':
				text_color = 'darkorange'
			elif label == '0.95-1: ':
				text_color = 'darkgreen'
			elif 'not found' in label:
				text_color = 'black'
				first_not_found = True
			ratio = round(count / n_polyphens * 100)
			if ratio > 0:
				plt.annotate(f"{label}{ratio}%\n", xy=(0.65, height), xycoords='axes fraction', fontsize=11, bbox=dict(facecolor='white', edgecolor='none'), color=text_color)
		else:
			if pred_print == 0:
				ax.annotate(f"Predictions:\n", xy=(0.65, height - 0.02), xycoords='axes fraction', fontsize=11, fontweight='bold')
				height = height - 0.07
			if 'probably' in label.lower():
				text_color = 'darkgreen'
			elif 'possibly' in label.lower():
				text_color = 'darkorange'
			elif 'benign' in label.lower():
				text_color = 'darkred'
			elif 'not found' in label:
				text_color = 'black'
			else:
				text_color = 'black'
			ratio = round(count / n_polyphens * 100)
			if ratio > 0:
				plt.annotate(f"{label}{ratio}%\n", xy=(0.65, height), xycoords='axes fraction', fontsize=11, bbox=dict(facecolor='white', edgecolor='none'), color=text_color)
			pred_print += 1
		height = height - 0.05

	plt.axhline(y=polyphen_threshold, color='darkred', linestyle='--')
	ax = plt.gca()
	custom_padding = max(polyphen_scores) - min(polyphen_scores)
	ax.annotate(str(polyphen_threshold), xy=(0.008, polyphen_threshold + custom_padding), xycoords=('axes fraction', 'data'), fontsize=9, color='darkred')

	if 'PNG' in dict_para['S_polyphen_protein_impacts_boxplot_format(s)'].upper():
		plt.savefig(sample_path.split('stats')[0] + 'Polyphen2_protein_impact_scores.png', dpi=400)
	if 'SVG' in dict_para['S_polyphen_protein_impacts_boxplot_format(s)'].upper():
		plt.savefig(sample_path.split('stats')[0] + 'Polyphen2_protein_impact_scores.svg', dpi=400)
	if 'PDF' in dict_para['S_polyphen_protein_impacts_boxplot_format(s)'].upper():
		plt.savefig(sample_path.split('stats')[0] + 'Polyphen2_protein_impact_scores.pdf', dpi=400)
	if 'JPG' in dict_para['S_polyphen_protein_impacts_boxplot_format(s)'].upper():
		plt.savefig(sample_path.split('stats')[0] + 'Polyphen2_protein_impact_scores.jpg', dpi=400)
	plt.close()


# sift_color = '#8B0000'
	# polyphen2_color = '#00008B'
	#
	# fig, ax = plt.subplots()
	# bars = ax.bar(true_categories, values)
	#
	# for i, category in enumerate(true_categories):
	# 	if 'SIFT' in category:
	# 		bars[i].set_color(sift_color)
	# 	elif 'Polyphen2' in category:
	# 		bars[i].set_color(polyphen2_color)
	#
	# for bar in bars:
	# 	height = bar.get_height()
	# 	ax.text(bar.get_x() + bar.get_width() / 2, height, str(int(height)), ha='center', va='bottom')
	#
	# ax.set_ylim(0, max(values) * 1.1)  # Adjust the multiplier (1.2) to control the range
	# ax.set_ylabel('Count', labelpad=15)
	# ax.set_title('Mutation Analysis')
	#
	# plt.xticks(rotation=45, ha='right', fontsize=10)
	# plt.subplots_adjust(bottom=0.4)
	# fig.set_size_inches(10, 6)
	#
	# sample_path = dict_para['output_path_sample'].split('.vcf')[0] + '/passed_stats.txt'
	# sample_name = sample_path.split('/passed_sta')[0].split('samples/')[1]
	# plt.title("Protein impact for each mutation found in \n" + sample_name, pad=12, fontsize=12)
	#
	# if 'PNG' in dict_para['S_protein_impacts_format(s)'].upper():
	# 	plt.savefig(sample_path.split('stats')[0] + 'protein_impacts.png', dpi=400)
	# if 'SVG' in dict_para['S_protein_impacts_format(s)'].upper():
	# 	plt.savefig(sample_path.split('stats')[0] + 'protein_impacts.svg', dpi=400)
	# if 'PDF' in dict_para['S_protein_impacts_format(s)'].upper():
	# 	plt.savefig(sample_path.split('stats')[0] + 'protein_impacts.pdf', dpi=400)
	# if 'JPG' in dict_para['S_protein_impacts_format(s)'].upper():
	# 	plt.savefig(sample_path.split('stats')[0] + 'protein_impacts.jpg', dpi=400)
	#
	# plt.close()

def summary(last, dict_para, dict_colors, output_path, vcf_file_filter: str, vcf_file_pass: str, genome_file: str,
			out_stats: str, out_genes: str, SNP_profile: str, out_indel: str, logger, enrichment: bool):
	'''
	Summarizes the vcf file
	Input : path to the vcf file from filter module (optional), path to the vcf file containing only pass variants from filter module, path to genome (fasta or pickle),
	outputs path, logger and boolean for writing gene list enrichments analysis
	Output : stat summary file, file containing the impacted genes, mutation profile graph, indel size graph
	'''
	if dict_para['verbose_prints'].upper() == 'TRUE':
		print('Extracting information from reference genome and vcf file(s)...')
	genome = get_genome_dict(dict_para, genome_file, logger)

	####################
	# Get stats and data
	counter_deletion_size, counter_insertion_size, snp_count_pct, snp_count, genes_list, stats, dict_impacts, dict_multiple_alts = count_variants(
		dict_para, dict_colors, vcf_file_filter, vcf_file_pass, genome, logger)

	print('Creating stats file, mutations table(s) and plots...')
	##################
	# Write stats file
	write_stats(genes_list, dict_para, vcf_file_filter, vcf_file_pass, out_stats, stats, dict_impacts, logger)

	create_mutations_types_barplot(dict_para)
	create_mutations_subtypes_barplot(dict_para)

	##########################################
	# Write genes tables
	complete_df, sorted_df = create_dataframe_from_gene(dict_para, genes_list)
	write_impacted_genes(dict_para, out_genes, sorted_df, logger)

	create_variants_table(dict_para, complete_df, dict_impacts, dict_multiple_alts)

	################################################
	# Biological process enrichment using the genes list with the ToppGene and Panther API
	if 'YES' in enrichment or 'TRUE' in enrichment or 'PANTHER' in enrichment or 'TOPPGENE' in enrichment:
		enrichment = True
	if enrichment:
		if len(genes_list.keys()) == 1:
			concern = f'gene is concerned ({next(iter(genes_list))})'
		else:
			concern = f'genes are concerned'
		print(f"\033[1m{len(genes_list.keys())}\033[0m{dict_colors['yellow3']} {concern}.")
		if len(genes_list.keys()) < 1500:
			print('Computing ToppGene GOEA analysis...')
			toppgene_name = 'GO_ToppGene'
			ToppGene_GOEA('summarise', dict_para, genes_list, toppgene_name, logger)
		else:
			print("The VCF is heavy, too many genes are concerned for ToppGene GO to be run.")
		if len(genes_list.keys()) < 2000:
			print('Computing Panther GOEA analysis...')
			panther_name = 'GO_Panther'
			Panther_GOEA('summarise', dict_para, genes_list, panther_name, logger)
		else:
			print("The VCF is heavy, too many genes are concerned for Panther GO to be run.")

	#################
	# Create snp mutation types plot and indel size plot
	create_snp_plot(dict_para, snp_count_pct, snp_count, SNP_profile, vcf_file_pass, logger, output_path)

	if str(counter_insertion_size) != 'Counter()':

		create_indel_plot(dict_para, counter_deletion_size, counter_insertion_size, out_indel, vcf_file_pass, logger)
	else:
		if dict_para['verbose_prints'].upper() == 'TRUE':
			print('No insertion in the vcf file !')
		else:
			pass

	if last:
		create_mutation_types_tables(dict_para)

	if not all(not bool(d) for d in dict_impacts.values()):
		create_protein_impacts_plots(dict_para, dict_impacts)

def main(dict_para, dict_colors, args, vcf_filtered, vcf_passed, output_path, last):
	vcf_file_filter = vcf_filtered
	vcf_file_pass = vcf_passed
	working_directory = output_path
	genome_file = args['reference_genome']
	if not "." in genome_file:
		genome_file = genome_file + '.pk'
	out_stats = output_path + 'passed_stats'
	format = dict_para['S_mutated_genes_table_format']
	format = "".join(format.split())
	l_formats = format.split(',') if ',' in format else [format]
	table_name = 'mutated_genes'
	if "." in table_name:
		table_name = re.sub(r"\..*", "", table_name)
	out_genes = [output_path + table_name + "." + format for format in l_formats]
	SNP_profile = output_path + 'SNP_profile'
	out_indel = output_path + 'indel_profile'
	enrichment = args['S_enrichment']

	# Logger configuration
	logger = logging.getLogger('LOTUS summarise')
	logger.setLevel(logging.DEBUG)
	fh = logging.FileHandler(args['log'])
	fh.setLevel(logging.DEBUG)
	formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
	fh.setFormatter(formatter)
	logger.addHandler(fh)

	try:
		logger.info('Verification of outputs file')
		verif_output(out_stats)
		verif_output(out_indel)
		logger.info('- Outputs file ok -')
	except ValueError:
		if dict_para['verbose_prints'].upper() == 'TRUE':
			print('Problem with one or more output files: ', sys.exc_info()[0])
		logger.error('- Problem with output files -')
		exit(1)

	# Start

	logger.info(
		'**************************************************************************************************************')
	logger.info('*** LOTUS summarise module ***')
	no_argument = ''
	if enrichment.upper() == 'TRUE':
		no_argument += ' --enrichment'
	logger.info(
		f'** cmd line : python lotus.py summarise -v {str(vcf_file_filter)} -vp {str(vcf_file_pass)} -g {str(genome_file)} -s {str(out_stats)} -genes {str(out_genes)} -p {str(SNP_profile)} -i {str(out_indel)}' + str(
			no_argument) + ' **')

	logger.info('* Start summarizing *')
	logger.info(f'Working directory (vcf files folder) : {working_directory}')
	logger.info(f'Current directory : {Path().absolute()}')

	summary(last, args, dict_colors, output_path, vcf_file_filter, vcf_file_pass, genome_file, out_stats, out_genes,
			SNP_profile, out_indel, logger, enrichment)

	logger.info('* End summarizing *')
	logger.info(f'File created :\t{out_stats}\t{SNP_profile}')
	logger.info(
		'**************************************************************************************************************')

# End
