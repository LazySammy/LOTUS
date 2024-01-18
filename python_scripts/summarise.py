#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import openpyxl
import numpy as np
import os
import pyfastx
import pickle as pk
from pathlib import Path
import re
import sys
import logging
from tqdm import tqdm
from itertools import dropwhile
from collections import OrderedDict, Counter
from copy import deepcopy
import warnings
import pandas as pd
from python_scripts.check_files import verif_input_vcf, verif_output
from python_scripts.toppgene_api import ToppGene_GOEA
from python_scripts.panther_api import Panther_GOEA
from python_scripts.path_modification import true_stem
from python_scripts.read_vcf import read_vcf
import matplotlib.pyplot as plt

TRANSLATE = {'A>C': 'T>G', 'A>G': 'T>C', 'A>T': 'T>A', 'C>A': 'C>A', 'C>G': 'C>G', 'C>T': 'C>T',
             'G>A': 'C>T', 'G>C': 'C>G', 'G>T': 'C>A', 'T>A': 'T>A', 'T>C': 'T>C', 'T>G': 'T>G'}
TAB = str.maketrans("ACGT", "TGCA")


def create_snp_dict():
    '''
    Create the dictionnary to count the SNP variant and create the vcf SNP profile. Only the C>* and T>* value are used because the variant sens isn't known.
    Input : None:
    Output : dictionnary with empty counts
    '''
    c_words = {}
    t_words = {}
    for i in 'ATCG':
        word = i
        for j in 'CT':
            word2 = word + j
            for k in 'ATCG':
                word3 = word2 + k
                if word3[1] == 'C':
                    c_words[word3] = 0
                else:
                    t_words[word3] = 0
    save = {}
    for snp in set(TRANSLATE.values()):
        if snp[0] == 'C':
            save[snp] = deepcopy(c_words)
        else:
            save[snp] = deepcopy(t_words)
    return save


def create_dataframe_from_gene(dict_para, d):
    '''
    Take a dictionary containing the genes list with information for each genes and transform it in pandas dataframe
    Input : dictionary containing genes (d)
    Output : dataframe corresponding to the dictionary
    '''

    col = ['Number of variants', 'Details (snp, dnp, tnp, onp, ins, del)', 'Ref', 'Alt variant(s)', 'Chromosome',
           'Position(s)', 'DP', 'QUALITY', '']  # Columns name

    for i in d:
        del d[i][8]  # ignore subtypes for the moment

    d = OrderedDict(sorted(d.items()))
    df = pd.DataFrame.from_dict(d, orient='index', columns=col)
    id = 'Gene name'
    df.index.name = id

    # Modification of the dataframe using a copy of it

    df2 = pd.DataFrame()
    df2['Details (snp, dnp, tnp, onp, ins, del)'] = [','.join(map(str, l)) for l in
                                                     df['Details (snp, dnp, tnp, onp, ins, del)']]
    df['Details (snp, dnp, tnp, onp, ins, del)'] = df2['Details (snp, dnp, tnp, onp, ins, del)'].values
    del df2
    df2 = pd.DataFrame()
    df2['Ref'] = [','.join(map(str, l)) for l in df['Ref']]
    df['Ref'] = df2['Ref'].values
    del df2
    df2 = pd.DataFrame()
    df2['Alt variant(s)'] = [','.join([str(l2) if len(l2) > 1 else str(l2[0]) for l2 in l]) for l in
                             df['Alt variant(s)']]
    df['Alt variant(s)'] = df2['Alt variant(s)'].values
    del df2
    df2 = pd.DataFrame()
    df2['DP'] = [','.join(map(str, l)) for l in df['DP']]
    df['DP'] = df2['DP'].values
    del df2
    df2 = pd.DataFrame()
    if dict_para['vcf_annotation_method'].upper() == 'ANNOVAR':
        df2['QUALITY'] = [','.join(map(str, l)) for l in df['QUALITY']]
        df['QUALITY'] = df2['QUALITY'].values
        del df2
        df2 = pd.DataFrame()
    # df2['AF'] = [','.join(map(str, l)) for l in df['AF']]
    # df['AF'] = df2['AF'].values
    del df2
    df2 = pd.DataFrame()
    df2['Position(s)'] = [','.join(map(str, l)) for l in df['Position(s)']]
    df['Position(s)'] = df2['Position(s)'].values
    del df2
    sorted_df = df.sort_values(by='Number of variants', ascending=False)
    return sorted_df


def create_ordered_dataframe(d):
    '''
    Take a dictionary contaning snp counts, ordered it using it and return the pandas dataframe corresponding
    Input : dictionary contaning snp counts
    Output : dataframe corresponding to the dictionary
    '''
    for k, v in d.items():
        d[k] = OrderedDict(sorted(v.items(), key=lambda x: x[0]))
    d = OrderedDict(sorted(d.items(), key=lambda t: t[0]))
    df = pd.DataFrame.from_dict(d, orient="index").stack().to_frame()
    return df


def graph_snp(dict_para, d, dcount, name, vcf_name, logger, output_path):
    '''
        Creation of the profile graph 
        Input : a dictionnary containing the count for all kind of SNP (d), same dictionnary but with counts instead of percentage, the output file name and the logger
        Output : write a .svg file 
        '''
    if dict_para['verbose_prints'].upper() == 'TRUE':
        print(f'Draw profile in {name}...')
    logger.info(f'Draw profile in {name}')

    df = create_ordered_dataframe(d)

    df.columns = [str(true_stem(vcf_name))]

    # color for graph

    white_96 = ['white'] * 96
    color_6 = ['darkblue', 'blue', 'lightblue', 'darkgreen', 'green', 'lightgreen']
    color_96 = []
    for i in color_6:
        color_96 += [i] * 16
    bars = [index[1] for index, row in df.iterrows()]
    height = [float(ser.iloc[0]) for _, ser in df.iterrows()]
    group = []
    for index, row in df.iterrows():
        if index[0] not in group:
            group.append(index[0])
    x_pos = np.arange(len(bars))

    # Create bars
    plt.figure(figsize=(15, 10))
    ax1 = plt.subplot(1, 1, 1)
    plt.bar(x_pos, height, color=color_96)

    # Create names on the x-axis
    plt.xticks(x_pos, bars, rotation=90, fontsize=8)
    ax1.set_ylabel('Fraction of each mutation-type')

    # Set scond x-axis
    ax2 = ax1.twiny()
    newlabel = group
    newpos = [8, 24, 40, 56, 72, 88]
    ax2.set_xticks(newpos)
    ax2.set_xticklabels(newlabel)
    # set the position of the second x-axis to bottom
    ax2.xaxis.set_ticks_position('bottom')
    ax2.xaxis.set_label_position('bottom')
    ax2.spines['bottom'].set_position(('outward', 36))
    ax2.set_xlabel('Mutation types')
    ax2.set_xlim(ax1.get_xlim())
    plt.tick_params(
        axis='x',  # changes apply to the x-axis
        which='both',  # both major and minor ticks are affected
        bottom=False,  # ticks along the bottom edge are off
        top=False,  # ticks along the top edge are off
        labelbottom=True)  # label along the bottom edge are on
    for xtick, color in zip(ax2.get_xticklabels(), color_6):
        xtick.set_color(color)
    ax2.spines['bottom'].set_visible(False)

    plt.draw()
    ylabels = [
        str(round(float(ytick.get_text()), 2)) if (ytick.get_text()[0] == '0' or ytick.get_text()[0] == '1') else str(
            round(float(ytick.get_text()[1:]), 2)) for ytick in ax1.get_yticklabels()]
    ax1.set_yticks(ax1.get_yticks())
    ax1.set_yticklabels([str(round(float(i) * 100, 1)) for i in ylabels])

    if 'PNG' in dict_para['SNP_profile'].upper():
        png = name + '.png'
        plt.savefig(png)
    if 'SVG' in dict_para['SNP_profile'].upper():
        svg = name + '.svg'
        plt.savefig(svg)
    if 'JPG' in dict_para['SNP_profile'].upper():
        jpg = name + '.jpg'
        plt.savefig(jpg)

    plt.close()

    counts = []
    for i in df.index:
        counts.append(dcount[i[0]][i[1]])
    df["Associated count " + str(true_stem(vcf_name))] = counts

    if "XLSX" in dict_para['SNP_profile'].upper():
        df.to_excel(Path(name).with_suffix(".xlsx"), index=False)
    name = os.path.join(os.path.dirname(name), 'new_filename')
    if "CSV" in dict_para['SNP_profile'].upper():
        df.to_csv(Path(name).with_suffix(".csv"), sep=',')
    new_path = os.path.join(os.path.dirname(name), 'SNP_profile.tsv')
    df.to_csv(Path(new_path), sep='\t')


def graph_indel(dict_para, deletion, insertion, name, vcf_name, logger):
    '''
    Creation of the indel size graph
    Input : two dictionnaries containing the count for insertion/deletion of different size, the output file name and the logger
    Output : write a .svg file
    '''

    insert = True
    delet = True
    if deletion == Counter() and insertion == Counter():  # If no Indel
        print('Warning ! No indel in vcf !')
        logger.warning(f'No indel in vcf !')
        return None
    elif insertion == Counter():  # If no insertion
        print('Warning ! No insertion in vcf !')
        logger.warning(f'No insertion in vcf !')
        insert = False
    elif deletion == Counter():  # If no deletion
        print('Warning ! No deletion in vcf !')
        logger.warning(f'No deletion in vcf !')
        delet = False
    if dict_para['verbose_prints'].upper() == 'TRUE':
        print(f'Draw indel size barplot in {name}...')
    logger.info(f'Draw indel size barplot in {name}')

    width = 0.25

    sum_del = 0
    sum_ins = 0
    if delet:
        sum_del = sum(deletion.values())
        df_del = pd.DataFrame.from_dict(deletion, orient='index').sort_index()
        height_del = list([float(i) / sum_del for i in df_del[0]])
        r_del = list(df_del.index)
        max_del = max([float(i) for i in r_del])
    if insert:
        sum_ins = sum(insertion.values())
        df_ins = pd.DataFrame.from_dict(insertion, orient='index').sort_index()
        height_ins = list([float(i) / sum_ins for i in df_ins[0]])
        r_ins = list(df_ins.index)
        max_ins = max([float(i) for i in r_ins])

    if delet and insert:
        maximum = max([int(max_del) + 1, int(max_ins) + 1])
    elif delet:
        maximum = int(max_del) + 1
    elif not delet:
        maximum = int(max_ins) + 1
    x = [0] + [i + 1 for i in range(maximum)]

    ax = plt.axes()
    if delet and insert:
        plt.bar([float(i) - (width * 0.65) for i in r_del], height_del, color='k', width=width, edgecolor='k',
                label='Deletion')
        plt.bar([float(i) + (width * 0.65) for i in r_ins], height_ins, color='r', width=width, edgecolor='r',
                label='Insertion')
    elif not insert:
        plt.bar([float(i) for i in r_del], height_del, color='k', width=width, edgecolor='k', label='Deletion')
    elif not delet:
        plt.bar([float(i) for i in r_ins], height_ins, color='r', width=width, edgecolor='r', label='Insertion')

    plt.xticks(x, fontsize=9)
    plt.yticks()

    if maximum > 10:
        ##############################################
        # Adaptation of the figure size to the x range

        plt.gca().margins(x=0)
        plt.gcf().canvas.draw()
        tl = plt.gca().get_xticklabels()
        maxsize = max([t.get_window_extent().width for t in tl]) + 1
        m = 0.5  # inch margin
        s = maxsize / plt.gcf().dpi * maximum + 2 * m
        margin = m / plt.gcf().get_size_inches()[0]
        plt.gcf().subplots_adjust(left=margin, right=1. - margin)
        plt.gcf().set_size_inches(s * 1.5, plt.gcf().get_size_inches()[1])

        plot_margin = 1
        x0, x1, y0, y1 = plt.axis()
        plt.axis((x0 - plot_margin, x1 + plot_margin * 2, y0, y1))

    plt.draw()
    ylabels = [
        str(round(float(ytick.get_text()), 2)) if (ytick.get_text()[0] == '0' or ytick.get_text()[0] == '1') else str(
            round(float(ytick.get_text()[1:]), 2)) for ytick in ax.get_yticklabels()]
    ax.set_yticks(ax.get_yticks())
    ax.set_yticklabels([str(round(float(i) * 100, 1)) for i in ylabels])
    plt.xlabel("Indel size (bp)")
    plt.ylabel("Indel percentage")
    plt.legend()

    if "SVG" in dict_para['indel_profile'].upper():
        svg = name + '.svg'
        plt.savefig(svg)
    if "PNG" in dict_para['indel_profile'].upper():
        png = name + '.png'
        plt.savefig(png)
    if "JPG" in dict_para['indel_profile'].upper():
        jpg = name + '.jpg'
        plt.savefig(jpg)

    plt.close()

    if dict_para['verbose_prints'].upper() == 'TRUE':
        print(f'Saving indel counts in {Path(name).with_suffix(".tsv")}...')
    logger.info(f'Save indel counts in {Path(name).with_suffix(".tsv")}')

    if delet and insert:
        df_del.columns = [str(Path(vcf_name).stem)]
        df_ins.columns = [str(Path(vcf_name).stem)]
        if not 'FALSE' in dict_para['indel_profile_deletion'].upper() and not 'FALSE' in dict_para[
            'indel_profile_deletion'].upper():
            tsv = name.rsplit('/', 1)[0] + '/deletion_profile'
            if 'TSV' in dict_para['indel_profile_deletion'].upper():
                df_del.to_csv(Path(tsv).with_suffix(".tsv"), sep='\t')
            if 'CSV' in dict_para['indel_profile_deletion'].upper():
                df_del.to_csv(Path(tsv).with_suffix(".csv"), sep=',')
            if 'XLSX' in dict_para['indel_profile_deletion'].upper():
                df_del.to_excel(Path(tsv).with_suffix(".xlsx"), index=False)
        if not 'FALSE' in dict_para['indel_profile_insertion'].upper() and not 'FALSE' in dict_para[
            'indel_profile_insertion'].upper():
            csv = name.rsplit('/', 1)[0] + '/insertion_profile'
            if 'TSV' in dict_para['indel_profile_insertion'].upper():
                df_ins.to_csv(Path(csv).with_suffix(".tsv"), sep='\t')
            if 'CSV' in dict_para['indel_profile_insertion'].upper():
                df_ins.to_csv(Path(csv).with_suffix(".csv"), sep=',')
            if 'XLSX' in dict_para['indel_profile_insertion'].upper():
                df_ins.to_excel(Path(csv).with_suffix(".xlsx"), index=False)


def is_fasta(filename: str) -> bool:
    '''
    Is the file a fasta file?
    Input : path to the file
    Output : True, False (if the file is a pickle file for example) or raise an error if file doesn't exist
    '''
    try:
        fa = pyfastx.Fastx(filename)
        fasta = [content[0] for content in fa]
        return any(fasta)
    except RuntimeError as runerr:
        return False
    except UnicodeDecodeError as uderr:
        return False
    except FileNotFoundError as fnferr:
        print(f'\nFile {filename} doesn\'t exists: {fnferr}\n')
        raise FileNotFoundError(f'\nFile {filename} doesn\'t exists\n')
    except FileExistsError as feerr:
        print(f'\nFile {filename} doesn\'t exists: {feerr}\n')
        raise FileExistsError(f'\nFile {filename} doesn\'t exists\n')
    except:
        print(f'\nUnexpected error: {sys.exc_info()[0]}\n')
        raise


def is_pickle(filename: str) -> bool:
    '''
    Is the file a pickle file?
    Input : path to the file
    Output : True, False (if the file is a fatsa file for example) or raise an error if file doesn't exist
    '''
    try:
        with open(filename, 'rb') as f:
            genome = pk.load(f)
            return any(genome)
    except EOFError as eoferr:
        return False
    except UnicodeDecodeError as uderr:
        return False
    except FileNotFoundError as fnferr:
        print(f'\nFile {filename} doesn\'t exist: {fnferr}\n')
        raise FileNotFoundError(f'\nFile {filename} doesn\'t exists\n')
    except pk.UnpicklingError:
        return False
    except:
        print(f'\nUnexpected error: {sys.exc_info()[0]}\n')
        raise


def get_genome_dict(dict_parameters, genome_file, logger):
    '''
    Check the genome file and (1) read the fasta file and create the pickle dictionary or (2) load the pickle dictionary and return this dictionary
    Input : path to the genome file and logger
    Output : dictionnary containing the genome
    '''
    if dict_parameters['verbose_prints'].upper() == 'TRUE':
        print(f'Check input genome file {genome_file}...')
    logger.info(f'Check input genome file {genome_file}')
    if is_fasta(genome_file):
        genome = {}
        path = Path(genome_file).with_suffix('.pk')
        fa = pyfastx.Fastx(genome_file)
        print("chr", "length")
        for content in fa:
            name = content[0]
            seq = content[1]
            genome[name] = seq
            print(name, len(seq))

        if not path.exists():
            print(f'Create the pickle genome dictionary in {path}')
            logger.info(f'Create the pickle genome dictionary in {path}')
            with open(path, 'wb') as out_pickle:
                pk.dump(genome, out_pickle)
        else:
            print(f'File {path} already exists !')
            logger.warning(f'File {path} already exists !')
    elif is_pickle(genome_file):
        if dict_parameters['verbose_prints'].upper() == 'TRUE':
            print('Charge the pickle genome dictionary...')
        with open(genome_file, 'rb') as f:
            genome = pk.load(f)
    else:
        raise ValueError(f'{genome_file} is not a fasta or a pickle file !')

    return genome


def add_snp(snp_count, ref, alt, triplet):
    '''
    Add 1 to the SNP in counter
    As in VCF the variant is be given for the leading (forward, 5'-3') strand, if we get G>* or A>* the reference, the variant and the triplet are reversed complement
    Input : SNP counter, the reference allele, the variant and the triplet containing the reference
    Output : SNP counter
    '''
    ref = ref.upper()
    alt = alt.upper()
    triplet = triplet.upper()
    if TRANSLATE[str(ref + ">" + alt)] == str(ref + ">" + alt):
        snp_count[str(ref + ">" + alt)][triplet] += 1
    else:
        snp_count[TRANSLATE[str(ref + ">" + alt)]][triplet.translate(TAB)[::-1]] += 1
    return snp_count


def size_indel(ref, alt):
    '''
    Get the indel size
    Input : reference allele and the alternative variant
    Output : size of the indel variant
    '''
    size = 0
    if len(ref) > len(alt):
        size = len(ref) - len(alt)
    elif len(alt) > len(ref):
        size = len(alt) - len(ref)
    else:
        print('Not an indel !')
        raise ValueError('Not an indel !')
    return size


def count_variants(dict_parameters, dict_colors, vcf_file_filter, vcf_file_pass, genome, logger):
    '''
    Read the vcf file(s) and create the different counters to summary the files
    Input : path to the vcf file from filter module (optional), path to the vcf file containing only pass variants from filter module, genome in a dictionnary and the output path file and the logger
    Output : counters of (1) the deletion size, (2) the insertion size, (3) the snp (in tripet), the impacted genes list and a dictiionnary contaning general stats from the vcf
    '''
    if dict_parameters['verbose_prints'] == 'TRUE':
        print('Counting variants and get genes...')
    logger.info(f'Compute stats on variants and genes')

    ##################
    # Counter/list creation
    stats = {}
    stats['Total'] = 0
    stats['germline'] = 0
    stats['PON'] = 0
    stats['non functional'] = 0
    stats['germline+PON'] = 0
    stats['germline+non functional'] = 0
    stats['PON+non functional'] = 0
    stats['germline+PON+non functional'] = 0
    stats['PASS'] = 0
    stats['SNP'] = [0, set()]  # Single Nucleotide Polymorphism
    stats['DNP'] = [0, set()]  # Double Nucleotide Polymorphism
    stats['TNP'] = [0, set()]  # Triple Nucleotide Polymorphism
    stats['ONP'] = [0, set()]  # Quadruple or more Nucleotide Polymorphism
    stats['INSERTION'] = [0, set()]
    stats['DELETION'] = [0, set()]
    genes_list = {}
    gene_info = [0, [0, 0, 0, 0, 0, 0], [], [], '', [], [], [], [],
                 {}]  # [gene, [variants number (tumor burden),[snp, dnp, tnp, onp, insertion, deletion]], chromosome, reference, alternative, position]
    idx = {}
    snp_count = create_snp_dict()
    counter_deletion_size = Counter()
    counter_insertion_size = Counter()
    annotation = dict_parameters['vcf_annotation_method'].upper()

    f = open(vcf_file_pass, 'r', encoding='latin1')
    nb_lines_pass = len(
        list(dropwhile(lambda x: x[0] == '#', (line for line in f))))  # passed.vcf content (number of lines / variants)

    ####################
    # Read pass vcf file
    if dict_parameters['verbose_prints'] == 'TRUE':
        print(f'Read {vcf_file_pass}...')
    logger.info(f'Read {vcf_file_pass}')
    color = dict_colors['yellow3']
    if dict_parameters['colored_execution'].upper() == 'FALSE' or dict_parameters[
        'colored_execution'].upper() == 'NONE':
        color = ''
    with tqdm(total=nb_lines_pass, bar_format='{l_bar}{bar:15}{r_bar}', ncols=130, smoothing=1) as pbar:
        pbar.set_description(color + f' -> parsing passed.vcf file:')  # progress bar
        vcf_reader_pass = read_vcf(vcf_file_pass)  # create the read file generator
        for line in vcf_reader_pass:
            if type(line) == type({}):  # get index position
                idx = line
            else:
                # Get values
                chr = line[idx['idx_chr']]
                ref = line[idx['idx_ref']]
                alts = line[idx['idx_alts']].split(',')
                nb_variants = len(alts)
                pos = int(line[idx['idx_pos']])
                match_DP = re.search(r'DP=(\d+)', str(line))
                DP_value = match_DP.group(1)

                triplet = str(genome[chr][
                              pos - 2:pos + 1])  # Get triplet by using the genome and variant position (better definition of SNP with its environment
                infos = [tuple(infos.split('=')) if len(infos.split('=')) > 1 else (infos, '') for infos in
                         line[idx['idx_info']].split(';')]

                if annotation == 'FUNCOTATOR':  # si le variant est annotée par Funcotator
                    idx_funcotation = list(zip(*infos))[0].index('FUNCOTATION')
                    gene = list(zip(*infos))[1][idx_funcotation].split('|')[0].lstrip('[')
                    l_format = line[-2].split(':')
                    l_format_values = line[-1].split(':')
                    format_dict = dict(zip(l_format, l_format_values))
                    # AF_value = format_dict['AF']

                elif annotation == 'ANNOVAR':
                    # match_AF = re.search(r'AF=(\d+)', str(line))
                    # AF_value = match_AF.group(1)
                    str_infos = str(infos)
                    pattern = r"'Gene.refGene', '(.*?)'"
                    match = re.search(pattern, str_infos)
                    if match:  # enlever - et après???
                        gene = match.group(1)
                        if ')' in gene:
                            gene = gene.split(')')[0]
                        if "'" in gene:
                            gene = gene.replace("'", "")
                        # if gene.startswith('LOC'):
                        #     gene = 'error'
                        if '\\' in gene:
                            gene = gene.split("\\")[0]
                    else:
                        print("error : no gene name found")
                    # print(gene)
                    # if len(gene) > 10:
                    # 	print('a')

                if gene != 'NONE' and not gene.startswith('LINC'): # if ANNOVAR doesn't find any gene for this variant or if lncRNA
                    if not gene in genes_list.keys():
                        if len(gene) > 15:
                            print('gene name length > 15')
                        genes_list[gene] = deepcopy(gene_info)
                    if genes_list[gene][4] == '':
                        genes_list[gene][4] = chr
                    genes_list[gene][2].append(ref)
                    genes_list[gene][3].append(alts)
                    genes_list[gene][5].append(pos)
                    qual = line[5]
                    genes_list[gene][6].append(DP_value)
                    genes_list[gene][7].append(qual)
                    # genes_list[gene][8].append(AF_value)
                    # exonic function / mutation subtype
                    match = re.search(r'ExonicFunc\.refGene=(.*?);', str(line))
                    if match:
                        mutation_subtype = match.group(1)
                        if not mutation_subtype in genes_list[gene][9].keys():
                            genes_list[gene][9][mutation_subtype] = 0
                        genes_list[gene][9][mutation_subtype] += 1

                    for i, alt in enumerate(alts):
                        stats['PASS'] += 1
                        genes_list[gene][0] += 1

                        if is_snp(len(ref), len(alt)):
                            stats['SNP'][0] += 1
                            snp_count = add_snp(snp_count, ref, alt, triplet)
                            genes_list[gene][1][0] += 1
                            stats['SNP'][1].add(gene)

                        elif is_dnp(len(ref), len(alt)):
                            stats['DNP'][0] += 1
                            genes_list[gene][1][1] += 1
                            stats['DNP'][1].add(gene)

                        elif is_tnp(len(ref), len(alt)):
                            stats['TNP'][0] += 1
                            genes_list[gene][1][2] += 1
                            stats['TNP'][1].add(gene)

                        elif is_onp(len(ref), len(alt)):
                            stats['ONP'][0] += 1
                            genes_list[gene][1][3] += 1
                            stats['ONP'][1].add(gene)

                        elif is_insertion(len(ref), len(alt)):
                            counter_insertion_size.update([size_indel(ref, alt)])
                            stats['INSERTION'][0] += 1
                            genes_list[gene][1][4] += 1
                            stats['INSERTION'][1].add(gene)

                        elif is_deletion(len(ref), len(alt)):
                            counter_deletion_size.update([size_indel(ref, alt)])
                            stats['DELETION'][0] += 1
                            genes_list[gene][1][5] += 1
                            stats['DELETION'][1].add(gene)
                pbar.update(1)

    #####################################################
    # If whole vcf (not just pass variants) file is given
    if vcf_file_filter:
        f = open(vcf_file_filter, 'r', encoding='latin1')
        nb_lines = len(list(dropwhile(lambda x: x[0] == '#',
                                      (line for line in f))))  # filtered.vcf content (number of lines / variants)
    f.close()
    if vcf_file_filter:
        if dict_parameters['verbose_prints'] == 'TRUE':
            print(f'Read {vcf_file_filter}...')
        logger.info(f'Read {vcf_file_filter}')
        with tqdm(total=nb_lines, bar_format='{l_bar}{bar:30}{r_bar}', ncols=150, smoothing=1) as pbar:
            pbar.set_description(color + f' -> getting supplementary information parsing filtered.vcf')  # progress bar
            vcf_reader_filter = read_vcf(vcf_file_filter)  # create the read file generator
            for line in vcf_reader_filter:
                if type(line) != type({}):
                    non_functional = True
                    alts = line[idx['idx_alts']].split(',')  # get variants section
                    nb_variants = len(
                        alts)  # get the number of variants (if > 1 -> alternative variant = nucleotide that can transform into several ones)
                    filters = line[idx['idx_filt']].split(',')[0].split(';')  # filter section of vcf_file
                    if ''.join(filters) != 'PASS':
                        for i, alt in enumerate(alts):
                            stats['Total'] += 1
                            if 'LOTUS_filter' in filters:
                                # Test if NOT_FUNCTIONAL is present in filter for the variant
                                # for information in line[idx['idx_info']].split(';'):
                                #	if information.split('=')[0] == 'OTHER_FILTER':
                                #		nf = information.split('=')[1].split(',')[-nb_variants:][i]
                                #		if 'NOT_FUNCTIONAL' in nf:
                                #			non_functional = True
                                # [True for nf in [information.split('=')[1].split(',')[-nb_variants:] for information in line[idx['idx_info']].split(';') if information.split('=')[0] == 'OTHER_FILTER'] if 'NOT_FUNCTIONAL' in nf]
                                non_functional = False
                                for information in line[idx['idx_info']].split(';'):
                                    if information.split('=')[0] == 'OTHER_FILTER':
                                        nf_list = information.split('=')[1].split(',')[-nb_variants:]
                                        for nf in nf_list:
                                            if 'NOT_FUNCTIONAL' in nf:
                                                non_functional = True
                                                break
                                    if non_functional:
                                        break

                            else:
                                non_functional = False

                            if 'germline' in filters and 'panel_of_normals' in filters and non_functional:
                                stats['germline+PON+non functional'] += 1
                            elif 'germline' in filters and 'panel_of_normals' in filters:
                                stats['germline+PON'] += 1
                            elif 'germline' in filters and non_functional:
                                stats['germline+non functional'] += 1
                            elif 'panel_of_normals' in filters and non_functional:
                                stats['PON+non functional'] += 1
                            elif 'germline' in filters:
                                stats['germline'] += 1
                            elif 'panel_of_normals' in filters:
                                stats['PON'] += 1
                            elif non_functional:
                                stats['non functional'] += 1
                    else:
                        for i, alt in enumerate(alts):
                            stats['Total'] += 1
                    pbar.update(1)

    # Get percentage from count for pass SNP
    snp_count_pct = deepcopy(snp_count)
    tot = sum([sum(val.values()) for key, val in snp_count.items()])
    if tot != 0:
        for key, val in snp_count.items():
            for k, v in val.items():
                snp_count_pct[key][k] = round((v / tot), 8)  # get the percentage of each SNP in the total number of SNP
    else:
        print('NO SNP IN SAMPLE ' + vcf_file_pass + ' !')
        for key, val in snp_count.items():
            for k, v in val.items():
                snp_count_pct[key][k] = 0

    return counter_deletion_size, counter_insertion_size, snp_count_pct, snp_count, genes_list, stats


def is_snp(ref_length, alt_length):
    '''ex: A > T '''
    if ref_length == 1:
        return ref_length == alt_length
    else:
        return False


def is_dnp(ref_length, alt_length):
    '''ex: AT > GA '''
    if ref_length == 2:
        return ref_length == alt_length
    else:
        return False


def is_tnp(ref_length, alt_length):
    '''ex: AAG > TGC '''
    if ref_length == 3:
        return ref_length == alt_length
    else:
        return False


def is_onp(ref_length, alt_length):
    '''ex: AATGG > GTCAA '''
    if ref_length > 3:
        return ref_length == alt_length
    else:
        return False


def is_deletion(ref_length, alt_length):
    '''ex: TAAG > T '''
    return ref_length > alt_length


def is_insertion(ref_length, alt_length):
    '''ex: T > TAAG '''
    return ref_length < alt_length


def write_stats(dict_subtypes, dict_parameters, vcf_file_filter: str, vcf_file_pass: str, out_stats: str,
                stats: Counter, logger):
    '''
    Writes a simple statistics file on the variants of the VCF file
    Input : path to the vcf file from filter module (optional), path to the vcf file containing only pass variants from filter module, counter containing vcf stats and the output path file
    Output : Write the vcf stats file
    '''
    if "TXT" in dict_parameters['Stats_file_format'].upper():
        out_stats = out_stats + ".txt"
    if dict_parameters['verbose_prints'].upper() == 'TRUE':
        print('Write stats file...')
    logger.info('Write stats file')
    with open(out_stats, 'w') as o:
        if vcf_file_filter and dict_parameters['vcf_annotation_method'].upper() == 'FUNCOTATOR':
            o.write(f'########################### {vcf_file_filter} ###########################\n')
            o.write(f'\n###########################\nTotal variants : {stats["Total"]}\n###########################\n')
            o.write(
                f'germline: {stats["germline"]}\t|\tPON: {stats["PON"]}\t|\tnot functional: {stats["non functional"]}\n')
            o.write(
                f'germline & PON: {stats["germline+PON"]}\t|\tgermline &  not functional: {stats["germline+non functional"]}\t|\tPON &  not functional: {stats["PON+non functional"]}\n')
            o.write(f'germline & PON & not functional: {stats["germline+PON+non functional"]}\n\n')
        elif dict_parameters['vcf_annotation_method'].upper() == 'ANNOVAR':
            mutation_counts = {'synonymous_SNV', 'nonsynonymous_SNV', 'stopgain', 'startloss', 'stoploss', 'nonframeshift_insertion',
                               'frameshift_insertion', 'nonframeshift_deletion', 'frameshift_deletion',
                               'nonframeshift_substitution', 'frameshift_substitution'}
            mutation_counts = {mutation_type: 0 for mutation_type in mutation_counts}
            for values in dict_subtypes.values():
                mutations = values[-1]  # Obtient le dictionnaire des mutations pour chaque clé
                for mutation_type, count in mutations.items():
                    mutation_counts[mutation_type] = mutation_counts.get(mutation_type, 0) + count
        o.write(f'########################### {vcf_file_pass} ###########################\n')
        o.write(f'number of variants: {stats["PASS"]}\n###########################\n')
        o.write(f'\n----------------\nMUTATION TYPES\n----------------\n')
        o.write(f'SNP: {stats["SNP"][0]}\t\tDNP: {stats["DNP"][0]}\tTNP: {stats["TNP"][0]}\tONP: {stats["ONP"][0]}\n')
        o.write(
            f'INDEL: {stats["INSERTION"][0] + stats["DELETION"][0]}\tINSERTION: {stats["INSERTION"][0]}, DELETION: {stats["DELETION"][0]}\n')
        if dict_parameters['vcf_annotation_method'].upper() == 'ANNOVAR':
            o.write(f'\n----------------\nMUTATION SUBTYPES\n----------------\n')
            o.write(
                f'synonymous SNV: {mutation_counts["synonymous_SNV"]}\tnonsynonymous SNV: {mutation_counts["nonsynonymous_SNV"]}\n')
            o.write(
                f'frameshift substitution: {mutation_counts["frameshift_substitution"]}\tnon frameshift substitution: {mutation_counts["nonframeshift_substitution"]}\n')
            o.write(
                f'stopgain: {mutation_counts["stopgain"]}\tstoploss: {mutation_counts["stoploss"]}\tstartloss: {mutation_counts["startloss"]}\n')
            o.write(
                f'frameshift insertion: {mutation_counts["frameshift_insertion"]}     non frameshift insertion: {mutation_counts["nonframeshift_insertion"]}\n')
            o.write(
                f'frameshift deletion: {mutation_counts["frameshift_deletion"]}     non frameshift deletion: {mutation_counts["nonframeshift_deletion"]}\n')
            o.write(f'unknown: {mutation_counts["."]}\n')
        # o.write(f'---\nUnique impacted genes GS: { len(stats["SNP"][1]) + len(stats["DNP"][1]) + len(stats["TNP"][1]) + len(stats["ONP"][1]) + len(stats["DELETION"][1]) + len(stats["INSERTION"][1]) } (list in {out_genes})\n')
        concatenated = [stats[key][1] for key in ["SNP", "DNP", "TNP", "ONP", "DELETION",
                                                  "INSERTION"]]  # get the list of impacted genes for each category
        concatenated_list = [value for sublist in concatenated for value in sublist]  # flatten the list
        unique_genes = set(concatenated_list)  # get the unique genes
        l_duplicate_genes = [gene for gene, count in Counter(concatenated_list).items() if count > 1]
        duplicate_genes = (', '.join(l_duplicate_genes))
        o.write(f'\n----------------\nIMPACTED GENES\n----------------\n')
        o.write(
            f'-> number of impacted genes: {len(unique_genes)}\n')  # write the number of unique genes (not in two or more categories of mutations)
        o.write(
            f'{len(l_duplicate_genes)} of those genes are impacted by 2 types of mutations (or more): \n{duplicate_genes}\n')
        o.write(f'---\n-> number of impacted genes for each mutation category: \n')
        o.write(
            f'SNP: {len(stats["SNP"][1])}\t\tDNP: {len(stats["DNP"][1])}\tTNP: {len(stats["TNP"][1])}\tONP: {len(stats["ONP"][1])}\n')
        o.write(
            f'INDEL: {len(stats["INSERTION"][1]) + len(stats["DELETION"][1])} \tINSERTION: {len(stats["INSERTION"][1])}\tDELETION: {len(stats["DELETION"][1])}\n')


def create_mutation_types_tables(dict_para):
    TMB, SNP, DNP, TNP, ONP, INDEL, INS, DEL  = [], [], [], [], [], [], [], []
    dict_info = {'Number of variants': TMB, 'SNP': SNP, 'DNP': DNP, 'TNP': TNP, 'ONP': ONP, 'INDEL': INDEL, 'INSERTION': INS,
                 'DELETION': DEL}
    samples_path = os.path.join(dict_para['output_path'], 'samples/')
    for folder in os.listdir(samples_path):
        for file in os.listdir(samples_path + folder + '/'):
            if file == 'stats.txt':
                with (open(samples_path + folder + '/' + file, 'r') as file):
                    for line in file:  # avant d'arriver à 'Impacted genes'
                        line = line.strip().replace('\t', ' ')
                        if 'impacted' in line:
                            break
                        elif not line.startswith("#") and not line.startswith("---"):
                            if 'SNP: ' in line:
                                match = re.search(r'SNP: (\d+)', line)
                                value = match.group(1) if match else None
                                SNP.append(value)
                            if 'DNP: ' in line:
                                match = re.search(r'DNP: (\d+)', line)
                                value = match.group(1) if match else None
                                DNP.append(value)
                            if 'TNP: ' in line:
                                match = re.search(r'TNP: (\d+)', line)
                                value = match.group(1) if match else None
                                TNP.append(value)
                            if 'ONP: ' in line:
                                match = re.search(r'ONP: (\d+)', line)
                                value = match.group(1) if match else None
                                ONP.append(value)
                            if 'INSERTION: ' in line:
                                match = re.search(r'INSERTION: (\d+)', line)
                                value = match.group(1) if match else None
                                INS.append(value)
                            if 'DELETION: ' in line:
                                match = re.search(r'DELETION: (\d+)', line)
                                value = match.group(1) if match else None
                                DEL.append(value)
                            if 'INDEL: ' in line:
                                match = re.search(r'INDEL: (\d+)', line)
                                value = match.group(1) if match else None
                                INDEL.append(value)
                            if 'variants: ' in line:
                                match = re.search(r'variants: (\d+)', line)
                                value = match.group(1) if match else None
                                TMB.append(value)
    samples = os.listdir(samples_path)
    df = pd.DataFrame.from_dict(dict_info)
    df = df.set_index(pd.Index(samples))

    if not 'FALSE' in dict_para['mutations_types_table_format'].upper() and not 'NONE' in dict_para[
        'mutations_types_table_format'].upper():
        if 'TSV' in dict_para['mutations_types_table_format'].upper():
            table_path = dict_para['output_path'] + dict_para['mutations_types_table_name'] + '.tsv'
            df.to_csv(table_path, sep='\t')
        if 'CSV' in dict_para['mutations_types_table_format'].upper():
            table_path = dict_para['output_path'] + dict_para['mutations_types_table_name'] + '.csv'
            df.to_csv(table_path, sep=',')
        if 'XLSX' in dict_para['mutations_types_table_format'].upper():
            table_path = dict_para['output_path'] + dict_para['mutations_types_table_name'] + '.xlsx'
            df.to_excel(table_path)

            workbook = openpyxl.load_workbook(table_path)
            sheet = workbook.active

            for row in sheet.iter_rows():
                for cell in row:
                    cell.alignment = openpyxl.styles.Alignment(horizontal='center')
                    cell.border = openpyxl.styles.Border(
                        left=openpyxl.styles.Side(style='thin'),
                        right=openpyxl.styles.Side(style='thin'),
                        top=openpyxl.styles.Side(style='thin'),
                        bottom=openpyxl.styles.Side(style='thin')
                    )
                    if cell.column == 1 or cell.row == 1:
                        if cell.column == 1 and cell.row == 1:
                            cell.fill = openpyxl.styles.PatternFill(fill_type="none")
                        else:
                            cell.fill = openpyxl.styles.PatternFill(start_color="FFFFCC", end_color="FFFFCC",
                                                                    fill_type="solid")

            for column_cells in sheet.columns:
                max_length = 0
                column = column_cells[0].column_letter
                for cell in column_cells:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                sheet.column_dimensions[column].width = adjusted_width

            workbook.save(table_path)

    if not 'FALSE' in dict_para['mutations_subtypes_table_name'].upper() and not 'NONE' in dict_para[
        'mutations_subtypes_table_name'].upper():

        s_SNV, ns_SNV, f_sub, nf_sub, stopgain, stoploss, startloss, f_ins, nf_ins, f_del, nf_del, unknown = [], [], [], [], [], [], [], [], [], [], [], []
        dic_subtypes = {'synonymous SNV': s_SNV, 'nonsynonymous SNV': ns_SNV, 'frameshift substitution': f_sub,
                            'non frameshift substitution': nf_sub, 'stopgain': stopgain, 'stoploss': stoploss,
                            'startloss': startloss, 'frameshift insertion': f_ins,
                            'non frameshift insertion': nf_ins, 'frameshift deletion': f_del,
                            'non frameshift deletion': nf_del, 'unknown': unknown}
        for folder in os.listdir(samples_path):
            for file in os.listdir(samples_path + folder + '/'):
                if file == 'stats.txt':
                    with (open(samples_path + folder + '/' + file, 'r') as file):
                        for line in file:
                            line = line.strip().replace('\t', ' ')
                            if 'impacted' in line:
                                break
                            elif not line.startswith("#") and not line.startswith("---"):
                                if 'synonymous SNV' in line:
                                    match = re.search(r'synonymous SNV: (\d+)', line)
                                    value = match.group(1) if match else None
                                    s_SNV.append(value)
                                if 'nonsynonymous SNV' in line:
                                    match = re.search(r'nonsynonymous SNV: (\d+)', line)
                                    value = match.group(1) if match else None
                                    ns_SNV.append(value)
                                if 'frameshift substitution' in line:
                                    match = re.search(r'frameshift substitution: (\d+)', line)
                                    value = match.group(1) if match else None
                                    f_sub.append(value)
                                if 'non frameshift substitution' in line:
                                    match = re.search(r'non frameshift substitution: (\d+)', line)
                                    value = match.group(1) if match else None
                                    nf_sub.append(value)
                                if 'stopgain' in line:
                                    match = re.search(r'stopgain: (\d+)', line)
                                    value = match.group(1) if match else None
                                    stopgain.append(value)
                                if 'stoploss' in line:
                                    match = re.search(r'stoploss: (\d+)', line)
                                    value = match.group(1) if match else None
                                    stoploss.append(value)
                                if 'startloss' in line:
                                    match = re.search(r'startloss: (\d+)', line)
                                    value = match.group(1) if match else None
                                    startloss.append(value)
                                if 'frameshift insertion' in line:
                                    match = re.search(r'frameshift insertion: (\d+)', line)
                                    value = match.group(1) if match else None
                                    f_ins.append(value)
                                if 'non frameshift insertion' in line:
                                    match = re.search(r'non frameshift insertion: (\d+)', line)
                                    value = match.group(1) if match else None
                                    nf_ins.append(value)
                                if 'frameshift deletion' in line:
                                    match = re.search(r'frameshift deletion: (\d+)', line)
                                    value = match.group(1) if match else None
                                    f_del.append(value)
                                if 'non frameshift deletion' in line:
                                    match = re.search(r'non frameshift deletion: (\d+)', line)
                                    value = match.group(1) if match else None
                                    nf_del.append(value)
                                if 'unknown' in line:
                                    match = re.search(r'unknown: (\d+)', line)
                                    value = match.group(1) if match else None
                                    unknown.append(value)
        samples = os.listdir(samples_path)
        df = pd.DataFrame.from_dict(dic_subtypes)
        df = df.set_index(pd.Index(samples))

    if not 'FALSE' in dict_para['mutations_subtypes_table_format'].upper() and not 'NONE' in dict_para[
        'mutations_types_table_format'].upper():
        if 'TSV' in dict_para['mutations_subtypes_table_format'].upper():
            table_path = dict_para['output_path'] + dict_para['mutations_subtypes_table_name'] + '.tsv'
            df.to_csv(table_path, sep='\t')
        if 'CSV' in dict_para['mutations_subtypes_table_format'].upper():
            table_path = dict_para['output_path'] + dict_para['mutations_subtypes_table_name'] + '.csv'
            df.to_csv(table_path, sep=',')
        if 'XLSX' in dict_para['mutations_subtypes_table_format'].upper():
            table_path = dict_para['output_path'] + dict_para['mutations_subtypes_table_name'] + '.xlsx'
            df.to_excel(table_path)


def write_impacted_genes(dict_para, out_genes: str, genes_list: pd.DataFrame, logger):
    '''
    Writes an xlsx file containing the list of genes impacted by the variants from the VCF file
    Input : name use for the output files (.xlsx), the genes list and the logger
    Output : Write the genes list in two different format
    '''
    if dict_para['verbose_prints'].upper() == 'TRUE':
        print(f'Write genes list file... {format})')
    logger.info(f'Write genes list file... {format})')
    for path in out_genes:
        if path.endswith('.xlsx'):
            genes_list.to_excel(Path(path).with_suffix('.xlsx'))
        elif path.endswith('.tsv'):
            genes_list.to_csv(Path(path).with_suffix('.tsv'), sep='\t')
        elif path.endswith('.csv'):
            genes_list.to_csv(Path(path).with_suffix('.csv'), sep=',')

def create_mutations_types_barplot(dict_para):
    dict_info = {'SNP': 0, 'DNP': 0, 'TNP': 0, 'ONP': 0, 'INDEL': 0, 'INSERTION': 0, 'DELETION': 0}
    sample_path = dict_para['output_path_sample'].split('.vcf')[0] + '/stats.txt'
    with (open(sample_path, 'r') as file):
        for line in file:
            line = line.strip().replace('\t', ' ')
            if 'impacted' in line:
                break
            elif not line.startswith("#") and not line.startswith("---"):
                if 'SNP: ' in line:
                    match = re.search(r'SNP: (\d+)', line)
                    value = match.group(1) if match else None
                    dict_info['SNP'] = value
                if 'DNP: ' in line:
                    match = re.search(r'DNP: (\d+)', line)
                    value = match.group(1) if match else None
                    dict_info['DNP'] = value
                if 'TNP: ' in line:
                    match = re.search(r'TNP: (\d+)', line)
                    value = match.group(1) if match else None
                    dict_info['TNP'] = value
                if 'ONP: ' in line:
                    match = re.search(r'ONP: (\d+)', line)
                    value = match.group(1) if match else None
                    dict_info['ONP'] = value
                if 'INSERTION: ' in line:
                    match = re.search(r'INSERTION: (\d+)', line)
                    value = match.group(1) if match else None
                    dict_info['INSERTION'] = value
                if 'DELETION: ' in line:
                    match = re.search(r'DELETION: (\d+)', line)
                    value = match.group(1) if match else None
                    dict_info['DELETION'] = value
                if 'INDEL: ' in line:
                    match = re.search(r'INDEL: (\d+)', line)
                    value = match.group(1) if match else None
                    dict_info['INDEL'] = value

    categories = list(dict_info.keys())
    values = [0 if value is None else int(value) for value in dict_info.values()]
    categories_filtered = [cat for cat, val in zip(categories, values) if val != 0]
    colors = {'SNP': '#77c3ec',
              'DNP': '#89cff0',
              'TNP': '#9dd9f3',
              'ONP': '#b8e2f2',
              'INDEL': '#95b89b',
              'INSERTION': '#aec9aa',
              'DELETION': '#bed8c0'}

    values_filtered = [val for val in values if val != 0]
    plt.bar(categories_filtered, values_filtered, color=[colors[cat] for cat in categories_filtered], width=0.4)
    sample_name = sample_path.split('/sta')[0].split('samples/')[1]
    plt.title("Mutation counts, for each mutation type found in " + sample_name, pad=10)
    plt.ylabel("Count", labelpad=10)

    for i, val in enumerate(values_filtered):
        plt.text(i, val, str(val), ha='center', va='bottom', fontsize=10.5)

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        plt.tight_layout()

    if 'PNG' in dict_para['S_types_barplot_format(s)'].upper():
        plt.savefig(sample_path.split('stats')[0] + dict_para['S_types_barplot_name'] + '.png', dpi=400)
    if 'SVG' in dict_para['S_types_barplot_format(s)'].upper():
        plt.savefig(sample_path.split('stats')[0] + dict_para['S_types_barplot_name'] + '.svg', dpi=400)
    if 'PDF' in dict_para['S_types_barplot_format(s)'].upper():
        plt.savefig(sample_path.split('stats')[0] + dict_para['S_types_barplot_name'] + '.pdf', dpi=400)
    if 'JPG' in dict_para['S_types_barplot_format(s)'].upper():
        plt.savefig(sample_path.split('stats')[0] + dict_para['S_types_barplot_name'] + '.jpg', dpi=400)

def create_mutations_subtypes_barplot(dict_para):
    dic_subtypes = {'synonymous SNV': 0, 'nonsynonymous SNV': 0, 'frameshift substitution': 0,
                    'non frameshift substitution': 0, 'stopgain': 0, 'stoploss': 0,
                    'startloss': 0, 'frameshift insertion': 0,
                    'non frameshift insertion': 0, 'frameshift deletion': 0,
                    'non frameshift deletion': 0, 'unknown': 0}
    sample_path = dict_para['output_path_sample'].split('.vcf')[0] + '/stats.txt'
    with open(sample_path, 'r') as file:
        for line in file:
            line = line.strip().replace('\t', ' ')
            if 'impacted' in line:
                break
            elif not line.startswith("#") and not line.startswith("---"):
                if 'synonymous SNV' in line:
                    match = re.search(r'synonymous SNV: (\d+)', line)
                    value = match.group(1) if match else None
                    dic_subtypes['synonymous SNV'] = value
                if 'nonsynonymous SNV' in line:
                    match = re.search(r'nonsynonymous SNV: (\d+)', line)
                    value = match.group(1) if match else None
                    dic_subtypes['nonsynonymous SNV'] = value
                if 'frameshift substitution' in line:
                    match = re.search(r'frameshift substitution: (\d+)', line)
                    value = match.group(1) if match else None
                    dic_subtypes['frameshift substitution'] = value
                if 'non frameshift substitution' in line:
                    match = re.search(r'non frameshift substitution: (\d+)', line)
                    value = match.group(1) if match else None
                    dic_subtypes['non frameshift substitution'] = value
                if 'stopgain' in line:
                    match = re.search(r'stopgain: (\d+)', line)
                    value = match.group(1) if match else None
                    dic_subtypes['stopgain'] = value
                if 'stoploss' in line:
                    match = re.search(r'stoploss: (\d+)', line)
                    value = match.group(1) if match else None
                    dic_subtypes['stoploss'] = value
                if 'startloss' in line:
                    match = re.search(r'startloss: (\d+)', line)
                    value = match.group(1) if match else None
                    dic_subtypes['startloss'] = value
                if 'frameshift insertion' in line:
                    match = re.search(r'frameshift insertion: (\d+)', line)
                    value = match.group(1) if match else None
                    dic_subtypes['frameshift insertion'] = value
                if 'non frameshift insertion' in line:
                    match = re.search(r'non frameshift insertion: (\d+)', line)
                    value = match.group(1) if match else None
                    dic_subtypes['non frameshift insertion'] = value
                if 'frameshift deletion' in line:
                    match = re.search(r'frameshift deletion: (\d+)', line)
                    value = match.group(1) if match else None
                    dic_subtypes['frameshift deletion'] = value
                if 'non frameshift deletion' in line:
                    match = re.search(r'non frameshift deletion: (\d+)', line)
                    value = match.group(1) if match else None
                    dic_subtypes['non frameshift deletion'] = value
                if 'unknown' in line:
                    match = re.search(r'unknown: (\d+)', line)
                    value = match.group(1) if match else None
                    dic_subtypes['unknown'] = value

    categories = list(dic_subtypes.keys())
    values = [0 if value is None else int(value) for value in dic_subtypes.values()]
    categories_filtered = [cat for cat, val in zip(categories, values) if val != 0]
    colors = {'synonymous SNV': '#77c3ec',  # Blue
              'nonsynonymous SNV': '#89cff0',  # Light Blue
              'frameshift substitution': '#aec9aa',  # Green
              'non frameshift substitution': '#95b89b',  # Light Green
              'stopgain': '#e68a00',  # Dark Orange
              'stoploss': '#ff9933',  # Light Orange
              'startloss': '#ffcc66',  # Lighter Orange
              'frameshift insertion': '#ff6666',  # Red
              'non frameshift insertion': '#ff9999',  # Light Red
              'frameshift deletion': '#b366ff',  # Purple
              'non frameshift deletion': '#cc99ff',  # Light Purple
              'unknown': '#4d4d4d'}

    values_filtered = [val for val in values if val != 0]
    plt.clf()
    plt.figure(figsize=(12, 6))  # Adjust the figure size as needed
    plt.bar(categories_filtered, values_filtered, color=[colors[cat] for cat in categories_filtered], label='Mutation Subtypes', width=0.3)
    sample_name = sample_path.split('/sta')[0].split('samples/')[1]
    plt.title("Mutation counts for each mutation subtype found in " + sample_name, pad=10)
    plt.ylabel("Count", labelpad=10)

    for i, val in enumerate(values_filtered):
        plt.text(i, val, str(val), ha='center', va='bottom', fontsize=9)

    plt.xticks(rotation=90)
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        plt.tight_layout()

    if 'PNG' in dict_para['S_subtypes_barplot_format(s)'].upper():
        plt.savefig(sample_path.split('stats')[0] + dict_para['S_subtypes_barplot_name'] + '.png', dpi=400)
    if 'SVG' in dict_para['S_subtypes_barplot_format(s)'].upper():
        plt.savefig(sample_path.split('stats')[0] + dict_para['S_subtypes_barplot_name'] + '.svg', dpi=400)
    if 'PDF' in dict_para['S_subtypes_barplot_format(s)'].upper():
        plt.savefig(sample_path.split('stats')[0] + dict_para['S_subtypes_barplot_name'] + '.pdf', dpi=400)
    if 'JPG' in dict_para['S_subtypes_barplot_format(s)'].upper():
        plt.savefig(sample_path.split('stats')[0] + dict_para['S_subtypes_barplot_name'] + '.jpg', dpi=400)


def summary(last, dict_para, dict_colors, output_path, vcf_file_filter: str, vcf_file_pass: str, genome_file: str,
            out_stats: str, out_genes: str, SNP_profile: str, out_indel: str, logger, enrichment: bool):
    '''
    Summarizes the vcf file
    Input : path to the vcf file from filter module (optional), path to the vcf file containing only pass variants from filter module, path to genome (fasta or pickle),
    outputs path, logger and boolean for writing gene list enrichments analysis
    Output : stat summary file, file containing the impacted genes, mutation profile graph, indel size graph
    '''
    if dict_para['verbose_prints'].upper() == 'TRUE':
        print('Extracting information from reference genome and vcf file(s)...')
    genome = get_genome_dict(dict_para, genome_file, logger)

    ####################
    # Get stats and data
    counter_deletion_size, counter_insertion_size, snp_count_pct, snp_count, genes_list, stats = count_variants(
        dict_para, dict_colors, vcf_file_filter, vcf_file_pass, genome, logger)

    print('Creating stats file, mutations table and plots...')
    ##################
    # Write stats file
    if dict_para['Stats_file_name'].upper() != 'NONE' or dict_para['Stats_file_name'].upper() != 'FALSE' or dict_para['Stats_file_name'].upper() != 'NO':
        write_stats(genes_list, dict_para, vcf_file_filter, vcf_file_pass, out_stats, stats, logger)

    if dict_para['S_types_barplot_name'].upper() != 'NONE' or dict_para['S_types_barplot_name'].upper() != 'FALSE' or dict_para['S_types_barplot_name'].upper() != 'NO':
        create_mutations_types_barplot(dict_para)

    if dict_para['S_subtypes_barplot_name'].upper() != 'NONE' or dict_para['S_subtypes_barplot_name'].upper() != 'FALSE' or dict_para['S_subtypes_barplot_name'].upper() != 'NO':
        create_mutations_subtypes_barplot(dict_para)


    ##########################################
    # Write genes file in tsv and wlxs formats
    if dict_para['S_MutatedGenes'].upper() != 'NONE' or dict_para['S_MutatedGenes'].upper() != 'FALSE':
        genes_df = create_dataframe_from_gene(dict_para, genes_list)
        write_impacted_genes(dict_para, out_genes, genes_df, logger)

    ################################################
    # Biological process enrichment using the genes list with the ToppGene and Panther API
    discard = dict_para['discard_weak_variants'].upper()
    if enrichment.upper() == 'TRUE' and (discard == 'TRUE' or discard == 'YES'):
        print(f"\033[1m{len(genes_list.keys())}\033[0m{dict_colors['yellow3']} genes are concerned.")
        if len(genes_list.keys()) < 1500:
            print('Computing ToppGene GOEA analysis...')
            toppgene_name = dict_para['S_ToppGene_name']
            ToppGene_GOEA('summarise', dict_para, genes_list, toppgene_name, logger)
        else:
            print("The VCF is heavy, too many genes are concerned for ToppGene GO to be run.")
        if len(genes_list.keys()) < 2000:
            print('Computing Panther GOEA analysis...')
            panther_name = dict_para['S_Panther_name']
            Panther_GOEA('summarise', dict_para, genes_list, panther_name, logger)
        else:
            print("The VCF is heavy, too many genes are concerned for Panther GO to be run.")

    #################
    # Create snp mutation types plot and indel size plot
    if not "NONE" in dict_para['SNP_profile'].upper() and not "FALSE" in dict_para['SNP_profile'].upper():
        if "." in SNP_profile:
            SNP_profile = re.sub(r"\..*", "", SNP_profile)
        graph_snp(dict_para, snp_count_pct, snp_count, SNP_profile, vcf_file_pass, logger, output_path)

    if not "NONE" in dict_para['indel_profile'].upper() and not "FALSE" in dict_para['indel_profile'].upper():
        if str(counter_insertion_size) != 'Counter()':
            if "." in out_indel:
                out_indel = re.sub(r"\..*", "", out_indel)
            graph_indel(dict_para, counter_deletion_size, counter_insertion_size, out_indel, vcf_file_pass, logger)
        else:
            if dict_para['verbose_prints'].upper() == 'TRUE':
                print('No insertion in the vcf file !')
            else:
                pass

    if last:
        create_mutation_types_tables(dict_para)


def main(dict_para, dict_colors, args, vcf_filtered, vcf_passed, output_path, last):
    vcf_file_filter = vcf_filtered
    vcf_file_pass = vcf_passed
    working_directory = output_path
    genome_file = args['genome']
    if not "." in genome_file:
        genome_file = genome_file + '.pk'
    out_stats = output_path + args['Stats_file_name']
    if "." in out_stats:
        out_stats = re.sub(r"\..*", "", out_stats)
    format = dict_para['S_MutatedGenes']
    format = "".join(format.split())
    l_formats = format.split(',') if ',' in format else [format]
    table_name = dict_para['S_genes']
    if "." in table_name:
        table_name = re.sub(r"\..*", "", table_name)
    out_genes = [output_path + table_name + "." + format for format in l_formats]
    SNP_profile = output_path + args['profile']
    if "." in SNP_profile:
        SNP_profile = re.sub(r"\..*", "", SNP_profile)
    out_indel = output_path + args['indel']
    if "." in out_indel:
        out_indel = re.sub(r"\..*", "", out_indel)
    enrichment = args['enrichment']

    # Logger configuration
    logger = logging.getLogger('LOTUS summarise')
    logger.setLevel(logging.DEBUG)
    fh = logging.FileHandler(args['log'])
    fh.setLevel(logging.DEBUG)
    formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
    fh.setFormatter(formatter)
    logger.addHandler(fh)

    vcf_file_stats = os.stat(vcf_file_filter)
    vcf_file_pass_stats = os.stat(vcf_file_pass)

    try:
        logger.info('Verification of outputs file')
        verif_output(out_stats)
        verif_output(out_indel)
        logger.info('- Outputs file ok -')
    except ValueError:
        if dict_para['verbose_prints'].upper() == 'TRUE':
            print('Problem with one or more output files: ', sys.exc_info()[0])
        logger.error('- Problem with output files -')
        exit(1)

    # Start

    logger.info(
        '**************************************************************************************************************')
    logger.info('*** LOTUS summarise module ***')
    no_argument = ''
    if enrichment.upper() == 'TRUE':
        no_argument += ' --enrichment'
    logger.info(
        f'** cmd line : python lotus.py summarise -v {str(vcf_file_filter)} -vp {str(vcf_file_pass)} -g {str(genome_file)} -s {str(out_stats)} -genes {str(out_genes)} -p {str(SNP_profile)} -i {str(out_indel)}' + str(
            no_argument) + ' **')

    logger.info('* Start summarizing *')
    logger.info(f'Working directory (vcf files folder) : {working_directory}')
    logger.info(f'Current directory : {Path().absolute()}')

    summary(last, args, dict_colors, output_path, vcf_file_filter, vcf_file_pass, genome_file, out_stats, out_genes,
            SNP_profile, out_indel, logger, enrichment)

    logger.info('* End summarizing *')
    logger.info(f'File created :\t{out_stats}\t{SNP_profile}')
    logger.info(
        '**************************************************************************************************************')

# End
